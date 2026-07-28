from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import Workbook

from samchat.budgets.service import (
    collect_workbook_project_configs,
    _iter_budget_concept_account_mapping_rows,
    _iter_budget_concept_catalog_rows,
    _iter_budget_concept_catalog_phase_rows,
    _match_budget_concept_for_account_mapping,
    _map_catalog_phase_label_to_etapa,
    _merge_budget_concept_metadata,
    build_budget_artifact_snapshot,
    build_budget_executive_comparison,
    build_budget_scenario_player,
    budget_alias_candidates,
    budget_concept_matches_fase,
    budget_concept_scope_summary,
    build_budget_concept_scope_metadata,
    bulk_save_budget_concepts,
    create_budget_concept,
    ensure_missing_cuentas_contables_from_workbook,
    import_budget_lines_upload,
    list_budget_concepts,
    list_budget_lines,
    list_budget_tournament_commitments,
    normalize_budget_line_direction,
    resolve_budget_concept,
    sync_budget_concepts_from_partidas_workbook,
    sync_budget_projects_from_partidas_workbook,
    update_budget_concept,
)


def test_budget_alias_candidates_match_known_tournaments():
    aliases = budget_alias_candidates("Liga Telmex Telcel de Beisbol 2026", "beisbol")
    assert "LTTB" in aliases

    aliases = budget_alias_candidates("Copa Club América", "cca")
    assert "CCA" in aliases


def test_normalize_budget_line_direction_accepts_spanish_aliases():
    assert normalize_budget_line_direction("ingresos") == "income"
    assert normalize_budget_line_direction("gastos") == "expense"
    assert normalize_budget_line_direction("") == "expense"


@pytest.mark.asyncio
async def test_list_budget_concepts_filters_by_budget_direction(monkeypatch):
    class _Result:
        def mappings(self):
            return self

        def all(self):
            return []

    class _Session:
        def __init__(self):
            self.executed = []

        async def execute(self, statement, params=None):
            self.executed.append((str(statement), params or {}))
            return _Result()

    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    session = _Session()

    assert (
        await list_budget_concepts(
            session,
            tournament_id="tournament-1",
            budget_direction="income",
        )
        == []
    )

    statement, params = session.executed[0]
    assert "COALESCE(bc.budget_direction, 'expense') = :budget_direction" in statement
    assert params["budget_direction"] == "income"


@pytest.mark.asyncio
async def test_list_budget_lines_filters_by_line_direction(monkeypatch):
    class _Result:
        def mappings(self):
            return self

        def all(self):
            return []

    class _Session:
        def __init__(self):
            self.executed = []

        async def execute(self, statement, params=None):
            self.executed.append((str(statement), params or {}))
            return _Result()

    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    session = _Session()

    assert (
        await list_budget_lines(
            session,
            version_id="version-1",
            line_direction="ingresos",
        )
        == []
    )

    statement, params = session.executed[0]
    assert "COALESCE(l.line_direction, 'expense') = :line_direction" in statement
    assert params["line_direction"] == "income"


def test_iter_budget_concept_account_mapping_rows_reads_catalogo_final(tmp_path):
    workbook_path = tmp_path / "partidas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catálogo Final"
    sheet.append(
        [
            "Proyecto",
            "Subproyecto",
            "Partida",
            "Cuenta contable",
            "Nombre Cuenta contable",
        ]
    )
    sheet.append(
        [
            "Copa Telmex Telcel de Fútbol",
            "Fase Nacional",
            "Alimentos",
            "5300-010-016",
            "ALIMENTACION JUGADORES",
        ]
    )
    workbook.save(workbook_path)

    rows = _iter_budget_concept_account_mapping_rows(workbook_path)

    assert rows == [
        {
            "partida": "Alimentos",
            "concept_key": "alimentos",
            "proyecto": "Copa Telmex Telcel de Fútbol",
            "subproyecto": "Fase Nacional",
            "cuenta_contable_codigo": "5300-010-016",
            "cuenta_contable_nombre": "ALIMENTACION JUGADORES",
            "sheet_row_index": 2,
        }
    ]


def test_collect_workbook_project_configs_preserves_subproject_order(tmp_path):
    workbook_path = tmp_path / "partidas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catálogo Final"
    sheet.append(
        [
            "Proyecto",
            "Subproyecto",
            "Partida",
            "Cuenta contable",
            "Nombre Cuenta contable",
        ]
    )
    sheet.append(["Proyecto A", "Fase Nacional", "Alimentos", "5300-001", "Cuenta A"])
    sheet.append(["Proyecto A", "Fase Estatal", "Hospedaje", "5300-002", "Cuenta B"])
    sheet.append(["Proyecto A", "Fase Nacional", "Uniformes", "5300-003", "Cuenta C"])
    sheet.append(["Proyecto B", "General", "Papelería", "5300-004", "Cuenta D"])
    workbook.save(workbook_path)

    configs = collect_workbook_project_configs(workbook_path)

    assert configs == [
        {"proyecto": "Proyecto A", "etapas": ["Fase Nacional", "Fase Estatal"]},
        {"proyecto": "Proyecto B", "etapas": ["General"]},
    ]


@pytest.mark.asyncio
async def test_ensure_missing_cuentas_contables_updates_workbook_account_names(
    monkeypatch,
) -> None:
    execute_calls: list[tuple[str, dict[str, Any]]] = []

    class _ExistingRows:
        def mappings(self):
            return self

        def all(self):
            return [
                {"codigo": "5300-010-016", "nombre": "Nombre anterior", "activo": False}
            ]

    class _Session:
        async def execute(self, statement, params=None):
            sql = " ".join(str(statement).split())
            payload = dict(params or {})
            execute_calls.append((sql, payload))
            if "FROM cuentas_contables" in sql:
                return _ExistingRows()
            return MagicMock(rowcount=1)

        async def commit(self):
            execute_calls.append(("COMMIT", {}))

    monkeypatch.setattr(
        "samchat.budgets.service.collect_workbook_cuentas_contables",
        lambda _path: {"5300-010-016": "ALIMENTACION JUGADORES"},
    )

    report = await ensure_missing_cuentas_contables_from_workbook(
        _Session(),
        workbook_path="/tmp/partidas.xlsx",
    )

    update_calls = [
        params for sql, params in execute_calls if sql.startswith("UPDATE cuentas_contables")
    ]
    assert report["created_count"] == 0
    assert report["updated_count"] == 1
    assert update_calls == [
        {
            "codigo": "5300-010-016",
            "nombre": "ALIMENTACION JUGADORES",
            "activo": True,
        }
    ]


def test_build_budget_artifact_snapshot_uses_explicit_path_and_filters_tournament():
    rows = [
        {
            "torneo_codigo": "LTTB",
            "torneo": "Liga Telmex Telcel de Beisbol",
            "concepto": "Uniformes",
            "cuenta_contable_sugerida": "6100-01",
            "cuenta_contable_final": "6100-99",
            "presupuesto_2026": "1000",
            "importe_referencia_total": "800",
            "prioridad": "alta",
            "etapa": "estatal",
            "responsable": "Operaciones Norte",
        },
        {
            "torneo_codigo": "CCA",
            "torneo": "Copa Club America",
            "concepto": "Hospedaje",
            "cuenta_contable_sugerida": "6200-01",
            "cuenta_contable_final": "6200-01",
            "presupuesto_2026": "500",
            "importe_referencia_total": "450",
            "prioridad": "media",
            "etapa": "nacional",
        },
    ]

    snapshot = build_budget_artifact_snapshot(
        rows,
        artifact_path=Path("/tmp/presupuesto.csv"),
        tournament_name="Liga Telmex Telcel de Beisbol 2026",
        edition_year=2026,
    )

    assert snapshot["artifact_path"] == "/tmp/presupuesto.csv"
    assert snapshot["summary"]["tournaments_count"] == 1
    assert snapshot["summary"]["budget_total"] == 1000.0
    assert snapshot["tournaments"][0]["tournament_code"] == "LTTB"
    assert (
        snapshot["tournaments"][0]["top_concepts"][0]["cuenta_contable_final"]
        == "6100-99"
    )
    assert snapshot["breakdowns"]["by_concept"][0]["label"] == "Uniformes"
    assert snapshot["breakdowns"]["by_phase"][0]["label"] == "estatal"
    assert snapshot["breakdowns"]["by_owner"][0]["label"] == "Operaciones Norte"
    assert snapshot["breakdowns"]["by_account"][0]["label"] == "6100-99"
    assert snapshot["breakdowns"]["by_provider"] == []
    assert snapshot["scenarios"]["base"]["label"] == "Base"
    assert (
        snapshot["scenarios"]["optimistic"]["projected_close_total"]
        <= snapshot["scenarios"]["base"]["projected_close_total"]
    )
    assert (
        snapshot["scenarios"]["stressed"]["projected_close_total"]
        >= snapshot["scenarios"]["base"]["projected_close_total"]
    )


def test_build_budget_executive_comparison_exposes_core_finance_ladder():
    comparison = build_budget_executive_comparison(
        {
            "budget_total": 1000.0,
            "requested_total": 700.0,
            "committed_total": 650.0,
            "paid_total": 400.0,
            "actual_total": 420.0,
            "pending_to_pay_total": 250.0,
        },
        {
            "projected_close_total": 1100.0,
        },
    )

    labels = [item["label"] for item in comparison]
    assert labels == [
        "Presupuesto",
        "Solicitado",
        "Comprometido",
        "Pagado",
        "Real",
        "Pendiente por pagar",
        "Cierre proyectado",
    ]
    assert comparison[1]["pct_of_budget"] == 70.0
    assert comparison[2]["variance_to_budget"] == 350.0
    assert comparison[-1]["variance_to_budget"] == -100.0


def test_build_budget_scenario_player_recalculates_without_mutation_contract():
    player = build_budget_scenario_player(
        {
            "budget_total": 1000.0,
            "paid_total": 250.0,
            "actual_total": 400.0,
            "committed_total": 650.0,
        },
        {"projected_close_total": 900.0},
        run_rate_delta_pct=10,
        discretionary_cut_pct=5,
        added_commitments=75,
        cash_acceleration=100,
    )

    assert player["read_only"] is True
    assert player["run_rate_impact"] == 90.0
    assert player["cut_impact"] == 50.0
    assert player["projected_close_total"] == 1015.0
    assert player["projected_variance"] == -15.0
    assert player["projected_cash_need"] == 665.0
    assert player["health"] == "over_budget"


class _BudgetCommitmentsResult:
    def __init__(self, rows):
        self._rows = rows

    def mappings(self):
        return self

    def all(self):
        return self._rows


class _BudgetCommitmentsSession:
    async def execute(self, statement, params=None):
        sql = " ".join(str(statement).split())
        assert "FROM documentos d" in sql
        return _BudgetCommitmentsResult(
            [
                {
                    "documento_id": "doc-1",
                    "numero_referencia": "SOL-0001",
                    "estado": "aprobado",
                    "concepto_pago": "Uniformes regionales",
                    "monto_solicitado": 1200.0,
                    "monto_total": 1200.0,
                    "fecha_pago": None,
                    "creado_en": None,
                    "gasto_generado_id": "gasto-1",
                    "gasto_generado_referencia": "GAS-001",
                    "gasto_generado_concepto": (
                        "Reembolso personal - Uniformes regionales"
                    ),
                    "gasto_generado_total": 1200.0,
                    "gasto_generado_fecha": None,
                    "gasto_generado_estado": "activo",
                    "gasto_generado_actor": "Operaciones Norte",
                    "related_expense_count": 1,
                    "related_expense_total": 1200.0,
                    "related_expense_latest_date": None,
                    "proveedor_nombre": "Proveedor Uno",
                    "torneo_nombre": "Liga Telmex Telcel de Beisbol",
                }
            ]
        )


@pytest.mark.asyncio
async def test_list_budget_tournament_commitments_returns_document_rows():
    rows = await list_budget_tournament_commitments(
        _BudgetCommitmentsSession(),
        edition_year=2026,
        tournament_id="torneo-1",
        tournament_name="Liga Telmex Telcel de Beisbol",
        tournament_code="LTTB",
        limit=20,
    )

    assert rows == [
        {
            "documento_id": "doc-1",
            "numero_referencia": "SOL-0001",
            "estado": "aprobado",
            "concepto_pago": "Uniformes regionales",
            "budget_concept_name": None,
            "monto_solicitado": 1200.0,
            "monto_total": 1200.0,
            "fecha_pago": None,
            "creado_en": None,
            "gasto_generado_id": "gasto-1",
            "gasto_generado_referencia": "GAS-001",
            "gasto_generado_concepto": "Reembolso personal - Uniformes regionales",
            "gasto_generado_budget_concept_name": None,
            "gasto_generado_total": 1200.0,
            "gasto_generado_fecha": None,
            "gasto_generado_estado": "activo",
            "gasto_generado_actor": "Operaciones Norte",
            "related_expense_count": 1,
            "related_expense_total": 1200.0,
            "related_expense_latest_date": None,
            "proveedor_nombre": "Proveedor Uno",
            "torneo_nombre": "Liga Telmex Telcel de Beisbol",
        }
    ]


def test_iter_budget_concept_catalog_rows_reads_unique_concepts(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LTTB"
    sheet["A1"] = "Liga Telmex Telcel de Beisbol"
    sheet["A3"] = "Concepto"
    sheet["B3"] = "Enero"
    sheet["A4"] = "Hospedaje"
    sheet["B4"] = 100
    sheet["A5"] = "Hospedaje"
    sheet["B5"] = 200
    sheet["A6"] = "Uniformes"
    sheet["B6"] = 300
    path = tmp_path / "catalogo.xlsx"
    workbook.save(path)

    rows = _iter_budget_concept_catalog_rows(path)

    assert [row["concept_name"] for row in rows] == ["Hospedaje", "Uniformes"]
    assert rows[0]["tournament_code"] == "LTTB"
    assert rows[0]["metadata"]["sheet_headers"] == ["Enero"]


def test_iter_budget_concept_catalog_rows_collects_phase_scope(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CTT"
    sheet["A1"] = "Copa Telmex Telcel"
    sheet["A4"] = "Concepto"
    sheet["B4"] = "Varonil y Femenil"
    sheet["A5"] = "Scouting"
    sheet["A6"] = "FASE ESTATAL"
    sheet["A7"] = "Uniformes"
    sheet["A8"] = "FASE NACIONAL"
    sheet["A9"] = "Uniformes"
    path = tmp_path / "catalogo_fases.xlsx"
    workbook.save(path)

    rows = _iter_budget_concept_catalog_rows(path)
    by_name = {row["concept_name"]: row for row in rows}

    assert by_name["Scouting"]["metadata"]["applicable_phase_keys"] == []
    assert by_name["Uniformes"]["metadata"]["applicable_phase_keys"] == [
        "estatal",
        "fase_estatal",
        "fase_nacional",
        "nacional",
    ]


def test_iter_budget_concept_catalog_phase_rows_assigns_rows_above_break(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CTT"
    sheet["A1"] = "Copa Telmex Telcel"
    sheet["A4"] = "Concepto"
    sheet["A5"] = "Scouting - Transporte Aéreo"
    sheet["A6"] = "Papelería y enseres"
    sheet["A7"] = "FASE COLECTIVA"
    sheet["A8"] = "Uniformes"
    sheet["A9"] = "FASE ESTATAL"
    path = tmp_path / "catalogo_fases_above.xlsx"
    workbook.save(path)

    rows = _iter_budget_concept_catalog_phase_rows(path, sheet_name="CTT")
    by_name = {(row["concept_name"], row["catalog_phase_break_label"]): row for row in rows}

    assert by_name[("Scouting - Transporte Aéreo", "FASE COLECTIVA")]["concept_key"].endswith(
        "fase_colectiva"
    )
    assert by_name[("Papelería y enseres", "FASE COLECTIVA")]["concept_key"].endswith(
        "fase_colectiva"
    )
    assert by_name[("Uniformes", "FASE ESTATAL")]["concept_key"].endswith("fase_estatal")


def test_map_catalog_phase_label_to_etapa_uses_tournament_labels() -> None:
    mapped = _map_catalog_phase_label_to_etapa(
        "FASE COLECTIVA",
        ["Fase Colectiva", "Fase Estatal", "Fase Nacional", "Viaje de Campeones"],
    )

    assert mapped == "Fase Colectiva"


def test_map_catalog_phase_label_to_etapa_maps_bimbo_phases() -> None:
    etapas = ["Municipal", "Estatal", "Regional", "Final Nacional"]

    assert _map_catalog_phase_label_to_etapa("FASE LOCAL", etapas) == "Municipal"
    assert _map_catalog_phase_label_to_etapa("FASE ESTATAL", etapas) == "Estatal"
    assert _map_catalog_phase_label_to_etapa("FASE REGIONAL", etapas) == "Regional"
    assert _map_catalog_phase_label_to_etapa("FASE NACIONAL", etapas) == "Final Nacional"


def test_budget_concept_matches_fase_uses_catalog_scope_keys() -> None:
    concept = {
        "metadata": {
            "applicable_phase_keys": ["estatal", "fase_estatal"],
            "applicable_subproject_keys": [],
        }
    }

    assert budget_concept_matches_fase(concept, "Estatal") is True
    assert budget_concept_matches_fase(concept, "FASE ESTATAL") is True
    assert budget_concept_matches_fase(concept, "Nacional") is False


def test_build_budget_concept_scope_metadata_empty_is_global() -> None:
    metadata = build_budget_concept_scope_metadata([])

    assert metadata["applicable_phase_labels"] == []
    assert metadata["applicable_phase_keys"] == []


def test_build_budget_concept_scope_metadata_collects_phase_aliases() -> None:
    metadata = build_budget_concept_scope_metadata(["Estatal"])

    assert metadata["applicable_phase_labels"] == ["Estatal"]
    assert "estatal" in metadata["applicable_phase_keys"]
    assert "fase_estatal" in metadata["applicable_phase_keys"]


def test_budget_concept_scope_summary_renders_labels() -> None:
    assert (
        budget_concept_scope_summary(
            {"applicable_phase_labels": ["Estatal", "Nacional"]}
        )
        == "Estatal, Nacional"
    )
    assert (
        budget_concept_scope_summary({})
        == "Todas las fases / subproyectos"
    )


@pytest.mark.asyncio
async def test_create_budget_concept_persists_scope(monkeypatch) -> None:
    class _EmptyResult:
        def mappings(self):
            return self

        def first(self):
            return None

    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._resolve_tournament_for_budget_concept",
        AsyncMock(
            return_value={
                "tournament_id": "tor-1",
                "tournament_code": "CTT",
                "tournament_name": "Copa Telmex Telcel",
            }
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_concept",
        AsyncMock(
            return_value={
                "id": "concept-1",
                "concept_name": "Uniformes",
                "metadata": build_budget_concept_scope_metadata(["Estatal"]),
            }
        ),
    )

    session = AsyncMock()
    session.execute = AsyncMock(return_value=_EmptyResult())
    session.commit = AsyncMock()

    concept = await create_budget_concept(
        session,
        tournament_id="tor-1",
        concept_name="Uniformes",
        scope_labels=["Estatal"],
        actor_empleado_id="emp-1",
    )

    assert concept["concept_name"] == "Uniformes"
    assert session.execute.await_count == 2
    assert session.commit.await_count == 1


@pytest.mark.asyncio
async def test_create_budget_concept_persists_pasivo_account(monkeypatch) -> None:
    class _EmptyResult:
        def mappings(self):
            return self

        def first(self):
            return None

    monkeypatch.setattr("samchat.budgets.service.ensure_budget_schema", AsyncMock())
    monkeypatch.setattr(
        "samchat.budgets.service._resolve_tournament_for_budget_concept",
        AsyncMock(
            return_value={
                "tournament_id": "tor-1",
                "tournament_code": "FB",
                "tournament_name": "Futbolito Bimbo",
            }
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.validate_active_cuenta_contable_id",
        AsyncMock(side_effect=lambda _session, cuenta_id: cuenta_id),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_concept",
        AsyncMock(return_value={"id": "concept-1", "concept_name": "Uniformes"}),
    )

    execute_calls: list[tuple[str, dict[str, Any]]] = []

    async def capture_execute(query, params=None):
        execute_calls.append((str(query), dict(params or {})))
        return _EmptyResult()

    session = AsyncMock()
    session.execute = AsyncMock(side_effect=capture_execute)
    session.commit = AsyncMock()

    await create_budget_concept(
        session,
        tournament_id="tor-1",
        concept_name="Uniformes",
        cuenta_contable_id="gasto-1",
        pasivo_cuenta_contable_id="pasivo-1",
    )

    insert_params = [
        params for query, params in execute_calls if "INSERT INTO budget_concepts" in query
    ][0]
    assert insert_params["cuenta_contable_id"] == "gasto-1"
    assert insert_params["pasivo_cuenta_contable_id"] == "pasivo-1"


@pytest.mark.asyncio
async def test_bulk_save_budget_concepts_creates_and_updates(monkeypatch) -> None:
    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    create_mock = AsyncMock(return_value={"id": "new"})
    update_mock = AsyncMock(return_value={"id": "existing"})
    monkeypatch.setattr("samchat.budgets.service.create_budget_concept", create_mock)
    monkeypatch.setattr("samchat.budgets.service.update_budget_concept", update_mock)

    session = AsyncMock()
    session.commit = AsyncMock()

    result = await bulk_save_budget_concepts(
        session,
        rows=[
            {
                "concept_id": "existing-id",
                "concept_name": "Hospedaje",
                "tournament_id": "tor-1",
                "sub_proyecto": "Estatal",
                "cuenta_contable_id": "gasto-1",
                "pasivo_cuenta_contable_id": "pasivo-1",
            },
            {
                "concept_id": None,
                "concept_name": "Uniformes",
                "tournament_id": "tor-1",
                "sub_proyecto": "",
                "cuenta_contable_id": "gasto-2",
                "pasivo_cuenta_contable_id": "pasivo-2",
            },
            {"concept_id": "", "concept_name": "", "tournament_id": "", "sub_proyecto": ""},
        ],
        actor_empleado_id="emp-1",
    )

    assert result == {"created": 1, "updated": 1, "total": 2}
    update_mock.assert_awaited_once()
    create_mock.assert_awaited_once()
    update_kwargs = update_mock.await_args.kwargs
    assert update_kwargs["pasivo_cuenta_contable_id"] == "pasivo-1"
    assert update_kwargs["pasivo_cuenta_contable_provided"] is True
    create_kwargs = create_mock.await_args.kwargs
    assert create_kwargs["pasivo_cuenta_contable_id"] == "pasivo-2"
    session.commit.assert_awaited_once()


@pytest.mark.asyncio
async def test_update_budget_concept_rejects_missing_record(monkeypatch) -> None:
    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_concept",
        AsyncMock(return_value=None),
    )

    with pytest.raises(ValueError, match="no encontrada"):
        await update_budget_concept(
            AsyncMock(),
            concept_id="missing",
            concept_name="Hospedaje",
        )


@pytest.mark.asyncio
async def test_update_budget_concept_changes_and_clears_pasivo_account(monkeypatch) -> None:
    class _EmptyResult:
        def mappings(self):
            return self

        def first(self):
            return None

    current = {
        "id": "concept-1",
        "concept_name": "Uniformes",
        "tournament_id": "tor-1",
        "tournament_code": "FB",
        "metadata": {},
    }
    monkeypatch.setattr("samchat.budgets.service.ensure_budget_schema", AsyncMock())
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_concept",
        AsyncMock(side_effect=[current, {**current, "pasivo_cuenta_contable_id": None}]),
    )

    execute_calls: list[tuple[str, dict[str, Any]]] = []

    async def capture_execute(query, params=None):
        execute_calls.append((str(query), dict(params or {})))
        return _EmptyResult()

    session = AsyncMock()
    session.execute = AsyncMock(side_effect=capture_execute)
    session.commit = AsyncMock()

    await update_budget_concept(
        session,
        concept_id="concept-1",
        pasivo_cuenta_contable_id="",
        pasivo_cuenta_contable_provided=True,
    )

    update_params = [
        params for query, params in execute_calls if "UPDATE budget_concepts" in query
    ][0]
    assert update_params["pasivo_cuenta_contable_id"] is None


@pytest.mark.asyncio
async def test_update_budget_concept_scope_only_preserves_catalog_concept_key(
    monkeypatch,
) -> None:
    class _EmptyResult:
        def mappings(self):
            return self

        def first(self):
            return None

    current = {
        "id": "concept-1",
        "concept_name": "Director General",
        "concept_key": "director_general__final_nacional",
        "tournament_id": "tor-1",
        "tournament_code": "FB",
        "tournament_name": "Futbolito Bimbo",
        "metadata": {"applicable_subproject_labels": ["Final Nacional"]},
    }
    updated = {
        **current,
        "metadata": build_budget_concept_scope_metadata(["Honorarios"]),
    }

    monkeypatch.setattr(
        "samchat.budgets.service.ensure_budget_schema",
        AsyncMock(),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_concept",
        AsyncMock(side_effect=[current, updated]),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._resolve_tournament_for_budget_concept",
        AsyncMock(
            return_value={
                "tournament_id": "tor-1",
                "tournament_code": "FB",
                "tournament_name": "Futbolito Bimbo",
                "etapas": ["Final Nacional", "Honorarios"],
            }
        ),
    )

    execute_calls: list[tuple[str, dict[str, Any]]] = []

    async def capture_execute(query, params=None):
        execute_calls.append((str(query), dict(params or {})))
        return _EmptyResult()

    session = AsyncMock()
    session.execute = AsyncMock(side_effect=capture_execute)
    session.commit = AsyncMock()

    result = await update_budget_concept(
        session,
        concept_id="concept-1",
        concept_name="Director General",
        scope_labels=["Honorarios"],
    )

    update_calls = [
        params
        for query, params in execute_calls
        if "UPDATE budget_concepts" in query
    ]
    duplicate_checks = [
        params
        for query, params in execute_calls
        if "concept_key = :concept_key" in query and "UPDATE" not in query
    ]

    assert len(update_calls) == 1
    assert "concept_key" not in update_calls[0]
    assert "concept_name" not in update_calls[0]
    assert "metadata" in update_calls[0]
    assert duplicate_checks == []
    assert result["concept_name"] == "Director General"


@pytest.mark.asyncio
async def test_resolve_budget_concept_filters_by_tournament(monkeypatch):
    monkeypatch.setattr(
        "samchat.budgets.service.list_budget_concepts",
        AsyncMock(
            return_value=[
                {
                    "id": "concept-1",
                    "tournament_id": "tor-1",
                    "tournament_code": "LTTB",
                    "tournament_name": "Liga Telmex Telcel",
                    "concept_name": "Hospedaje",
                }
            ]
        ),
    )

    concept = await resolve_budget_concept(
        object(),
        budget_concept_id="concept-1",
        tournament_id="tor-1",
    )
    wrong_tournament = await resolve_budget_concept(
        object(),
        budget_concept_id="concept-1",
        tournament_id="tor-2",
    )

    assert concept is not None
    assert concept["id"] == "concept-1"
    assert wrong_tournament is None


@pytest.mark.asyncio
async def test_resolve_budget_concept_filters_by_fase(monkeypatch):
    monkeypatch.setattr(
        "samchat.budgets.service.list_budget_concepts",
        AsyncMock(
            return_value=[
                {
                    "id": "concept-1",
                    "tournament_id": "tor-1",
                    "metadata": {
                        "applicable_phase_keys": ["estatal", "fase_estatal"],
                        "applicable_subproject_keys": [],
                    },
                }
            ]
        ),
    )

    concept = await resolve_budget_concept(
        object(),
        budget_concept_id="concept-1",
        tournament_id="tor-1",
        fase="Estatal",
    )
    wrong_fase = await resolve_budget_concept(
        object(),
        budget_concept_id="concept-1",
        tournament_id="tor-1",
        fase="Nacional",
    )

    assert concept is not None
    assert wrong_fase is None


@pytest.mark.asyncio
async def test_import_budget_lines_upload_creates_catalog_backed_line(monkeypatch):
    created = {}

    async def _fake_create(*_args, **kwargs):
        created.update(kwargs)
        return {"id": "line-1", **kwargs}

    monkeypatch.setattr("samchat.budgets.service.ensure_budget_schema", AsyncMock())
    monkeypatch.setattr(
        "samchat.budgets.service.get_budget_version",
        AsyncMock(
            return_value={"id": "ver-1", "status": "draft"}
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._load_tabular_upload_rows",
        lambda **_kwargs: [
            {
                "torneo": "LTTB",
                "partida_presupuestal": "Hospedaje",
                "monto_anual": "1500",
                "fase": "Regional",
            }
        ],
    )
    monkeypatch.setattr(
        "samchat.budgets.service.list_budget_concepts",
        AsyncMock(
            return_value=[
                {
                    "id": "concept-1",
                    "tournament_code": "LTTB",
                    "tournament_name": "Liga Telmex Telcel",
                    "concept_name": "Hospedaje",
                    "concept_key": "hospedaje",
                }
            ]
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.list_budget_lines",
        AsyncMock(return_value=[]),
    )
    monkeypatch.setattr("samchat.budgets.service.create_budget_line", _fake_create)

    result = await import_budget_lines_upload(
        object(),
        version_id="ver-1",
        actor_empleado_id="emp-1",
        file_bytes=b"fake",
        filename="presupuesto.xlsx",
    )

    assert result["created"] == 1
    assert created["budget_concept_id"] == "concept-1"
    assert created["phase"] == "Regional"


@pytest.mark.asyncio
async def test_sync_budget_projects_from_partidas_workbook_aligns_only_workbook_projects(
    monkeypatch,
) -> None:
    execute_calls: list[tuple[str, dict[str, Any]]] = []

    class _ScalarResult:
        def scalar(self):
            return 7

    class _Session:
        async def execute(self, statement, params=None):
            sql = " ".join(str(statement).split())
            payload = dict(params or {})
            execute_calls.append((sql, payload))
            if "MAX(display_order)" in sql:
                return _ScalarResult()
            return MagicMock(rowcount=1)

        async def commit(self):
            execute_calls.append(("COMMIT", {}))

    existing_rows = [
        {
            "id": "tor-ctt",
            "name": "Copa Telmex Telcel de Fútbol",
            "active": True,
            "etapas": ["Fase Estatal"],
        },
        {
            "id": "tor-other",
            "name": "Proyecto fuera del archivo",
            "active": True,
            "etapas": ["No tocar"],
        },
    ]
    monkeypatch.setattr(
        "samchat.budgets.service._load_tournament_rows",
        AsyncMock(return_value=existing_rows),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.collect_workbook_project_configs",
        lambda _path: [
            {
                "proyecto": "Copa Telmex Telcel de Fútbol",
                "etapas": ["Fase Nacional", "Fase Estatal"],
            },
            {
                "proyecto": "Homeless World Cup México",
                "etapas": ["Operación"],
            },
        ],
    )
    authority_guard = AsyncMock(return_value=None)
    monkeypatch.setattr(
        "samchat.budgets.service.require_ungoverned_gastos_project",
        authority_guard,
    )

    report = await sync_budget_projects_from_partidas_workbook(
        _Session(),
        workbook_path="/tmp/partidas.xlsx",
    )

    update_calls = [
        params for sql, params in execute_calls if sql.startswith("UPDATE tournaments")
    ]
    insert_calls = [
        params for sql, params in execute_calls if sql.startswith("INSERT INTO tournaments")
    ]
    assert report["workbook_projects_count"] == 2
    assert report["projects_matched"] == 1
    assert report["projects_created"] == 1
    assert report["projects_updated"] == 1
    assert update_calls == [
        {
            "tournament_id": "tor-ctt",
            "name": "Copa Telmex Telcel de Fútbol",
            "etapas": '["Fase Nacional", "Fase Estatal"]',
        }
    ]
    assert len(insert_calls) == 1
    assert insert_calls[0]["name"] == "Homeless World Cup México"
    assert insert_calls[0]["display_order"] == 8
    assert "tor-other" not in str(update_calls)
    assert authority_guard.await_count == 1
    assert authority_guard.await_args.args[1] == "tor-ctt"


@pytest.mark.asyncio
async def test_sync_budget_projects_refuses_governed_target_before_update_or_commit(
    monkeypatch,
) -> None:
    from devnous.gastos.services.tournament_authority_service import (
        GovernedGastosProjectError,
    )

    execute_calls: list[str] = []

    class _ScalarResult:
        def scalar(self):
            return 7

    class _Session:
        async def execute(self, statement, params=None):
            sql = " ".join(str(statement).split())
            execute_calls.append(sql)
            return _ScalarResult()

        async def commit(self):
            execute_calls.append("COMMIT")

    monkeypatch.setattr(
        "samchat.budgets.service._load_tournament_rows",
        AsyncMock(
            return_value=[
                {
                    "id": "20000000-0000-0000-0000-000000000055",
                    "name": "Proyecto gobernado",
                    "active": True,
                    "etapas": ["Anterior"],
                }
            ]
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.collect_workbook_project_configs",
        lambda _path: [{"proyecto": "Proyecto gobernado", "etapas": ["Nueva"]}],
    )
    monkeypatch.setattr(
        "samchat.budgets.service.require_ungoverned_gastos_project",
        AsyncMock(
            side_effect=GovernedGastosProjectError(
                case_id="analyst_case_" + "5" * 32,
                case_version=4,
                application_hash="sha256:" + "a" * 64,
            )
        ),
    )

    with pytest.raises(GovernedGastosProjectError):
        await sync_budget_projects_from_partidas_workbook(
            _Session(), workbook_path="/tmp/partidas.xlsx"
        )

    assert not any(sql.startswith("UPDATE tournaments") for sql in execute_calls)
    assert not any(sql.startswith("INSERT INTO tournaments") for sql in execute_calls)
    assert "COMMIT" not in execute_calls


@pytest.mark.asyncio
async def test_sync_budget_concepts_deactivates_stale_only_for_matched_projects(
    monkeypatch,
) -> None:
    execute_calls: list[tuple[str, dict[str, Any]]] = []

    class _AccountRows:
        def mappings(self):
            return self

        def all(self):
            return [
                {
                    "id": "account-1",
                    "codigo": "5300-010-016",
                    "nombre": "ALIMENTACION JUGADORES",
                    "activo": True,
                }
            ]

    class _UpdateResult:
        rowcount = 3

    class _Session:
        async def execute(self, statement, params=None):
            sql = " ".join(str(statement).split())
            payload = dict(params or {})
            execute_calls.append((sql, payload))
            if "FROM cuentas_contables" in sql:
                return _AccountRows()
            if "SET active = FALSE" in sql:
                return _UpdateResult()
            return MagicMock(rowcount=1)

        async def commit(self):
            execute_calls.append(("COMMIT", {}))

    update_mock = AsyncMock(
        return_value={
            "id": "concept-present",
            "concept_name": "Alimentos",
            "concept_key": "alimentos__fase_nacional",
        }
    )

    monkeypatch.setattr("samchat.budgets.service.ensure_budget_schema", AsyncMock())
    monkeypatch.setattr(
        "samchat.budgets.service.sync_budget_projects_from_partidas_workbook",
        AsyncMock(
            return_value={
                "workbook_projects_count": 1,
                "projects_matched": 1,
                "projects_created": 0,
                "projects_updated": 0,
            }
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.ensure_missing_cuentas_contables_from_workbook",
        AsyncMock(return_value={}),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._iter_budget_concept_account_mapping_rows",
        lambda _path: [
            {
                "partida": "Alimentos",
                "concept_key": "alimentos",
                "proyecto": "Copa Telmex Telcel de Fútbol",
                "subproyecto": "Fase Nacional",
                "cuenta_contable_codigo": "5300-010-016",
                "cuenta_contable_nombre": "ALIMENTACION JUGADORES",
                "sheet_row_index": 2,
            }
        ],
    )
    monkeypatch.setattr(
        "samchat.budgets.service.list_budget_concepts",
        AsyncMock(
            return_value=[
                {
                    "id": "concept-present",
                    "tournament_id": "tor-ctt",
                    "tournament_code": "CTT",
                    "tournament_name": "Copa Telmex Telcel",
                    "concept_name": "Alimentos",
                    "concept_key": "alimentos__fase_nacional",
                    "active": True,
                    "metadata": {},
                },
                {
                    "id": "concept-stale",
                    "tournament_id": "tor-ctt",
                    "tournament_code": "CTT",
                    "tournament_name": "Copa Telmex Telcel",
                    "concept_name": "Vieja",
                    "concept_key": "vieja",
                    "active": True,
                    "metadata": {},
                },
                {
                    "id": "concept-other-project",
                    "tournament_id": "tor-other",
                    "tournament_code": "BIMBO",
                    "tournament_name": "Futbolito Bimbo",
                    "concept_name": "Debe quedar intacta",
                    "concept_key": "debe_quedar_intacta",
                    "active": True,
                    "metadata": {},
                },
            ]
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._load_tournament_rows",
        AsyncMock(
            return_value=[
                {"id": "tor-ctt", "name": "Copa Telmex Telcel"},
                {"id": "tor-other", "name": "Futbolito Bimbo"},
            ]
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service._resolve_tournament_for_budget_concept",
        AsyncMock(
            return_value={
                "tournament_id": "tor-ctt",
                "tournament_code": "CTT",
                "tournament_name": "Copa Telmex Telcel",
                "etapas": ["Fase Nacional"],
                "categorias": [],
            }
        ),
    )
    monkeypatch.setattr(
        "samchat.budgets.service.update_budget_concept",
        update_mock,
    )

    result = await sync_budget_concepts_from_partidas_workbook(
        _Session(),
        workbook_path="/tmp/partidas.xlsx",
        create_missing_cuentas=False,
    )

    stale_updates = [
        params for sql, params in execute_calls if "SET active = FALSE" in sql
    ]
    assert result["updated"] == 1
    assert result["created"] == 0
    assert result["matched_tournaments"] == 1
    assert result["deactivated_stale"] == 3
    assert result["workbook_projects_count"] == 1
    assert result["projects_matched"] == 1
    assert stale_updates == [
        {
            "tournament_id": "tor-ctt",
            "workbook_keys": ["alimentos__fase_nacional"],
        }
    ]
    assert "tor-other" not in str(stale_updates)


def test_match_budget_concept_for_account_mapping_keeps_subproject_scope() -> None:
    concepts = [
        {
            "id": "concept-catering-nacional",
            "tournament_name": "Copa Telmex Telcel de Fútbol",
            "tournament_code": "CTT",
            "concept_name": "Catering",
            "concept_key": "catering__fase_nacional",
            "metadata": build_budget_concept_scope_metadata(["Fase Nacional"]),
        }
    ]
    row = {
        "partida": "Catering",
        "concept_key": "catering",
        "proyecto": "Copa Telmex Telcel de Fútbol",
        "subproyecto": "Cena de integración / Bienvenida",
        "sheet_row_index": 23,
    }

    assert (
        _match_budget_concept_for_account_mapping(
            concepts,
            row=row,
            tournament_code="CTT",
        )
        is None
    )


def test_merge_budget_concept_metadata_skips_catalog_for_admin_ui() -> None:
    stored = {"applicable_phase_keys": ["estatal"]}
    catalog = {"applicable_phase_keys": ["nacional"], "sheet_row_indexes": [1]}

    merged = _merge_budget_concept_metadata(stored, catalog, source="admin_ui")

    assert merged == stored


def test_merge_budget_concept_metadata_fills_missing_catalog_keys() -> None:
    stored = {"applicable_phase_keys": ["estatal"]}
    catalog = {
        "applicable_subproject_keys": ["juvenil"],
        "sheet_row_indexes": [1],
    }

    merged = _merge_budget_concept_metadata(stored, catalog, source="default_catalog_xlsx")

    assert merged["applicable_phase_keys"] == ["estatal"]
    assert merged["applicable_subproject_keys"] == ["juvenil"]
    assert merged["sheet_row_indexes"] == [1]


@pytest.mark.asyncio
async def test_hide_budget_concept_sets_active_false(monkeypatch) -> None:
    from samchat.budgets.service import hide_budget_concept

    update_mock = AsyncMock(
        return_value={"id": "concept-1", "concept_name": "Hospedaje", "active": False}
    )
    monkeypatch.setattr("samchat.budgets.service.update_budget_concept", update_mock)
    session = AsyncMock()

    result = await hide_budget_concept(
        session,
        concept_id="concept-1",
        actor_empleado_id="emp-1",
    )

    assert result["active"] is False
    update_mock.assert_awaited_once_with(
        session,
        concept_id="concept-1",
        active=False,
        actor_empleado_id="emp-1",
        commit=True,
    )


@pytest.mark.asyncio
async def test_clear_budget_concept_scope_for_tournament_resets_scoped_metadata(
    monkeypatch,
) -> None:
    from uuid import uuid4
    from unittest.mock import MagicMock

    from samchat.budgets.service import clear_budget_concept_scope_for_tournament

    concept_id = uuid4()
    tournament_id = uuid4()
    session = AsyncMock()
    select_result = MagicMock()
    select_result.mappings.return_value.all.return_value = [
        {
            "id": concept_id,
            "metadata": {
                "applicable_phase_labels": ["Estatal"],
                "applicable_phase_keys": ["estatal"],
            },
        }
    ]
    session.execute = AsyncMock(side_effect=[select_result, AsyncMock()])
    monkeypatch.setattr("samchat.budgets.service.ensure_budget_schema", AsyncMock())

    updated = await clear_budget_concept_scope_for_tournament(
        session,
        tournament_id=str(tournament_id),
    )

    assert updated == 1
    assert session.execute.await_count == 2
