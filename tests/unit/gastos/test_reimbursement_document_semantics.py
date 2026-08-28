from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import UUID, uuid4

from openpyxl import load_workbook

from devnous.gastos.models import Documento
from devnous.gastos.routes import user_routes
from devnous.gastos.services import documento_telegram as tg
from devnous.gastos.services.documento_semantics import (
    account_uses_provider_beneficiary,
    approval_subject_empleado,
    approval_subject_empleado_id,
    effective_account_beneficiary_id,
    effective_account_provider_beneficiary_id,
    is_employee_reimbursement,
    reimbursement_concept_from_cuenta,
)
from devnous.gastos.utils.excel_exports import (
    INFORME_AUTORIZADO_ROW,
    create_informe_excel,
)


def test_effective_account_beneficiary_prefers_selected_employee():
    requester_id = uuid4()
    beneficiary_id = uuid4()
    cuenta = SimpleNamespace(
        empleado_id=requester_id,
        beneficiario_empleado_id=beneficiary_id,
    )

    assert effective_account_beneficiary_id(cuenta) == beneficiary_id


def test_effective_account_beneficiary_falls_back_to_requester():
    requester_id = uuid4()
    cuenta = SimpleNamespace(
        empleado_id=requester_id,
        beneficiario_empleado_id=None,
    )

    assert effective_account_beneficiary_id(cuenta) == requester_id



def test_effective_account_provider_beneficiary_detects_operator():
    provider_id = uuid4()
    cuenta = SimpleNamespace(
        empleado_id=uuid4(),
        beneficiario_empleado_id=None,
        beneficiario_proveedor_cliente_id=provider_id,
    )

    assert effective_account_provider_beneficiary_id(cuenta) == provider_id
    assert account_uses_provider_beneficiary(cuenta) is True


def test_effective_account_provider_beneficiary_is_absent_for_employee_report():
    cuenta = SimpleNamespace(
        empleado_id=uuid4(),
        beneficiario_empleado_id=uuid4(),
        beneficiario_proveedor_cliente_id=None,
    )

    assert effective_account_provider_beneficiary_id(cuenta) is None
    assert account_uses_provider_beneficiary(cuenta) is False

def test_approval_subject_prefers_employee_beneficiary_over_requester():
    requester = SimpleNamespace(id=uuid4(), nombre="Juan Pablo", aprobador_id=uuid4())
    beneficiary = SimpleNamespace(id=uuid4(), nombre="Roberto Rogers", aprobador_id=uuid4())
    documento = SimpleNamespace(
        empleado=requester,
        empleado_id=requester.id,
        beneficiario_empleado=beneficiary,
        beneficiario_empleado_id=beneficiary.id,
    )

    assert approval_subject_empleado(documento) is beneficiary
    assert approval_subject_empleado_id(documento) == beneficiary.id


def test_approval_subject_falls_back_to_requester_for_direct_documents():
    requester = SimpleNamespace(id=uuid4(), nombre="Juan Pablo", aprobador_id=uuid4())
    documento = SimpleNamespace(
        empleado=requester,
        empleado_id=requester.id,
        beneficiario_empleado=None,
        beneficiario_empleado_id=None,
    )

    assert approval_subject_empleado(documento) is requester
    assert approval_subject_empleado_id(documento) == requester.id


def test_pending_approval_contract_routes_by_beneficiary_and_shows_flash_columns():
    route_source = Path("src/devnous/gastos/routes/user_routes.py").read_text()
    workflow_source = Path("src/devnous/gastos/services/documento_workflow_service.py").read_text()

    assert "beneficiario_alias.aprobador_id == current_empleado.id" in route_source
    assert "Documento.beneficiario_empleado_id.is_(None)" in route_source
    assert "<th>Torneo</th>" in route_source
    assert "Solicitante / beneficiario" in route_source
    assert "<th>Solicitante</th>" in route_source
    assert "<th>Beneficiario/Proveedor</th>" in route_source
    assert "<th>Acciones</th>" in route_source
    assert '/documentos/{documento.id}/aprobar' in route_source
    assert '/documentos/{documento.id}/rechazar' in route_source
    assert 'getattr(cuenta, "nombre", None)' in route_source
    assert "approval_subject = approval_subject_empleado(documento)" in workflow_source




def test_document_reporting_description_uses_expense_report_purpose_for_informes():
    documento = SimpleNamespace(
        id=uuid4(),
        numero_referencia="I-138907",
        tipo="INFORME",
        concepto_pago=None,
        referencia_pago=None,
        referencia_operaciones="56",
        monto_solicitado=0,
        monto_total=5098.66,
        currency="MXN",
        estado="enviado",
        creado_en=None,
        enviado_en=None,
        aprobado_en=None,
        pagado_en=None,
        empleado=SimpleNamespace(nombre="Bibiana Roman"),
        beneficiario_empleado=SimpleNamespace(nombre="Bibiana Roman"),
        proveedor_cliente=None,
        beneficiario_empleado_id=uuid4(),
        proveedor_cliente_id=None,
        cuenta_gastos=SimpleNamespace(nombre="Viaje a Leon scouting FN CTT"),
    )

    row_values = user_routes._documentos_todos_reporting_row_values(documento)

    assert row_values["concepto"] == "Viaje a Leon scouting FN CTT"




def test_reimbursement_concept_from_cuenta_includes_report_description():
    cuenta = SimpleNamespace(
        referencia_base="773309",
        nombre="LTTB Envio de uniformes de la Fases Estatal Campeche BRA",
    )

    assert reimbursement_concept_from_cuenta(cuenta) == (
        "Reembolso de saldo a favor — "
        "LTTB Envio de uniformes de la Fases Estatal Campeche BRA — I-773309"
    )


def test_document_reporting_description_expands_legacy_reimbursement_from_linked_report():
    documento = SimpleNamespace(
        id=uuid4(),
        numero_referencia="S-26000143",
        tipo="SOLICITUD",
        concepto_pago="Reembolso de saldo a favor — I-773309",
        referencia_pago=None,
        referencia_operaciones="92",
        monto_solicitado=1495.21,
        monto_total=1495.21,
        currency="MXN",
        estado="control_presupuestal",
        creado_en=None,
        enviado_en=None,
        aprobado_en=None,
        pagado_en=None,
        empleado=SimpleNamespace(nombre="Bibiana Roman"),
        beneficiario_empleado=SimpleNamespace(nombre="Bibiana Roman"),
        proveedor_cliente=None,
        beneficiario_empleado_id=uuid4(),
        proveedor_cliente_id=None,
        cuenta_gastos_id=uuid4(),
        cuenta_gastos=SimpleNamespace(
            referencia_base="773309",
            nombre="LTTB Envio de uniformes de la Fases Estatal Campeche BRA",
        ),
    )

    row_values = user_routes._documentos_todos_reporting_row_values(documento)

    assert row_values["concepto"] == (
        "Reembolso de saldo a favor — "
        "LTTB Envio de uniformes de la Fases Estatal Campeche BRA — I-773309"
    )

def test_document_reporting_description_falls_back_to_payment_concept_for_requests():
    documento = SimpleNamespace(
        id=uuid4(),
        numero_referencia="S-26000109",
        tipo="SOLICITUD",
        concepto_pago="SERVICIO DE TRANSPORTE PARA REVISION DE CAMPOS",
        referencia_pago=None,
        referencia_operaciones="58",
        monto_solicitado=8700,
        monto_total=8700,
        currency="MXN",
        estado="enviado",
        creado_en=None,
        enviado_en=None,
        aprobado_en=None,
        pagado_en=None,
        empleado=SimpleNamespace(nombre="Alicia Zuniga"),
        beneficiario_empleado=None,
        proveedor_cliente=SimpleNamespace(nombre="Veronica Jimenez"),
        beneficiario_empleado_id=None,
        proveedor_cliente_id=uuid4(),
        cuenta_gastos=None,
    )

    row_values = user_routes._documentos_todos_reporting_row_values(documento)

    assert row_values["concepto"] == "SERVICIO DE TRANSPORTE PARA REVISION DE CAMPOS"


def test_workflow_notification_recipients_use_beneficiary_approval_subject() -> None:
    source = Path(tg.__file__).read_text()
    start = source.index("async def resolve_workflow_approval_notification_recipients")
    end = source.index("async def _record_failed_workflow_notification", start)
    block = source[start:end]

    assert "approval_subject = approval_subject_empleado(documento)" in block
    assert "subject_approver_id" in block
    assert "owner = documento.empleado" not in block
    assert "owner_approver_id" not in block


def test_documento_approver_display_uses_beneficiary_approval_subject() -> None:
    source = Path("src/devnous/gastos/services/documento_service.py").read_text()
    start = source.index("async def fetch_documento_aprobador_display_batch")
    block = source[start:]

    assert "subject_id = approval_subject_empleado_id(doc)" in block
    assert "assigned = empleado.aprobador" in block
    assert "empleado = doc.empleado" not in block

def test_telegram_approver_queue_uses_effective_beneficiary_approver():
    odilon_id = uuid4()
    federico_id = uuid4()
    requester = SimpleNamespace(nombre="Juan Pablo", aprobador_id=federico_id)
    beneficiary = SimpleNamespace(nombre="Roberto Rogers", aprobador_id=odilon_id)
    documento = SimpleNamespace(
        estado="enviado",
        empleado=requester,
        beneficiario_empleado=beneficiary,
    )

    assert tg.approver_can_see_document_in_queue(
        SimpleNamespace(id=odilon_id, rol="finanzas"), documento
    ) is True
    assert tg.approver_can_see_document_in_queue(
        SimpleNamespace(id=federico_id, rol="finanzas"), documento
    ) is False


def test_legacy_system_reimbursement_is_not_a_third_party_request():
    documento = SimpleNamespace(
        tipo="SOLICITUD",
        cuenta_gastos_id=uuid4(),
        beneficiario_empleado_id=uuid4(),
        proveedor_cliente_id=uuid4(),
        concepto_pago="Reembolso de saldo a favor — I-793655",
    )

    assert is_employee_reimbursement(documento) is True


def test_regular_supplier_request_is_not_an_employee_reimbursement():
    documento = SimpleNamespace(
        tipo="SOLICITUD",
        cuenta_gastos_id=None,
        beneficiario_empleado_id=None,
        proveedor_cliente_id=uuid4(),
        concepto_pago="Compra de uniformes",
    )

    assert is_employee_reimbursement(documento) is False


def test_reimbursement_telegram_omits_provider_and_separates_requester():
    provider = MagicMock()
    provider.nombre = "Cuenta adaptadora de Alicia"
    beneficiary = MagicMock()
    beneficiary.nombre = "Bibiana Roman"

    documento = MagicMock(spec=Documento)
    documento.numero_referencia = "S-26000052"
    documento.tipo = "SOLICITUD"
    documento.estado = "enviado"
    documento.cuenta_gastos_id = UUID("d324fc55-ca8f-44a9-bb44-8382eb6f8ff5")
    documento.beneficiario_empleado_id = UUID(
        "435825e1-0bd0-45c1-a7cc-c97cd18a2b15"
    )
    documento.beneficiario_empleado = beneficiary
    documento.proveedor_cliente = provider
    documento.concepto_pago = "Reembolso de saldo a favor — I-793655"
    documento.notas = None
    documento.referencia_operaciones = "9"
    documento.enviado_en = None
    documento.aprobado_en = None

    text = tg.format_documento_resumen_es(
        documento,
        context={
            "solicitante": "Alicia Zuñiga",
            "proyecto": "Operaciones",
            "etapa": "Artículos varios",
            "monto_line": "$628.00 MXN",
            "referencia_operaciones": "9",
        },
    )

    assert text.startswith("*Tipo de pago* *Reembolso a empleado*")
    assert "*Beneficiario del reembolso* *Bibiana Roman*" in text
    assert "*Solicitante* Alicia Zuñiga" in text
    assert "Proveedor" not in text
    assert "Cuenta adaptadora de Alicia" not in text


def test_telegram_solicitud_includes_project_phase_and_keeps_approval_keyboard() -> None:
    documento_id = uuid4()
    provider = MagicMock()
    provider.nombre = "HK DISENO SA DE CV"
    empleado = MagicMock()
    empleado.nombre = "Alicia Zuniga"

    documento = MagicMock(spec=Documento)
    documento.id = documento_id
    documento.numero_referencia = "S-26000051"
    documento.tipo = "SOLICITUD"
    documento.estado = "enviado"
    documento.empleado = empleado
    documento.proveedor_cliente = provider
    documento.beneficiario_empleado = None
    documento.beneficiario_empleado_id = None
    documento.cuenta_gastos_id = None
    documento.concepto_pago = "Estampado de playeras"
    documento.notas = "Fase Nacional"
    documento.referencia_operaciones = "10"
    documento.enviado_en = None
    documento.aprobado_en = None

    text = tg.format_documento_resumen_es(
        documento,
        context={
            "solicitante": "Alicia Zuniga",
            "proyecto": "Copa Telmex 2026",
            "etapa": "Fase Nacional",
            "monto_line": "$1,397.22 MXN",
            "referencia_operaciones": "10",
        },
        include_actions_hint=True,
    )
    keyboard = tg.approval_inline_keyboard(documento_id)

    assert "*Proyecto* Copa Telmex 2026" in text
    assert "*Etapa / subproyecto* Fase Nacional" in text
    assert "Usa los botones de abajo" in text
    assert keyboard["inline_keyboard"][0][0]["text"].endswith("Aprobar")
    assert keyboard["inline_keyboard"][0][1]["text"].endswith("Rechazar")
    assert keyboard["inline_keyboard"][0][0]["callback_data"] == f"{tg.CB_APPROVE}{documento_id}"
    assert keyboard["inline_keyboard"][0][1]["callback_data"] == f"{tg.CB_REJECT}{documento_id}"


def test_telegram_informe_includes_project_phase_from_context() -> None:
    beneficiary = MagicMock()
    beneficiary.nombre = "Bibiana Roman"
    empleado = MagicMock()
    empleado.nombre = "Bibiana Roman"

    documento = MagicMock(spec=Documento)
    documento.numero_referencia = "I-793655"
    documento.tipo = "INFORME"
    documento.estado = "abierta"
    documento.empleado = empleado
    documento.proveedor_cliente = None
    documento.beneficiario_empleado = beneficiary
    documento.beneficiario_empleado_id = uuid4()
    documento.cuenta_gastos_id = uuid4()
    documento.concepto_pago = None
    documento.notas = "Compra de articulos de papeleria"
    documento.enviado_en = None
    documento.aprobado_en = None

    text = tg.format_documento_resumen_es(
        documento,
        context={
            "solicitante": "Bibiana Roman",
            "proyecto": "Gastos Administrativos - Operaciones",
            "etapa": "Articulos varios",
            "monto_solicitado": "$0.00 MXN",
            "monto_gastado": "$500.00 MXN",
            "saldo_line": "$500.00 MXN - A favor del empleado - Reembolso pendiente",
        },
    )

    assert "*Proyecto* Gastos Administrativos - Operaciones" in text
    assert "*Etapa / subproyecto* Articulos varios" in text
    assert "*Monto gastado* $500.00 MXN" in text

def test_telegram_informe_for_employee_beneficiary_does_not_call_employee_provider() -> None:
    beneficiary = MagicMock()
    beneficiary.nombre = "JOSE ODILON TRUJILLO MACEDO"
    provider = MagicMock()
    provider.nombre = "JOSE ODILON TRUJILLO MACEDO"

    documento = MagicMock(spec=Documento)
    documento.numero_referencia = "I-128355"
    documento.tipo = "INFORME"
    documento.estado = "enviado"
    documento.proveedor_cliente = provider
    documento.beneficiario_empleado = beneficiary
    documento.beneficiario_empleado_id = uuid4()
    documento.cuenta_gastos_id = uuid4()
    documento.concepto_pago = None
    documento.notas = None
    documento.enviado_en = None
    documento.aprobado_en = None

    text = tg.format_documento_resumen_es(
        documento,
        context={
            "solicitante": "ALICIA EDITH ZUNIGA SALAZAR",
            "proyecto": "Copa Telmex Telcel de Futbol",
            "etapa": "?",
            "monto_solicitado": "$0.00 MXN",
            "monto_gastado": "$0.00 MXN",
            "saldo_line": "$0.00 MXN - Saldado - Entregado y gastos de bolsillo coinciden.",
        },
    )

    assert text.startswith("*Tipo de documento* *Informe de gastos*")
    assert "*Persona que comprueba* *JOSE ODILON TRUJILLO MACEDO*" in text
    assert "*Solicitante* ALICIA EDITH ZUNIGA SALAZAR" in text
    assert "Proveedor" not in text


def test_project_and_phase_labels_tolerates_deferred_attribute_failures() -> None:
    class DeferredDocumento:
        @property
        def cuenta_gastos(self):
            raise RuntimeError("deferred relationship unavailable")

        @property
        def torneo(self):
            raise RuntimeError("deferred relationship unavailable")

        @property
        def proyecto_otro(self):
            raise RuntimeError("deferred column unavailable")

        @property
        def fase(self):
            raise RuntimeError("deferred column unavailable")

    assert tg._project_and_phase_labels(DeferredDocumento()) == ("—", "—")


def test_document_detail_source_prefers_employee_beneficiary_for_reimbursement() -> None:
    source = Path(user_routes.__file__).read_text()
    detail_block_start = source.index('solicitud_transferencia_html = ""')
    detail_block_end = source.index('st_monto = documento.monto_solicitado', detail_block_start)
    detail_block = source[detail_block_start:detail_block_end]

    assert 'if is_employee_reimbursement(documento):' in detail_block
    assert 'documento.beneficiario_empleado.nombre' in detail_block
    assert 'elif documento.proveedor_cliente_id and documento.proveedor_cliente:' in detail_block
    assert detail_block.index('if is_employee_reimbursement(documento):') < detail_block.index(
        'elif documento.proveedor_cliente_id and documento.proveedor_cliente:'
    )


def test_pending_documents_view_shows_operations_reference_column() -> None:
    source = Path(user_routes.__file__).read_text()
    route_start = source.index('async def documentos_pendientes(')
    route_end = source.index('@router.get("/documentos/historial-aprobador"', route_start)
    route_block = source[route_start:route_end]

    assert "referencia_operaciones = escape(" in route_block
    assert "documento.referencia_operaciones" in route_block
    assert "<th>Referencia Operaciones</th>" in route_block
    assert "<td>{referencia_operaciones}</td>" in route_block


def test_approval_history_view_shows_operations_reference_column() -> None:
    source = Path(user_routes.__file__).read_text()
    route_start = source.index('async def historial_aprobador(')
    route_end = source.index("def _documentos_todos_reporting_type", route_start)
    route_block = source[route_start:route_end]

    assert "Eventos registrados" in route_block
    assert "referencia_operaciones = escape(" in route_block
    assert "documento.referencia_operaciones" in route_block
    assert "<th>Referencia Operaciones</th>" in route_block
    assert "<td>{referencia_operaciones}</td>" in route_block


def test_telegram_workflow_approval_triggers_odilon_finance_alert() -> None:
    source = Path(tg.__file__).read_text()
    workflow_start = source.index("async def run_document_workflow_telegram_notifications")
    workflow_end = source.index("def _sum_active_expense_amounts", workflow_start)
    workflow_block = source[workflow_start:workflow_end]
    approve_start = workflow_block.index('elif normalized == "approve":')
    reject_start = workflow_block.index('elif normalized == "reject":')
    approve_block = workflow_block[approve_start:reject_start]

    assert "notify_requester_decision" in approve_block
    assert "notify_finance_when_odilon_approves(session, documento, actor)" in approve_block


def test_workflow_send_notifies_all_resolved_recipients() -> None:
    source = Path(tg.__file__).read_text()
    notify_start = source.index("async def notify_assigned_approver_new_request")
    notify_end = source.index("async def _find_monitor_alert_recipient", notify_start)
    notify_block = source[notify_start:notify_end]

    assert "resolve_workflow_approval_notification_recipients" in notify_block
    assert "for recipient in recipients:" in notify_block
    assert 'notification_type="workflow_send_approver"' in notify_block
    assert "recipient_empleado_id=recipient.id" in notify_block


def test_workflow_monitor_retries_and_alerts_francisco_after_cutoff() -> None:
    source = Path(tg.__file__).read_text()
    alert_start = source.index("async def _send_workflow_monitor_alert")
    monitor_start = source.index("async def monitor_workflow_telegram_notifications")
    monitor_end = source.index("async def notify_requester_decision", monitor_start)
    alert_block = source[alert_start:monitor_start]
    monitor_block = source[monitor_start:monitor_end]

    assert "WORKFLOW_NOTIFICATION_MONITOR_MINUTES = 5" in source
    assert 'WORKFLOW_NOTIFICATION_ALERT_MATCHER = "francisco"' in source
    assert "await notify_assigned_approver_new_request(session, documento)" in monitor_block
    assert "await _send_workflow_monitor_alert(" in monitor_block
    assert 'notification_type="workflow_notification_monitor_alert"' in alert_block
    assert "existing is not None and existing.status == \"sent\"" in alert_block
    assert "status_by_recipient.get(recipient.id) != \"sent\"" in monitor_block


def test_outbox_idempotency_has_service_and_schema_guards() -> None:
    outbox_source = Path(
        "src/devnous/gastos/services/telegram_outbox_service.py"
    ).read_text()
    schema_source = Path("src/devnous/gastos/schema_guard.py").read_text()

    assert "async def find_outbox_entry" in outbox_source
    assert "if existing is not None and existing.status == \"sent\"" in outbox_source
    assert "except IntegrityError" in outbox_source
    assert "ux_telegram_notification_outbox_logical_recipient" in schema_source
    assert (
        "PARTITION BY notification_type, documento_id, recipient_empleado_id"
        in schema_source
    )


def test_support_status_exposes_incomplete_workflow_notifications() -> None:
    source = Path("src/devnous/gastos/routes/support_routes.py").read_text()

    assert "async def _workflow_telegram_incomplete_rows" in source
    assert "resolve_workflow_approval_notification_recipients" in source
    assert "Solicitudes con aviso de aprobación incompleto" in source
    assert "WORKFLOW_NOTIFICATION_MONITOR_MINUTES" in source


def test_systemd_timer_runs_workflow_monitor_every_five_minutes() -> None:
    timer = Path(
        "deployment/systemd/samchat-telegram-workflow-monitor.timer"
    ).read_text()
    service = Path(
        "deployment/systemd/samchat-telegram-workflow-monitor.service"
    ).read_text()

    assert "OnUnitActiveSec=5min" in timer
    assert "scripts/monitor_telegram_workflow_notifications.py" in service
    assert "SAMCHAT_ENV_FILE=/etc/samchat/samchat.env" in service
    assert "--older-than-minutes 5" in service



def test_document_detail_approval_actions_use_beneficiary_subject() -> None:
    source = Path("src/devnous/gastos/routes/user_routes.py").read_text()
    start = source.index("async def ver_documento")
    end = source.index("    # Load torneo if linked.", start)
    block = source[start:end]

    assert "approval_subject = approval_subject_empleado(documento) or empleado" in block
    assert "approval_subject.aprobador_id == current_empleado.id" in block
    assert "empleado.aprobador_id == current_empleado.id" not in block


def test_document_detail_inherits_project_context_from_expense_account() -> None:
    source = Path("src/devnous/gastos/routes/user_routes.py").read_text()
    start = source.index("async def ver_documento")
    end = source.index("authorization_strategy_html =", start)
    block = source[start:end]

    assert "selectinload(Documento.cuenta_gastos)" in block
    assert ".undefer(CuentaDeGastos.torneo_id)" in block
    assert ".undefer(CuentaDeGastos.fase)" in block
    assert "selectinload(CuentaDeGastos.torneo)" in block
    assert 'torneo = getattr(cuenta_vinculada, "torneo", None)' in block
    assert 'fase_doc_display = str(getattr(cuenta_vinculada, "fase", None)' in block
    assert "project_doc_display = documento_project_name(documento, torneo)" in block



def test_document_workflow_allows_assigned_non_admin_approver_before_role_gate() -> None:
    source = Path("src/devnous/gastos/services/documento_workflow_service.py").read_text()
    approve_start = source.index('    elif normalized_action == "approve":')
    reject_start = source.index('    elif normalized_action == "reject":')
    withdraw_start = source.index('    elif normalized_action == "withdraw":')
    approve_block = source[approve_start:reject_start]
    reject_block = source[reject_start:withdraw_start]

    for block in (approve_block, reject_block):
        subject_pos = block.index("approval_subject = approval_subject_empleado(documento)")
        role_gate_pos = block.index("elif actor.rol not in APPROVER_ROLES")
        assert subject_pos < role_gate_pos
        assert "es_aprobador_asignado = approval_subject.aprobador_id == actor.id" in block
        assert "es_superadmin = actor.rol in" in block


def test_operator_beneficiary_reimbursement_is_preserved_in_service_and_route():
    route_source = Path("src/devnous/gastos/routes/user_routes.py").read_text()
    service_source = Path("src/devnous/gastos/services/documento_service.py").read_text()

    assert "effective_account_provider_beneficiary_id(cuenta)" in route_source
    assert "_html_single_proveedor_bank_account_options" in route_source
    assert "Documento.beneficiario_proveedor_cliente_id == provider_beneficiary_id" in route_source
    assert "cuenta_provider_beneficiary_id" in service_source
    assert "beneficiario_empleado_id=(" in service_source
    assert "beneficiario_proveedor_cliente_id=cuenta_provider_beneficiary_id" in service_source
    assert "Seleccione la cuenta bancaria del operador regional beneficiario" in service_source


def test_informe_excel_uses_real_authorizer_not_requester() -> None:
    workbook = load_workbook(
        BytesIO(
            create_informe_excel(
                numero_referencia="I-775379",
                empleado_nombre="Bibiana Roman",
                beneficiario_nombre="Carlos Lozano",
                autorizado_por_nombre="Odilon Rodriguez",
                expenses=[
                    {
                        "concepto": "Taxi",
                        "fecha": "2026-08-26",
                        "no_factura": "ABC",
                        "importe_sin_iva": 100,
                        "iva": 16,
                        "total": 116,
                    }
                ],
            )
        )
    )

    assert workbook.active.cell(row=INFORME_AUTORIZADO_ROW, column=4).value == (
        "Odilon Rodriguez"
    )


def test_informe_excel_leaves_authorizer_blank_before_approval() -> None:
    workbook = load_workbook(
        BytesIO(
            create_informe_excel(
                numero_referencia="I-775379",
                empleado_nombre="Bibiana Roman",
                beneficiario_nombre="Carlos Lozano",
                expenses=[
                    {
                        "concepto": "Taxi",
                        "fecha": "2026-08-26",
                        "no_factura": "ABC",
                        "importe_sin_iva": 100,
                        "iva": 16,
                        "total": 116,
                    }
                ],
            )
        )
    )

    assert workbook.active.cell(row=INFORME_AUTORIZADO_ROW, column=4).value in (
        None,
        "",
    )


def test_reimbursement_request_is_created_when_closed_informe_is_approved() -> None:
    route_source = Path("src/devnous/gastos/routes/user_routes.py").read_text()
    service_source = Path("src/devnous/gastos/services/documento_service.py").read_text()

    assert (
        "async def _ensure_reembolso_solicitud_for_approved_informe"
        in route_source
    )
    assert 'informe_doc.estado != "aprobado"' in route_source
    assert 'saldo_raw >= -0.005' in route_source
    assert (
        'Documento.concepto_pago.like("Reembolso de saldo a favor%")'
        in route_source
    )
    assert "allow_closed_cuenta=True" in route_source
    assert "await _ensure_reembolso_solicitud_for_approved_informe(" in route_source
    assert "allow_closed_cuenta: bool = False" in service_source
    assert (
        'cuenta.estado == "cerrada" and not payload.allow_closed_cuenta'
        in service_source
    )
