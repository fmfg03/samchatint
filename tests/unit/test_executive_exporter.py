from io import BytesIO

from openpyxl import load_workbook

from samchat.executive import generate_executive_export_xlsx


def test_generate_executive_export_xlsx_builds_expected_workbook() -> None:
    payload = generate_executive_export_xlsx(
        finance_platform={
            "period": {"year": 2026, "month": 9},
            "summary": {"open_actions": 3},
            "payment_run": {"payable_total": 1200, "payable_count": 2},
            "cash_control_center": {
                "paid_total": 500,
                "paid_documents_count": 1,
                "income_total": 700,
                "income_polizas_count": 4,
            },
            "accounting_close_center": {
                "pending_coi_expenses_count": 5,
                "unbalanced_count": 1,
            },
            "tax_readiness": {"diot_blockers_count": 2, "status": "Atención"},
        },
        budget_snapshot={
            "summary": {
                "budget_total": 10000,
                "requested_total": 1500,
                "committed_total": 2000,
                "paid_total": 500,
                "actual_total": 450,
                "pending_to_pay_total": 300,
            },
            "forecast": {
                "projected_close_total": 9500,
                "projected_cash_need": 800,
            },
            "executive_comparison": [
                {"label": "Torneo A", "total": 2500, "detail": "Revisar"}
            ],
        },
        ar_payload={
            "summary": {
                "expected_income_total": 9000,
                "invoiced_total": 4000,
                "collected_total": 2500,
                "balance_total": 1500,
            }
        },
        ar_rows=[
            {
                "operational_status": "Vencido",
                "payer_name": "Cliente A",
                "balance_amount": 1500,
                "ar_item_id": "ar-1",
            }
        ],
        alerts=[
            {
                "severity": "high",
                "module": "Pagos",
                "title": "Pago urgente",
                "detail": "Documento pendiente",
                "owner": "Contabilidad",
                "source": "Finance Platform",
                "href": "/admin/finanzas",
            }
        ],
        source_notes=["Fuente parcial de prueba."],
    )

    wb = load_workbook(BytesIO(payload))

    assert wb.sheetnames == [
        "Resumen ejecutivo",
        "Presupuesto",
        "Flujo y pagos",
        "Cuentas por cobrar",
        "Alertas",
        "Fuentes y limites",
    ]
    assert wb["Resumen ejecutivo"]["A1"].value == "Export ejecutivo SamChat"
    assert wb["Resumen ejecutivo"]["A2"].value == "Periodo: 9/2026"
    assert "Saldo pendiente CxC" in [
        wb["Resumen ejecutivo"].cell(row=row, column=1).value
        for row in range(1, wb["Resumen ejecutivo"].max_row + 1)
    ]
    assert wb["Cuentas por cobrar"]["A6"].value == "Vencido"
    assert wb["Alertas"]["C2"].value == "Pago urgente"
    assert wb["Fuentes y limites"]["B9"].value == "Fuente parcial de prueba."


def test_generate_executive_export_xlsx_handles_empty_payloads() -> None:
    payload = generate_executive_export_xlsx(
        finance_platform={},
        budget_snapshot={},
        ar_payload={},
        ar_rows=[],
        alerts=[],
    )

    wb = load_workbook(BytesIO(payload))

    assert wb["Resumen ejecutivo"]["A1"].value == "Export ejecutivo SamChat"
    assert wb["Alertas"].max_row == 1
    assert wb["Fuentes y limites"]["A8"].value == "Límite"
