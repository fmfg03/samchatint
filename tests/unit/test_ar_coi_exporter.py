from __future__ import annotations

import io

from openpyxl import load_workbook

from samchat.ar.coi_exporter import (
    build_ar_coi_ready_policy_rows,
    generate_ar_coi_ready_xlsx,
)


def _ready_invoice_row() -> dict:
    return {
        "source": "issued_linked",
        "ar_item_id": "linked:invoice-1",
        "cfdi_uuid": "uuid-invoice-1",
        "payer_name": "Cliente SA",
        "payer_rfc": "CLI010101AAA",
        "concept_name": "Patrocinio",
        "issued_amount": 116,
        "iva_amount": 16,
        "account_code_final": "4100-001-004",
    }


def _ready_collection_row() -> dict:
    return {
        **_ready_invoice_row(),
        "ar_item_id": "linked:collection-1",
        "collection_match_id": "match-1",
        "collected_amount": 116,
    }


def test_build_ar_coi_ready_policy_rows_includes_ready_invoice_and_collection():
    rows = build_ar_coi_ready_policy_rows(
        [_ready_invoice_row(), _ready_collection_row()]
    )

    assert [row["policy_type"] for row in rows].count("factura") == 6
    assert [row["policy_type"] for row in rows].count("cobro") == 2
    assert ("1150-001-003", 116.0, 0.0) in {
        (row["account_code"], row["debe"], row["haber"]) for row in rows
    }
    assert ("1120-001-001", 116.0, 0.0) in {
        (row["account_code"], row["debe"], row["haber"]) for row in rows
    }
    assert ("4100-001-004", 0.0, 100.0) in {
        (row["account_code"], row["debe"], row["haber"]) for row in rows
    }


def test_build_ar_coi_ready_policy_rows_excludes_incomplete_and_unmatched():
    rows = build_ar_coi_ready_policy_rows(
        [
            {
                "source": "issued_linked",
                "ar_item_id": "linked:missing-account",
                "issued_amount": 100,
            },
            {
                "source": "expected_income",
                "ar_item_id": "expected:1",
                "expected_income_amount": 100,
            },
        ]
    )

    assert rows == []


def test_generate_ar_coi_ready_xlsx_contains_coi_and_trace_sheets():
    payload = generate_ar_coi_ready_xlsx([_ready_collection_row()])
    wb = load_workbook(io.BytesIO(payload))

    assert wb.sheetnames == ["COI CxC", "Trazabilidad"]
    coi_values = [
        cell
        for row in wb["COI CxC"].iter_rows(values_only=True)
        for cell in row
        if cell not in (None, "")
    ]
    trace_values = [
        cell
        for row in wb["Trazabilidad"].iter_rows(values_only=True)
        for cell in row
        if cell not in (None, "")
    ]

    assert "Dr" in coi_values
    assert "Ig" in coi_values
    assert "INICIO_CFDI" in coi_values
    assert "uuid-invoice-1" in coi_values
    assert "4100-001-004" in trace_values
    assert "1120-001-001" in trace_values


def test_generate_ar_coi_ready_xlsx_handles_empty_workbook_without_500():
    payload = generate_ar_coi_ready_xlsx([])
    wb = load_workbook(io.BytesIO(payload))

    assert wb["COI CxC"]["A1"].value == "Sin prepólizas CxC listas para exportar"
    assert wb["Trazabilidad"]["G2"].value == "Sin prepólizas CxC listas"
