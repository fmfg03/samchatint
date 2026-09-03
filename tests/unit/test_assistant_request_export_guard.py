from samchat.assistant.request_intent import detect_request_intent
from samchat.assistant.request_reports import RequestReportResult
from samchat.assistant.request_response import build_request_trace
from samchat.assistant.request_router import route_request
from samchat.assistant.router import (
    AssistantReportExportRequest,
    _maybe_append_export_prompt,
    _report_csv_bytes,
    _report_xlsx_bytes,
)


def _budget_template_report():
    return {
        "report_type": "budget_vs_actual",
        "title": "Presupuesto vs Real",
        "subtitle": "Junio / Enero-Junio",
        "currency_scale": "mxn",
        "columns": [
            "Segmento",
            "Presupuesto Junio",
            "Presupuesto Enero-Junio",
            "Real Junio",
            "Real Enero-Junio",
            "Variación Junio",
            "Variación Enero-Junio",
        ],
        "rows": [
            {
                "segment": "Ingresos",
                "budget_month": 2400,
                "budget_accumulated": 13850,
                "real_month": 0,
                "real_accumulated": 11450,
                "variance_month": 2400,
                "variance_accumulated": 2400,
            }
        ],
        "summary": {"variance_accumulated_total": 2400},
        "source_notes": ["template source note"],
    }


def test_timeout_with_stale_exportable_trace_never_offers_export():
    stale_trace = [
        {
            "tool": "executive.realtime_report",
            "result": {"rows": [{"concepto": "Uniformes", "amount": 100}]},
        }
    ]
    message = (
        "El proveedor del asistente tardó demasiado en responder. "
        "No ejecuté acciones ni cambios; intenta de nuevo con una "
        "consulta más "
        "corta."
    )

    assert _maybe_append_export_prompt(message, stale_trace) == message


def test_successful_request_trace_with_rows_allows_export_prompt():
    intent = detect_request_intent("Compara gasto 2026 vs 2025 por concepto")
    route = route_request(intent)
    result = RequestReportResult(
        status="success",
        title="Comparación",
        summary="ok",
        columns=["concepto", "gasto_2025", "gasto_2026"],
        rows=[{"concepto": "Uniformes", "gasto_2025": 100, "gasto_2026": 150}],
        caveats=[],
        exportable=True,
        provider_called=False,
        actions_executed=[],
        canonical_action="finance.read_only_comparison",
    )

    message = "Comparación generada."
    trace = build_request_trace(intent=intent, route=route, result=result)

    assert "¿Quieres que te lo exporte ahora?" in _maybe_append_export_prompt(
        message,
        trace,
    )


def test_template_report_payload_allows_export_prompt():
    trace = [
        {
            "tool": "assistant_finance_read",
            "result": {
                "ok": True,
                "payload": {
                    "report_type": "budget_vs_actual",
                    "title": "Presupuesto vs Real",
                    "rows": [{"segment": "Ingresos", "variance_accumulated": 100}],
                    "summary": {"variance_accumulated_total": 100},
                },
            },
        }
    ]

    assert "¿Quieres que te lo exporte ahora?" in _maybe_append_export_prompt(
        "Presupuesto vs Real generado.",
        trace,
    )


def test_assistant_report_export_request_accepts_xlsx_format():
    payload = AssistantReportExportRequest(
        conversation_id="conversation-1",
        format="xlsx",
        report_data=_budget_template_report(),
    )

    assert payload.format == "xlsx"


def test_template_report_csv_uses_report_columns():
    data = _report_csv_bytes(_budget_template_report()).decode("utf-8")

    assert "Presupuesto vs Real" in data
    assert "Segmento,Presupuesto Junio,Presupuesto Enero-Junio" in data
    assert "Ingresos,2400,13850,0,11450,2400,2400" in data
    assert "template source note" in data


def test_template_report_xlsx_uses_report_columns():
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(_report_xlsx_bytes(_budget_template_report())))
    ws = wb["Reporte ejecutivo"]

    assert ws["A1"].value == "Presupuesto vs Real"
    assert ws["A2"].value == "Junio / Enero-Junio"
    headers = [ws.cell(row=9, column=idx).value for idx in range(1, 8)]
    assert headers == [
        "Segmento",
        "Presupuesto Junio",
        "Presupuesto Enero-Junio",
        "Real Junio",
        "Real Enero-Junio",
        "Variación Junio",
        "Variación Enero-Junio",
    ]
    assert ws["A10"].value == "Ingresos"
    assert wb["Fuentes y limites"]["A4"].value == "template source note"


def test_unavailable_request_trace_does_not_allow_export_prompt():
    intent = detect_request_intent("Compara gasto 2026 vs 2025 por concepto")
    route = route_request(intent)
    result = RequestReportResult(
        status="data_source_unavailable",
        title="Sin fuente",
        summary="No disponible",
        columns=[],
        rows=[],
        caveats=[],
        exportable=False,
        provider_called=False,
        actions_executed=[],
        canonical_action="finance.read_only_comparison",
    )

    message = (
        "No encontré una fuente de datos financiera disponible. "
        "No ejecuté cambios."
    )
    trace = build_request_trace(intent=intent, route=route, result=result)

    assert _maybe_append_export_prompt(message, trace) == message
