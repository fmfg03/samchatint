"""Tests for budget monthly plan and YoY helpers."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest

from samchat.budgets.service import (
    build_budget_monthly_actuals,
    _merge_monthly_actual,
    _monthly_actual_store,
    build_budget_yoy_comparison,
    distribute_even_monthly_allocations,
)


def test_distribute_even_monthly_allocations_sums_to_total():
    allocations = distribute_even_monthly_allocations(1200.0)
    assert len(allocations) == 52
    assert round(sum(allocations.values()), 2) == 1200.0




def test_merge_monthly_actual_accepts_budget_week_52_and_ignores_53():
    store = _monthly_actual_store()
    _merge_monthly_actual(
        store,
        concept_key="concept-1",
        month_number=52,
        real_expense_cash=75.0,
    )
    _merge_monthly_actual(
        store,
        concept_key="concept-1",
        month_number=53,
        real_expense_cash=25.0,
    )

    assert store["concept-1"][52]["real_expense_cash"] == 75.0
    assert 53 not in store["concept-1"]

def test_merge_monthly_actual_buckets_by_concept_and_month():
    store = _monthly_actual_store()
    _merge_monthly_actual(
        store,
        concept_key="concept-1",
        month_number=3,
        real_expense_cash=100.0,
    )
    _merge_monthly_actual(
        store,
        concept_key="concept-1",
        month_number=3,
        real_expense_cash=50.0,
        committed_unpaid=25.0,
    )
    bucket = store["concept-1"][3]
    assert bucket["real_expense_cash"] == 150.0
    assert bucket["committed_unpaid"] == 25.0
    assert bucket["real_income"] == 0.0


@pytest.mark.asyncio
async def test_build_budget_monthly_actuals_includes_active_cfdi_income_links():
    class _Result:
        def __init__(self, rows):
            self._rows = rows

        def mappings(self):
            return self

        def all(self):
            return self._rows

    class _Session:
        async def execute(self, statement, _params=None):
            sql = str(statement)
            if "FROM budget_cfdi_income_links" in sql:
                return _Result(
                    [
                        {
                            "concept_key": "concept-1",
                            "month_number": 3,
                            "income_total": 2500,
                        }
                    ]
                )
            return _Result([])

    actuals = await build_budget_monthly_actuals(
        _Session(),
        edition_year=2026,
        version_id="11111111-1111-1111-1111-111111111111",
        tournament_id="22222222-2222-2222-2222-222222222222",
    )

    assert actuals["concept-1"][3]["real_income"] == 2500.0
    assert actuals["concept-1"][3]["real_expense_cash"] == 0.0


@pytest.mark.asyncio
async def test_build_budget_yoy_comparison_marks_new_and_retired():
    current_lines = [
        {
            "budget_concept_id": "c-new",
            "concept_name": "Streaming",
            "budget_amount": 500.0,
        }
    ]
    prior_lines = [
        {
            "budget_concept_id": "c-old",
            "concept_name": "Impresos",
            "budget_amount": 300.0,
        }
    ]

    class _Session:
        async def execute(self, _stmt, _params=None):
            class _Result:
                def mappings(self):
                    return self

                def all(self):
                    return [
                        {
                            "id": "c-new",
                            "concept_name": "Streaming",
                            "lineage_key": "streaming",
                            "concept_key": "streaming",
                        },
                        {
                            "id": "c-old",
                            "concept_name": "Impresos",
                            "lineage_key": "impresos",
                            "concept_key": "impresos",
                        },
                    ]

            return _Result()

    with patch(
        "samchat.budgets.service.list_budget_lines",
        new=AsyncMock(side_effect=[current_lines, prior_lines]),
    ):
        rows = await build_budget_yoy_comparison(
            _Session(),
            current_version_id="v-current",
            prior_version_id="v-prior",
            tournament_code="LTTB",
        )

    statuses = {row["concept_name"]: row["status"] for row in rows}
    assert statuses["Streaming"] == "new"
    assert statuses["Impresos"] == "retired"


def test_render_add_tournament_line_form_uses_phase_select_and_fetch():
    from devnous.gastos.routes.admin_budget_ui import render_add_tournament_line_form

    html = render_add_tournament_line_form(
        version_id="version-1",
        tournament_key="copatest",
        tournament_id="11111111-1111-1111-1111-111111111111",
        tournament_code="COPA",
        tournament_name="Copa Test",
        phase_labels=["Estatal", "Nacional"],
        cuentas_contables=[
            {
                "id": "cuenta-1",
                "codigo": "5300-001-006",
                "nombre": "AGUINALDO",
                "tipo": "Gasto",
            }
        ],
    )

    assert "Agregar partida al torneo" in html
    assert "Agregar partida de ingreso" not in html
    assert "Ej. Hospedaje" in html
    assert 'id="add-line-phase"' in html
    assert 'name="phase"' in html
    assert 'name="line_direction" value="expense"' in html
    assert 'name="tournament_id"' in html
    assert 'name="cuenta_contable_id"' in html
    assert "5300-001-006" in html
    assert 'name="budget_amount"' not in html
    assert "Presupuesto anual" not in html
    assert "Estatal" in html
    assert "Nacional" in html
    assert "/api/torneos/" in html
    assert 'for="add-line-concept"' in html
    assert "workspace-section-subtitle" in html


def test_render_add_tournament_line_form_without_tournament_id_disables_phase():
    from devnous.gastos.routes.admin_budget_ui import render_add_tournament_line_form

    html = render_add_tournament_line_form(
        version_id="version-1",
        tournament_key="manual-key",
        tournament_id=None,
        tournament_code="MANUAL",
        tournament_name="Manual",
        phase_labels=[],
    )

    assert "Torneos y proyectos" in html
    assert "disabled" in html
    assert "/api/torneos/" not in html


def test_render_add_tournament_line_form_income_variant_has_no_phase_field():
    from devnous.gastos.routes.admin_budget_ui import render_add_tournament_line_form

    html = render_add_tournament_line_form(
        version_id="version-1",
        tournament_key="copatest",
        tournament_id="11111111-1111-1111-1111-111111111111",
        tournament_code="COPA",
        tournament_name="Copa Test",
        phase_labels=["Estatal", "Nacional"],
        line_direction="income",
        show_phase_field=False,
        section_id="presupuesto-ingresos",
        cuentas_contables=[
            {
                "id": "cuenta-1",
                "codigo": "4100-001",
                "nombre": "PATROCINIOS",
                "tipo": "Ingreso",
            }
        ],
    )

    assert 'id="presupuesto-ingresos"' in html
    assert "Agregar partida de ingreso" in html
    assert "Agregar partida al torneo" not in html
    assert "Ej. Inscripción, patrocinio, recuperación" in html
    assert 'name="line_direction" value="income"' in html
    assert 'name="cuenta_contable_id"' in html
    assert "4100-001" in html
    assert 'name="budget_amount"' not in html
    assert 'id="add-line-phase"' not in html
    assert 'name="phase"' not in html
    assert "Fase / subproyecto" not in html
    assert "/api/torneos/" not in html
    assert "sin fase/subproyecto" in html


def test_filter_budget_lines_by_phase():
    from devnous.gastos.routes.admin_budget_ui import (
        GENERAL_PHASE_FILTER,
        budget_line_phase_key,
        collect_matrix_phase_filter_options,
        filter_budget_lines_by_phase,
    )

    lines = [
        {"id": "1", "phase": "Estatal", "concept_name": "A"},
        {"id": "2", "phase": "", "concept_name": "B"},
        {"id": "3", "phase": "Nacional", "concept_name": "C"},
    ]

    assert budget_line_phase_key(lines[1]) == GENERAL_PHASE_FILTER
    assert len(filter_budget_lines_by_phase(lines, "Estatal")) == 1
    assert len(filter_budget_lines_by_phase(lines, GENERAL_PHASE_FILTER)) == 1
    assert len(filter_budget_lines_by_phase(lines, None)) == 3

    options = collect_matrix_phase_filter_options(lines, ["Regional"])
    labels = {label for _, label in options}
    assert "Estatal" in labels
    assert "General" in labels
    assert "Regional" in labels


def test_render_budget_matrix_filters_includes_year_and_phase_controls():
    from devnous.gastos.routes.admin_budget_ui import render_budget_matrix_filters

    html = render_budget_matrix_filters(
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-2026",
        all_versions=[{"edition_year": 2026}, {"edition_year": 2027}],
        phase_options=[("Estatal", "Estatal"), ("__general__", "General")],
        selected_phase_filter="Estatal",
        visible_count=2,
        total_count=5,
    )

    assert "matrix-edition-year" in html
    assert "matrix-phase-filter" in html
    assert "Mostrando 2 de 5 partidas" in html
    assert "2027" in html
    assert 'name="version_id" value="version-2026"' in html
    assert "matrix-version" not in html
    assert "Copiar del año anterior" not in html


def test_render_cfdi_income_bridge_panel_uses_searchable_line_inputs_without_phase_controls():
    from devnous.gastos.routes.admin_budget_ui import render_cfdi_income_bridge_panel

    html = render_cfdi_income_bridge_panel(
        tournament_key="copatest",
        edition_year=2026,
        lines=[
            {
                "id": "line-estatal",
                "phase": "Estatal",
                "concept_name": "Inscripciones",
                "account_code_final": "4100-001",
                "budget_amount": 1500,
            },
            {
                "id": "line-nacional",
                "phase": "Nacional",
                "concept_name": "Patrocinios",
                "account_code_suggested": "4100-002",
                "budget_amount": 2000,
            },
        ],
        candidates=[
            {
                "id": "cfdi-1",
                "cfdi_uuid": "UUID-1",
                "emisor_rfc": "PSP010101AA1",
                "total": 1500,
                "fecha": "2026-07-01",
            }
        ],
        links=[],
        can_edit=True,
    )

    assert 'data-cfdi-income-field="cfdi"' in html
    assert 'data-cfdi-income-field="line"' in html
    assert 'name="cfdi_report_id"' in html
    assert 'name="budget_line_id"' in html
    assert 'data-cfdi-income-phase="existing"' not in html
    assert 'data-cfdi-income-phase="upload"' not in html
    assert "data-phase-select=" not in html
    assert 'id="cfdi-income-existing-phase"' not in html
    assert 'id="cfdi-income-upload-phase"' not in html
    assert "Estatal / Inscripciones / 4100-001 / $1,500.00" in html
    assert "Nacional / Patrocinios / 4100-002 / $2,000.00" in html
    assert "visibleLines" not in html
    assert "Sin partidas de ingreso disponibles" in html
    assert "Selecciona una opción de la lista." in html
    assert "Primero agrega o importa partidas de ingreso" not in html
    assert "Vincular CFDI SAT existente" in html
    assert "CFDI emitido por PSP" in html
    assert 'href="/admin/gastos/sat"' in html
    assert 'href="/admin/gastos/cfdis/matching"' in html


def test_render_cfdi_income_bridge_panel_explains_missing_income_lines():
    from devnous.gastos.routes.admin_budget_ui import render_cfdi_income_bridge_panel

    html = render_cfdi_income_bridge_panel(
        tournament_key="copatest",
        edition_year=2026,
        lines=[],
        candidates=[],
        links=[],
        can_edit=True,
    )

    assert "Primero agrega o importa partidas de ingreso" in html
    assert "Sin partidas de ingreso disponibles" in html



def test_render_budget_executive_dashboard_rolls_up_monthly_expense_view():
    from devnous.gastos.routes.admin_budget_ui import render_budget_executive_dashboard

    html = render_budget_executive_dashboard(
        [
            {
                "id": "line-1",
                "concept_name": "Balones",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={
            "line-1": {
                1: {"budget_expense_amount": 1000.0},
                2: {"budget_expense_amount": 500.0},
            }
        },
        actuals_map={
            "concept-1": {
                1: {"real_expense_cash": 200.0, "committed_unpaid": 100.0},
                2: {"real_expense_cash": 50.0, "committed_unpaid": 25.0},
            }
        },
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-1",
        budget_view="expenses",
        budget_period="monthly",
    )

    assert "Tablero ejecutivo de presupuesto" in html
    assert "Estado ejecutivo" in html
    assert "Controlado" in html
    assert "Presupuesto autorizado" in html
    assert "Ejercido real" in html
    assert "Comprometido pendiente" in html
    assert "% utilizado" in html
    assert "Lectura por periodo" in html
    assert "read-only" not in html
    assert "Mensual" in html
    assert "Enero 2026" in html
    assert "$1,500.00" in html
    assert "$250.00" in html
    assert "$125.00" in html
    assert "$1,125.00" in html
    assert "25.0%" in html
    assert 'name="budget_period"' in html
    assert 'value="monthly" selected' in html


def test_render_budget_executive_dashboard_counts_unassigned_actuals_once():
    from devnous.gastos.routes.admin_budget_ui import render_budget_executive_dashboard

    html = render_budget_executive_dashboard(
        [
            {
                "id": "line-1",
                "concept_name": "Balones",
                "budget_concept_id": "concept-1",
            },
            {
                "id": "line-2",
                "concept_name": "Uniformes",
                "budget_concept_id": "concept-2",
            },
        ],
        plan_map={
            "line-1": {1: {"budget_expense_amount": 100.0}},
            "line-2": {1: {"budget_expense_amount": 100.0}},
        },
        actuals_map={
            "__unassigned__": {
                1: {"real_expense_cash": 50.0, "committed_unpaid": 20.0}
            }
        },
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-1",
        budget_view="expenses",
        budget_period="weekly",
    )

    assert "Ejercido real</span><strong>$50.00</strong>" in html
    assert "Comprometido pendiente</span><strong>$20.00</strong>" in html
    assert "35.0%" in html


def test_render_budget_executive_dashboard_supports_quarter_semester_annual_options():
    from devnous.gastos.routes.admin_budget_ui import render_budget_executive_dashboard

    html = render_budget_executive_dashboard(
        [],
        plan_map={},
        actuals_map={},
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-1",
        budget_view="income",
        budget_period="annual",
    )

    assert "Semanal" in html
    assert "Mensual" in html
    assert "Trimestral" in html
    assert "Semestral" in html
    assert "Anual" in html
    assert "Variación" in html
    assert "2026" in html
    assert "$0.00" in html
    assert "Controlado" in html
    assert "read-only" not in html


def test_render_budget_executive_dashboard_status_warns_near_budget():
    from devnous.gastos.routes.admin_budget_ui import render_budget_executive_dashboard

    html = render_budget_executive_dashboard(
        [
            {
                "id": "line-1",
                "concept_name": "Hospedaje",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={"line-1": {1: {"budget_expense_amount": 1000.0}}},
        actuals_map={
            "concept-1": {
                1: {"real_expense_cash": 800.0, "committed_unpaid": 100.0},
            }
        },
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-1",
        budget_view="expenses",
        budget_period="weekly",
    )

    assert "En observación" in html
    assert "90.0%" in html


def test_render_budget_executive_dashboard_status_flags_over_budget():
    from devnous.gastos.routes.admin_budget_ui import render_budget_executive_dashboard

    html = render_budget_executive_dashboard(
        [
            {
                "id": "line-1",
                "concept_name": "Traslados",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={"line-1": {1: {"budget_expense_amount": 1000.0}}},
        actuals_map={
            "concept-1": {
                1: {"real_expense_cash": 1000.0, "committed_unpaid": 50.0},
            }
        },
        tournament_key="copatest",
        edition_year=2026,
        version_id="version-1",
        budget_view="expenses",
        budget_period="weekly",
    )

    assert "Excedido" in html
    assert "105.0%" in html


def test_render_tournament_dashboard_cards_show_controlled_executive_state():
    from devnous.gastos.routes.admin_budget_ui import render_tournament_dashboard_cards

    html = render_tournament_dashboard_cards(
        [
            {
                "tournament_id": "torneo-1",
                "tournament_name": "Copa Prueba",
                "tournament_code": "CP",
                "line_count": 4,
                "comparison": {"paid_total": 200.0, "committed_total": 100.0},
            }
        ],
        edition_year=2026,
        version_id="version-1",
        tournament_rollups={
            "torneo-1": {
                "budget_expense_total": 1000.0,
                "expected_income_total": 500.0,
                "real_income_total": 250.0,
            }
        },
    )

    assert "Estado ejecutivo" in html
    assert "Controlado" in html
    assert "30.0%" in html
    assert "Presupuesto autorizado" in html
    assert "Ejercido real" in html
    assert "Comprometido pendiente" in html
    assert "Abrir detalle" in html
    assert "/admin/presupuestos/torneo/torneo-1" in html
    assert "read-only" not in html


def test_render_tournament_dashboard_cards_show_observation_state():
    from devnous.gastos.routes.admin_budget_ui import render_tournament_dashboard_cards

    html = render_tournament_dashboard_cards(
        [
            {
                "tournament_id": "torneo-1",
                "tournament_name": "Copa Prueba",
                "comparison": {"paid_total": 800.0, "committed_total": 100.0},
            }
        ],
        edition_year=2026,
        version_id="version-1",
        tournament_rollups={"torneo-1": {"budget_expense_total": 1000.0}},
    )

    assert "En observación" in html
    assert "90.0%" in html


def test_render_tournament_dashboard_cards_show_exceeded_state():
    from devnous.gastos.routes.admin_budget_ui import render_tournament_dashboard_cards

    html = render_tournament_dashboard_cards(
        [
            {
                "tournament_id": "torneo-1",
                "tournament_name": "Copa Prueba",
                "comparison": {"paid_total": 1000.0, "committed_total": 50.0},
            }
        ],
        edition_year=2026,
        version_id="version-1",
        tournament_rollups={"torneo-1": {"budget_expense_total": 1000.0}},
    )

    assert "Excedido" in html
    assert "105.0%" in html


def test_render_tournament_dashboard_cards_prefers_detail_rollup_for_real_spend():
    from devnous.gastos.routes.admin_budget_ui import render_tournament_dashboard_cards

    html = render_tournament_dashboard_cards(
        [
            {
                "tournament_id": "torneo-1",
                "tournament_name": "Copa Prueba",
                "comparison": {
                    "actual_total": 250.0,
                    "paid_total": 200.0,
                    "committed_total": 100.0,
                },
            }
        ],
        edition_year=2026,
        version_id="version-1",
        tournament_rollups={
            "torneo-1": {
                "budget_expense_total": 1000.0,
                "real_expense_total": 75.0,
                "committed_pending_total": 25.0,
            }
        },
    )

    assert "Ejercido real" in html
    assert "$75.00" in html
    assert "$25.00" in html
    assert "10.0%" in html


def test_render_tournament_dashboard_cards_falls_back_to_snapshot_actual_total():
    from devnous.gastos.routes.admin_budget_ui import render_tournament_dashboard_cards

    html = render_tournament_dashboard_cards(
        [
            {
                "tournament_id": "torneo-1",
                "tournament_name": "Copa Prueba",
                "comparison": {
                    "actual_total": 250.0,
                    "paid_total": 200.0,
                    "committed_total": 100.0,
                },
            }
        ],
        edition_year=2026,
        version_id="version-1",
        tournament_rollups={"torneo-1": {"budget_expense_total": 1000.0}},
    )

    assert "Ejercido real" in html
    assert "$250.00" in html
    assert "$100.00" in html
    assert "35.0%" in html


def test_render_budget_partida_matrix_includes_editable_cuenta_search():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [
            {
                "id": "line-1",
                "concept_name": "Aguinaldo",
                "budget_amount": 1000,
                "phase": "Estatal",
                "budget_concept_id": "concept-1",
                "cuenta_contable_id": "cuenta-1",
                "cuenta_contable_codigo": "5300-001-006",
                "cuenta_contable_nombre": "AGUINALDO",
            }
        ],
        plan_map={"line-1": {1: {"budget_expense_amount": 100.0}}},
        actuals_map={"concept-1": {}},
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        cuentas_contables=[
            {
                "id": "cuenta-1",
                "codigo": "5300-001-006",
                "nombre": "AGUINALDO",
                "tipo": "Gasto",
            }
        ],
    )

    assert "matrix-cuenta-search" in html
    assert 'name="cuenta_contable_id"' in html
    assert 'name="account_code_final"' in html
    assert "5300-001-006 - AGUINALDO" in html
    assert "Guardar plan por semanas" in html
    assert "Distribuir total" in html
    assert 'name="budget_amount"' in html
    assert 'data-budget-week-kind="expense"' in html
    assert "Semana 52" in html
    assert 'name="month_52_expense"' in html


def test_render_budget_partida_matrix_does_not_fallback_to_unassigned_for_concept():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [
            {
                "id": "line-1",
                "concept_name": "Hospedaje",
                "budget_amount": 1000,
                "phase": "Estatal",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={"line-1": {1: {"budget_expense_amount": 100.0}}},
        actuals_map={
            "__unassigned__": {
                1: {"real_expense_cash": 50.0, "committed_unpaid": 20.0}
            }
        },
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        matrix_mode="expenses",
    )

    assert "Sin partida asignada" in html
    assert "Gasto real: <strong>$0.00</strong>" in html
    assert html.count("$50.00") == 2
    assert html.count("$20.00") == 2


def test_render_budget_partida_matrix_monthly_view_is_aggregated_readonly():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [
            {
                "id": "line-1",
                "concept_name": "Hospedaje",
                "budget_amount": 1000,
                "phase": "Estatal",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={"line-1": {1: {"budget_expense_amount": 100.0}}},
        actuals_map={
            "concept-1": {
                1: {"real_expense_cash": 25.0, "committed_unpaid": 10.0}
            }
        },
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        matrix_mode="expenses",
        budget_period="monthly",
    )

    assert "Vista agregada por periodo" in html
    assert "Enero 2026" in html
    assert "Semana 52" not in html
    assert 'name="month_52_expense"' not in html
    assert "Distribuir total" not in html


def test_budget_detail_route_passes_period_to_partida_matrix():
    source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()
    start = source.index("gastos_matrix_html = render_budget_partida_matrix(")
    end = source.index("active_visible_count = (", start)
    block = source[start:end]

    assert block.count("budget_period=budget_period") == 2


def test_budget_dashboard_route_summarizes_actuals_for_cards():
    source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()

    assert "summarize_budget_actuals_for_lines" in source
    assert '"real_income_total": round(real_income, 2)' in source
    assert 'line_direction="expense"' in source


def test_render_budget_partida_matrix_expenses_mode_hides_income_rows():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [
            {
                "id": "line-1",
                "concept_name": "Hospedaje",
                "budget_amount": 1200,
                "phase": "Estatal",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={
            "line-1": {
                1: {
                    "budget_expense_amount": 100.0,
                    "expected_income_amount": 50.0,
                }
            }
        },
        actuals_map={
            "concept-1": {
                1: {
                    "real_expense_cash": 25.0,
                    "real_income": 15.0,
                }
            }
        },
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        matrix_mode="expenses",
    )

    assert "Presupuesto gasto" in html
    assert "Gasto real (caja)" in html
    assert "Ingreso esperado" not in html
    assert "Ingreso real" not in html
    assert "Guardar gasto por semanas" in html


def test_render_budget_partida_matrix_income_mode_hides_expense_rows():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [
            {
                "id": "line-1",
                "concept_name": "Patrocinio",
                "budget_amount": 1200,
                "phase": "Estatal",
                "budget_concept_id": "concept-1",
            }
        ],
        plan_map={
            "line-1": {
                1: {
                    "budget_expense_amount": 100.0,
                    "expected_income_amount": 50.0,
                }
            }
        },
        actuals_map={
            "concept-1": {
                1: {
                    "real_expense_cash": 25.0,
                    "real_income": 15.0,
                }
            }
        },
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        matrix_mode="income",
    )

    assert "Ingreso esperado" in html
    assert "Ingreso real" in html
    assert 'name="month_1_income"' in html
    assert "Presupuesto gasto" not in html
    assert "Gasto real (caja)" not in html
    assert "Comprometido no pagado" not in html
    assert "Guardar ingreso por semanas" in html


def test_render_budget_partida_matrix_income_mode_empty_state_is_income_specific():
    from devnous.gastos.routes.admin_budget_ui import render_budget_partida_matrix

    html = render_budget_partida_matrix(
        [],
        plan_map={},
        actuals_map={},
        version_id="version-1",
        tournament_key="copatest",
        can_edit=True,
        matrix_mode="income",
    )

    assert "partidas presupuestales de ingresos" in html


def test_render_budget_detail_section_nav_has_gastos_ingresos_controls():
    from devnous.gastos.routes.admin_budget_ui import render_budget_detail_section_nav

    html = render_budget_detail_section_nav()

    assert 'href="#presupuesto-gastos"' in html
    assert 'href="#presupuesto-ingresos"' in html
    assert ">Gastos<" in html
    assert ">Ingresos<" in html


def test_tournament_budget_detail_defaults_to_current_year():
    route_source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()

    assert "resolved_year = edition_year or date.today().year" in route_source
    assert (
        'int(all_versions[0]["edition_year"])'
        not in route_source[
            route_source.index(
                "async def admin_presupuestos_tournament_detail"
            ) : route_source.index(
                '@router.get("/admin/presupuestos/torneo/{tournament_key}/cfdi-ingresos")'
            )
        ]
    )


def test_tournament_budget_detail_splits_lines_by_direction():
    route_source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()
    detail_source = route_source[
        route_source.index(
            "async def admin_presupuestos_tournament_detail"
        ) : route_source.index(
            '@router.get("/admin/presupuestos/torneo/{tournament_key}/cfdi-ingresos")'
        )
    ]

    assert 'line_direction="expense"' in detail_source
    assert 'line_direction="income"' in detail_source
    assert "lines=income_lines" in detail_source
    assert "create_income_line_form" in detail_source
    assert "show_phase_field=False" in detail_source
    assert "cuentas_contables=cuentas_contables" in detail_source


def test_tournament_budget_detail_preserves_income_expense_view():
    route_source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()
    ui_source = Path("src/devnous/gastos/routes/admin_budget_ui.py").read_text()
    detail_source = route_source[
        route_source.index(
            "async def admin_presupuestos_tournament_detail"
        ) : route_source.index(
            '@router.post("/admin/presupuestos/torneo/{tournament_key}/ingresos/import")'
        )
    ]

    assert 'budget_view: Optional[str] = Query("expenses")' in detail_source
    assert 'selected_budget_view = (' in detail_source
    assert "active_budget_section_html" in detail_source
    assert "budget_view=selected_budget_view" in detail_source
    assert "budget_view=\"income\"" in route_source
    assert 'name="budget_view"' in ui_source
    assert "budget_view: Optional[str] = Form(None)" in route_source


def test_budget_and_accounting_surfaces_link_to_sat_cfdi_workflows():
    budget_source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()
    accounting_source = Path("src/devnous/gastos/routes/admin_routes.py").read_text()

    dashboard_source = budget_source[
        budget_source.index("async def admin_presupuestos_dashboard") : budget_source.index(
            "async def admin_presupuestos_tournament_detail"
        )
    ]
    detail_source = budget_source[
        budget_source.index("async def admin_presupuestos_tournament_detail") : budget_source.index(
            '@router.post("/admin/presupuestos/torneo/{tournament_key}/ingresos/import")'
        )
    ]
    cleanup_source = accounting_source[
        accounting_source.index("async def gastos_sin_cuenta_contable") : accounting_source.index(
            '@router.post("/admin/gastos/{gasto_id}/cleanup-contable")'
        )
    ]

    for source in (dashboard_source, detail_source, cleanup_source):
        assert "/admin/gastos/sat" in source
        assert "/admin/gastos/cfdis/matching" in source

    assert "SAT / CFDI" in dashboard_source
    assert "SAT / CFDI" in detail_source
    assert "SAT / CFDI" in cleanup_source


def test_create_budget_line_route_creates_concept_with_direction_and_account():
    source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()
    route_source = source[
        source.index(
            '@router.post("/admin/presupuestos/versiones/{version_id}/lineas/create")'
        ) : source.index('@router.post("/admin/presupuestos/lineas/{line_id}/update")')
    ]

    assert "create_budget_concept" in route_source
    assert "tournament_id: Optional[str] = Form(None)" in route_source
    assert "cuenta_contable_id: Optional[str] = Form(None)" in route_source
    assert "line_direction: Optional[str] = Form(None)" in route_source
    assert "budget_direction=line_direction" in route_source
    assert "cuenta_contable_id=cuenta_contable_id" in route_source
    assert "line_direction=line_direction" in route_source


def test_budget_catalog_editor_supports_pasivo_account():
    source = Path("src/devnous/gastos/routes/admin_routes.py").read_text()
    active_route_source = Path(
        "src/devnous/gastos/routes/admin_budget_routes.py"
    ).read_text()

    assert "Contracuenta presupuestal" in active_route_source
    assert "Tipo" in active_route_source
    assert "list_budget_concepts" in active_route_source
    assert "/admin/presupuestos/conceptos/bulk-save" in active_route_source
    assert "budget_directions: List[str] = Form([])" in source
    assert 'name="budget_directions"' in active_route_source
    assert "pasivo_cuenta_contable_ids: List[str] = Form([])" in source
    assert 'field_name="pasivo_cuenta_contable_ids"' in active_route_source
    assert 'field_name="pasivo_cuenta_contable_ids"' in source
    assert "DEFAULT_BUDGET_CONCEPT_PASIVO_ACCOUNT_CODE" in source
    assert "DEFAULT_BUDGET_CONCEPT_PASIVO_ACCOUNT_CODE" in active_route_source


def test_budget_catalog_import_export_routes_are_wired():
    source = Path("src/devnous/gastos/routes/admin_routes.py").read_text()
    active_route_source = Path(
        "src/devnous/gastos/routes/admin_budget_routes.py"
    ).read_text()

    assert "/admin/presupuestos/conceptos/export.xlsx" in active_route_source
    assert "/admin/presupuestos/conceptos/import" in active_route_source
    assert "archivo_catalogo" in active_route_source
    assert 'access.get("export")' in active_route_source
    assert 'access.get("line_update")' in active_route_source

    assert '@router.get("/admin/presupuestos/conceptos/export.xlsx")' in source
    assert '_require_budget_access(current_empleado, "export")' in source
    assert "generate_budget_concepts_catalog_xlsx" in source
    assert '@router.post("/admin/presupuestos/conceptos/import")' in source
    assert "archivo_catalogo: UploadFile = File(...)" in source
    assert '_require_budget_access(current_empleado, "line_update")' in source
    assert "import_budget_concepts_upload" in source
    assert "await session.rollback()" in source


def test_budget_catalog_tournament_filter_is_wired():
    source = Path("src/devnous/gastos/routes/admin_routes.py").read_text()
    active_route_source = Path(
        "src/devnous/gastos/routes/admin_budget_routes.py"
    ).read_text()

    assert 'catalog_scope: Optional[str] = Query("none")' in active_route_source
    assert "catalog_tournament_ids: list[str] = Query([])" in active_route_source
    assert "selected_catalog_tournament_ids" in active_route_source
    assert "catalog_tournament_ids" in active_route_source
    assert "Aplicar filtro" in active_route_source
    assert "Todos" in active_route_source
    assert "Ninguno" in active_route_source
    assert "Selecciona torneos para cargar partidas" in active_route_source
    assert "partidas cargadas de" in active_route_source
    assert "selected_tournament_aliases" in active_route_source
    assert "tournaments = [" in active_route_source

    assert "catalog_scope: Optional[str] = None" in active_route_source
    assert "catalog_tournament_ids: Optional[list[str]] = None" in active_route_source
    assert "catalog_tournament_ids: List[str] = Form([])" in source
    assert "catalog_scope={quote(str(catalog_scope))}" in active_route_source
    assert "catalog_tournament_ids={quote(str(tournament_id))}" in active_route_source


def test_budget_catalog_upload_supports_direction_counterpart_and_active_fields():
    service_source = Path("src/samchat/budgets/service.py").read_text()
    import_source = service_source[
        service_source.index(
            "async def import_budget_concepts_upload("
        ) : service_source.index("async def import_budget_concepts_catalog(")
    ]

    assert "generate_budget_concepts_catalog_xlsx" in service_source
    assert "import_budget_concepts_upload" in service_source
    assert '"tipo"' in service_source
    assert '"cuenta_presupuestal"' in service_source
    assert '"contracuenta_presupuestal"' in service_source
    assert '"cuenta_pasivo"' in service_source
    assert '"activo"' in service_source
    assert '"cuenta_contable_ingresos"' in service_source
    assert '"cuenta_contable_cuenta_por_cobrar"' in service_source
    assert "resolve_active_cuenta_contable_id_by_code" in service_source
    assert "_budget_catalog_active_value" in service_source
    assert "_budget_catalog_direction_value" in service_source
    assert "required_headers" in service_source
    assert "El archivo no contiene las columnas requeridas" in service_source
    assert '"pasivo_cuenta_contable_id": pasivo_id' in service_source
    assert '"budget_direction": budget_direction' in service_source
    assert '"active": active' in service_source
    assert import_source.index("existing = concepts_by_id.get") < import_source.index(
        "_match_tournament_id"
    )
    assert (
        'tournament_id = _safe_str((existing or {}).get("tournament_id"))'
        in import_source
    )


def test_budget_schema_adds_pasivo_account_and_default_backfill():
    service_source = Path("src/samchat/budgets/service.py").read_text()
    guard_source = Path("src/devnous/gastos/schema_guard.py").read_text()

    assert "pasivo_cuenta_contable_id UUID NULL" in service_source
    assert "ix_budget_concepts_pasivo_cuenta_contable_id" in service_source
    assert "2120-002-099" in service_source
    assert (
        'RequiredColumn("budget_concepts", "pasivo_cuenta_contable_id")' in guard_source
    )
    assert "budget_concepts_pasivo_default_backfill" in guard_source


def test_create_budget_concept_accepts_budget_direction():
    source = Path("src/samchat/budgets/service.py").read_text()
    function_source = source[
        source.index("async def create_budget_concept(") : source.index(
            "async def update_budget_concept("
        )
    ]

    assert "budget_direction: Optional[str] = None" in function_source
    assert (
        "clean_direction = normalize_budget_line_direction(budget_direction)"
        in function_source
    )
    assert "budget_direction, active, source" in function_source
    assert '"budget_direction": clean_direction' in function_source


def test_income_budget_import_uses_income_direction_and_expected_months():
    source = Path("src/samchat/budgets/service.py").read_text()
    import_source = source[
        source.index("async def import_budget_lines_upload(") : source.index(
            "def load_budget_artifact_rows("
        )
    ]

    assert "line_direction: Optional[str] = None" in import_source
    assert "default_direction" in import_source
    assert (
        "clean_direction = default_direction or normalize_budget_line_direction"
        in import_source
    )
    assert '"expected_income_amount"' in import_source
    assert '"budget_expense_amount"' in import_source
    assert "line_direction=clean_direction" in import_source
    assert "replace_budget_line_monthly_plan" in import_source
    assert "concepts_by_scope" in import_source
    assert "clean_direction" in import_source


def test_active_budget_income_import_export_routes_are_wired():
    source = Path("src/devnous/gastos/routes/admin_budget_routes.py").read_text()

    assert (
        '@router.post("/admin/presupuestos/torneo/{tournament_key}/ingresos/import")'
        in source
    )
    assert (
        '@router.get("/admin/presupuestos/torneo/{tournament_key}/ingresos/export.xlsx")'
        in source
    )
    assert 'line_direction="income"' in source
    assert "generate_budget_income_xlsx" in source
    assert "list_psp_cfdi_income_candidates" in source
    assert "list_budget_cfdi_income_links" in source
    assert "Descargar ingresos" in source
    assert "Importar ingresos" in source
    assert 'id="presupuesto-ingresos"' in source


def test_generate_budget_income_xlsx_contains_expected_real_and_pending_cfdi():
    from io import BytesIO

    from openpyxl import load_workbook

    from samchat.budgets.exporter import generate_budget_income_xlsx

    payload = generate_budget_income_xlsx(
        lines=[
            {
                "id": "line-1",
                "budget_concept_id": "concept-1",
                "tournament_code": "TOR",
                "tournament_name": "Torneo",
                "phase": "Regional",
                "concept_name": "Patrocinio",
                "account_code_final": "4100-001",
                "budget_amount": 1200,
            }
        ],
        plan_map={
            "line-1": {
                1: {"expected_income_amount": 100},
                2: {"expected_income_amount": 100},
            }
        },
        actuals_map={
            "concept-1": {
                1: {"real_income": 80},
                2: {"real_income": 150},
            }
        },
        links=[
            {
                "cfdi_uuid": "UUID-1",
                "emisor_rfc": "PSP010101AAA",
                "emisor_nombre": "PSP",
                "receptor_rfc": "SAM010101AAA",
                "concept_name": "Patrocinio",
                "phase": "Regional",
                "amount": 230,
                "income_date": "2026-02-15",
                "unlinked_at": None,
            }
        ],
        candidates=[
            {
                "cfdi_uuid": "UUID-PENDING",
                "fecha": "2026-03-01",
                "emisor_rfc": "PSP010101AAA",
                "emisor_nombre": "PSP",
                "receptor_rfc": "SAM010101AAA",
                "total": 500,
            }
        ],
        selected_version={
            "version_name": "Presupuesto operativo 2026",
            "status": "draft",
        },
        tournament_context={"tournament_name": "Torneo", "tournament_code": "TOR"},
        edition_year=2026,
    )

    workbook = load_workbook(BytesIO(payload), data_only=True)
    assert workbook.sheetnames == [
        "Ingresos",
        "Mensual",
        "CFDI vinculados",
        "CFDI sin clasificar",
    ]
    income_sheet = workbook["Ingresos"]
    assert income_sheet["D9"].value == "Patrocinio"
    assert income_sheet["G9"].value == 200
    assert income_sheet["H9"].value == 230
    assert income_sheet["I9"].value == 30
    assert income_sheet["J9"].value == 115
    linked_sheet = workbook["CFDI vinculados"]
    assert linked_sheet["A2"].value == "UUID-1"
    pending_sheet = workbook["CFDI sin clasificar"]
    assert pending_sheet["A2"].value == "UUID-PENDING"
