from __future__ import annotations

import csv
import io
import json
import zipfile
from typing import Any, Dict, List, Optional, Sequence
from xml.etree import ElementTree as ET

from fastapi import HTTPException

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover
    fitz = None


def _normalize_col_name(value: str) -> str:
    normalized = "".join(
        ch if ch.isalnum() else "_" for ch in (value or "").strip().lower()
    )
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def dataframe_records(df: Any) -> List[Dict[str, Any]]:
    try:
        filled = df.fillna("")
        return filled.to_dict(orient="records")
    except (AttributeError, TypeError, ValueError):
        return []


def decode_bytes_text(
    raw: bytes, *, encodings: Sequence[str] = ("utf-8", "latin-1")
) -> Optional[str]:
    for encoding in encodings:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def decode_spreadsheet_csv(raw: bytes) -> List[Dict[str, Any]]:
    text = decode_bytes_text(raw, encodings=("utf-8-sig", "utf-8", "latin-1"))
    if text is None:
        raise HTTPException(
            status_code=400,
            detail="No se pudo leer el CSV. Usa UTF-8, UTF-8 BOM o Latin-1.",
        )
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="El CSV no contiene encabezados.")
    return [dict(row) for row in reader]


def decode_spreadsheet_xlsx(raw: bytes) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise HTTPException(
            status_code=500, detail="XLSX parser not available in backend"
        )
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"No se pudo leer el archivo XLSX: {exc}"
        ) from exc

    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = None
    records: List[Dict[str, Any]] = []
    for raw_row in rows_iter:
        values = list(raw_row or [])
        if headers is None:
            if not any(str(cell or "").strip() for cell in values):
                continue
            headers = [str(cell or "").strip() for cell in values]
            continue
        if not any(str(cell or "").strip() for cell in values):
            continue
        row_dict: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = values[idx] if idx < len(values) else ""
        if row_dict:
            records.append(row_dict)
    return records


def spreadsheet_records_from_bytes(
    *, raw: bytes, filename: str, content_type: str
) -> List[Dict[str, Any]]:
    fname = (filename or "").lower()
    buffer = io.BytesIO(raw)
    try:
        if fname.endswith(".csv") or "text/csv" in (content_type or "").lower():
            if pd is not None:
                df = pd.read_csv(buffer)
                return dataframe_records(df)
            return decode_spreadsheet_csv(raw)
        if fname.endswith(".xls"):
            raise HTTPException(
                status_code=400,
                detail=(
                    "Archivos .xls no soportados. "
                    "Convierte el archivo a .xlsx o CSV."
                ),
            )
        if pd is not None:
            df = pd.read_excel(buffer)
            return dataframe_records(df)
        return decode_spreadsheet_xlsx(raw)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"No se pudo leer el archivo tabular: {exc}"
        ) from exc


def normalize_spreadsheet_records(
    records: List[Dict[str, Any]],
) -> List[Dict[str, str]]:
    normalized_records: List[Dict[str, str]] = []
    for row in records:
        normalized_row: Dict[str, str] = {}
        for key, value in (row or {}).items():
            normalized_key = _normalize_col_name(str(key))
            if not normalized_key:
                continue
            normalized_row[normalized_key] = str(value or "").strip()
        if normalized_row:
            normalized_records.append(normalized_row)
    return normalized_records


def spreadsheet_looks_like_roster(records: List[Dict[str, Any]]) -> bool:
    normalized_records = normalize_spreadsheet_records(records[:5])
    keyset = set()
    for row in normalized_records:
        keyset.update(row.keys())
    roster_markers = {
        "nombre",
        "first_name",
        "apellido",
        "last_name",
        "fecha_nacimiento",
        "birth_date",
        "equipo",
        "team_name",
        "categoria",
        "category_name",
    }
    return len(keyset.intersection(roster_markers)) >= 2


def spreadsheet_looks_like_balance(records: List[Dict[str, Any]]) -> bool:
    normalized_records = normalize_spreadsheet_records(records[:5])
    keyset = set()
    for row in normalized_records:
        keyset.update(row.keys())
    if "cuenta" not in keyset:
        return False
    balance_markers = {
        "descripcion_de_la_cuenta",
        "descripcin_de_la_cuenta",
        "saldo_inicial",
        "total_de_cargos",
        "total_de_abonos",
        "saldo_final",
    }
    return len(keyset.intersection(balance_markers)) >= 2


def spreadsheet_preview_text(
    *, records: List[Dict[str, Any]], filename: str, note: Optional[str]
) -> str:
    normalized_records = normalize_spreadsheet_records(records)
    headers = list(normalized_records[0].keys()) if normalized_records else []
    preview_rows = normalized_records[: min(len(normalized_records), 20)]
    kind = (
        "balanza/contabilidad" if spreadsheet_looks_like_balance(records) else "tabular"
    )
    note_block = f"\nNota del usuario: {note.strip()}" if (note or "").strip() else ""
    guidance = (
        "Usa esta tabla como contexto del usuario. "
        "Si parece balanza o reporte contable, "
        "contesta con base en sus filas y no lo trates como roster."
    )
    return (
        f"Entrada spreadsheet procesada para analisis {kind}.\n"
        f"Archivo: {filename or 'spreadsheet'}\n"
        f"Rows parsed: {len(normalized_records)}\n"
        f"Headers: {', '.join(headers[:20])}\n"
        f"Preview JSON:\n{json.dumps(preview_rows, ensure_ascii=False)}\n"
        f"{note_block}\n\n"
        f"{guidance}"
    )


def extract_document_text_from_bytes(
    *,
    raw: bytes,
    filename: Optional[str],
    mime_type: Optional[str],
    max_bytes: Optional[int] = None,
    allow_pdf: bool = False,
    allow_spreadsheet: bool = False,
) -> str:
    if not raw:
        return ""
    if max_bytes is not None and len(raw) > max_bytes:
        raise HTTPException(
            status_code=400, detail=f"El archivo excede {max_bytes // (1024 * 1024)}MB"
        )

    name = str(filename or "").lower()
    mime = str(mime_type or "").lower()

    if allow_pdf and (name.endswith(".pdf") or mime == "application/pdf"):
        if fitz is None:
            raise HTTPException(
                status_code=400,
                detail="Este servidor no tiene extractor PDF disponible.",
            )
        try:
            with fitz.open(stream=raw, filetype="pdf") as doc:
                pages = []
                for page_index in range(min(len(doc), 40)):
                    page = doc[page_index]
                    pages.append(page.get_text("text") or "")
            return "\n\n".join(pages).strip()
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"No se pudo extraer texto del PDF: {exc}"
            ) from exc

    if name.endswith(".docx") or "wordprocessingml.document" in mime:
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                xml_bytes = zf.read("word/document.xml")
            root = ET.fromstring(xml_bytes)
            nodes = root.findall(
                ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
            )
            return "\n".join(
                (node.text or "").strip() for node in nodes if (node.text or "").strip()
            ).strip()
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"No se pudo leer DOCX: {exc}"
            ) from exc

    if allow_spreadsheet and (
        name.endswith((".xlsx", ".xlsm")) or "spreadsheetml.sheet" in mime
    ):
        if load_workbook is None:
            raise HTTPException(
                status_code=400,
                detail="Este servidor no tiene extractor XLSX disponible.",
            )
        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            lines: List[str] = []
            for sheet_name in workbook.sheetnames[:8]:
                sheet = workbook[sheet_name]
                lines.append(f"Hoja: {sheet_name}")
                for row in sheet.iter_rows(max_row=80, max_col=30, values_only=True):
                    values = [str(value).strip() for value in row if value is not None]
                    if values:
                        lines.append(" | ".join(values))
            return "\n".join(lines).strip()
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"No se pudo leer XLSX: {exc}"
            ) from exc

    return (decode_bytes_text(raw) or "").strip()
