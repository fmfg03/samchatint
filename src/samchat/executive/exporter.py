"""Excel export for the unified executive center."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def _safe_float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    start_row: int = 1,
) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, _header in enumerate(headers, start=1):
        max_len = max(
            len(_safe_text(ws.cell(row=row, column=col_idx).value))
            for row in range(start_row, start_row + len(rows) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 12),
            55,
        )
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    return start_row + len(rows) + 2


def _summary_metric_rows(
    *,
    finance_platform: dict[str, Any],
    budget_snapshot: dict[str, Any],
    ar_payload: dict[str, Any],
    alerts: list[dict[str, Any]],
) -> list[list[Any]]:
    finance_summary = finance_platform.get("summary") or {}
    budget_summary = budget_snapshot.get("summary") or {}
    budget_forecast = budget_snapshot.get("forecast") or {}
    ar_summary = ar_payload.get("summary") or {}
    high_alerts = sum(1 for item in alerts if item.get("severity") == "high")
    medium_alerts = sum(
        1 for item in alerts if item.get("severity") == "medium"
    )
    return [
        ["Alertas ejecutivas", len(alerts)],
        ["Alertas alta prioridad", high_alerts],
        ["Alertas media prioridad", medium_alerts],
        [
            "Acciones financieras abiertas",
            _safe_int(finance_summary.get("open_actions")),
        ],
        ["Presupuesto total", _safe_float(budget_summary.get("budget_total"))],
        ["Comprometido", _safe_float(budget_summary.get("committed_total"))],
        ["Pagado", _safe_float(budget_summary.get("paid_total"))],
        [
            "Cierre proyectado",
            _safe_float(budget_forecast.get("projected_close_total")),
        ],
        [
            "Ingreso presupuestado",
            _safe_float(ar_summary.get("expected_income_total")),
        ],
        ["Facturado CxC", _safe_float(ar_summary.get("invoiced_total"))],
        [
            "Cobrado comprobado CxC",
            _safe_float(ar_summary.get("collected_total")),
        ],
        ["Saldo pendiente CxC", _safe_float(ar_summary.get("balance_total"))],
    ]


def _budget_rows(snapshot: dict[str, Any]) -> list[list[Any]]:
    summary = snapshot.get("summary") or {}
    forecast = snapshot.get("forecast") or {}
    rows = [
        [
            "Presupuesto",
            _safe_float(summary.get("budget_total")),
            "Bolsa aprobada",
        ],
        [
            "Solicitado",
            _safe_float(summary.get("requested_total")),
            "Solicitudes",
        ],
        [
            "Comprometido",
            _safe_float(summary.get("committed_total")),
            "Compromisos",
        ],
        [
            "Pagado",
            _safe_float(summary.get("paid_total")),
            "Salida ejecutada",
        ],
        ["Real", _safe_float(summary.get("actual_total")), "Gasto observado"],
        [
            "Pendiente por pagar",
            _safe_float(summary.get("pending_to_pay_total")),
            "Presión operativa",
        ],
        [
            "Necesidad de caja",
            _safe_float(forecast.get("projected_cash_need")),
            "Forecast",
        ],
    ]
    for item in list(snapshot.get("executive_comparison") or [])[:12]:
        rows.append(
            [
                _safe_text(item.get("label")),
                _safe_float(item.get("total")),
                _safe_text(item.get("detail")),
            ]
        )
    return rows


def _cashflow_rows(platform: dict[str, Any]) -> list[list[Any]]:
    cash = platform.get("cash_control_center") or {}
    payment = platform.get("payment_run") or {}
    accounting = platform.get("accounting_close_center") or {}
    tax = platform.get("tax_readiness") or {}
    return [
        [
            "Pagos pendientes",
            _safe_float(payment.get("payable_total")),
            _safe_int(payment.get("payable_count")),
        ],
        [
            "Pagado",
            _safe_float(cash.get("paid_total")),
            _safe_int(cash.get("paid_documents_count")),
        ],
        [
            "Ingresos contabilizados",
            _safe_float(cash.get("income_total")),
            _safe_int(cash.get("income_polizas_count")),
        ],
        [
            "COI pendiente",
            _safe_int(accounting.get("pending_coi_expenses_count")),
            "gastos",
        ],
        [
            "Pólizas descuadradas",
            _safe_int(accounting.get("unbalanced_count")),
            "pólizas",
        ],
        [
            "DIOT/CFDI bloqueado",
            _safe_int(tax.get("diot_blockers_count")),
            _safe_text(tax.get("status")),
        ],
    ]


def _ar_rows(
    ar_payload: dict[str, Any],
    ar_rows: list[dict[str, Any]],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    summary = ar_payload.get("summary") or {}
    rows.append(
        [
            "Resumen",
            "Ingreso presupuestado",
            _safe_float(summary.get("expected_income_total")),
            "",
        ]
    )
    rows.append(
        [
            "Resumen",
            "Facturado",
            _safe_float(summary.get("invoiced_total")),
            "",
        ]
    )
    rows.append(
        [
            "Resumen",
            "Cobrado comprobado",
            _safe_float(summary.get("collected_total")),
            "",
        ]
    )
    rows.append(
        [
            "Resumen",
            "Saldo pendiente",
            _safe_float(summary.get("balance_total")),
            "",
        ]
    )
    for item in ar_rows[:100]:
        rows.append(
            [
                _safe_text(item.get("operational_status")),
                _safe_text(
                    item.get("payer_name") or item.get("tournament_name")
                ),
                _safe_float(
                    item.get("balance_amount") or item.get("issued_amount")
                ),
                _safe_text(item.get("ar_item_id")),
            ]
        )
    return rows


def _alert_rows(alerts: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            _safe_text(item.get("severity")),
            _safe_text(item.get("module")),
            _safe_text(item.get("title")),
            _safe_text(item.get("detail")),
            _safe_text(item.get("owner")),
            _safe_text(item.get("source")),
            _safe_text(item.get("href")),
        ]
        for item in alerts[:100]
    ]


def generate_executive_export_xlsx(
    *,
    finance_platform: dict[str, Any],
    budget_snapshot: dict[str, Any],
    ar_payload: dict[str, Any],
    ar_rows: list[dict[str, Any]],
    alerts: list[dict[str, Any]],
    source_notes: list[str] | None = None,
) -> bytes:
    """Build a read-only workbook for executive review."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen ejecutivo"
    generated_at = (
        datetime.now(UTC)
        .isoformat(timespec="seconds")
        .replace("+00:00", "Z")
    )
    period = finance_platform.get("period") or {}

    ws["A1"] = "Export ejecutivo SamChat"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = (
        f"Periodo: {period.get('month') or ''}/{period.get('year') or ''}"
    )
    ws["A3"] = f"Generado: {generated_at}"
    _write_table(
        ws,
        ["Métrica", "Valor"],
        _summary_metric_rows(
            finance_platform=finance_platform,
            budget_snapshot=budget_snapshot,
            ar_payload=ar_payload,
            alerts=alerts,
        ),
        start_row=5,
    )

    ws_budget = wb.create_sheet("Presupuesto")
    _write_table(
        ws_budget,
        ["Métrica", "Valor", "Detalle"],
        _budget_rows(budget_snapshot),
    )

    ws_cash = wb.create_sheet("Flujo y pagos")
    _write_table(
        ws_cash,
        ["Métrica", "Valor", "Conteo/estado"],
        _cashflow_rows(finance_platform),
    )

    ws_ar = wb.create_sheet("Cuentas por cobrar")
    _write_table(
        ws_ar,
        ["Estado", "Cliente/proyecto", "Monto", "Referencia"],
        _ar_rows(ar_payload, ar_rows),
    )

    ws_alerts = wb.create_sheet("Alertas")
    _write_table(
        ws_alerts,
        [
            "Severidad",
            "Módulo",
            "Alerta",
            "Detalle",
            "Responsable",
            "Fuente",
            "Ruta",
        ],
        _alert_rows(alerts),
    )

    ws_sources = wb.create_sheet("Fuentes y limites")
    ws_sources["A1"] = "Fuentes y límites"
    ws_sources["A1"].font = _BOLD_FONT
    _write_table(
        ws_sources,
        ["Fuente", "Estado"],
        [
            ["Finance Platform", "snapshot read-only"],
            ["Presupuestos", "snapshot read-only"],
            ["Cuentas por cobrar", "read model AR"],
            ["Alertas ejecutivas", "consolidación read-only"],
            ["Límite", "No ejecuta pagos, no crea pólizas, no cambia saldos."],
            *[["Aviso", note] for note in source_notes or []],
        ],
        start_row=3,
    )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
