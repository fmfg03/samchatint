from __future__ import annotations

import csv
import json
import uuid
from functools import lru_cache
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Optional
import re
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession


_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_BUDGET_ARTIFACT = (
    _ROOT / "Conta2025" / "reportes_2025" / "borrador_presupuesto_2026.csv"
)
DEFAULT_BUDGET_CONCEPT_CATALOG = _ROOT / "docs" / "Catálogo de Torneos.xlsx"
_BUDGET_STATUS_ORDER = {
    "frozen": 1,
    "approved": 2,
    "submitted": 3,
    "reforecast": 4,
    "draft": 5,
    "closed": 6,
}
_BUDGET_ALLOWED_TRANSITIONS = {
    "draft": {"submitted", "closed"},
    "submitted": {"draft", "approved", "closed"},
    "approved": {"frozen", "reforecast", "closed"},
    "frozen": {"reforecast", "closed"},
    "reforecast": {"submitted", "approved", "frozen", "closed"},
    "closed": set(),
}
_BUDGET_ALIAS_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "LTTB",
        ("liga telmex telcel de beisbol", "liga telmex telcel", "beisbol", "béisbol", "lttb"),
    ),
    ("DCC", ("de la calle a la cancha", "calle a la cancha", "dcc")),
    ("BIMBO", ("futbolito bimbo", "bimbo")),
    ("LA MERCED", ("la merced",)),
    ("CCA", ("copa club america", "club america", "club américa", "cca")),
    ("CTTVC", ("viaje de campeones", "cttvc")),
    ("CTT", ("copa telmex telcel", "copa telmex", "ctt")),
)


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _safe_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(str(value).replace(",", "").strip())


def _normalize_text(value: Any) -> str:
    return _safe_str(value).lower()


def _normalize_budget_key(value: Any) -> str:
    raw = _safe_str(value)
    if not raw:
        return ""
    ascii_text = unicodedata.normalize("NFKD", raw)
    ascii_text = "".join(ch for ch in ascii_text if not unicodedata.combining(ch))
    ascii_text = re.sub(r"[^a-zA-Z0-9]+", " ", ascii_text).strip().lower()
    return re.sub(r"\s+", "_", ascii_text)


def _normalize_budget_scope_key(value: Any) -> str:
    normalized = _normalize_budget_key(value)
    if normalized.startswith("fase_"):
        normalized = normalized[len("fase_") :]
    return normalized


def _budget_scope_aliases(*values: Any) -> set[str]:
    aliases: set[str] = set()
    for value in values:
        normalized = _normalize_budget_scope_key(value)
        if not normalized:
            continue
        aliases.add(normalized)
        aliases.add(f"fase_{normalized}")
    return aliases


def _is_budget_phase_break_label(label: Any, trailing_values: list[Any]) -> bool:
    raw = _safe_str(label)
    if not raw or any(_safe_str(value) for value in trailing_values):
        return False
    compact = "".join(ch for ch in raw if ch.isalpha() or ch.isspace()).strip()
    if not compact or compact != compact.upper():
        return False
    normalized = _normalize_budget_scope_key(raw)
    return raw.startswith("FASE ") or normalized in {"viaje_de_campeones", "no_aplica"}


def _active_budget_label(value: Optional[str]) -> str:
    return _safe_str(value) or "Sin partida asignada"


def _normalize_budget_status(value: Optional[str]) -> str:
    normalized = _safe_str(value).lower()
    return normalized or "draft"


def _editable_version_status(status: Optional[str]) -> bool:
    return _normalize_budget_status(status) in {"draft", "reforecast"}


def _edition_bounds(edition_year: int) -> tuple[date, date]:
    return date(edition_year, 1, 1), date(edition_year, 12, 31)


def _build_budget_forecast(
    *,
    edition_year: int,
    budget_total: float,
    comparison: dict[str, Any],
) -> dict[str, Any]:
    today = date.today()
    period_start = date(edition_year, 1, 1)
    period_end = date(edition_year, 12, 31)
    effective_today = min(max(today, period_start), period_end)
    elapsed_days = max(1, (effective_today - period_start).days + 1)
    total_days = max(1, (period_end - period_start).days + 1)
    remaining_days = max(0, total_days - elapsed_days)

    actual_total = float(comparison.get("actual_total") or 0)
    committed_total = float(comparison.get("committed_total") or 0)
    paid_total = float(comparison.get("paid_total") or 0)
    pending_to_pay_total = float(comparison.get("pending_to_pay_total") or 0)
    run_rate_daily = actual_total / elapsed_days
    projected_actual_close = round(run_rate_daily * total_days, 2)
    projected_close_total = round(
        max(projected_actual_close, actual_total + pending_to_pay_total),
        2,
    )
    remaining_budget = round(budget_total - max(actual_total, committed_total), 2)
    projected_variance = round(budget_total - projected_close_total, 2)
    projected_cash_need = round(max(projected_close_total - paid_total, 0), 2)
    health = _budget_health_for_close(
        budget_total=budget_total,
        projected_close_total=projected_close_total,
    )

    return {
        "edition_year": edition_year,
        "as_of_date": effective_today.isoformat(),
        "elapsed_days": elapsed_days,
        "remaining_days": remaining_days,
        "total_days": total_days,
        "run_rate_daily": round(run_rate_daily, 2),
        "projected_actual_close": projected_actual_close,
        "projected_close_total": projected_close_total,
        "projected_variance": projected_variance,
        "projected_cash_need": projected_cash_need,
        "remaining_budget": remaining_budget,
        "health": health,
    }


def _budget_health_for_close(*, budget_total: float, projected_close_total: float) -> str:
    health = "healthy"
    if projected_close_total > budget_total:
        health = "over_budget"
    elif projected_close_total > (budget_total * 0.92):
        health = "at_risk"
    return health


def _build_budget_scenarios(
    *,
    edition_year: int,
    budget_total: float,
    comparison: dict[str, Any],
    forecast: dict[str, Any],
) -> dict[str, Any]:
    actual_total = float(comparison.get("actual_total") or 0)
    paid_total = float(comparison.get("paid_total") or 0)
    pending_to_pay_total = float(comparison.get("pending_to_pay_total") or 0)
    projected_actual_close = float(forecast.get("projected_actual_close") or 0)
    base_close_total = float(forecast.get("projected_close_total") or 0)
    base_run_rate = float(forecast.get("run_rate_daily") or 0)
    remaining_days = int(forecast.get("remaining_days") or 0)
    scenario_specs = {
        "optimistic": {
            "label": "Optimista",
            "run_rate_factor": 0.94,
            "pending_factor": 0.82,
            "assumption": "Menor ritmo de gasto y mejor contención de compromisos abiertos.",
        },
        "base": {
            "label": "Base",
            "run_rate_factor": 1.0,
            "pending_factor": 1.0,
            "assumption": "Continuidad del ritmo actual y cumplimiento normal de compromisos ya abiertos.",
        },
        "stressed": {
            "label": "Estresado",
            "run_rate_factor": 1.12,
            "pending_factor": 1.18,
            "assumption": "Mayor presión de gasto y materialización más alta de compromisos pendientes.",
        },
    }
    scenarios: dict[str, Any] = {}
    for key, spec in scenario_specs.items():
        projected_close_total = round(
            max(
                actual_total + (pending_to_pay_total * float(spec["pending_factor"])),
                projected_actual_close * float(spec["run_rate_factor"]),
            ),
            2,
        )
        if key == "base":
            projected_close_total = round(base_close_total, 2)
        projected_variance = round(budget_total - projected_close_total, 2)
        projected_cash_need = round(max(projected_close_total - paid_total, 0), 2)
        scenarios[key] = {
            "scenario_key": key,
            "label": spec["label"],
            "edition_year": edition_year,
            "remaining_days": remaining_days,
            "adjustment_vs_base_pct": round(
                ((projected_close_total - base_close_total) / base_close_total * 100)
                if base_close_total
                else 0,
                2,
            ),
            "run_rate_daily": round(base_run_rate * float(spec["run_rate_factor"]), 2),
            "projected_close_total": projected_close_total,
            "projected_variance": projected_variance,
            "projected_cash_need": projected_cash_need,
            "remaining_budget": round(
                budget_total
                - max(
                    actual_total,
                    projected_close_total,
                ),
                2,
            ),
            "health": _budget_health_for_close(
                budget_total=budget_total,
                projected_close_total=projected_close_total,
            ),
            "assumption": spec["assumption"],
        }
    return scenarios


def build_budget_executive_alerts(
    summary: dict[str, Any],
    forecast: dict[str, Any],
    scenarios: dict[str, Any],
) -> list[dict[str, str]]:
    alerts: list[dict[str, str]] = []

    forecast_health = _safe_str(forecast.get("health")).lower()
    forecast_health_counts = summary.get("forecast_health_counts") or {}
    over_budget_count = int(forecast_health_counts.get("over_budget") or 0)
    at_risk_count = int(forecast_health_counts.get("at_risk") or 0)
    overdue_open_total = float(summary.get("overdue_open_total") or 0)
    due_next_30_total = float(summary.get("due_next_30_total") or 0)
    pending_to_pay_total = float(summary.get("pending_to_pay_total") or 0)
    stressed_close_total = float(
        ((scenarios.get("stressed") or {}).get("projected_close_total")) or 0
    )
    budget_total = float(summary.get("budget_total") or 0)

    if forecast_health == "over_budget" or over_budget_count > 0:
        alerts.append(
            {
                "severity": "critical",
                "title": "Forecast por encima del presupuesto",
                "detail": (
                    f"{over_budget_count} torneo(s) over budget en el snapshot actual."
                    if over_budget_count
                    else "La proyección consolidada rebasa el presupuesto."
                ),
                "playbook": "Congelar gasto discrecional, reasignar bolsa y validar recorte inmediato.",
            }
        )
    elif forecast_health == "at_risk" or at_risk_count > 0:
        alerts.append(
            {
                "severity": "warning",
                "title": "Forecast en zona de riesgo",
                "detail": (
                    f"{at_risk_count} torneo(s) cerca del techo presupuestal."
                    if at_risk_count
                    else "La proyección consolidada está cerca del techo."
                ),
                "playbook": "Revisar concentración por concepto y exigir plan de contención quincenal.",
            }
        )

    if overdue_open_total > 0:
        alerts.append(
            {
                "severity": "warning",
                "title": "Pendientes vencidos abiertos",
                "detail": f"Hay ${overdue_open_total:,.2f} vencidos sin resolver.",
                "playbook": "Depurar vencidos y escalar responsables antes del siguiente corte.",
            }
        )

    if due_next_30_total > 0 or pending_to_pay_total > 0:
        alerts.append(
            {
                "severity": "info" if due_next_30_total <= 0 else "warning",
                "title": "Salida de caja próxima",
                "detail": (
                    f"${due_next_30_total:,.2f} vence en 30 días; "
                    f"${pending_to_pay_total:,.2f} sigue pendiente por pagar."
                ),
                "playbook": "Calendarizar caja, priorizar pagos críticos y anticipar faltantes.",
            }
        )

    if budget_total > 0 and stressed_close_total > budget_total:
        alerts.append(
            {
                "severity": "warning",
                "title": "Escenario estresado rebasa el techo",
                "detail": (
                    f"El escenario estresado cerraría en ${stressed_close_total:,.2f} "
                    f"vs ${budget_total:,.2f} presupuestados."
                ),
                "playbook": "Definir contingencia y recortes activables si el escenario estresado materializa.",
            }
        )

    return alerts


def build_budget_executive_comparison(
    summary: dict[str, Any],
    forecast: dict[str, Any],
) -> list[dict[str, Any]]:
    budget_total = float(summary.get("budget_total") or 0)
    requested_total = float(summary.get("requested_total") or 0)
    committed_total = float(summary.get("committed_total") or 0)
    paid_total = float(summary.get("paid_total") or 0)
    actual_total = float(summary.get("actual_total") or 0)
    pending_to_pay_total = float(summary.get("pending_to_pay_total") or 0)
    projected_close_total = float(forecast.get("projected_close_total") or 0)

    def _pct(value: float) -> float:
        if budget_total <= 0:
            return 0.0
        return round((value / budget_total) * 100, 2)

    metrics = [
        (
            "Presupuesto",
            budget_total,
            None,
            "Base aprobada del snapshot activo.",
        ),
        (
            "Solicitado",
            requested_total,
            round(budget_total - requested_total, 2),
            "Monto ya solicitado contra la bolsa total.",
        ),
        (
            "Comprometido",
            committed_total,
            round(budget_total - committed_total, 2),
            "Compromisos abiertos y documentos ya amarrados.",
        ),
        (
            "Pagado",
            paid_total,
            round(budget_total - paid_total, 2),
            "Salida de caja ya ejecutada.",
        ),
        (
            "Real",
            actual_total,
            round(budget_total - actual_total, 2),
            "Gasto observado al corte actual.",
        ),
        (
            "Pendiente por pagar",
            pending_to_pay_total,
            round(budget_total - pending_to_pay_total, 2),
            "Compromiso reconocido que todavía no sale de caja.",
        ),
        (
            "Cierre proyectado",
            projected_close_total,
            round(budget_total - projected_close_total, 2),
            "Proyección de cierre usando run rate y pendientes.",
        ),
    ]

    rows: list[dict[str, Any]] = []
    for label, total, variance_to_budget, detail in metrics:
        row = {
            "label": label,
            "total": round(total, 2),
            "pct_of_budget": _pct(total),
            "detail": detail,
        }
        if variance_to_budget is not None:
            row["variance_to_budget"] = round(variance_to_budget, 2)
        rows.append(row)
    return rows


def build_budget_scenario_player(
    summary: dict[str, Any],
    forecast: dict[str, Any],
    *,
    run_rate_delta_pct: Any = 0,
    discretionary_cut_pct: Any = 0,
    added_commitments: Any = 0,
    cash_acceleration: Any = 0,
) -> dict[str, Any]:
    """Calculate a non-persistent what-if scenario over the active snapshot."""

    budget_total = _safe_decimal(summary.get("budget_total"))
    paid_total = _safe_decimal(summary.get("paid_total"))
    base_close_total = _safe_decimal(forecast.get("projected_close_total"))
    if base_close_total <= 0:
        base_close_total = max(
            _safe_decimal(summary.get("actual_total")),
            _safe_decimal(summary.get("committed_total")),
            _safe_decimal(summary.get("requested_total")),
        )

    run_rate_pct = _safe_decimal(run_rate_delta_pct)
    cut_pct = max(0.0, _safe_decimal(discretionary_cut_pct))
    added = _safe_decimal(added_commitments)
    accelerated_cash = max(0.0, _safe_decimal(cash_acceleration))

    run_rate_impact = round(base_close_total * (run_rate_pct / 100), 2)
    cut_impact = round(budget_total * (cut_pct / 100), 2)
    projected_close_total = round(base_close_total + run_rate_impact - cut_impact + added, 2)
    projected_variance = round(budget_total - projected_close_total, 2)
    projected_cash_need = round(max(projected_close_total - paid_total - accelerated_cash, 0), 2)
    health = _budget_health_for_close(
        budget_total=budget_total,
        projected_close_total=projected_close_total,
    )

    if health == "over_budget":
        recommendation = "Requiere recorte, reasignación o aprobación explícita de excedente."
    elif health == "at_risk":
        recommendation = "Mantener vigilancia semanal y bloquear gasto discrecional no crítico."
    else:
        recommendation = "Escenario dentro de rango; documentar supuestos y monitorear ejecución."

    return {
        "label": "Scenario player",
        "budget_total": round(budget_total, 2),
        "base_close_total": round(base_close_total, 2),
        "run_rate_delta_pct": round(run_rate_pct, 2),
        "run_rate_impact": run_rate_impact,
        "discretionary_cut_pct": round(cut_pct, 2),
        "cut_impact": cut_impact,
        "added_commitments": round(added, 2),
        "cash_acceleration": round(accelerated_cash, 2),
        "projected_close_total": projected_close_total,
        "projected_variance": projected_variance,
        "projected_cash_need": projected_cash_need,
        "health": health,
        "recommendation": recommendation,
        "read_only": True,
    }


def _new_breakdown_store() -> dict[str, dict[str, Any]]:
    return {}


def _merge_breakdown_row(
    store: dict[str, dict[str, Any]],
    *,
    label: Any,
    budget_total: float = 0.0,
    reference_total: float = 0.0,
    variance_total: float = 0.0,
    requested_total: float = 0.0,
    committed_total: float = 0.0,
    paid_total: float = 0.0,
    actual_total: float = 0.0,
    pending_to_pay_total: float = 0.0,
    line_count: int = 0,
    document_count: int = 0,
    expense_count: int = 0,
) -> None:
    bucket = _safe_str(label) or "Sin dato"
    entry = store.setdefault(
        bucket,
        {
            "label": bucket,
            "budget_total": 0.0,
            "reference_total": 0.0,
            "variance_total": 0.0,
            "requested_total": 0.0,
            "committed_total": 0.0,
            "paid_total": 0.0,
            "actual_total": 0.0,
            "pending_to_pay_total": 0.0,
            "line_count": 0,
            "document_count": 0,
            "expense_count": 0,
        },
    )
    entry["budget_total"] += float(budget_total or 0)
    entry["reference_total"] += float(reference_total or 0)
    entry["variance_total"] += float(variance_total or 0)
    entry["requested_total"] += float(requested_total or 0)
    entry["committed_total"] += float(committed_total or 0)
    entry["paid_total"] += float(paid_total or 0)
    entry["actual_total"] += float(actual_total or 0)
    entry["pending_to_pay_total"] += float(pending_to_pay_total or 0)
    entry["line_count"] += int(line_count or 0)
    entry["document_count"] += int(document_count or 0)
    entry["expense_count"] += int(expense_count or 0)


def _finalize_breakdown_store(
    store: dict[str, dict[str, Any]],
    *,
    limit: int = 6,
) -> list[dict[str, Any]]:
    rows = []
    for entry in store.values():
        rows.append(
            {
                "label": _safe_str(entry.get("label")) or "Sin dato",
                "budget_total": round(float(entry.get("budget_total") or 0), 2),
                "reference_total": round(float(entry.get("reference_total") or 0), 2),
                "variance_total": round(float(entry.get("variance_total") or 0), 2),
                "requested_total": round(float(entry.get("requested_total") or 0), 2),
                "committed_total": round(float(entry.get("committed_total") or 0), 2),
                "paid_total": round(float(entry.get("paid_total") or 0), 2),
                "actual_total": round(float(entry.get("actual_total") or 0), 2),
                "pending_to_pay_total": round(
                    float(entry.get("pending_to_pay_total") or 0), 2
                ),
                "line_count": int(entry.get("line_count") or 0),
                "document_count": int(entry.get("document_count") or 0),
                "expense_count": int(entry.get("expense_count") or 0),
            }
        )
    rows.sort(
        key=lambda item: (
            -max(
                float(item.get("budget_total") or 0),
                float(item.get("actual_total") or 0),
                float(item.get("committed_total") or 0),
                float(item.get("requested_total") or 0),
                float(item.get("paid_total") or 0),
                float(item.get("reference_total") or 0),
            ),
            item["label"],
        )
    )
    return rows[: max(1, limit)]


def _build_budget_line_breakdowns(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_concept = _new_breakdown_store()
    by_phase = _new_breakdown_store()
    by_entity = _new_breakdown_store()
    by_owner = _new_breakdown_store()
    by_account = _new_breakdown_store()
    for row in rows:
        budget_total = _safe_decimal(row.get("budget_amount"))
        reference_total = _safe_decimal(row.get("reference_amount"))
        variance_total = _safe_decimal(row.get("variance_amount"))
        _merge_breakdown_row(
            by_concept,
            label=row.get("concept_name"),
            budget_total=budget_total,
            reference_total=reference_total,
            variance_total=variance_total,
            line_count=1,
        )
        _merge_breakdown_row(
            by_phase,
            label=row.get("phase") or "Sin fase",
            budget_total=budget_total,
            reference_total=reference_total,
            variance_total=variance_total,
            line_count=1,
        )
        _merge_breakdown_row(
            by_entity,
            label=row.get("entity_name") or "Sin entidad",
            budget_total=budget_total,
            reference_total=reference_total,
            variance_total=variance_total,
            line_count=1,
        )
        _merge_breakdown_row(
            by_owner,
            label=row.get("owner_name") or "Sin responsable",
            budget_total=budget_total,
            reference_total=reference_total,
            variance_total=variance_total,
            line_count=1,
        )
        _merge_breakdown_row(
            by_account,
            label=row.get("account_code_final") or row.get("account_code_suggested") or "Sin cuenta final",
            budget_total=budget_total,
            reference_total=reference_total,
            variance_total=variance_total,
            line_count=1,
        )
    return {
        "by_concept": _finalize_breakdown_store(by_concept),
        "by_phase": _finalize_breakdown_store(by_phase),
        "by_entity": _finalize_breakdown_store(by_entity),
        "by_owner": _finalize_breakdown_store(by_owner),
        "by_account": _finalize_breakdown_store(by_account),
    }


def _build_artifact_breakdowns(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    normalized_rows = [
        {
            "concept_name": row.get("concepto"),
            "phase": row.get("etapa"),
            "entity_name": row.get("entidad") or row.get("entity_name"),
            "owner_name": row.get("responsable"),
            "account_code_final": row.get("cuenta_contable_final"),
            "account_code_suggested": row.get("cuenta_contable_sugerida"),
            "budget_amount": row.get("presupuesto_2026"),
            "reference_amount": row.get("importe_referencia_total"),
            "variance_amount": _safe_decimal(row.get("presupuesto_2026"))
            - _safe_decimal(row.get("importe_referencia_total")),
        }
        for row in rows
    ]
    breakdowns = _build_budget_line_breakdowns(normalized_rows)
    breakdowns["by_provider"] = []
    return breakdowns


def _build_budget_scope_filters(
    *,
    edition_year: int,
    tournament_id: Optional[str],
    tournament_name: Optional[str],
    tournament_code: Optional[str],
) -> tuple[list[str], list[str], dict[str, Any]]:
    aliases = budget_alias_candidates(
        _safe_str(tournament_name),
        _safe_str(tournament_code),
    )
    date_from, date_to = _edition_bounds(edition_year)
    document_filter = [
        "d.tipo = 'SOLICITUD'",
        "DATE(d.creado_en) >= :date_from",
        "DATE(d.creado_en) <= :date_to",
    ]
    expense_filter = [
        "e.estado_gasto != 'cancelado'",
        "DATE(e.fecha) >= :date_from",
        "DATE(e.fecha) <= :date_to",
    ]
    params: dict[str, Any] = {"date_from": date_from, "date_to": date_to}
    if tournament_id:
        document_filter.append("d.torneo_id = :tournament_id")
        expense_filter.append("(d.torneo_id = :tournament_id)")
        params["tournament_id"] = tournament_id
    elif aliases:
        alias_clauses = []
        expense_alias_clauses = []
        for index, alias in enumerate(sorted(aliases)):
            key = f"alias_{index}"
            params[key] = f"%{alias.lower()}%"
            alias_clauses.append(
                f"(LOWER(COALESCE(t.name, '')) LIKE :{key} OR LOWER(COALESCE(d.notas, '')) LIKE :{key} OR LOWER(COALESCE(d.concepto_pago, '')) LIKE :{key})"
            )
            expense_alias_clauses.append(
                f"(LOWER(COALESCE(t.name, '')) LIKE :{key} OR LOWER(COALESCE(e.proyecto, '')) LIKE :{key} OR LOWER(COALESCE(e.concepto, '')) LIKE :{key})"
            )
        document_filter.append(f"({' OR '.join(alias_clauses)})")
        expense_filter.append(f"({' OR '.join(expense_alias_clauses)})")
    return document_filter, expense_filter, params


async def _build_budget_finance_breakdowns(
    session: AsyncSession,
    *,
    edition_year: int,
    tournament_id: Optional[str],
    tournament_name: Optional[str],
    tournament_code: Optional[str],
) -> dict[str, list[dict[str, Any]]]:
    document_filter, expense_filter, params = _build_budget_scope_filters(
        edition_year=edition_year,
        tournament_id=tournament_id,
        tournament_name=tournament_name,
        tournament_code=tournament_code,
    )
    by_provider = _new_breakdown_store()
    by_concept = _new_breakdown_store()

    document_provider_rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(NULLIF(TRIM(pc.nombre), ''), 'Sin proveedor asignado') AS label,
                    COALESCE(SUM(CASE WHEN d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS requested_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS committed_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('pagado', 'cerrado') OR d.pagado_en IS NOT NULL THEN COALESCE(d.monto_total, d.monto_solicitado, 0) ELSE 0 END), 0) AS paid_total,
                    COUNT(*) FILTER (WHERE d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado')) AS document_count
                FROM documentos d
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                LEFT JOIN proveedores_clientes pc ON pc.id = d.proveedor_cliente_id
                WHERE {' AND '.join(document_filter)}
                GROUP BY 1
                """
            ),
            params,
        )
    ).mappings().all()
    for row in document_provider_rows:
        requested_total = _safe_decimal(row.get("requested_total"))
        committed_total = _safe_decimal(row.get("committed_total"))
        paid_total = _safe_decimal(row.get("paid_total"))
        _merge_breakdown_row(
            by_provider,
            label=row.get("label"),
            requested_total=requested_total,
            committed_total=committed_total,
            paid_total=paid_total,
            pending_to_pay_total=max(committed_total - paid_total, 0),
            document_count=int(row.get("document_count") or 0),
        )

    expense_provider_rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(NULLIF(TRIM(pc.nombre), ''), 'Sin proveedor asignado') AS label,
                    COALESCE(SUM(e.gasto_cantidad), 0) AS actual_total,
                    COUNT(*) AS expense_count
                FROM expense_reports e
                LEFT JOIN documentos d ON d.id = e.documento_id
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                LEFT JOIN proveedores_clientes pc ON pc.id = d.proveedor_cliente_id
                WHERE {' AND '.join(expense_filter)}
                GROUP BY 1
                """
            ),
            params,
        )
    ).mappings().all()
    for row in expense_provider_rows:
        _merge_breakdown_row(
            by_provider,
            label=row.get("label"),
            actual_total=_safe_decimal(row.get("actual_total")),
            expense_count=int(row.get("expense_count") or 0),
        )

    document_concept_rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(NULLIF(TRIM(bc.concept_name), ''), 'Sin partida asignada') AS label,
                    COALESCE(SUM(CASE WHEN d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS requested_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS committed_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('pagado', 'cerrado') OR d.pagado_en IS NOT NULL THEN COALESCE(d.monto_total, d.monto_solicitado, 0) ELSE 0 END), 0) AS paid_total,
                    COUNT(*) FILTER (WHERE d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado')) AS document_count
                FROM documentos d
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                LEFT JOIN budget_concepts bc ON bc.id = d.budget_concept_id
                WHERE {' AND '.join(document_filter)}
                GROUP BY 1
                """
            ),
            params,
        )
    ).mappings().all()
    for row in document_concept_rows:
        requested_total = _safe_decimal(row.get("requested_total"))
        committed_total = _safe_decimal(row.get("committed_total"))
        paid_total = _safe_decimal(row.get("paid_total"))
        _merge_breakdown_row(
            by_concept,
            label=row.get("label"),
            requested_total=requested_total,
            committed_total=committed_total,
            paid_total=paid_total,
            pending_to_pay_total=max(committed_total - paid_total, 0),
            document_count=int(row.get("document_count") or 0),
        )

    expense_concept_rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(NULLIF(TRIM(bc.concept_name), ''), 'Sin partida asignada') AS label,
                    COALESCE(SUM(e.gasto_cantidad), 0) AS actual_total,
                    COUNT(*) AS expense_count
                FROM expense_reports e
                LEFT JOIN documentos d ON d.id = e.documento_id
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                LEFT JOIN budget_concepts bc ON bc.id = e.budget_concept_id
                WHERE {' AND '.join(expense_filter)}
                GROUP BY 1
                """
            ),
            params,
        )
    ).mappings().all()
    for row in expense_concept_rows:
        _merge_breakdown_row(
            by_concept,
            label=row.get("label"),
            actual_total=_safe_decimal(row.get("actual_total")),
            expense_count=int(row.get("expense_count") or 0),
        )

    return {
        "by_provider": _finalize_breakdown_store(by_provider),
        "by_concept": _finalize_breakdown_store(by_concept),
    }


async def list_budget_tournament_commitments(
    session: AsyncSession,
    *,
    edition_year: int,
    tournament_id: Optional[str],
    tournament_name: Optional[str],
    tournament_code: Optional[str],
    limit: int = 20,
) -> list[dict[str, Any]]:
    document_filter, _expense_filter, params = _build_budget_scope_filters(
        edition_year=edition_year,
        tournament_id=tournament_id,
        tournament_name=tournament_name,
        tournament_code=tournament_code,
    )
    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    CAST(d.id AS text) AS documento_id,
                    d.numero_referencia,
                    d.estado,
                    d.concepto_pago,
                    COALESCE(NULLIF(TRIM(bc_doc.concept_name), ''), NULL) AS budget_concept_name,
                    d.monto_solicitado,
                    d.monto_total,
                    d.fecha_pago,
                    d.creado_en,
                    CAST(d.gasto_generado_id AS text) AS gasto_generado_id,
                    er.numero_referencia AS gasto_generado_referencia,
                    er.concepto AS gasto_generado_concepto,
                    COALESCE(NULLIF(TRIM(bc_exp.concept_name), ''), NULL) AS gasto_generado_budget_concept_name,
                    er.gasto_cantidad AS gasto_generado_total,
                    er.fecha AS gasto_generado_fecha,
                    er.estado_gasto AS gasto_generado_estado,
                    COALESCE(NULLIF(TRIM(er.usuario_nombre), ''), NULLIF(TRIM(eg.nombre), ''), 'Sin gasto generado') AS gasto_generado_actor,
                    COALESCE(expense_link.related_expense_count, 0) AS related_expense_count,
                    COALESCE(expense_link.related_expense_total, 0) AS related_expense_total,
                    expense_link.related_expense_latest_date,
                    COALESCE(NULLIF(TRIM(pc.nombre), ''), 'Sin proveedor asignado') AS proveedor_nombre,
                    COALESCE(NULLIF(TRIM(t.name), ''), 'Sin torneo') AS torneo_nombre
                FROM documentos d
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                LEFT JOIN proveedores_clientes pc ON pc.id = d.proveedor_cliente_id
                LEFT JOIN budget_concepts bc_doc ON bc_doc.id = d.budget_concept_id
                LEFT JOIN expense_reports er ON er.id = d.gasto_generado_id
                LEFT JOIN budget_concepts bc_exp ON bc_exp.id = er.budget_concept_id
                LEFT JOIN empleados eg ON eg.id = er.empleado_id
                LEFT JOIN LATERAL (
                    SELECT
                        COUNT(*) FILTER (WHERE e.estado_gasto != 'cancelado') AS related_expense_count,
                        COALESCE(
                            SUM(
                                CASE
                                    WHEN e.estado_gasto != 'cancelado' THEN COALESCE(e.gasto_cantidad, 0)
                                    ELSE 0
                                END
                            ),
                            0
                        ) AS related_expense_total,
                        MAX(e.fecha) FILTER (WHERE e.estado_gasto != 'cancelado') AS related_expense_latest_date
                    FROM expense_reports e
                    WHERE (
                        e.id = d.gasto_generado_id
                        OR e.documento_id = d.id
                        OR e.solicitud_documento_id = d.id
                        OR e.informe_documento_id = d.id
                        OR (d.cuenta_gastos_id IS NOT NULL AND e.cuenta_gastos_id = d.cuenta_gastos_id)
                    )
                ) expense_link ON TRUE
                WHERE {' AND '.join(document_filter)}
                ORDER BY d.creado_en DESC
                LIMIT :row_limit
                """
            ),
            {**params, "row_limit": int(limit)},
        )
    ).mappings().all()
    return [
        {
            "documento_id": _safe_str(row.get("documento_id")) or None,
            "numero_referencia": _safe_str(row.get("numero_referencia")) or None,
            "estado": _safe_str(row.get("estado")) or None,
            "concepto_pago": _safe_str(row.get("concepto_pago")) or None,
            "budget_concept_name": _safe_str(row.get("budget_concept_name")) or None,
            "monto_solicitado": round(_safe_decimal(row.get("monto_solicitado")), 2),
            "monto_total": round(_safe_decimal(row.get("monto_total")), 2),
            "fecha_pago": row.get("fecha_pago").isoformat() if row.get("fecha_pago") else None,
            "creado_en": row.get("creado_en").isoformat() if row.get("creado_en") else None,
            "gasto_generado_id": _safe_str(row.get("gasto_generado_id")) or None,
            "gasto_generado_referencia": _safe_str(row.get("gasto_generado_referencia")) or None,
            "gasto_generado_concepto": _safe_str(row.get("gasto_generado_concepto")) or None,
            "gasto_generado_budget_concept_name": _safe_str(
                row.get("gasto_generado_budget_concept_name")
            )
            or None,
            "gasto_generado_total": round(_safe_decimal(row.get("gasto_generado_total")), 2),
            "gasto_generado_fecha": row.get("gasto_generado_fecha").isoformat() if row.get("gasto_generado_fecha") else None,
            "gasto_generado_estado": _safe_str(row.get("gasto_generado_estado")) or None,
            "gasto_generado_actor": _safe_str(row.get("gasto_generado_actor")) or None,
            "related_expense_count": int(row.get("related_expense_count") or 0),
            "related_expense_total": round(_safe_decimal(row.get("related_expense_total")), 2),
            "related_expense_latest_date": row.get("related_expense_latest_date").isoformat() if row.get("related_expense_latest_date") else None,
            "proveedor_nombre": _safe_str(row.get("proveedor_nombre")) or "Sin proveedor asignado",
            "torneo_nombre": _safe_str(row.get("torneo_nombre")) or "Sin torneo",
        }
        for row in rows
    ]


def budget_alias_candidates(*values: str) -> set[str]:
    haystack = " ".join(_normalize_text(value) for value in values if value).strip()
    aliases: set[str] = set()
    for code, terms in _BUDGET_ALIAS_RULES:
        if any(term in haystack for term in terms):
            aliases.add(code)
    return aliases


def _status_sort_value(status: Optional[str]) -> int:
    return _BUDGET_STATUS_ORDER.get(_normalize_budget_status(status), 999)


def _matches_aliases(
    *,
    aliases: set[str],
    tournament_code: Optional[str],
    tournament_name: Optional[str],
) -> bool:
    if not aliases:
        return True
    row_code = _safe_str(tournament_code).upper()
    if row_code and row_code in aliases:
        return True
    return bool(
        budget_alias_candidates(
            _safe_str(tournament_name),
            _safe_str(tournament_code),
        )
        & aliases
    )


def build_budget_commitment_expense_preview(commitment: Optional[dict[str, Any]]) -> dict[str, Any]:
    item = commitment or {}
    generated_expense_id = str(item.get("gasto_generado_id") or "").strip() or None
    generated_expense_reference = (
        str(item.get("gasto_generado_referencia") or "").strip() or None
    )
    generated_expense_state = (
        str(item.get("gasto_generado_estado") or "").strip() or None
    )
    generated_expense_actor = (
        str(item.get("gasto_generado_actor") or "").strip() or None
    )
    generated_expense_concept = (
        str(item.get("gasto_generado_concepto") or "").strip() or None
    )
    related_expense_count = int(item.get("related_expense_count") or 0)
    related_expense_total = round(float(item.get("related_expense_total") or 0), 2)
    related_expense_latest_date = (
        str(item.get("related_expense_latest_date") or "").strip() or None
    )
    generated_expense_total = round(float(item.get("gasto_generado_total") or 0), 2)

    return {
        "has_generated_expense": bool(generated_expense_id),
        "generated_expense_id": generated_expense_id,
        "generated_expense_reference": generated_expense_reference,
        "generated_expense_state": generated_expense_state,
        "generated_expense_actor": generated_expense_actor,
        "generated_expense_concept": generated_expense_concept,
        "generated_expense_total": generated_expense_total,
        "generated_expense_href": (
            f"/gastos/{generated_expense_id}" if generated_expense_id else None
        ),
        "related_expense_count": related_expense_count,
        "related_expense_total": related_expense_total,
        "related_expense_latest_date": related_expense_latest_date,
    }


async def _load_tournament_rows(session: AsyncSession) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            text(
                """
                SELECT id, name
                FROM tournaments
                ORDER BY active DESC, display_order ASC, name ASC
                """
            )
        )
    ).mappings().all()
    return [dict(row) for row in rows]


def _match_tournament_id(
    tournament_rows: list[dict[str, Any]],
    *,
    tournament_code: Optional[str],
    tournament_name: Optional[str],
) -> Optional[str]:
    target_name = _normalize_text(tournament_name)
    target_aliases = budget_alias_candidates(
        _safe_str(tournament_code), _safe_str(tournament_name)
    )
    for row in tournament_rows:
        row_name = _normalize_text(row.get("name"))
        if target_name and row_name == target_name:
            return _safe_str(row.get("id")) or None
        row_aliases = budget_alias_candidates(_safe_str(row.get("name")))
        if target_aliases and row_aliases & target_aliases:
            return _safe_str(row.get("id")) or None
    return None


def _derive_budget_tournament_code(sheet_name: str, tournament_name: str) -> str:
    aliases = sorted(
        budget_alias_candidates(sheet_name, tournament_name),
        key=lambda item: (0 if item == _safe_str(sheet_name).upper() else 1, len(item), item),
    )
    if aliases:
        return aliases[0]
    fallback = re.sub(r"[^A-Z0-9]+", "", _safe_str(sheet_name).upper())
    return fallback[:40] or "GENERAL"


def _map_catalog_phase_label_to_etapa(
    catalog_label: Optional[str],
    tournament_etapas: Optional[list[str]],
) -> str:
    clean_label = _safe_str(catalog_label)
    if not clean_label:
        return ""
    etapas = [_safe_str(value) for value in (tournament_etapas or []) if _safe_str(value)]
    catalog_aliases = _budget_scope_aliases(clean_label)
    for etapa in etapas:
        if _normalize_budget_scope_key(etapa) == _normalize_budget_scope_key(clean_label):
            return etapa
        if catalog_aliases & _budget_scope_aliases(etapa):
            return etapa
    normalized = _normalize_budget_scope_key(clean_label)
    crosswalk_keys = {
        "local": {"municipal", "fase_municipal"},
        "fase_local": {"municipal", "fase_municipal"},
        "nacional": {"final_nacional", "fase_final_nacional"},
        "fase_nacional": {
            "final_nacional",
            "fase_final_nacional",
            "nacional",
            "fase_nacional",
        },
    }
    if normalized in crosswalk_keys:
        for etapa in etapas:
            etapa_aliases = _budget_scope_aliases(etapa)
            if etapa_aliases & crosswalk_keys[normalized]:
                return etapa
    fallback_map = {
        "fase_colectiva": "Fase Colectiva",
        "fase_estatal": "Fase Estatal",
        "fase_nacional": "Fase Nacional",
        "viaje_de_campeones": "Viaje de Campeones",
    }
    if normalized in fallback_map and fallback_map[normalized] in etapas:
        return fallback_map[normalized]
    return clean_label.title()


async def validate_active_cuenta_contable_id(
    session: AsyncSession,
    cuenta_contable_id: str,
) -> str:
    """Return the cuenta id when it exists and is active; raise ValueError otherwise."""
    clean_id = _safe_str(cuenta_contable_id)
    if not clean_id:
        raise ValueError("La cuenta contable seleccionada no es válida.")
    row = (
        await session.execute(
            text(
                """
                SELECT id
                FROM cuentas_contables
                WHERE CAST(id AS text) = :cuenta_contable_id
                  AND activo = TRUE
                LIMIT 1
                """
            ),
            {"cuenta_contable_id": clean_id},
        )
    ).mappings().first()
    if not row:
        raise ValueError(
            "La cuenta contable seleccionada no existe o está inactiva."
        )
    return clean_id


DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING = (
    Path(__file__).resolve().parents[2]
    / "docs"
    / "Catálogo de Partidas Presupuestales vs Cuenta Contable.xlsx"
)


def _iter_budget_concept_account_mapping_rows(
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING,
    *,
    sheet_name: str = "Catálogo General",
) -> list[dict[str, Any]]:
    path = Path(workbook_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = None
    column_map: dict[str, int] = {}
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        headers = {
            _normalize_budget_key(_safe_str(sheet.cell(row=row_index, column=col).value)): col
            for col in range(1, sheet.max_column + 1)
            if _safe_str(sheet.cell(row=row_index, column=col).value)
        }
        if "partida" in headers and "cuenta_contable" in headers:
            header_row = row_index
            column_map = headers
            break
    if header_row is None:
        return []

    def _col(*aliases: str) -> Optional[int]:
        for alias in aliases:
            key = _normalize_budget_key(alias)
            if key in column_map:
                return column_map[key]
        return None

    partida_col = _col("partida")
    proyecto_col = _col("proyecto")
    subproyecto_col = _col("subproyecto")
    cuenta_col = _col("cuenta_contable", "cuenta contable")
    cuenta_nombre_col = _col("nombre_cuenta_contable", "nombre cuenta contable")
    if not partida_col or not cuenta_col:
        return []

    rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        partida = _safe_str(sheet.cell(row=row_index, column=partida_col).value)
        cuenta_codigo = _safe_str(sheet.cell(row=row_index, column=cuenta_col).value)
        if not partida or not cuenta_codigo:
            continue
        proyecto = (
            _safe_str(sheet.cell(row=row_index, column=proyecto_col).value)
            if proyecto_col
            else ""
        )
        subproyecto = (
            _safe_str(sheet.cell(row=row_index, column=subproyecto_col).value)
            if subproyecto_col
            else ""
        )
        cuenta_nombre = (
            _safe_str(sheet.cell(row=row_index, column=cuenta_nombre_col).value)
            if cuenta_nombre_col
            else ""
        )
        rows.append(
            {
                "partida": partida,
                "concept_key": _normalize_budget_key(partida),
                "proyecto": proyecto,
                "subproyecto": subproyecto,
                "cuenta_contable_codigo": cuenta_codigo,
                "cuenta_contable_nombre": cuenta_nombre,
                "sheet_row_index": row_index,
            }
        )
    return rows


def collect_workbook_cuentas_contables(
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING,
) -> dict[str, str]:
    """Return unique workbook account codes mapped to their display names."""
    catalog: dict[str, str] = {}
    for row in _iter_budget_concept_account_mapping_rows(workbook_path):
        codigo = _safe_str(row.get("cuenta_contable_codigo"))
        nombre = _safe_str(row.get("cuenta_contable_nombre"))
        if not codigo:
            continue
        if codigo not in catalog and nombre:
            catalog[codigo] = nombre
        elif codigo not in catalog:
            catalog[codigo] = codigo
    return catalog


async def ensure_missing_cuentas_contables_from_workbook(
    session: AsyncSession,
    *,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING,
    tipo: str = "gasto",
    activo: bool = True,
    commit: bool = True,
) -> dict[str, Any]:
    """Create active cuentas contables referenced by the workbook but missing in DB."""
    from devnous.gastos.models import CuentaContable

    workbook_accounts = collect_workbook_cuentas_contables(workbook_path)
    if not workbook_accounts:
        return {
            "workbook_accounts_count": 0,
            "created_count": 0,
            "existing_count": 0,
            "created": [],
        }

    existing_rows = (
        await session.execute(
            text(
                """
                SELECT codigo
                FROM cuentas_contables
                WHERE codigo = ANY(:codigos)
                """
            ),
            {"codigos": sorted(workbook_accounts)},
        )
    ).scalars().all()
    existing_codes = {_safe_str(code) for code in existing_rows if _safe_str(code)}

    created: list[dict[str, str]] = []
    for codigo in sorted(workbook_accounts):
        if codigo in existing_codes:
            continue
        nombre = workbook_accounts[codigo]
        session.add(
            CuentaContable(
                codigo=codigo,
                nombre=nombre,
                tipo=_safe_str(tipo) or "gasto",
                activo=bool(activo),
            )
        )
        created.append({"codigo": codigo, "nombre": nombre, "tipo": _safe_str(tipo) or "gasto"})

    if commit and created:
        await session.commit()

    return {
        "workbook_accounts_count": len(workbook_accounts),
        "created_count": len(created),
        "existing_count": len(existing_codes),
        "created": created,
    }


def _match_budget_concept_for_account_mapping(
    concepts: list[dict[str, Any]],
    *,
    row: dict[str, Any],
    tournament_code: str = "",
) -> Optional[dict[str, Any]]:
    """Resolve a workbook row to a budget concept using scoped keys and SSOT metadata."""
    base_key = _safe_str(row.get("concept_key"))
    subproyecto = _safe_str(row.get("subproyecto"))
    proyecto = _safe_str(row.get("proyecto"))
    sheet_row_index = row.get("sheet_row_index")
    if not base_key:
        return None

    def _proyecto_matches(item: dict[str, Any]) -> bool:
        metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
        if not proyecto and not tournament_code:
            return True
        if tournament_code and (
            _safe_str(item.get("tournament_code") or "").upper() == tournament_code.upper()
        ):
            return True
        if _safe_str(item.get("tournament_name") or "").lower() == proyecto.lower():
            return True
        if _safe_str(metadata.get("ssot_proyecto")).lower() == proyecto.lower():
            return True
        row_aliases = budget_alias_candidates(proyecto)
        item_aliases = budget_alias_candidates(
            _safe_str(item.get("tournament_code")),
            _safe_str(item.get("tournament_name")),
            _safe_str(metadata.get("ssot_proyecto")),
        )
        return bool(row_aliases & item_aliases)

    scoped_key = (
        f"{base_key}__{_normalize_budget_key(subproyecto)}" if subproyecto else base_key
    )
    exact = [
        item
        for item in concepts
        if _safe_str(item.get("concept_key")) == scoped_key and _proyecto_matches(item)
    ]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1 and subproyecto:
        phased = [
            item for item in exact if budget_concept_matches_fase(item, subproyecto)
        ]
        if len(phased) == 1:
            return phased[0]

    ssot_matches = [
        item
        for item in concepts
        if isinstance(item.get("metadata"), dict)
        and item["metadata"].get("ssot_row_index") == sheet_row_index
        and (
            _safe_str(item["metadata"].get("ssot_proyecto")).lower() == proyecto.lower()
            or _proyecto_matches(item)
        )
    ]
    if len(ssot_matches) == 1:
        return ssot_matches[0]

    name_matches = [
        item
        for item in concepts
        if _normalize_budget_key(item.get("concept_name") or "") == base_key
        and _proyecto_matches(item)
    ]
    if subproyecto:
        scoped_name = [
            item for item in name_matches if budget_concept_matches_fase(item, subproyecto)
        ]
        if len(scoped_name) == 1:
            return scoped_name[0]
        if len(scoped_name) > 1:
            name_matches = scoped_name
    if len(name_matches) == 1:
        return name_matches[0]

    loose = [
        item
        for item in concepts
        if base_key in _safe_str(item.get("concept_key"))
        and _proyecto_matches(item)
    ]
    if subproyecto:
        loose = [
            item for item in loose if budget_concept_matches_fase(item, subproyecto)
        ]
    if len(loose) == 1:
        return loose[0]
    return None


async def preview_budget_concept_account_mappings(
    session: AsyncSession,
    *,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING,
) -> dict[str, Any]:
    """Dry-run report for partida→cuenta contable mappings from the client workbook."""
    await ensure_budget_schema(session)
    mapping_rows = _iter_budget_concept_account_mapping_rows(workbook_path)
    concepts = await list_budget_concepts(session, active_only=False, limit=5000)
    tournament_rows = await _load_tournament_rows(session)
    account_rows = (
        await session.execute(
            text(
                """
                SELECT id, codigo, nombre, activo
                FROM cuentas_contables
                ORDER BY codigo ASC
                """
            )
        )
    ).mappings().all()
    accounts_by_code = {
        _normalize_budget_key(_safe_str(row.get("codigo"))): row for row in account_rows
    }

    matched: list[dict[str, Any]] = []
    skipped_no_concept: list[dict[str, Any]] = []
    skipped_missing_account: list[dict[str, Any]] = []
    skipped_inactive_account: list[dict[str, Any]] = []

    for row in mapping_rows:
        tournament_id = _match_tournament_id(
            tournament_rows,
            tournament_code=None,
            tournament_name=row.get("proyecto") or "",
        )
        tournament_code = ""
        if tournament_id:
            tournament = next(
                (item for item in tournament_rows if str(item.get("id")) == tournament_id),
                None,
            )
            if tournament:
                tournament_code = _derive_budget_tournament_code(
                    tournament.get("name") or "",
                    row.get("proyecto") or "",
                )
        subproyecto = _safe_str(row.get("subproyecto"))
        concept = _match_budget_concept_for_account_mapping(
            concepts,
            row=row,
            tournament_code=tournament_code,
        )
        account = accounts_by_code.get(
            _normalize_budget_key(_safe_str(row.get("cuenta_contable_codigo")))
        )
        if concept is None:
            skipped_no_concept.append(row)
            continue
        if account is None:
            skipped_missing_account.append({**row, "concept_id": concept.get("id")})
            continue
        if not bool(account.get("activo")):
            skipped_inactive_account.append(
                {**row, "concept_id": concept.get("id"), "account_id": account.get("id")}
            )
            continue
        matched.append(
            {
                **row,
                "concept_id": concept.get("id"),
                "concept_name": concept.get("concept_name"),
                "cuenta_contable_id": _safe_str(account.get("id")),
                "cuenta_contable_nombre": _safe_str(account.get("nombre")),
            }
        )

    return {
        "workbook_path": str(Path(workbook_path)),
        "rows_total": len(mapping_rows),
        "matched_count": len(matched),
        "skipped_no_concept_count": len(skipped_no_concept),
        "skipped_missing_account_count": len(skipped_missing_account),
        "skipped_inactive_account_count": len(skipped_inactive_account),
        "matched": matched,
        "skipped_no_concept": skipped_no_concept,
        "skipped_missing_account": skipped_missing_account,
        "skipped_inactive_account": skipped_inactive_account,
    }


async def import_budget_concept_account_mappings(
    session: AsyncSession,
    *,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_ACCOUNT_MAPPING,
    dry_run: bool = True,
    create_missing_cuentas: bool = False,
) -> dict[str, Any]:
    cuenta_bootstrap: dict[str, Any] = {}
    if create_missing_cuentas and not dry_run:
        cuenta_bootstrap = await ensure_missing_cuentas_contables_from_workbook(
            session,
            workbook_path=workbook_path,
            commit=True,
        )
    preview = await preview_budget_concept_account_mappings(
        session,
        workbook_path=workbook_path,
    )
    if dry_run:
        return preview
    updated = 0
    for item in preview.get("matched") or []:
        concept_id = _safe_str(item.get("concept_id"))
        cuenta_id = _safe_str(item.get("cuenta_contable_id"))
        if not concept_id:
            continue
        await update_budget_concept(
            session,
            concept_id=concept_id,
            cuenta_contable_id=cuenta_id or None,
            cuenta_contable_provided=True,
            commit=False,
        )
        updated += 1
    await session.commit()
    return {
        **preview,
        **cuenta_bootstrap,
        "updated": updated,
        "dry_run": False,
        "create_missing_cuentas": create_missing_cuentas,
    }


def _iter_budget_concept_catalog_phase_rows(
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_CATALOG,
    *,
    sheet_name: Optional[str] = None,
) -> list[dict[str, Any]]:
    """
    Parse catalog rows assigning each partida to the subproyecto/fase row that
    appears immediately below it in the spreadsheet.
    """
    path = Path(workbook_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    worksheets = []
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheets = [workbook[sheet_name]]
    else:
        worksheets = list(workbook.worksheets)

    rows: list[dict[str, Any]] = []
    for sheet in worksheets:
        title_candidates = [
            _safe_str(sheet.cell(row=row_index, column=1).value)
            for row_index in range(1, min(sheet.max_row, 4) + 1)
            if _safe_str(sheet.cell(row=row_index, column=1).value)
        ]
        tournament_name = title_candidates[0] if title_candidates else _safe_str(sheet.title)
        tournament_code = _derive_budget_tournament_code(sheet.title, tournament_name)
        header_row = None
        dimension_headers: list[str] = []
        for row_index in range(1, sheet.max_row + 1):
            first_value = _safe_str(sheet.cell(row=row_index, column=1).value)
            if first_value.lower() == "concepto":
                header_row = row_index
                dimension_headers = [
                    _safe_str(sheet.cell(row=row_index, column=col_index).value)
                    for col_index in range(2, sheet.max_column + 1)
                    if _safe_str(sheet.cell(row=row_index, column=col_index).value)
                ]
                break
        if header_row is None:
            continue

        pending_rows: list[tuple[int, str, list[str]]] = []

        def _flush_pending(phase_label: Optional[str]) -> None:
            for row_index, concept_name, trailing_values in pending_rows:
                base_key = _normalize_budget_key(concept_name)
                if not base_key:
                    continue
                phase_key = _normalize_budget_key(phase_label or "general")
                concept_key = (
                    f"{base_key}__{phase_key}" if phase_label else base_key
                )
                metadata: dict[str, Any] = {
                    "sheet_name": sheet.title,
                    "sheet_headers": dimension_headers,
                    "sheet_row_index": row_index,
                    "catalog_phase_break_label": _safe_str(phase_label) or None,
                }
                if any(trailing_values):
                    metadata["sheet_dimension_values"] = trailing_values
                rows.append(
                    {
                        "tournament_code": tournament_code,
                        "tournament_name": tournament_name,
                        "concept_name": concept_name,
                        "concept_key": concept_key,
                        "catalog_phase_break_label": _safe_str(phase_label) or None,
                        "metadata": metadata,
                    }
                )
            pending_rows.clear()

        for row_index in range(header_row + 1, sheet.max_row + 1):
            concept_name = _safe_str(sheet.cell(row=row_index, column=1).value)
            trailing_values = [
                _safe_str(sheet.cell(row=row_index, column=col_index).value)
                for col_index in range(2, sheet.max_column + 1)
            ]
            if not concept_name:
                continue
            if _is_budget_phase_break_label(concept_name, trailing_values):
                _flush_pending(concept_name)
                continue
            pending_rows.append((row_index, concept_name, trailing_values))
        _flush_pending(None)
    return rows


async def replace_tournament_budget_concepts_from_catalog_sheet(
    session: AsyncSession,
    *,
    tournament_id: str,
    sheet_name: str,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_CATALOG,
    actor_empleado_id: Optional[str] = None,
    source: str = "catalog_sheet",
) -> dict[str, Any]:
    """Replace all budget partidas for one project from a catalog worksheet."""
    await ensure_budget_schema(session)
    tournament = await _resolve_tournament_for_budget_concept(
        session,
        tournament_id=tournament_id,
    )
    catalog_rows = _iter_budget_concept_catalog_phase_rows(
        workbook_path,
        sheet_name=sheet_name,
    )
    if not catalog_rows:
        raise ValueError(
            f"No se encontraron partidas en la hoja «{sheet_name}» del catálogo."
        )
    if _safe_str(catalog_rows[0].get("tournament_code")).upper() != _safe_str(
        tournament.get("tournament_code")
    ).upper() and _safe_str(sheet_name).upper() != _safe_str(
        tournament.get("tournament_code")
    ).upper():
        # Allow explicit sheet selection even when code differs from derived alias.
        pass

    await session.execute(
        text(
            """
            DELETE FROM budget_concepts
            WHERE tournament_id = CAST(:tournament_id AS uuid)
               OR (
                    COALESCE(tournament_code, '') = :tournament_code
                    AND tournament_id IS NULL
               )
            """
        ),
        {
            "tournament_id": tournament["tournament_id"],
            "tournament_code": tournament["tournament_code"] or "",
        },
    )

    created = 0
    for row in catalog_rows:
        mapped_phase = _map_catalog_phase_label_to_etapa(
            row.get("catalog_phase_break_label"),
            tournament.get("etapas"),
        )
        metadata = build_budget_concept_scope_metadata(
            [mapped_phase] if mapped_phase else [],
            tournament_etapas=tournament.get("etapas"),
            tournament_categorias=[],
        )
        metadata.update(
            {
                "sheet_name": sheet_name,
                "sheet_row_index": row.get("metadata", {}).get("sheet_row_index"),
                "catalog_phase_break_label": row.get("catalog_phase_break_label"),
            }
        )
        await session.execute(
            text(
                """
                INSERT INTO budget_concepts (
                    id, tournament_id, tournament_code, tournament_name,
                    concept_name, concept_key, active, source, metadata,
                    created_by_empleado_id, created_at, updated_at
                ) VALUES (
                    :id, CAST(:tournament_id AS uuid), :tournament_code, :tournament_name,
                    :concept_name, :concept_key, TRUE, :source, CAST(:metadata AS jsonb),
                    :created_by_empleado_id, NOW(), NOW()
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "tournament_id": tournament["tournament_id"],
                "tournament_code": tournament["tournament_code"],
                "tournament_name": tournament["tournament_name"],
                "concept_name": row.get("concept_name"),
                "concept_key": row.get("concept_key"),
                "source": _safe_str(source) or "catalog_sheet",
                "metadata": json.dumps(metadata, ensure_ascii=False),
                "created_by_empleado_id": actor_empleado_id,
            },
        )
        created += 1

    await session.commit()
    return {
        "ok": True,
        "tournament_id": tournament["tournament_id"],
        "tournament_name": tournament["tournament_name"],
        "sheet_name": sheet_name,
        "created": created,
    }


def _iter_budget_concept_catalog_rows(
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_CATALOG,
) -> list[dict[str, Any]]:
    path = Path(workbook_path)
    if not path.exists():
        return []
    workbook = load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        title_candidates = [
            _safe_str(sheet.cell(row=row_index, column=1).value)
            for row_index in range(1, min(sheet.max_row, 4) + 1)
            if _safe_str(sheet.cell(row=row_index, column=1).value)
        ]
        tournament_name = title_candidates[0] if title_candidates else _safe_str(sheet.title)
        tournament_code = _derive_budget_tournament_code(sheet.title, tournament_name)
        header_row = None
        dimension_headers: list[str] = []
        for row_index in range(1, sheet.max_row + 1):
            first_value = _safe_str(sheet.cell(row=row_index, column=1).value)
            if first_value.lower() == "concepto":
                header_row = row_index
                dimension_headers = [
                    _safe_str(sheet.cell(row=row_index, column=col_index).value)
                    for col_index in range(2, sheet.max_column + 1)
                    if _safe_str(sheet.cell(row=row_index, column=col_index).value)
                ]
                break
        if header_row is None:
            continue
        sheet_rows: list[tuple[int, str, list[str]]] = []
        has_phase_breaks = False
        for row_index in range(header_row + 1, sheet.max_row + 1):
            concept_name = _safe_str(sheet.cell(row=row_index, column=1).value)
            trailing_values = [
                _safe_str(sheet.cell(row=row_index, column=col_index).value)
                for col_index in range(2, sheet.max_column + 1)
            ]
            if not concept_name:
                continue
            if _is_budget_phase_break_label(concept_name, trailing_values):
                has_phase_breaks = True
            sheet_rows.append((row_index, concept_name, trailing_values))

        aggregated_rows: dict[str, dict[str, Any]] = {}
        current_phase_label: Optional[str] = None
        for row_index, concept_name, trailing_values in sheet_rows:
            if _is_budget_phase_break_label(concept_name, trailing_values):
                current_phase_label = concept_name
                continue
            concept_key = _normalize_budget_key(concept_name)
            if not concept_key:
                continue
            entry = aggregated_rows.setdefault(
                concept_key,
                {
                    "tournament_code": tournament_code,
                    "tournament_name": tournament_name,
                    "concept_name": concept_name,
                    "concept_key": concept_key,
                    "metadata": {
                        "sheet_name": sheet.title,
                        "sheet_headers": dimension_headers,
                        "sheet_row_indexes": [],
                        "applicable_phase_labels": [],
                        "applicable_phase_keys": [],
                        "applicable_subproject_labels": [],
                        "applicable_subproject_keys": [],
                    },
                },
            )
            metadata = entry["metadata"]
            metadata["sheet_row_indexes"].append(row_index)
            if has_phase_breaks:
                if current_phase_label:
                    phase_labels = metadata["applicable_phase_labels"]
                    if current_phase_label not in phase_labels:
                        phase_labels.append(current_phase_label)
                    phase_keys = metadata["applicable_phase_keys"]
                    for key in sorted(_budget_scope_aliases(current_phase_label)):
                        if key not in phase_keys:
                            phase_keys.append(key)
            elif dimension_headers:
                subproject_labels = metadata["applicable_subproject_labels"]
                subproject_keys = metadata["applicable_subproject_keys"]
                for header in dimension_headers:
                    if not header:
                        continue
                    if header not in subproject_labels:
                        subproject_labels.append(header)
                    for key in sorted(_budget_scope_aliases(header)):
                        if key not in subproject_keys:
                            subproject_keys.append(key)
            if any(trailing_values):
                metadata["sheet_dimension_values"] = trailing_values
        rows.extend(aggregated_rows.values())
    return rows


@lru_cache(maxsize=2)
def _budget_catalog_scope_index(
    workbook_path: str = str(DEFAULT_BUDGET_CONCEPT_CATALOG),
) -> dict[tuple[str, str], dict[str, Any]]:
    index: dict[tuple[str, str], dict[str, Any]] = {}
    for row in _iter_budget_concept_catalog_rows(workbook_path):
        tournament_code = _safe_str(row.get("tournament_code")).upper()
        concept_key = _safe_str(row.get("concept_key"))
        metadata = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        if tournament_code and concept_key and metadata:
            index[(tournament_code, concept_key)] = metadata
    return index


def _merge_budget_concept_metadata(
    stored_metadata: dict[str, Any],
    catalog_metadata: dict[str, Any],
    *,
    source: Optional[str] = None,
) -> dict[str, Any]:
    merged = dict(stored_metadata or {})
    if _safe_str(source) == "admin_ui":
        return merged
    for key in (
        "sheet_name",
        "sheet_headers",
        "sheet_row_indexes",
        "sheet_dimension_values",
        "applicable_phase_labels",
        "applicable_phase_keys",
        "applicable_subproject_labels",
        "applicable_subproject_keys",
    ):
        if key not in merged and key in catalog_metadata:
            merged[key] = catalog_metadata[key]
    return merged


def budget_concept_matches_fase(
    concept: Optional[dict[str, Any]],
    fase: Optional[str],
) -> bool:
    if concept is None:
        return False
    selected_aliases = _budget_scope_aliases(fase)
    if not selected_aliases:
        return True
    metadata = concept.get("metadata") if isinstance(concept.get("metadata"), dict) else {}
    applicable_keys = {
        _safe_str(key)
        for key in (
            list(metadata.get("applicable_phase_keys") or [])
            + list(metadata.get("applicable_subproject_keys") or [])
        )
        if _safe_str(key)
    }
    if not applicable_keys:
        return True
    return bool(applicable_keys & selected_aliases)


async def list_budget_concepts(
    session: AsyncSession,
    *,
    tournament_id: Optional[str] = None,
    tournament_code: Optional[str] = None,
    active_only: bool = True,
    limit: int = 2000,
) -> list[dict[str, Any]]:
    await ensure_budget_schema(session)
    filters = []
    params: dict[str, Any] = {"limit": max(1, min(limit, 5000))}
    if active_only:
        filters.append("bc.active = TRUE")
    if tournament_id:
        filters.append("CAST(bc.tournament_id AS text) = :tournament_id")
        params["tournament_id"] = tournament_id
    elif tournament_code:
        aliases = sorted(budget_alias_candidates(tournament_code))
        if aliases:
            filters.append("UPPER(COALESCE(bc.tournament_code, '')) = ANY(:aliases)")
            params["aliases"] = aliases
    where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    bc.id,
                    bc.tournament_id,
                    bc.tournament_code,
                    bc.tournament_name,
                    bc.concept_name,
                    bc.concept_key,
                    bc.active,
                    bc.source,
                    bc.metadata,
                    bc.cuenta_contable_id,
                    cc.codigo AS cuenta_contable_codigo,
                    cc.nombre AS cuenta_contable_nombre,
                    bc.created_by_empleado_id,
                    bc.created_at,
                    bc.updated_at
                FROM budget_concepts bc
                LEFT JOIN cuentas_contables cc
                    ON cc.id = bc.cuenta_contable_id
                {where_clause}
                ORDER BY bc.tournament_name ASC, bc.concept_name ASC
                LIMIT :limit
                """
            ),
            params,
        )
    ).mappings().all()
    return [
        {
            "id": _safe_str(row.get("id")) or None,
            "tournament_id": _safe_str(row.get("tournament_id")) or None,
            "tournament_code": _safe_str(row.get("tournament_code")) or None,
            "tournament_name": _safe_str(row.get("tournament_name")) or None,
            "concept_name": _safe_str(row.get("concept_name")) or None,
            "concept_key": _safe_str(row.get("concept_key")) or None,
            "active": bool(row.get("active")),
            "source": _safe_str(row.get("source")) or None,
            "cuenta_contable_id": _safe_str(row.get("cuenta_contable_id")) or None,
            "cuenta_contable_codigo": _safe_str(row.get("cuenta_contable_codigo")) or None,
            "cuenta_contable_nombre": _safe_str(row.get("cuenta_contable_nombre")) or None,
            "metadata": _merge_budget_concept_metadata(
                row.get("metadata") if isinstance(row.get("metadata"), dict) else {},
                _budget_catalog_scope_index().get(
                    (
                        _safe_str(row.get("tournament_code")).upper(),
                        _safe_str(row.get("concept_key")),
                    ),
                    {},
                ),
                source=_safe_str(row.get("source")) or None,
            ),
            "created_by_empleado_id": _safe_str(row.get("created_by_empleado_id")) or None,
            "created_at": row.get("created_at").isoformat() if row.get("created_at") else None,
            "updated_at": row.get("updated_at").isoformat() if row.get("updated_at") else None,
        }
        for row in rows
    ]


async def resolve_budget_concept(
    session: AsyncSession,
    *,
    budget_concept_id: Optional[str],
    tournament_id: Optional[str] = None,
    tournament_code: Optional[str] = None,
    fase: Optional[str] = None,
) -> Optional[dict[str, Any]]:
    concept_id = _safe_str(budget_concept_id)
    if not concept_id:
        return None
    concepts = await list_budget_concepts(session, active_only=True, limit=5000)
    concept = next((item for item in concepts if item["id"] == concept_id), None)
    if concept is None:
        return None
    if tournament_id and concept.get("tournament_id") not in {None, tournament_id}:
        return None
    if tournament_code:
        aliases = budget_alias_candidates(tournament_code)
        concept_aliases = budget_alias_candidates(
            concept.get("tournament_code") or "",
            concept.get("tournament_name") or "",
        )
        if aliases and concept_aliases and not (aliases & concept_aliases):
            return None
    if not budget_concept_matches_fase(concept, fase):
        return None
    return concept


def build_budget_concept_scope_metadata(
    scope_labels: Optional[list[str]] = None,
    *,
    tournament_etapas: Optional[list[str]] = None,
    tournament_categorias: Optional[list[str]] = None,
) -> dict[str, Any]:
    labels = [_safe_str(label) for label in (scope_labels or []) if _safe_str(label)]
    if not labels:
        return {
            "applicable_phase_labels": [],
            "applicable_phase_keys": [],
            "applicable_subproject_labels": [],
            "applicable_subproject_keys": [],
        }
    etapa_lookup = {
        _safe_str(label).lower(): _safe_str(label)
        for label in (tournament_etapas or [])
        if _safe_str(label)
    }
    categoria_lookup = {
        _safe_str(label).lower(): _safe_str(label)
        for label in (tournament_categorias or [])
        if _safe_str(label)
    }
    phase_labels: list[str] = []
    subproject_labels: list[str] = []
    for label in labels:
        normalized = label.lower()
        if normalized in categoria_lookup:
            canonical = categoria_lookup[normalized]
            if canonical not in subproject_labels:
                subproject_labels.append(canonical)
        elif normalized in etapa_lookup:
            canonical = etapa_lookup[normalized]
            if canonical not in phase_labels:
                phase_labels.append(canonical)
        else:
            if label not in phase_labels:
                phase_labels.append(label)
    phase_keys: list[str] = []
    for label in phase_labels:
        for key in sorted(_budget_scope_aliases(label)):
            if key not in phase_keys:
                phase_keys.append(key)
    subproject_keys: list[str] = []
    for label in subproject_labels:
        for key in sorted(_budget_scope_aliases(label)):
            if key not in subproject_keys:
                subproject_keys.append(key)
    return {
        "applicable_phase_labels": phase_labels,
        "applicable_phase_keys": phase_keys,
        "applicable_subproject_labels": subproject_labels,
        "applicable_subproject_keys": subproject_keys,
    }


def budget_concept_scope_summary(metadata: Optional[dict[str, Any]]) -> str:
    payload = metadata if isinstance(metadata, dict) else {}
    labels = [
        _safe_str(label)
        for label in (
            list(payload.get("applicable_phase_labels") or [])
            + list(payload.get("applicable_subproject_labels") or [])
        )
        if _safe_str(label)
    ]
    if not labels:
        return "Todas las fases / subproyectos"
    return ", ".join(labels)


_BUDGET_SCOPE_METADATA_KEYS = (
    "applicable_phase_labels",
    "applicable_phase_keys",
    "applicable_subproject_labels",
    "applicable_subproject_keys",
)


def cleared_budget_concept_scope_metadata(
    stored_metadata: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Preserve non-scope metadata while resetting partida to all fases/subproyectos."""
    merged = dict(stored_metadata or {})
    empty_scope = build_budget_concept_scope_metadata([])
    for key in _BUDGET_SCOPE_METADATA_KEYS:
        merged[key] = empty_scope[key]
    return merged


def _merge_budget_concept_scope_metadata(
    stored_metadata: dict[str, Any],
    scope_labels: Optional[list[str]],
    *,
    tournament_etapas: Optional[list[str]] = None,
    tournament_categorias: Optional[list[str]] = None,
) -> dict[str, Any]:
    merged = dict(stored_metadata or {})
    scope_metadata = build_budget_concept_scope_metadata(
        scope_labels,
        tournament_etapas=tournament_etapas,
        tournament_categorias=tournament_categorias,
    )
    for key in _BUDGET_SCOPE_METADATA_KEYS:
        merged[key] = scope_metadata[key]
    return merged


async def clear_budget_concept_scope_for_tournament(
    session: AsyncSession,
    *,
    tournament_id: str,
    commit: bool = False,
) -> int:
    """Reset scoped partidas for a tournament so they apply to all fases/subproyectos."""
    await ensure_budget_schema(session)
    clean_id = _safe_str(tournament_id)
    if not clean_id:
        return 0
    rows = (
        await session.execute(
            text(
                """
                SELECT id, metadata
                FROM budget_concepts
                WHERE CAST(tournament_id AS text) = :tournament_id
                """
            ),
            {"tournament_id": clean_id},
        )
    ).mappings().all()
    updated = 0
    for row in rows:
        stored = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        cleared = cleared_budget_concept_scope_metadata(stored)
        if all(not cleared.get(key) for key in _BUDGET_SCOPE_METADATA_KEYS) and all(
            not stored.get(key) for key in _BUDGET_SCOPE_METADATA_KEYS
        ):
            continue
        await session.execute(
            text(
                """
                UPDATE budget_concepts
                SET metadata = CAST(:metadata AS jsonb),
                    updated_at = NOW()
                WHERE id = :concept_id
                """
            ),
            {
                "concept_id": row.get("id"),
                "metadata": json.dumps(cleared, ensure_ascii=False),
            },
        )
        updated += 1
    if commit:
        await session.commit()
    return updated


async def _resolve_tournament_for_budget_concept(
    session: AsyncSession,
    *,
    tournament_id: str,
) -> dict[str, str]:
    clean_id = _safe_str(tournament_id)
    if not clean_id:
        raise ValueError("Debe seleccionar un proyecto.")
    row = (
        await session.execute(
            text(
                """
                SELECT id, name, etapas, categorias
                FROM tournaments
                WHERE id = :tournament_id
                LIMIT 1
                """
            ),
            {"tournament_id": clean_id},
        )
    ).mappings().first()
    if not row:
        raise ValueError("Proyecto no encontrado.")
    tournament_name = _safe_str(row.get("name")) or "Sin proyecto"
    tournament_code = _derive_budget_tournament_code(tournament_name, tournament_name)
    etapas = row.get("etapas") if isinstance(row.get("etapas"), list) else []
    categorias = row.get("categorias") if isinstance(row.get("categorias"), list) else []
    return {
        "tournament_id": _safe_str(row.get("id")) or clean_id,
        "tournament_code": tournament_code,
        "tournament_name": tournament_name,
        "etapas": [_safe_str(value) for value in etapas if _safe_str(value)],
        "categorias": [_safe_str(value) for value in categorias if _safe_str(value)],
    }


async def get_budget_concept(
    session: AsyncSession,
    *,
    concept_id: str,
) -> Optional[dict[str, Any]]:
    clean_id = _safe_str(concept_id)
    if not clean_id:
        return None
    concepts = await list_budget_concepts(session, active_only=False, limit=5000)
    return next((item for item in concepts if item.get("id") == clean_id), None)


async def create_budget_concept(
    session: AsyncSession,
    *,
    tournament_id: str,
    concept_name: str,
    scope_labels: Optional[list[str]] = None,
    cuenta_contable_id: Optional[str] = None,
    actor_empleado_id: Optional[str] = None,
    source: str = "admin_ui",
    commit: bool = True,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    tournament = await _resolve_tournament_for_budget_concept(
        session,
        tournament_id=tournament_id,
    )
    clean_name = _safe_str(concept_name)
    concept_key = _normalize_budget_key(clean_name)
    if not clean_name or not concept_key:
        raise ValueError("El nombre de la partida presupuestal es obligatorio.")
    existing = (
        await session.execute(
            text(
                """
                SELECT id
                FROM budget_concepts
                WHERE COALESCE(tournament_code, '') = :tournament_code
                  AND concept_key = :concept_key
                  AND active = TRUE
                LIMIT 1
                """
            ),
            {
                "tournament_code": tournament["tournament_code"] or "",
                "concept_key": concept_key,
            },
        )
    ).mappings().first()
    if existing:
        raise ValueError(
            "Ya existe una partida con ese nombre para el proyecto seleccionado."
        )
    concept_id = str(uuid.uuid4())
    metadata = build_budget_concept_scope_metadata(
        scope_labels,
        tournament_etapas=tournament.get("etapas"),
        tournament_categorias=[],
    )
    resolved_cuenta_id: Optional[str] = None
    clean_cuenta = _safe_str(cuenta_contable_id)
    if clean_cuenta:
        resolved_cuenta_id = await validate_active_cuenta_contable_id(
            session, clean_cuenta
        )
    await session.execute(
        text(
            """
            INSERT INTO budget_concepts (
                id, tournament_id, tournament_code, tournament_name,
                concept_name, concept_key, active, source, metadata,
                cuenta_contable_id, created_by_empleado_id, created_at, updated_at
            ) VALUES (
                :id, :tournament_id, :tournament_code, :tournament_name,
                :concept_name, :concept_key, TRUE, :source, CAST(:metadata AS jsonb),
                :cuenta_contable_id, :created_by_empleado_id, NOW(), NOW()
            )
            """
        ),
        {
            "id": concept_id,
            "tournament_id": tournament["tournament_id"],
            "tournament_code": tournament["tournament_code"],
            "tournament_name": tournament["tournament_name"],
            "concept_name": clean_name,
            "concept_key": concept_key,
            "source": _safe_str(source) or "admin_ui",
            "metadata": json.dumps(metadata, ensure_ascii=False),
            "cuenta_contable_id": resolved_cuenta_id,
            "created_by_empleado_id": actor_empleado_id,
        },
    )
    if commit:
        await session.commit()
    concept = await get_budget_concept(session, concept_id=concept_id)
    if concept is None:
        raise ValueError("No se pudo crear la partida presupuestal.")
    return concept


async def update_budget_concept(
    session: AsyncSession,
    *,
    concept_id: str,
    concept_name: Optional[str] = None,
    tournament_id: Optional[str] = None,
    scope_labels: Optional[list[str]] = None,
    cuenta_contable_id: Optional[str] = None,
    cuenta_contable_provided: bool = False,
    active: Optional[bool] = None,
    actor_empleado_id: Optional[str] = None,
    commit: bool = True,
) -> dict[str, Any]:
    del actor_empleado_id  # reserved for future audit events
    await ensure_budget_schema(session)
    current = await get_budget_concept(session, concept_id=concept_id)
    if current is None:
        raise ValueError("Partida presupuestal no encontrada.")
    updates: dict[str, Any] = {}
    if tournament_id is not None:
        tournament = await _resolve_tournament_for_budget_concept(
            session,
            tournament_id=tournament_id,
        )
        updates["tournament_id"] = tournament["tournament_id"]
        updates["tournament_code"] = tournament["tournament_code"]
        updates["tournament_name"] = tournament["tournament_name"]
    if concept_name is not None:
        clean_name = _safe_str(concept_name)
        current_name = _safe_str(current.get("concept_name"))
        if clean_name != current_name:
            concept_key = _normalize_budget_key(clean_name)
            if not clean_name or not concept_key:
                raise ValueError("El nombre de la partida presupuestal es obligatorio.")
            tournament_code = _safe_str(
                updates.get("tournament_code") or current.get("tournament_code")
            ) or ""
            duplicate = (
                await session.execute(
                    text(
                        """
                        SELECT id
                        FROM budget_concepts
                        WHERE COALESCE(tournament_code, '') = :tournament_code
                          AND concept_key = :concept_key
                          AND active = TRUE
                          AND id <> :concept_id
                        LIMIT 1
                        """
                    ),
                    {
                        "tournament_code": tournament_code,
                        "concept_key": concept_key,
                        "concept_id": concept_id,
                    },
                )
            ).mappings().first()
            if duplicate:
                raise ValueError(
                    "Ya existe otra partida con ese nombre para el mismo proyecto."
                )
            updates["concept_name"] = clean_name
            updates["concept_key"] = concept_key
    if scope_labels is not None:
        stored_metadata = (
            current.get("metadata")
            if isinstance(current.get("metadata"), dict)
            else {}
        )
        tournament_etapas = None
        if tournament_id is not None:
            tournament = await _resolve_tournament_for_budget_concept(
                session,
                tournament_id=tournament_id,
            )
            tournament_etapas = tournament.get("etapas")
        elif updates.get("tournament_id"):
            tournament = await _resolve_tournament_for_budget_concept(
                session,
                tournament_id=str(updates.get("tournament_id")),
            )
            tournament_etapas = tournament.get("etapas")
        else:
            current_tid = _safe_str(current.get("tournament_id"))
            if current_tid:
                tournament = await _resolve_tournament_for_budget_concept(
                    session,
                    tournament_id=current_tid,
                )
                tournament_etapas = tournament.get("etapas")
        updates["metadata"] = json.dumps(
            _merge_budget_concept_scope_metadata(
                stored_metadata,
                scope_labels,
                tournament_etapas=tournament_etapas,
                tournament_categorias=[],
            ),
            ensure_ascii=False,
        )
    if active is not None:
        updates["active"] = bool(active)
    if cuenta_contable_provided:
        clean_cuenta = _safe_str(cuenta_contable_id)
        if clean_cuenta:
            updates["cuenta_contable_id"] = await validate_active_cuenta_contable_id(
                session, clean_cuenta
            )
        else:
            updates["cuenta_contable_id"] = None
    if not updates:
        return current
    updates["source"] = "admin_ui"
    set_clause = ", ".join(f"{column} = :{column}" for column in updates)
    await session.execute(
        text(
            f"""
            UPDATE budget_concepts
            SET {set_clause},
                updated_at = NOW()
            WHERE id = :concept_id
            """
        ),
        {**updates, "concept_id": concept_id},
    )
    if commit:
        await session.commit()
    updated = await get_budget_concept(session, concept_id=concept_id)
    if updated is None:
        raise ValueError("No se pudo actualizar la partida presupuestal.")
    return updated


async def hide_budget_concept(
    session: AsyncSession,
    *,
    concept_id: str,
    actor_empleado_id: Optional[str] = None,
    commit: bool = True,
) -> dict[str, Any]:
    """Soft-hide a budget partida from selectors without deleting the row."""
    return await update_budget_concept(
        session,
        concept_id=concept_id,
        active=False,
        actor_empleado_id=actor_empleado_id,
        commit=commit,
    )


async def bulk_save_budget_concepts(
    session: AsyncSession,
    *,
    rows: list[dict[str, Any]],
    actor_empleado_id: Optional[str] = None,
) -> dict[str, Any]:
    created = 0
    updated = 0
    for row in rows:
        concept_id = _safe_str(row.get("concept_id"))
        concept_name = _safe_str(row.get("concept_name"))
        tournament_id = _safe_str(row.get("tournament_id"))
        sub_proyecto = _safe_str(row.get("sub_proyecto"))
        cuenta_contable_id = row.get("cuenta_contable_id")
        if not concept_name and not concept_id:
            continue
        if not concept_name or not tournament_id:
            raise ValueError(
                "Cada partida requiere nombre y proyecto. Revise las filas del catálogo."
            )
        scope_labels = [sub_proyecto] if sub_proyecto else []
        cuenta_kwargs = {
            "cuenta_contable_id": cuenta_contable_id,
            "cuenta_contable_provided": True,
        }
        if concept_id:
            await update_budget_concept(
                session,
                concept_id=concept_id,
                concept_name=concept_name,
                tournament_id=tournament_id,
                scope_labels=scope_labels,
                active=True,
                actor_empleado_id=actor_empleado_id,
                commit=False,
                **cuenta_kwargs,
            )
            updated += 1
        else:
            await create_budget_concept(
                session,
                tournament_id=tournament_id,
                concept_name=concept_name,
                scope_labels=scope_labels,
                cuenta_contable_id=_safe_str(cuenta_contable_id) or None,
                actor_empleado_id=actor_empleado_id,
                source="admin_ui",
                commit=False,
            )
            created += 1
    await session.commit()
    return {"created": created, "updated": updated, "total": created + updated}


async def import_budget_concepts_catalog(
    session: AsyncSession,
    *,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_CATALOG,
    actor_empleado_id: Optional[str] = None,
    source: str = "catalog_xlsx",
    force_overwrite: bool = False,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    concept_rows = _iter_budget_concept_catalog_rows(workbook_path)
    if not concept_rows:
        raise ValueError("No se encontraron conceptos presupuestales en el archivo.")
    tournament_rows = await _load_tournament_rows(session)
    created = 0
    updated = 0
    skipped = 0
    affected_codes: set[str] = set()
    for row in concept_rows:
        tournament_code = _safe_str(row.get("tournament_code")) or None
        tournament_name = _safe_str(row.get("tournament_name")) or "Sin torneo"
        concept_key = _safe_str(row.get("concept_key"))
        concept_name = _safe_str(row.get("concept_name"))
        if not concept_key or not concept_name:
            continue
        tournament_id = _match_tournament_id(
            tournament_rows,
            tournament_code=tournament_code,
            tournament_name=tournament_name,
        )
        if tournament_id:
            try:
                tournament = await _resolve_tournament_for_budget_concept(
                    session,
                    tournament_id=tournament_id,
                )
                tournament_name = tournament["tournament_name"]
            except ValueError:
                pass
        existing = (
            await session.execute(
                text(
                    """
                    SELECT id, source
                    FROM budget_concepts
                    WHERE COALESCE(tournament_code, '') = :tournament_code
                      AND concept_key = :concept_key
                    LIMIT 1
                    """
                ),
                {
                    "tournament_code": tournament_code or "",
                    "concept_key": concept_key,
                },
            )
        ).mappings().first()
        if existing and _safe_str(existing.get("source")) == "admin_ui" and not force_overwrite:
            skipped += 1
            continue
        payload = {
            "tournament_id": tournament_id,
            "tournament_code": tournament_code,
            "tournament_name": tournament_name,
            "concept_name": concept_name,
            "concept_key": concept_key,
            "source": _safe_str(source) or "catalog_xlsx",
            "metadata": json.dumps(row.get("metadata") or {}, ensure_ascii=False),
            "created_by_empleado_id": actor_empleado_id,
        }
        if existing:
            await session.execute(
                text(
                    """
                    UPDATE budget_concepts
                    SET tournament_id = :tournament_id,
                        tournament_code = :tournament_code,
                        tournament_name = :tournament_name,
                        concept_name = :concept_name,
                        source = :source,
                        metadata = CAST(:metadata AS jsonb),
                        active = TRUE,
                        updated_at = NOW()
                    WHERE id = :id
                    """
                ),
                {**payload, "id": existing["id"]},
            )
            updated += 1
        else:
            await session.execute(
                text(
                    """
                    INSERT INTO budget_concepts (
                        id, tournament_id, tournament_code, tournament_name,
                        concept_name, concept_key, active, source, metadata,
                        created_by_empleado_id, created_at, updated_at
                    ) VALUES (
                        :id, :tournament_id, :tournament_code, :tournament_name,
                        :concept_name, :concept_key, TRUE, :source, CAST(:metadata AS jsonb),
                        :created_by_empleado_id, NOW(), NOW()
                    )
                    """
                ),
                {**payload, "id": str(uuid.uuid4())},
            )
            created += 1
        if tournament_code:
            affected_codes.add(tournament_code)
    await session.commit()
    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "tournaments_count": len(affected_codes),
        "concepts_count": created + updated,
        "workbook_path": str(Path(workbook_path)),
    }


COPA_TELMEX_TELCEL_2026_TOURNAMENT_ID = "67c556df-f953-4f99-8cab-87ed4d111b05"


async def restore_budget_concepts_for_tournament(
    session: AsyncSession,
    *,
    tournament_id: str,
    workbook_path: str | Path = DEFAULT_BUDGET_CONCEPT_CATALOG,
    source: str = "catalog_snapshot",
    actor_empleado_id: Optional[str] = None,
) -> dict[str, Any]:
    """Re-apply catalog metadata for one project without touching admin_ui rows."""
    await ensure_budget_schema(session)
    tournament = await _resolve_tournament_for_budget_concept(
        session,
        tournament_id=tournament_id,
    )
    catalog_by_key = {
        _safe_str(row.get("concept_key")): row
        for row in _iter_budget_concept_catalog_rows(workbook_path)
        if _safe_str(row.get("tournament_code")).upper()
        == _safe_str(tournament["tournament_code"]).upper()
    }
    if not catalog_by_key:
        raise ValueError(
            f"No hay partidas de catálogo para el código {tournament['tournament_code']}."
        )
    rows = (
        await session.execute(
            text(
                """
                SELECT id, concept_key, source
                FROM budget_concepts
                WHERE tournament_id = CAST(:tournament_id AS uuid)
                   OR (
                        COALESCE(tournament_code, '') = :tournament_code
                        AND tournament_id IS NULL
                   )
                """
            ),
            {
                "tournament_id": tournament["tournament_id"],
                "tournament_code": tournament["tournament_code"] or "",
            },
        )
    ).mappings().all()
    restored = 0
    skipped = 0
    for row in rows:
        if _safe_str(row.get("source")) == "admin_ui":
            skipped += 1
            continue
        concept_key = _safe_str(row.get("concept_key"))
        catalog_row = catalog_by_key.get(concept_key)
        if catalog_row is None:
            continue
        await session.execute(
            text(
                """
                UPDATE budget_concepts
                SET tournament_id = CAST(:tournament_id AS uuid),
                    tournament_code = :tournament_code,
                    tournament_name = :tournament_name,
                    concept_name = :concept_name,
                    metadata = CAST(:metadata AS jsonb),
                    source = :source,
                    active = TRUE,
                    updated_at = NOW()
                WHERE id = :id
                """
            ),
            {
                "id": row["id"],
                "tournament_id": tournament["tournament_id"],
                "tournament_code": tournament["tournament_code"],
                "tournament_name": tournament["tournament_name"],
                "concept_name": _safe_str(catalog_row.get("concept_name")),
                "metadata": json.dumps(
                    catalog_row.get("metadata") or {},
                    ensure_ascii=False,
                ),
                "source": _safe_str(source) or "catalog_snapshot",
            },
        )
        restored += 1
    await session.commit()
    return {
        "ok": True,
        "tournament_id": tournament["tournament_id"],
        "tournament_name": tournament["tournament_name"],
        "restored": restored,
        "skipped_admin_ui": skipped,
        "catalog_rows": len(catalog_by_key),
    }


def _load_tabular_upload_rows(
    *,
    file_bytes: bytes,
    filename: str,
) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        text_data = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text_data))
        return [dict(row) for row in reader]
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        headers = [_safe_str(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(
                {
                    headers[index] or f"column_{index+1}": values[index]
                    for index in range(min(len(headers), len(values)))
                }
            )
        if rows and not any(key for key in headers):
            raise ValueError("El archivo no tiene encabezados válidos.")
        return rows
    raise ValueError("Solo se permiten archivos CSV o XLSX.")


def _pick_upload_value(row: dict[str, Any], *keys: str) -> Any:
    normalized_row = {
        _normalize_budget_key(key): value for key, value in row.items() if _safe_str(key)
    }
    for key in keys:
        normalized_key = _normalize_budget_key(key)
        if normalized_key in normalized_row and normalized_row[normalized_key] not in (None, ""):
            return normalized_row[normalized_key]
    return None


async def import_budget_lines_upload(
    session: AsyncSession,
    *,
    version_id: str,
    actor_empleado_id: Optional[str],
    file_bytes: bytes,
    filename: str,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    current = await get_budget_version(session, version_id=version_id)
    if not _editable_version_status(current.get("status")):
        raise ValueError("Only draft or reforecast versions allow line imports")
    rows = _load_tabular_upload_rows(file_bytes=file_bytes, filename=filename)
    if not rows:
        raise ValueError("El archivo no contiene líneas presupuestales.")
    concepts = await list_budget_concepts(session, active_only=True, limit=5000)
    concepts_by_scope: dict[tuple[str, str], dict[str, Any]] = {}
    for concept in concepts:
        concepts_by_scope[
            (
                _safe_str(concept.get("tournament_code")),
                _safe_str(concept.get("concept_key")),
            )
        ] = concept
    existing_lines = await list_budget_lines(session, version_id=version_id, limit=5000)
    existing_by_scope = {
        (
            _safe_str(line.get("tournament_code")),
            _safe_str(line.get("budget_concept_id") or ""),
        ): line
        for line in existing_lines
    }
    created = 0
    updated = 0
    for row in rows:
        tournament_value = _safe_str(
            _pick_upload_value(row, "torneo", "tournament", "torneo_codigo", "tournament_code")
        )
        concept_value = _safe_str(
            _pick_upload_value(row, "partida_presupuestal", "concepto", "partida", "concept_name")
        )
        amount_value = _pick_upload_value(row, "monto_anual", "presupuesto", "budget_amount")
        if not tournament_value or not concept_value:
            raise ValueError("Cada fila debe incluir torneo y partida presupuestal.")
        tournament_code = _derive_budget_tournament_code(tournament_value, tournament_value)
        concept_key = _normalize_budget_key(concept_value)
        concept = concepts_by_scope.get((tournament_code, concept_key))
        if concept is None:
            raise ValueError(
                f"No existe la partida presupuestal '{concept_value}' para '{tournament_value}'."
            )
        budget_amount = _safe_decimal(amount_value)
        line_key = (tournament_code, _safe_str(concept.get("id")))
        line_updates = {
            "budget_concept_id": _safe_str(concept.get("id")),
            "concept_name": _safe_str(concept.get("concept_name")),
            "account_code_final": _safe_str(
                _pick_upload_value(row, "cuenta_final", "account_code_final")
            )
            or None,
            "phase": _safe_str(_pick_upload_value(row, "fase", "phase")) or None,
            "owner_name": _safe_str(
                _pick_upload_value(row, "responsable", "owner_name")
            )
            or None,
            "priority": _safe_str(_pick_upload_value(row, "prioridad", "priority")) or None,
            "criteria_note": _safe_str(
                _pick_upload_value(row, "criterio", "criteria_note")
            )
            or None,
            "observations": _safe_str(
                _pick_upload_value(row, "observaciones", "observations")
            )
            or None,
            "budget_amount": budget_amount,
        }
        existing_line = existing_by_scope.get(line_key)
        if existing_line:
            updated_line = await update_budget_line(
                session,
                line_id=_safe_str(existing_line.get("id")),
                actor_empleado_id=actor_empleado_id,
                updates=line_updates,
            )
            existing_by_scope[line_key] = updated_line
            updated += 1
            continue
        created_line = await create_budget_line(
            session,
            version_id=version_id,
            actor_empleado_id=actor_empleado_id,
            tournament_code=_safe_str(concept.get("tournament_code")) or tournament_code,
            tournament_name=_safe_str(concept.get("tournament_name")) or tournament_value,
            budget_concept_id=_safe_str(concept.get("id")),
            concept_name=_safe_str(concept.get("concept_name")) or concept_value,
            account_code_final=line_updates["account_code_final"],
            phase=line_updates["phase"],
            owner_name=line_updates["owner_name"],
            priority=line_updates["priority"],
            budget_amount=budget_amount,
            criteria_note=line_updates["criteria_note"],
            observations=line_updates["observations"],
        )
        existing_by_scope[line_key] = created_line
        created += 1
    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "rows_processed": created + updated,
        "filename": filename,
    }


def load_budget_artifact_rows(
    artifact_path: str | Path = DEFAULT_BUDGET_ARTIFACT,
) -> list[dict[str, Any]]:
    path = Path(artifact_path)
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def build_budget_artifact_snapshot(
    rows: list[dict[str, Any]],
    *,
    artifact_path: str | Path = DEFAULT_BUDGET_ARTIFACT,
    tournament_name: Optional[str] = None,
    tournament_slug: Optional[str] = None,
    edition_year: int = 2026,
) -> dict[str, Any]:
    aliases = budget_alias_candidates(tournament_name or "", tournament_slug or "")
    filtered_rows = [
        row
        for row in rows
        if _matches_aliases(
            aliases=aliases,
            tournament_code=_safe_str(row.get("torneo_codigo")),
            tournament_name=_safe_str(row.get("torneo")),
        )
    ]

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in filtered_rows:
        grouped[_safe_str(row.get("torneo")) or "Torneo sin nombre"].append(row)

    tournaments: list[dict[str, Any]] = []
    total_budget = 0.0
    total_reference = 0.0
    total_lines = 0

    for tournament, items in sorted(grouped.items()):
        budget_total = sum(_safe_decimal(item.get("presupuesto_2026")) for item in items)
        reference_total = sum(
            _safe_decimal(item.get("importe_referencia_total")) for item in items
        )
        total_budget += budget_total
        total_reference += reference_total
        total_lines += len(items)
        priorities = defaultdict(int)
        stages = defaultdict(int)
        for item in items:
            priorities[_safe_str(item.get("prioridad")) or "sin_prioridad"] += 1
            stages[_safe_str(item.get("etapa")) or "sin_etapa"] += 1
        comparison = {
            "requested_total": 0.0,
            "committed_total": 0.0,
            "paid_total": 0.0,
            "actual_total": 0.0,
            "pending_to_pay_total": 0.0,
        }
        forecast = _build_budget_forecast(
            edition_year=edition_year,
            budget_total=round(budget_total, 2),
            comparison=comparison,
        )
        tournaments.append(
            {
                "tournament_id": None,
                "tournament_code": _safe_str(items[0].get("torneo_codigo")),
                "tournament_name": tournament,
                "edition_year": edition_year,
                "line_count": len(items),
                "budget_total": round(budget_total, 2),
                "reference_total": round(reference_total, 2),
                "comparison": comparison,
                "forecast": forecast,
                "scenarios": _build_budget_scenarios(
                    edition_year=edition_year,
                    budget_total=round(budget_total, 2),
                    comparison=comparison,
                    forecast=forecast,
                ),
                "top_concepts": [
                    {
                        "concepto": _safe_str(item.get("concepto")),
                        "presupuesto_2026": round(
                            _safe_decimal(item.get("presupuesto_2026")), 2
                        ),
                        "cuenta_contable_final": _safe_str(
                            item.get("cuenta_contable_final")
                        )
                        or _safe_str(item.get("cuenta_contable_sugerida")),
                        "prioridad": _safe_str(item.get("prioridad")) or None,
                        "etapa": _safe_str(item.get("etapa")) or None,
                    }
                    for item in sorted(
                        items,
                        key=lambda entry: _safe_decimal(entry.get("presupuesto_2026")),
                        reverse=True,
                    )[:8]
                ],
                "priorities": dict(sorted(priorities.items())),
                "stages": dict(sorted(stages.items())),
                "breakdowns": _build_artifact_breakdowns(items),
            }
        )

    artifact_breakdowns = _build_artifact_breakdowns(filtered_rows)
    summary_comparison = {
        "requested_total": 0.0,
        "committed_total": 0.0,
        "paid_total": 0.0,
        "actual_total": 0.0,
        "pending_to_pay_total": 0.0,
    }
    summary_forecast = _build_budget_forecast(
        edition_year=edition_year,
        budget_total=round(total_budget, 2),
        comparison=summary_comparison,
    )
    return {
        "ok": True,
        "source": "budget_artifact_csv",
        "artifact_path": str(Path(artifact_path)),
        "summary": {
            "edition_year": edition_year,
            "tournaments_count": len(tournaments),
            "line_count": total_lines,
            "budget_total": round(total_budget, 2),
            "reference_total": round(total_reference, 2),
        },
        "comparison": summary_comparison,
        "forecast": summary_forecast,
        "scenarios": _build_budget_scenarios(
            edition_year=edition_year,
            budget_total=round(total_budget, 2),
            comparison=summary_comparison,
            forecast=summary_forecast,
        ),
        "breakdowns": artifact_breakdowns,
        "tournaments": tournaments,
    }


async def ensure_budget_schema(session: AsyncSession) -> None:
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_concepts (
                id UUID PRIMARY KEY,
                tournament_id UUID NULL REFERENCES tournaments(id) ON DELETE SET NULL,
                tournament_code VARCHAR(40) NULL,
                tournament_name VARCHAR(200) NOT NULL,
                concept_name VARCHAR(200) NOT NULL,
                concept_key VARCHAR(200) NOT NULL,
                active BOOLEAN NOT NULL DEFAULT TRUE,
                source VARCHAR(80) NOT NULL DEFAULT 'manual',
                metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_by_empleado_id UUID NULL REFERENCES empleados(id) ON DELETE SET NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            """
            ALTER TABLE budget_concepts
            ADD COLUMN IF NOT EXISTS cuenta_contable_id UUID NULL
            REFERENCES cuentas_contables(id) ON DELETE SET NULL
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_versions (
                id UUID PRIMARY KEY,
                edition_year INTEGER NOT NULL,
                version_name VARCHAR(120) NOT NULL,
                status VARCHAR(40) NOT NULL DEFAULT 'draft',
                source VARCHAR(80) NOT NULL DEFAULT 'manual',
                notes TEXT NULL,
                artifact_path TEXT NULL,
                created_by_empleado_id UUID NULL REFERENCES empleados(id),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            """
            ALTER TABLE expense_reports
            ADD COLUMN IF NOT EXISTS budget_concept_id UUID NULL
            REFERENCES budget_concepts(id) ON DELETE SET NULL
            """
        )
    )
    await session.execute(
        text(
            """
            ALTER TABLE documentos
            ADD COLUMN IF NOT EXISTS budget_concept_id UUID NULL
            REFERENCES budget_concepts(id) ON DELETE SET NULL
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_lines (
                id UUID PRIMARY KEY,
                budget_version_id UUID NOT NULL REFERENCES budget_versions(id) ON DELETE CASCADE,
                budget_concept_id UUID NULL REFERENCES budget_concepts(id) ON DELETE SET NULL,
                tournament_id UUID NULL REFERENCES tournaments(id) ON DELETE SET NULL,
                tournament_code VARCHAR(40) NULL,
                tournament_name VARCHAR(200) NOT NULL,
                sport VARCHAR(40) NULL,
                phase VARCHAR(80) NULL,
                entity_name VARCHAR(200) NULL,
                concept_name VARCHAR(200) NOT NULL,
                account_code_suggested VARCHAR(80) NULL,
                account_code_final VARCHAR(80) NULL,
                budget_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
                reference_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
                variance_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
                priority VARCHAR(40) NULL,
                owner_name VARCHAR(200) NULL,
                criteria_note TEXT NULL,
                observations TEXT NULL,
                metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            """
            ALTER TABLE budget_lines
            ADD COLUMN IF NOT EXISTS budget_concept_id UUID NULL
            REFERENCES budget_concepts(id) ON DELETE SET NULL
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_scenarios (
                id UUID PRIMARY KEY,
                budget_version_id UUID NOT NULL REFERENCES budget_versions(id) ON DELETE CASCADE,
                scenario_key VARCHAR(80) NOT NULL,
                scenario_name VARCHAR(120) NOT NULL,
                status VARCHAR(40) NOT NULL DEFAULT 'draft',
                assumptions JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                UNIQUE (budget_version_id, scenario_key)
            )
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_line_monthly_allocations (
                id UUID PRIMARY KEY,
                budget_line_id UUID NOT NULL REFERENCES budget_lines(id) ON DELETE CASCADE,
                month_number INTEGER NOT NULL,
                allocated_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
                metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                UNIQUE (budget_line_id, month_number)
            )
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_approvals (
                id UUID PRIMARY KEY,
                budget_version_id UUID NOT NULL REFERENCES budget_versions(id) ON DELETE CASCADE,
                approved_by_empleado_id UUID NULL REFERENCES empleados(id),
                approval_status VARCHAR(40) NOT NULL DEFAULT 'draft',
                approval_note TEXT NULL,
                approved_at TIMESTAMPTZ NULL,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_actuals_bridge (
                id UUID PRIMARY KEY,
                budget_line_id UUID NOT NULL REFERENCES budget_lines(id) ON DELETE CASCADE,
                expense_report_id UUID NULL REFERENCES expense_reports(id) ON DELETE SET NULL,
                documento_id UUID NULL REFERENCES documentos(id) ON DELETE SET NULL,
                actual_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
                source VARCHAR(40) NOT NULL DEFAULT 'expense_report',
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS budget_version_audit_log (
                id UUID PRIMARY KEY,
                budget_version_id UUID NOT NULL REFERENCES budget_versions(id) ON DELETE CASCADE,
                event_type VARCHAR(60) NOT NULL,
                actor_empleado_id UUID NULL REFERENCES empleados(id),
                from_status VARCHAR(40) NULL,
                to_status VARCHAR(40) NULL,
                note TEXT NULL,
                payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    await session.execute(
        text(
            "CREATE UNIQUE INDEX IF NOT EXISTS ux_budget_concepts_tournament_code_key "
            "ON budget_concepts(COALESCE(tournament_code, ''), concept_key)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_concepts_tournament_id "
            "ON budget_concepts(tournament_id)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_concepts_active "
            "ON budget_concepts(active)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_versions_edition_year "
            "ON budget_versions(edition_year)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_versions_status "
            "ON budget_versions(status)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_lines_version "
            "ON budget_lines(budget_version_id)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_lines_tournament_id "
            "ON budget_lines(tournament_id)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_lines_tournament_code "
            "ON budget_lines(tournament_code)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_lines_budget_concept "
            "ON budget_lines(budget_concept_id)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_line_monthly_allocations_line "
            "ON budget_line_monthly_allocations(budget_line_id)"
        )
    )
    await session.execute(
        text(
            "CREATE INDEX IF NOT EXISTS ix_budget_version_audit_log_version "
            "ON budget_version_audit_log(budget_version_id)"
        )
    )
    await session.commit()


async def _audit_budget_event(
    session: AsyncSession,
    *,
    budget_version_id: str,
    event_type: str,
    actor_empleado_id: Optional[str],
    from_status: Optional[str] = None,
    to_status: Optional[str] = None,
    note: Optional[str] = None,
    payload: Optional[dict[str, Any]] = None,
) -> None:
    await session.execute(
        text(
            """
            INSERT INTO budget_version_audit_log (
                id, budget_version_id, event_type, actor_empleado_id,
                from_status, to_status, note, payload, created_at
            ) VALUES (
                :id, :budget_version_id, :event_type, :actor_empleado_id,
                :from_status, :to_status, :note, CAST(:payload AS jsonb), NOW()
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "budget_version_id": budget_version_id,
            "event_type": event_type,
            "actor_empleado_id": actor_empleado_id,
            "from_status": from_status,
            "to_status": to_status,
            "note": note,
            "payload": json.dumps(payload or {}, ensure_ascii=False),
        },
    )


async def list_budget_audit_events(
    session: AsyncSession,
    *,
    version_id: Optional[str] = None,
    limit: int = 80,
) -> list[dict[str, Any]]:
    await ensure_budget_schema(session)
    filters = []
    params: dict[str, Any] = {"limit": max(1, min(limit, 300))}
    if version_id:
        filters.append("log.budget_version_id = :version_id")
        params["version_id"] = version_id
    where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    log.id,
                    log.budget_version_id,
                    v.version_name,
                    v.status AS version_status,
                    log.event_type,
                    log.actor_empleado_id,
                    actor.nombre AS actor_nombre,
                    actor.correo AS actor_correo,
                    log.from_status,
                    log.to_status,
                    log.note,
                    log.payload,
                    log.created_at
                FROM budget_version_audit_log log
                JOIN budget_versions v ON v.id = log.budget_version_id
                LEFT JOIN empleados actor ON actor.id = log.actor_empleado_id
                {where_clause}
                ORDER BY log.created_at DESC
                LIMIT :limit
                """
            ),
            params,
        )
    ).mappings().all()
    events: list[dict[str, Any]] = []
    for row in rows:
        payload = row["payload"] if isinstance(row["payload"], dict) else {}
        events.append(
            {
                "id": _safe_str(row["id"]),
                "budget_version_id": _safe_str(row["budget_version_id"]),
                "version_name": _safe_str(row["version_name"]) or None,
                "version_status": _safe_str(row["version_status"]) or None,
                "event_type": _safe_str(row["event_type"]) or None,
                "actor_empleado_id": _safe_str(row["actor_empleado_id"]) or None,
                "actor_nombre": _safe_str(row["actor_nombre"]) or "Sistema",
                "actor_correo": _safe_str(row["actor_correo"]) or None,
                "from_status": _safe_str(row["from_status"]) or None,
                "to_status": _safe_str(row["to_status"]) or None,
                "note": _safe_str(row["note"]) or None,
                "payload": payload,
                "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            }
        )
    return events


async def _link_budget_lines_to_tournaments(
    session: AsyncSession, *, budget_version_id: str
) -> None:
    tournament_rows = await _load_tournament_rows(session)
    if not tournament_rows:
        return
    line_rows = (
        await session.execute(
            text(
                """
                SELECT id, tournament_code, tournament_name
                FROM budget_lines
                WHERE budget_version_id = :budget_version_id
                """
            ),
            {"budget_version_id": budget_version_id},
        )
    ).mappings().all()
    for row in line_rows:
        tournament_id = _match_tournament_id(
            tournament_rows,
            tournament_code=_safe_str(row.get("tournament_code")),
            tournament_name=_safe_str(row.get("tournament_name")),
        )
        if not tournament_id:
            continue
        await session.execute(
            text(
                """
                UPDATE budget_lines
                SET tournament_id = :tournament_id, updated_at = NOW()
                WHERE id = :line_id
                """
            ),
            {"line_id": row["id"], "tournament_id": tournament_id},
        )


async def import_budget_artifact(
    session: AsyncSession,
    *,
    artifact_path: str | Path = DEFAULT_BUDGET_ARTIFACT,
    edition_year: int = 2026,
    version_name: str = "Presupuesto 2026 Draft",
    status: str = "draft",
    source: str = "artifact_csv",
    created_by_empleado_id: Optional[str] = None,
    replace_existing: bool = False,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    artifact_rows = load_budget_artifact_rows(artifact_path)
    if not artifact_rows:
        raise ValueError("No budget artifact rows were found")

    existing_version_id = (
        await session.execute(
            text(
                """
                SELECT id
                FROM budget_versions
                WHERE edition_year = :edition_year
                  AND version_name = :version_name
                ORDER BY created_at DESC
                LIMIT 1
                """
            ),
            {"edition_year": edition_year, "version_name": version_name},
        )
    ).scalar_one_or_none()
    if existing_version_id and not replace_existing:
        return {
            "ok": True,
            "created": False,
            "version_id": str(existing_version_id),
            "message": "Budget version already exists",
        }

    if existing_version_id and replace_existing:
        await session.execute(
            text("DELETE FROM budget_versions WHERE id = :version_id"),
            {"version_id": existing_version_id},
        )

    version_id = str(uuid.uuid4())
    normalized_status = _normalize_budget_status(status)
    await session.execute(
        text(
            """
            INSERT INTO budget_versions (
                id, edition_year, version_name, status, source, notes, artifact_path,
                created_by_empleado_id, created_at, updated_at
            ) VALUES (
                :id, :edition_year, :version_name, :status, :source, :notes, :artifact_path,
                :created_by_empleado_id, NOW(), NOW()
            )
            """
        ),
        {
            "id": version_id,
            "edition_year": edition_year,
            "version_name": version_name,
            "status": normalized_status,
            "source": source,
            "notes": "Importado desde artefacto presupuestal 2026.",
            "artifact_path": str(Path(artifact_path)),
            "created_by_empleado_id": created_by_empleado_id,
        },
    )

    for row in artifact_rows:
        metadata = {
            "confianza_base": _safe_str(row.get("confianza_base")) or None,
            "referencias_usadas": _safe_str(row.get("referencias_usadas")) or None,
            "importe_referencia_promedio": _safe_decimal(
                row.get("importe_referencia_promedio")
            ),
            "importe_referencia_maximo": _safe_decimal(
                row.get("importe_referencia_maximo")
            ),
            "proveedores_distintos": _safe_str(row.get("proveedores_distintos")) or None,
            "estatus_torneo": _safe_str(row.get("estatus_torneo")) or None,
        }
        await session.execute(
            text(
                """
                INSERT INTO budget_lines (
                    id, budget_version_id, tournament_code, tournament_name, phase,
                    concept_name, account_code_suggested, account_code_final,
                    budget_amount, reference_amount, variance_amount, priority,
                    owner_name, criteria_note, observations, metadata,
                    created_at, updated_at
                ) VALUES (
                    :id, :budget_version_id, :tournament_code, :tournament_name, :phase,
                    :concept_name, :account_code_suggested, :account_code_final,
                    :budget_amount, :reference_amount, :variance_amount, :priority,
                    :owner_name, :criteria_note, :observations, CAST(:metadata AS jsonb),
                    NOW(), NOW()
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "budget_version_id": version_id,
                "tournament_code": _safe_str(row.get("torneo_codigo")) or None,
                "tournament_name": _safe_str(row.get("torneo")) or "Torneo sin nombre",
                "phase": _safe_str(row.get("etapa")) or None,
                "concept_name": _safe_str(row.get("concepto")) or "Sin concepto",
                "account_code_suggested": _safe_str(
                    row.get("cuenta_contable_sugerida")
                )
                or None,
                "account_code_final": _safe_str(row.get("cuenta_contable_final"))
                or None,
                "budget_amount": _safe_decimal(row.get("presupuesto_2026")),
                "reference_amount": _safe_decimal(row.get("importe_referencia_total")),
                "variance_amount": _safe_decimal(row.get("ajuste_vs_referencia")),
                "priority": _safe_str(row.get("prioridad")) or None,
                "owner_name": _safe_str(row.get("responsable")) or None,
                "criteria_note": _safe_str(row.get("criterio_presupuesto")) or None,
                "observations": _safe_str(row.get("observaciones_presupuesto")) or None,
                "metadata": json.dumps(metadata, ensure_ascii=False),
            },
        )
    await _link_budget_lines_to_tournaments(session, budget_version_id=version_id)
    await _audit_budget_event(
        session,
        budget_version_id=version_id,
        event_type="imported",
        actor_empleado_id=created_by_empleado_id,
        to_status=normalized_status,
        note="Importación inicial del artefacto presupuestal.",
        payload={"artifact_path": str(Path(artifact_path)), "line_count": len(artifact_rows)},
    )
    await session.commit()
    return {
        "ok": True,
        "created": True,
        "version_id": version_id,
        "line_count": len(artifact_rows),
    }


async def create_budget_version(
    session: AsyncSession,
    *,
    edition_year: int = 2026,
    version_name: str,
    notes: Optional[str] = None,
    source: str = "ui_manual",
    created_by_empleado_id: Optional[str] = None,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    clean_name = _safe_str(version_name)
    if not clean_name:
        raise ValueError("Budget version name is required")

    version_id = str(uuid.uuid4())
    await session.execute(
        text(
            """
            INSERT INTO budget_versions (
                id, edition_year, version_name, status, source, notes, artifact_path,
                created_by_empleado_id, created_at, updated_at
            ) VALUES (
                :id, :edition_year, :version_name, 'draft', :source, :notes, NULL,
                :created_by_empleado_id, NOW(), NOW()
            )
            """
        ),
        {
            "id": version_id,
            "edition_year": edition_year,
            "version_name": clean_name,
            "source": _safe_str(source) or "ui_manual",
            "notes": notes,
            "created_by_empleado_id": created_by_empleado_id,
        },
    )
    await _audit_budget_event(
        session,
        budget_version_id=version_id,
        event_type="version_created",
        actor_empleado_id=created_by_empleado_id,
        to_status="draft",
        note=notes,
        payload={"source": _safe_str(source) or "ui_manual"},
    )
    await session.commit()
    return await get_budget_version(session, version_id=version_id)


async def list_budget_versions(
    session: AsyncSession,
    *,
    edition_year: Optional[int] = None,
) -> list[dict[str, Any]]:
    await ensure_budget_schema(session)
    filters = []
    params: dict[str, Any] = {}
    if edition_year is not None:
        filters.append("v.edition_year = :edition_year")
        params["edition_year"] = edition_year
    where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    v.id,
                    v.edition_year,
                    v.version_name,
                    v.status,
                    v.source,
                    v.notes,
                    v.artifact_path,
                    v.created_by_empleado_id,
                    v.created_at,
                    v.updated_at,
                    COUNT(l.id) AS line_count,
                    COALESCE(SUM(l.budget_amount), 0) AS budget_total,
                    COALESCE(SUM(l.reference_amount), 0) AS reference_total,
                    MAX(a.approved_at) AS latest_approved_at
                FROM budget_versions v
                LEFT JOIN budget_lines l ON l.budget_version_id = v.id
                LEFT JOIN budget_approvals a ON a.budget_version_id = v.id
                {where_clause}
                GROUP BY
                    v.id, v.edition_year, v.version_name, v.status, v.source,
                    v.notes, v.artifact_path, v.created_by_empleado_id,
                    v.created_at, v.updated_at
                ORDER BY v.edition_year DESC, v.updated_at DESC, v.created_at DESC
                """
            ),
            params,
        )
    ).mappings().all()
    return [
        {
            "id": _safe_str(row["id"]),
            "edition_year": int(row["edition_year"] or 0),
            "version_name": _safe_str(row["version_name"]),
            "status": _safe_str(row["status"]),
            "source": _safe_str(row["source"]),
            "notes": _safe_str(row["notes"]) or None,
            "artifact_path": _safe_str(row["artifact_path"]) or None,
            "created_by_empleado_id": _safe_str(row["created_by_empleado_id"]) or None,
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            "line_count": int(row["line_count"] or 0),
            "budget_total": round(_safe_decimal(row["budget_total"]), 2),
            "reference_total": round(_safe_decimal(row["reference_total"]), 2),
            "latest_approved_at": row["latest_approved_at"].isoformat()
            if row["latest_approved_at"]
            else None,
        }
        for row in rows
    ]


async def create_budget_line(
    session: AsyncSession,
    *,
    version_id: str,
    actor_empleado_id: Optional[str] = None,
    tournament_code: Optional[str] = None,
    tournament_name: Optional[str] = None,
    budget_concept_id: Optional[str] = None,
    concept_name: str,
    account_code_final: Optional[str] = None,
    phase: Optional[str] = None,
    owner_name: Optional[str] = None,
    priority: Optional[str] = None,
    budget_amount: Any = 0,
    reference_amount: Any = 0,
    criteria_note: Optional[str] = None,
    observations: Optional[str] = None,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    current = await get_budget_version(session, version_id=version_id)
    if not _editable_version_status(current.get("status")):
        raise ValueError("Only draft or reforecast versions allow line creation")

    clean_concept = _safe_str(concept_name)
    concept = await resolve_budget_concept(
        session,
        budget_concept_id=budget_concept_id,
        tournament_code=tournament_code,
    )
    if concept is not None:
        clean_concept = _safe_str(concept.get("concept_name"))
        tournament_code = _safe_str(concept.get("tournament_code")) or tournament_code
        tournament_name = _safe_str(concept.get("tournament_name")) or tournament_name
    if not clean_concept:
        raise ValueError("Budget line concept is required")
    amount = _safe_decimal(budget_amount)
    reference = _safe_decimal(reference_amount)
    line_id = str(uuid.uuid4())
    await session.execute(
        text(
            """
            INSERT INTO budget_lines (
                id, budget_version_id, budget_concept_id, tournament_code, tournament_name, phase,
                concept_name, account_code_suggested, account_code_final,
                budget_amount, reference_amount, variance_amount, priority,
                owner_name, criteria_note, observations, metadata,
                created_at, updated_at
            ) VALUES (
                :id, :budget_version_id, :budget_concept_id, :tournament_code, :tournament_name, :phase,
                :concept_name, NULL, :account_code_final,
                :budget_amount, :reference_amount, :variance_amount, :priority,
                :owner_name, :criteria_note, :observations, CAST(:metadata AS jsonb),
                NOW(), NOW()
            )
            """
        ),
        {
            "id": line_id,
            "budget_version_id": version_id,
            "budget_concept_id": (
                _safe_str(concept.get("id")) or None if concept else None
            ),
            "tournament_code": _safe_str(tournament_code) or None,
            "tournament_name": _safe_str(tournament_name) or "Presupuesto general",
            "phase": _safe_str(phase) or None,
            "concept_name": clean_concept,
            "account_code_final": _safe_str(account_code_final) or None,
            "budget_amount": amount,
            "reference_amount": reference,
            "variance_amount": round(amount - reference, 2),
            "priority": _safe_str(priority) or None,
            "owner_name": _safe_str(owner_name) or None,
            "criteria_note": _safe_str(criteria_note) or None,
            "observations": _safe_str(observations) or None,
            "metadata": json.dumps({"created_from": "ui_manual"}, ensure_ascii=False),
        },
    )
    await _link_budget_lines_to_tournaments(session, budget_version_id=version_id)
    await _audit_budget_event(
        session,
        budget_version_id=version_id,
        event_type="line_created",
        actor_empleado_id=actor_empleado_id,
        from_status=current.get("status"),
        to_status=current.get("status"),
        payload={
            "line_id": line_id,
            "budget_concept_id": (
                _safe_str(concept.get("id")) or None if concept else None
            ),
            "concept_name": clean_concept,
            "budget_amount": amount,
        },
    )
    await session.commit()
    lines = await list_budget_lines(session, version_id=version_id, limit=500)
    for line in lines:
        if line["id"] == line_id:
            return line
    raise ValueError("Budget line not found after creation")


async def transition_budget_version(
    session: AsyncSession,
    *,
    version_id: str,
    new_status: str,
    actor_empleado_id: Optional[str] = None,
    note: Optional[str] = None,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    current = (
        await session.execute(
            text(
                """
                SELECT id, status
                FROM budget_versions
                WHERE id = :version_id
                LIMIT 1
                """
            ),
            {"version_id": version_id},
        )
    ).mappings().first()
    if not current:
        raise ValueError("Budget version not found")
    from_status = _normalize_budget_status(current["status"])
    to_status = _normalize_budget_status(new_status)
    if to_status == from_status:
        versions = await list_budget_versions(session)
        return next(item for item in versions if item["id"] == version_id)
    allowed = _BUDGET_ALLOWED_TRANSITIONS.get(from_status, set())
    if to_status not in allowed:
        raise ValueError(f"Invalid budget status transition: {from_status} -> {to_status}")

    await session.execute(
        text(
            """
            UPDATE budget_versions
            SET status = :status, updated_at = NOW()
            WHERE id = :version_id
            """
        ),
        {"version_id": version_id, "status": to_status},
    )
    if to_status == "approved":
        await session.execute(
            text(
                """
                INSERT INTO budget_approvals (
                    id, budget_version_id, approved_by_empleado_id,
                    approval_status, approval_note, approved_at, created_at
                ) VALUES (
                    :id, :budget_version_id, :approved_by_empleado_id,
                    'approved', :approval_note, NOW(), NOW()
                )
                """
            ),
            {
                "id": str(uuid.uuid4()),
                "budget_version_id": version_id,
                "approved_by_empleado_id": actor_empleado_id,
                "approval_note": note,
            },
        )
    await _audit_budget_event(
        session,
        budget_version_id=version_id,
        event_type="status_changed",
        actor_empleado_id=actor_empleado_id,
        from_status=from_status,
        to_status=to_status,
        note=note,
        payload={"action": "transition"},
    )
    await session.commit()
    versions = await list_budget_versions(session)
    return next(item for item in versions if item["id"] == version_id)


async def get_budget_version(
    session: AsyncSession,
    *,
    version_id: str,
) -> dict[str, Any]:
    versions = await list_budget_versions(session)
    for version in versions:
        if version["id"] == version_id:
            return version
    raise ValueError("Budget version not found")


async def update_budget_version_metadata(
    session: AsyncSession,
    *,
    version_id: str,
    actor_empleado_id: Optional[str] = None,
    version_name: Optional[str] = None,
    notes: Optional[str] = None,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    current = await get_budget_version(session, version_id=version_id)
    if not _editable_version_status(current.get("status")):
        raise ValueError("Only draft or reforecast versions can be edited")
    updated_name = _safe_str(version_name) or current["version_name"]
    updated_notes = notes if notes is not None else current.get("notes")
    await session.execute(
        text(
            """
            UPDATE budget_versions
            SET version_name = :version_name,
                notes = :notes,
                updated_at = NOW()
            WHERE id = :version_id
            """
        ),
        {
            "version_id": version_id,
            "version_name": updated_name,
            "notes": updated_notes,
        },
    )
    await _audit_budget_event(
        session,
        budget_version_id=version_id,
        event_type="version_updated",
        actor_empleado_id=actor_empleado_id,
        from_status=current.get("status"),
        to_status=current.get("status"),
        payload={
            "before": {
                "version_name": current.get("version_name"),
                "notes": current.get("notes"),
            },
            "after": {
                "version_name": updated_name,
                "notes": updated_notes,
            },
        },
    )
    await session.commit()
    return await get_budget_version(session, version_id=version_id)


async def list_budget_lines(
    session: AsyncSession,
    *,
    version_id: str,
    tournament_id: Optional[str] = None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    await ensure_budget_schema(session)
    filters = ["l.budget_version_id = :version_id"]
    params: dict[str, Any] = {"version_id": version_id, "limit": max(1, min(limit, 500))}
    if tournament_id:
        filters.append("CAST(l.tournament_id AS text) = :tournament_id")
        params["tournament_id"] = tournament_id
    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    l.id,
                    l.budget_version_id,
                    l.budget_concept_id,
                    l.tournament_id,
                    l.tournament_code,
                    l.tournament_name,
                    l.sport,
                    l.phase,
                    l.entity_name,
                    l.concept_name,
                    l.account_code_suggested,
                    l.account_code_final,
                    l.budget_amount,
                    l.reference_amount,
                    l.variance_amount,
                    l.priority,
                    l.owner_name,
                    l.criteria_note,
                    l.observations,
                    l.metadata,
                    l.created_at,
                    l.updated_at
                FROM budget_lines l
                WHERE {' AND '.join(filters)}
                ORDER BY l.budget_amount DESC, l.tournament_name ASC, l.concept_name ASC
                LIMIT :limit
                """
            ),
            params,
        )
    ).mappings().all()
    return [
        {
            "id": _safe_str(row["id"]),
            "budget_version_id": _safe_str(row["budget_version_id"]),
            "budget_concept_id": _safe_str(row["budget_concept_id"]) or None,
            "tournament_id": _safe_str(row["tournament_id"]) or None,
            "tournament_code": _safe_str(row["tournament_code"]) or None,
            "tournament_name": _safe_str(row["tournament_name"]),
            "sport": _safe_str(row["sport"]) or None,
            "phase": _safe_str(row["phase"]) or None,
            "entity_name": _safe_str(row["entity_name"]) or None,
            "concept_name": _safe_str(row["concept_name"]),
            "account_code_suggested": _safe_str(row["account_code_suggested"]) or None,
            "account_code_final": _safe_str(row["account_code_final"]) or None,
            "budget_amount": round(_safe_decimal(row["budget_amount"]), 2),
            "reference_amount": round(_safe_decimal(row["reference_amount"]), 2),
            "variance_amount": round(_safe_decimal(row["variance_amount"]), 2),
            "priority": _safe_str(row["priority"]) or None,
            "owner_name": _safe_str(row["owner_name"]) or None,
            "criteria_note": _safe_str(row["criteria_note"]) or None,
            "observations": _safe_str(row["observations"]) or None,
            "metadata": row["metadata"] if isinstance(row["metadata"], dict) else {},
            "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
        }
        for row in rows
    ]


async def update_budget_line(
    session: AsyncSession,
    *,
    line_id: str,
    actor_empleado_id: Optional[str] = None,
    updates: dict[str, Any],
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    current = (
        await session.execute(
            text(
                """
                SELECT
                    l.id,
                    l.budget_version_id,
                    l.budget_concept_id,
                    l.tournament_code,
                    l.concept_name,
                    l.account_code_final,
                    l.account_code_suggested,
                    l.budget_amount,
                    l.reference_amount,
                    l.variance_amount,
                    l.priority,
                    l.owner_name,
                    l.phase,
                    l.criteria_note,
                    l.observations,
                    l.metadata,
                    v.status AS version_status
                FROM budget_lines l
                JOIN budget_versions v ON v.id = l.budget_version_id
                WHERE l.id = :line_id
                LIMIT 1
                """
            ),
            {"line_id": line_id},
        )
    ).mappings().first()
    if not current:
        raise ValueError("Budget line not found")
    if not _editable_version_status(current.get("version_status")):
        raise ValueError("Only draft or reforecast versions allow line edits")

    allowed_fields = {
        "budget_concept_id",
        "concept_name",
        "account_code_final",
        "budget_amount",
        "priority",
        "owner_name",
        "phase",
        "criteria_note",
        "observations",
    }
    requested_fields = {key: value for key, value in updates.items() if key in allowed_fields}
    if not requested_fields:
        raise ValueError("No editable budget line fields were provided")

    concept = None
    if "budget_concept_id" in requested_fields:
        concept = await resolve_budget_concept(
            session,
            budget_concept_id=_safe_str(requested_fields.get("budget_concept_id")),
            tournament_code=_safe_str(current.get("tournament_code")),
        )
        if concept is None:
            raise ValueError("Budget concept not found for this tournament")

    next_budget_amount = (
        _safe_decimal(requested_fields["budget_amount"])
        if "budget_amount" in requested_fields
        else _safe_decimal(current["budget_amount"])
    )
    reference_amount = _safe_decimal(current["reference_amount"])
    variance_amount = round(next_budget_amount - reference_amount, 2)
    update_payload = {
        "budget_concept_id": (
            _safe_str(concept.get("id"))
            if concept is not None
            else _safe_str(current["budget_concept_id"]) or None
        ),
        "concept_name": (
            _safe_str(concept.get("concept_name"))
            if concept is not None
            else _safe_str(requested_fields.get("concept_name"))
            or _safe_str(current["concept_name"])
        ),
        "account_code_final": _safe_str(requested_fields.get("account_code_final")) or _safe_str(current["account_code_final"]),
        "budget_amount": next_budget_amount,
        "priority": _safe_str(requested_fields.get("priority")) or _safe_str(current["priority"]) or None,
        "owner_name": _safe_str(requested_fields.get("owner_name")) or _safe_str(current["owner_name"]) or None,
        "phase": _safe_str(requested_fields.get("phase")) or _safe_str(current["phase"]) or None,
        "criteria_note": _safe_str(requested_fields.get("criteria_note")) or _safe_str(current["criteria_note"]) or None,
        "observations": _safe_str(requested_fields.get("observations")) or _safe_str(current["observations"]) or None,
        "variance_amount": variance_amount,
    }
    await session.execute(
        text(
            """
            UPDATE budget_lines
            SET budget_concept_id = :budget_concept_id,
                concept_name = :concept_name,
                account_code_final = :account_code_final,
                budget_amount = :budget_amount,
                priority = :priority,
                owner_name = :owner_name,
                phase = :phase,
                criteria_note = :criteria_note,
                observations = :observations,
                variance_amount = :variance_amount,
                updated_at = NOW()
            WHERE id = :line_id
            """
        ),
        {"line_id": line_id, **update_payload},
    )
    await _audit_budget_event(
        session,
        budget_version_id=_safe_str(current["budget_version_id"]),
        event_type="line_updated",
        actor_empleado_id=actor_empleado_id,
        from_status=_safe_str(current["version_status"]),
        to_status=_safe_str(current["version_status"]),
        payload={
            "line_id": line_id,
            "before": {
                "concept_name": _safe_str(current["concept_name"]),
                "budget_concept_id": _safe_str(current["budget_concept_id"]) or None,
                "account_code_final": _safe_str(current["account_code_final"])
                or _safe_str(current["account_code_suggested"]),
                "budget_amount": round(_safe_decimal(current["budget_amount"]), 2),
                "priority": _safe_str(current["priority"]) or None,
                "owner_name": _safe_str(current["owner_name"]) or None,
                "phase": _safe_str(current["phase"]) or None,
                "criteria_note": _safe_str(current["criteria_note"]) or None,
                "observations": _safe_str(current["observations"]) or None,
            },
            "after": update_payload,
        },
    )
    await session.commit()
    lines = await list_budget_lines(
        session,
        version_id=_safe_str(current["budget_version_id"]),
        limit=500,
    )
    for line in lines:
        if line["id"] == line_id:
            return line
    raise ValueError("Budget line not found after update")


def _budget_version_sort_key(item: dict[str, Any]) -> tuple[int, str, str]:
    return (
        _status_sort_value(item.get("status")),
        _safe_str(item.get("updated_at")) or _safe_str(item.get("created_at")),
    )


async def _select_budget_version(
    session: AsyncSession,
    *,
    edition_year: int,
    version_id: Optional[str] = None,
) -> Optional[dict[str, Any]]:
    versions = await list_budget_versions(session, edition_year=edition_year)
    if version_id:
        for version in versions:
            if version["id"] == version_id:
                return version
    if not versions:
        return None
    latest_by_priority = sorted(versions, key=_budget_version_sort_key)
    preferred_status = _safe_str(latest_by_priority[0].get("status"))
    same_status_versions = [
        version for version in versions if _safe_str(version.get("status")) == preferred_status
    ]
    return same_status_versions[0] if same_status_versions else latest_by_priority[0]


async def _build_budget_finance_comparison(
    session: AsyncSession,
    *,
    edition_year: int,
    tournament_id: Optional[str],
    tournament_name: Optional[str],
    tournament_code: Optional[str],
    budget_total: float,
) -> dict[str, Any]:
    document_filter, expense_filter, params = _build_budget_scope_filters(
        edition_year=edition_year,
        tournament_id=tournament_id,
        tournament_name=tournament_name,
        tournament_code=tournament_code,
    )
    today = date.today()
    next_30 = today + timedelta(days=30)

    document_row = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(SUM(CASE WHEN d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS requested_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('aprobado', 'pagado', 'cerrado') THEN COALESCE(d.monto_solicitado, d.monto_total, 0) ELSE 0 END), 0) AS committed_total,
                    COALESCE(SUM(CASE WHEN d.estado IN ('pagado', 'cerrado') OR d.pagado_en IS NOT NULL THEN COALESCE(d.monto_total, d.monto_solicitado, 0) ELSE 0 END), 0) AS paid_total,
                    COALESCE(SUM(CASE WHEN d.estado NOT IN ('pagado', 'cerrado') AND d.fecha_pago IS NOT NULL AND d.fecha_pago < :today THEN COALESCE(d.monto_total, d.monto_solicitado, 0) ELSE 0 END), 0) AS overdue_open_total,
                    COALESCE(SUM(CASE WHEN d.estado NOT IN ('pagado', 'cerrado') AND d.fecha_pago IS NOT NULL AND d.fecha_pago >= :today AND d.fecha_pago <= :next_30 THEN COALESCE(d.monto_total, d.monto_solicitado, 0) ELSE 0 END), 0) AS due_next_30_total,
                    COUNT(*) FILTER (WHERE d.estado IN ('enviado', 'aprobado', 'pagado', 'cerrado')) AS document_count,
                    COUNT(*) FILTER (WHERE d.estado NOT IN ('pagado', 'cerrado') AND d.fecha_pago IS NOT NULL AND d.fecha_pago < :today) AS overdue_open_count,
                    COUNT(*) FILTER (WHERE d.estado NOT IN ('pagado', 'cerrado') AND d.fecha_pago IS NOT NULL AND d.fecha_pago >= :today AND d.fecha_pago <= :next_30) AS due_next_30_count
                FROM documentos d
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                WHERE {' AND '.join(document_filter)}
                """
            ),
            {**params, "today": today.isoformat(), "next_30": next_30.isoformat()},
        )
    ).mappings().first()
    expense_row = (
        await session.execute(
            text(
                f"""
                SELECT
                    COALESCE(SUM(e.gasto_cantidad), 0) AS actual_total,
                    COUNT(*) AS expense_count
                FROM expense_reports e
                LEFT JOIN documentos d ON d.id = e.documento_id
                LEFT JOIN tournaments t ON t.id = d.torneo_id
                WHERE {' AND '.join(expense_filter)}
                """
            ),
            params,
        )
    ).mappings().first()

    requested_total = round(_safe_decimal(document_row["requested_total"] if document_row else 0), 2)
    committed_total = round(_safe_decimal(document_row["committed_total"] if document_row else 0), 2)
    paid_total = round(_safe_decimal(document_row["paid_total"] if document_row else 0), 2)
    actual_total = round(_safe_decimal(expense_row["actual_total"] if expense_row else 0), 2)
    return {
        "requested_total": requested_total,
        "committed_total": committed_total,
        "paid_total": paid_total,
        "actual_total": actual_total,
        "pending_to_pay_total": round(max(committed_total - paid_total, 0), 2),
        "variance_vs_actual": round(budget_total - actual_total, 2),
        "variance_vs_committed": round(budget_total - committed_total, 2),
        "variance_vs_paid": round(budget_total - paid_total, 2),
        "variance_vs_requested": round(budget_total - requested_total, 2),
        "document_count": int(document_row["document_count"] or 0) if document_row else 0,
        "expense_count": int(expense_row["expense_count"] or 0) if expense_row else 0,
        "overdue_open_total": round(_safe_decimal(document_row["overdue_open_total"] if document_row else 0), 2),
        "due_next_30_total": round(_safe_decimal(document_row["due_next_30_total"] if document_row else 0), 2),
        "overdue_open_count": int(document_row["overdue_open_count"] or 0) if document_row else 0,
        "due_next_30_count": int(document_row["due_next_30_count"] or 0) if document_row else 0,
    }


async def build_budget_snapshot(
    session: AsyncSession,
    *,
    tournament_id: Optional[str] = None,
    tournament_name: Optional[str] = None,
    tournament_slug: Optional[str] = None,
    edition_year: int = 2026,
    version_id: Optional[str] = None,
) -> dict[str, Any]:
    await ensure_budget_schema(session)
    selected_version = await _select_budget_version(
        session,
        edition_year=edition_year,
        version_id=version_id,
    )
    aliases = budget_alias_candidates(tournament_name or "", tournament_slug or "")

    if not selected_version:
        artifact_rows = load_budget_artifact_rows()
        artifact_snapshot = build_budget_artifact_snapshot(
            artifact_rows,
            artifact_path=DEFAULT_BUDGET_ARTIFACT,
            tournament_name=tournament_name,
            tournament_slug=tournament_slug,
            edition_year=edition_year,
        )
        artifact_snapshot["comparison"] = {
            "requested_total": 0.0,
            "committed_total": 0.0,
            "paid_total": 0.0,
            "actual_total": 0.0,
        }
        artifact_snapshot["forecast"] = _build_budget_forecast(
            edition_year=edition_year,
            budget_total=float(artifact_snapshot.get("summary", {}).get("budget_total") or 0),
            comparison=artifact_snapshot["comparison"],
        )
        artifact_snapshot["scenarios"] = _build_budget_scenarios(
            edition_year=edition_year,
            budget_total=float(artifact_snapshot.get("summary", {}).get("budget_total") or 0),
            comparison=artifact_snapshot["comparison"],
            forecast=artifact_snapshot["forecast"],
        )
        artifact_snapshot["executive_alerts"] = build_budget_executive_alerts(
            artifact_snapshot.get("summary", {}),
            artifact_snapshot.get("forecast", {}),
            artifact_snapshot.get("scenarios", {}),
        )
        artifact_snapshot["executive_comparison"] = build_budget_executive_comparison(
            artifact_snapshot.get("summary", {}),
            artifact_snapshot.get("forecast", {}),
        )
        return artifact_snapshot

    filters = ["l.budget_version_id = :version_id"]
    params: dict[str, Any] = {"version_id": selected_version["id"]}
    if tournament_id:
        filters.append(
            "(CAST(l.tournament_id AS text) = :tournament_id OR UPPER(COALESCE(l.tournament_code, '')) = ANY(:aliases))"
        )
        params["tournament_id"] = tournament_id
        params["aliases"] = list(sorted(aliases)) or [""]
    elif aliases:
        filters.append("UPPER(COALESCE(l.tournament_code, '')) = ANY(:aliases)")
        params["aliases"] = list(sorted(aliases))

    rows = (
        await session.execute(
            text(
                f"""
                SELECT
                    l.tournament_id,
                    l.tournament_code,
                    l.tournament_name,
                    l.concept_name,
                    l.account_code_final,
                    l.account_code_suggested,
                    l.phase,
                    l.priority,
                    l.owner_name,
                    l.budget_amount,
                    l.reference_amount,
                    l.variance_amount
                FROM budget_lines l
                WHERE {' AND '.join(filters)}
                ORDER BY l.tournament_name ASC, l.budget_amount DESC, l.concept_name ASC
                """
            ),
            params,
        )
    ).mappings().all()

    if not rows:
        artifact_rows = load_budget_artifact_rows()
        artifact_snapshot = build_budget_artifact_snapshot(
            artifact_rows,
            artifact_path=DEFAULT_BUDGET_ARTIFACT,
            tournament_name=tournament_name,
            tournament_slug=tournament_slug,
            edition_year=edition_year,
        )
        artifact_snapshot["version"] = {
            "id": selected_version["id"],
            "name": selected_version["version_name"],
            "status": selected_version["status"],
            "source": selected_version["source"],
            "artifact_path": selected_version["artifact_path"],
            "edition_year": selected_version["edition_year"],
        }
        artifact_snapshot["forecast"] = _build_budget_forecast(
            edition_year=edition_year,
            budget_total=float(artifact_snapshot.get("summary", {}).get("budget_total") or 0),
            comparison={
                "actual_total": 0.0,
                "committed_total": 0.0,
                "paid_total": 0.0,
                "pending_to_pay_total": 0.0,
            },
        )
        artifact_snapshot["scenarios"] = _build_budget_scenarios(
            edition_year=edition_year,
            budget_total=float(artifact_snapshot.get("summary", {}).get("budget_total") or 0),
            comparison={
                "actual_total": 0.0,
                "committed_total": 0.0,
                "paid_total": 0.0,
                "pending_to_pay_total": 0.0,
            },
            forecast=artifact_snapshot["forecast"],
        )
        artifact_snapshot["executive_alerts"] = build_budget_executive_alerts(
            artifact_snapshot.get("summary", {}),
            artifact_snapshot.get("forecast", {}),
            artifact_snapshot.get("scenarios", {}),
        )
        return artifact_snapshot

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[_safe_str(row["tournament_name"])].append(dict(row))

    tournaments: list[dict[str, Any]] = []
    total_budget = 0.0
    total_reference = 0.0
    summary_comparison = {
        "requested_total": 0.0,
        "committed_total": 0.0,
        "paid_total": 0.0,
        "actual_total": 0.0,
        "pending_to_pay_total": 0.0,
        "overdue_open_total": 0.0,
        "due_next_30_total": 0.0,
    }
    forecast_health_counts = {"healthy": 0, "at_risk": 0, "over_budget": 0}
    summary_breakdowns = _build_budget_line_breakdowns([dict(row) for row in rows])
    summary_finance_breakdowns = await _build_budget_finance_breakdowns(
        session,
        edition_year=edition_year,
        tournament_id=tournament_id,
        tournament_name=tournament_name,
        tournament_code=None,
    )
    by_concept_store = _new_breakdown_store()
    for item in summary_breakdowns.get("by_concept", []):
        _merge_breakdown_row(by_concept_store, **item)
    for item in summary_finance_breakdowns.get("by_concept", []):
        _merge_breakdown_row(by_concept_store, **item)
    for name, items in sorted(grouped.items()):
        budget_total = sum(_safe_decimal(item["budget_amount"]) for item in items)
        reference_total = sum(_safe_decimal(item["reference_amount"]) for item in items)
        total_budget += budget_total
        total_reference += reference_total
        comparison = await _build_budget_finance_comparison(
            session,
            edition_year=edition_year,
            tournament_id=_safe_str(items[0].get("tournament_id")) or tournament_id,
            tournament_name=name,
            tournament_code=_safe_str(items[0].get("tournament_code")) or None,
            budget_total=round(budget_total, 2),
        )
        for key in summary_comparison:
            summary_comparison[key] += float(comparison.get(key) or 0)
        forecast = _build_budget_forecast(
            edition_year=edition_year,
            budget_total=round(budget_total, 2),
            comparison=comparison,
        )
        forecast_health_counts[str(forecast.get("health") or "healthy")] = (
            forecast_health_counts.get(str(forecast.get("health") or "healthy"), 0)
            + 1
        )
        tournament_line_breakdowns = _build_budget_line_breakdowns(items)
        tournament_finance_breakdowns = await _build_budget_finance_breakdowns(
            session,
            edition_year=edition_year,
            tournament_id=_safe_str(items[0].get("tournament_id")) or tournament_id,
            tournament_name=name,
            tournament_code=_safe_str(items[0].get("tournament_code")) or None,
        )
        tournament_concept_store = _new_breakdown_store()
        for item in tournament_line_breakdowns.get("by_concept", []):
            _merge_breakdown_row(tournament_concept_store, **item)
        for item in tournament_finance_breakdowns.get("by_concept", []):
            _merge_breakdown_row(tournament_concept_store, **item)
        tournaments.append(
            {
                "tournament_id": _safe_str(items[0].get("tournament_id")) or None,
                "tournament_code": _safe_str(items[0].get("tournament_code")) or None,
                "tournament_name": name,
                "line_count": len(items),
                "budget_total": round(budget_total, 2),
                "reference_total": round(reference_total, 2),
                "comparison": comparison,
                "forecast": forecast,
                "scenarios": _build_budget_scenarios(
                    edition_year=edition_year,
                    budget_total=round(budget_total, 2),
                    comparison=comparison,
                    forecast=forecast,
                ),
                "top_concepts": [
                    {
                        "concepto": _safe_str(item.get("concept_name")),
                        "presupuesto_2026": round(_safe_decimal(item.get("budget_amount")), 2),
                        "cuenta_contable_final": _safe_str(
                            item.get("account_code_final") or item.get("account_code_suggested")
                        )
                        or None,
                        "prioridad": _safe_str(item.get("priority")) or None,
                        "etapa": _safe_str(item.get("phase")) or None,
                    }
                    for item in items[:8]
                ],
                "breakdowns": {
                    "by_concept": _finalize_breakdown_store(tournament_concept_store),
                    "by_phase": tournament_line_breakdowns.get("by_phase", []),
                    "by_entity": tournament_line_breakdowns.get("by_entity", []),
                    "by_owner": tournament_line_breakdowns.get("by_owner", []),
                    "by_account": tournament_line_breakdowns.get("by_account", []),
                    "by_provider": tournament_finance_breakdowns.get("by_provider", []),
                },
            }
        )

    summary_forecast = _build_budget_forecast(
        edition_year=edition_year,
        budget_total=round(total_budget, 2),
        comparison=summary_comparison,
    )

    scenarios = _build_budget_scenarios(
        edition_year=edition_year,
        budget_total=round(total_budget, 2),
        comparison=summary_comparison,
        forecast=summary_forecast,
    )
    summary = {
        "edition_year": edition_year,
        "tournaments_count": len(tournaments),
        "line_count": len(rows),
        "budget_total": round(total_budget, 2),
        "reference_total": round(total_reference, 2),
        "requested_total": round(summary_comparison["requested_total"], 2),
        "committed_total": round(summary_comparison["committed_total"], 2),
        "paid_total": round(summary_comparison["paid_total"], 2),
        "actual_total": round(summary_comparison["actual_total"], 2),
        "pending_to_pay_total": round(summary_comparison["pending_to_pay_total"], 2),
        "overdue_open_total": round(summary_comparison["overdue_open_total"], 2),
        "due_next_30_total": round(summary_comparison["due_next_30_total"], 2),
        "variance_vs_actual": round(total_budget - summary_comparison["actual_total"], 2),
        "variance_vs_committed": round(total_budget - summary_comparison["committed_total"], 2),
        "variance_vs_paid": round(total_budget - summary_comparison["paid_total"], 2),
        "forecast_health_counts": forecast_health_counts,
    }

    return {
        "ok": True,
        "source": "budget_db",
        "version": {
            "id": selected_version["id"],
            "name": selected_version["version_name"],
            "status": selected_version["status"],
            "source": selected_version["source"],
            "artifact_path": selected_version["artifact_path"],
            "edition_year": edition_year,
        },
        "summary": summary,
        "breakdowns": {
            "by_concept": _finalize_breakdown_store(by_concept_store),
            "by_phase": summary_breakdowns.get("by_phase", []),
            "by_entity": summary_breakdowns.get("by_entity", []),
            "by_owner": summary_breakdowns.get("by_owner", []),
            "by_account": summary_breakdowns.get("by_account", []),
            "by_provider": summary_finance_breakdowns.get("by_provider", []),
        },
        "forecast": summary_forecast,
        "scenarios": scenarios,
        "executive_alerts": build_budget_executive_alerts(
            summary,
            summary_forecast,
            scenarios,
        ),
        "executive_comparison": build_budget_executive_comparison(
            summary,
            summary_forecast,
        ),
        "tournaments": tournaments,
    }
