"""Excel exports for budget review and executive delivery."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_SUBHEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def _safe_float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _write_table(
    ws, headers: list[str], rows: list[list[Any]], start_row: int = 1
) -> int:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, _header in enumerate(headers, start=1):
        max_len = max(
            len(_safe_text(ws.cell(row=row, column=col_idx).value))
            for row in range(start_row, start_row + len(rows) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 12), 45
        )
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    return start_row + len(rows) + 2


def generate_budget_income_xlsx(
    *,
    lines: list[dict[str, Any]],
    plan_map: dict[str, dict[int, dict[str, Any]]],
    actuals_map: dict[str, dict[int, dict[str, Any]]],
    links: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    selected_version: dict[str, Any] | None,
    tournament_context: dict[str, Any],
    edition_year: int,
) -> bytes:
    """Build the income-budget mirror export for a tournament detail page."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos"

    ws["A1"] = "Presupuesto de ingresos"
    ws["A1"].font = Font(size=18, bold=True)
    generated_at = (
        datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
    )
    ws["A2"] = f"Generado: {generated_at}"
    ws["A3"] = f"Anio: {int(edition_year)}"
    ws["A4"] = f"Torneo: {_safe_text(tournament_context.get('tournament_name'))}"
    if selected_version:
        ws["A5"] = f"Version: {_safe_text(selected_version.get('version_name'))}"
        ws["A6"] = f"Estatus: {_safe_text(selected_version.get('status'))}"

    summary_rows: list[list[Any]] = []
    monthly_rows: list[list[Any]] = []
    month_labels = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]
    for line in lines:
        line_id = _safe_text(line.get("id"))
        concept_id = _safe_text(line.get("budget_concept_id"))
        actual_key = concept_id or "__unassigned__"
        plan = plan_map.get(line_id, {})
        actuals = actuals_map.get(actual_key, actuals_map.get("__unassigned__", {}))
        expected_total = sum(
            _safe_float(plan.get(month, {}).get("expected_income_amount"))
            for month in range(1, 13)
        )
        real_total = sum(
            _safe_float(actuals.get(month, {}).get("real_income"))
            for month in range(1, 13)
        )
        variance = round(real_total - expected_total, 2)
        compliance = (
            round((real_total / expected_total) * 100, 2) if expected_total else 0.0
        )
        summary_rows.append(
            [
                _safe_text(line.get("tournament_code")),
                _safe_text(line.get("tournament_name")),
                _safe_text(line.get("phase")),
                _safe_text(line.get("concept_name")),
                _safe_text(
                    line.get("account_code_final") or line.get("account_code_suggested")
                ),
                _safe_float(line.get("budget_amount")),
                expected_total,
                real_total,
                variance,
                compliance,
            ]
        )
        for month in range(1, 13):
            expected = _safe_float(plan.get(month, {}).get("expected_income_amount"))
            real = _safe_float(actuals.get(month, {}).get("real_income"))
            monthly_rows.append(
                [
                    _safe_text(line.get("tournament_code")),
                    _safe_text(line.get("tournament_name")),
                    _safe_text(line.get("phase")),
                    _safe_text(line.get("concept_name")),
                    month,
                    month_labels[month - 1],
                    expected,
                    real,
                    round(real - expected, 2),
                    round((real / expected) * 100, 2) if expected else 0.0,
                ]
            )

    _write_table(
        ws,
        [
            "Codigo torneo",
            "Torneo",
            "Fase",
            "Partida ingreso",
            "Cuenta contable",
            "Monto anual",
            "Ingreso esperado",
            "Ingreso real CFDI",
            "Variacion",
            "% cumplimiento",
        ],
        summary_rows,
        start_row=8,
    )

    ws_monthly = wb.create_sheet("Mensual")
    _write_table(
        ws_monthly,
        [
            "Codigo torneo",
            "Torneo",
            "Fase",
            "Partida ingreso",
            "Mes",
            "Etiqueta",
            "Ingreso esperado",
            "Ingreso real CFDI",
            "Variacion",
            "% cumplimiento",
        ],
        monthly_rows,
    )

    ws_links = wb.create_sheet("CFDI vinculados")
    _write_table(
        ws_links,
        [
            "UUID CFDI",
            "RFC emisor",
            "Emisor",
            "RFC receptor",
            "Partida ingreso",
            "Fase",
            "Monto",
            "Fecha ingreso",
            "Estado",
        ],
        [
            [
                _safe_text(item.get("cfdi_uuid")),
                _safe_text(item.get("emisor_rfc")),
                _safe_text(item.get("emisor_nombre")),
                _safe_text(item.get("receptor_rfc")),
                _safe_text(item.get("concept_name")),
                _safe_text(item.get("phase") or "General"),
                _safe_float(item.get("amount")),
                _safe_text(item.get("income_date"))[:10],
                "Desvinculado" if item.get("unlinked_at") else "Activo",
            ]
            for item in links
        ],
    )

    ws_pending = wb.create_sheet("CFDI sin clasificar")
    _write_table(
        ws_pending,
        ["UUID CFDI", "Fecha", "RFC emisor", "Emisor", "RFC receptor", "Total"],
        [
            [
                _safe_text(item.get("cfdi_uuid")),
                _safe_text(item.get("fecha"))[:10],
                _safe_text(item.get("emisor_rfc")),
                _safe_text(item.get("emisor_nombre")),
                _safe_text(item.get("receptor_rfc")),
                _safe_float(item.get("total")),
            ]
            for item in candidates
        ],
    )

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def generate_budget_review_xlsx(
    *,
    snapshot: dict[str, Any],
    versions: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    audit_events: list[dict[str, Any]] | None = None,
    selected_version: dict[str, Any] | None = None,
) -> bytes:
    """Build a read-only budget workbook for executive review."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen ejecutivo"

    summary = snapshot.get("summary") or {}
    forecast = snapshot.get("forecast") or {}
    scenarios = snapshot.get("scenarios") or {}
    comparison = list(snapshot.get("executive_comparison") or [])
    alerts = list(snapshot.get("executive_alerts") or [])
    breakdowns = snapshot.get("breakdowns") or {}
    tournaments = list(snapshot.get("tournaments") or [])

    ws["A1"] = "Presupuesto 2026"
    ws["A1"].font = Font(size=18, bold=True)
    generated_at = (
        datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
    )
    ws["A2"] = f"Generado: {generated_at}"
    ws["A3"] = f"Fuente: {_safe_text(snapshot.get('source') or 'sin fuente')}"
    if selected_version:
        ws["A4"] = f"Version: {_safe_text(selected_version.get('version_name'))}"
        ws["A5"] = f"Estatus: {_safe_text(selected_version.get('status'))}"

    _write_table(
        ws,
        ["Metrica", "Valor"],
        [
            ["Presupuesto", _safe_float(summary.get("budget_total"))],
            ["Solicitado", _safe_float(summary.get("requested_total"))],
            ["Comprometido", _safe_float(summary.get("committed_total"))],
            ["Pagado", _safe_float(summary.get("paid_total"))],
            ["Real", _safe_float(summary.get("actual_total"))],
            ["Pendiente por pagar", _safe_float(summary.get("pending_to_pay_total"))],
            ["Cierre proyectado", _safe_float(forecast.get("projected_close_total"))],
            ["Necesidad de caja", _safe_float(forecast.get("projected_cash_need"))],
            ["Lineas", int(summary.get("line_count") or 0)],
            ["Torneos", int(summary.get("tournaments_count") or 0)],
        ],
        start_row=7,
    )

    ws_compare = wb.create_sheet("Comparativo")
    _write_table(
        ws_compare,
        ["Metrica", "Total", "% presupuesto", "Varianza vs presupuesto", "Detalle"],
        [
            [
                _safe_text(item.get("label")),
                _safe_float(item.get("total")),
                _safe_float(item.get("pct_of_budget")),
                item.get("variance_to_budget"),
                _safe_text(item.get("detail")),
            ]
            for item in comparison
        ],
    )

    ws_tournaments = wb.create_sheet("Torneos")
    _write_table(
        ws_tournaments,
        [
            "Codigo",
            "Torneo",
            "Lineas",
            "Presupuesto",
            "Solicitado",
            "Comprometido",
            "Pagado",
            "Real",
            "Cierre proyectado",
            "Salud",
        ],
        [
            [
                _safe_text(item.get("tournament_code")),
                _safe_text(item.get("tournament_name")),
                int(item.get("line_count") or 0),
                _safe_float(item.get("budget_total")),
                _safe_float((item.get("comparison") or {}).get("requested_total")),
                _safe_float((item.get("comparison") or {}).get("committed_total")),
                _safe_float((item.get("comparison") or {}).get("paid_total")),
                _safe_float((item.get("comparison") or {}).get("actual_total")),
                _safe_float((item.get("forecast") or {}).get("projected_close_total")),
                _safe_text((item.get("forecast") or {}).get("health")),
            ]
            for item in tournaments
        ],
    )

    ws_lines = wb.create_sheet("Lineas")
    _write_table(
        ws_lines,
        [
            "Codigo torneo",
            "Torneo",
            "Fase",
            "Concepto",
            "Cuenta final",
            "Responsable",
            "Prioridad",
            "Presupuesto",
            "Referencia",
            "Varianza",
            "Criterio",
            "Observaciones",
        ],
        [
            [
                _safe_text(item.get("tournament_code")),
                _safe_text(item.get("tournament_name")),
                _safe_text(item.get("phase")),
                _safe_text(item.get("concept_name")),
                _safe_text(item.get("account_code_final")),
                _safe_text(item.get("owner_name")),
                _safe_text(item.get("priority")),
                _safe_float(item.get("budget_amount")),
                _safe_float(item.get("reference_amount")),
                _safe_float(item.get("variance_amount")),
                _safe_text(item.get("criteria_note")),
                _safe_text(item.get("observations")),
            ]
            for item in lines
        ],
    )

    ws_scenarios = wb.create_sheet("Escenarios")
    _write_table(
        ws_scenarios,
        [
            "Escenario",
            "Cierre proyectado",
            "Varianza",
            "Caja requerida",
            "Salud",
            "Supuesto",
        ],
        [
            [
                _safe_text(item.get("label")),
                _safe_float(item.get("projected_close_total")),
                _safe_float(item.get("projected_variance")),
                _safe_float(item.get("projected_cash_need")),
                _safe_text(item.get("health")),
                _safe_text(item.get("assumption")),
            ]
            for item in scenarios.values()
            if isinstance(item, dict)
        ],
    )

    ws_breakdowns = wb.create_sheet("Desgloses")
    row = 1
    for title, rows in [
        ("Conceptos", breakdowns.get("by_concept") or []),
        ("Proveedores", breakdowns.get("by_provider") or []),
        ("Fases", breakdowns.get("by_phase") or []),
        ("Entidades", breakdowns.get("by_entity") or []),
        ("Responsables", breakdowns.get("by_owner") or []),
        ("Cuentas", breakdowns.get("by_account") or []),
    ]:
        ws_breakdowns.cell(row=row, column=1, value=title).font = _BOLD_FONT
        ws_breakdowns.cell(row=row, column=1).fill = _SUBHEADER_FILL
        row = _write_table(
            ws_breakdowns,
            ["Etiqueta", "Lineas", "Presupuesto", "Comprometido", "Real"],
            [
                [
                    _safe_text(item.get("label")),
                    int(item.get("line_count") or 0),
                    _safe_float(item.get("budget_total")),
                    _safe_float(item.get("committed_total")),
                    _safe_float(item.get("actual_total")),
                ]
                for item in rows
            ],
            start_row=row + 1,
        )

    ws_versions = wb.create_sheet("Versiones")
    _write_table(
        ws_versions,
        [
            "Anio",
            "Version",
            "Estatus",
            "Fuente",
            "Lineas",
            "Presupuesto",
            "Aprobado",
            "Actualizado",
        ],
        [
            [
                int(item.get("edition_year") or 0),
                _safe_text(item.get("version_name")),
                _safe_text(item.get("status")),
                _safe_text(item.get("source")),
                int(item.get("line_count") or 0),
                _safe_float(item.get("budget_total")),
                _safe_text(item.get("latest_approved_at")),
                _safe_text(item.get("updated_at") or item.get("created_at")),
            ]
            for item in versions
        ],
    )

    ws_alerts = wb.create_sheet("Alertas")
    _write_table(
        ws_alerts,
        ["Severidad", "Titulo", "Detalle", "Playbook"],
        [
            [
                _safe_text(item.get("severity")),
                _safe_text(item.get("title")),
                _safe_text(item.get("detail")),
                _safe_text(item.get("playbook")),
            ]
            for item in alerts
        ],
    )

    ws_audit = wb.create_sheet("Auditoria")
    _write_table(
        ws_audit,
        ["Evento", "Version", "Actor", "Desde", "Hacia", "Fecha"],
        [
            [
                _safe_text(item.get("event_type")),
                _safe_text(item.get("version_name")),
                _safe_text(item.get("actor_nombre")),
                _safe_text(item.get("from_status")),
                _safe_text(item.get("to_status")),
                _safe_text(item.get("created_at")),
            ]
            for item in (audit_events or [])
        ],
    )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
