"""Read-only COI workbook export for ready CxC policy previews."""

from __future__ import annotations

import io
import re
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .service import build_ar_accounting_preview

COI_COLUMNS = 9
COI_MONEY_NUMBER_FORMAT = "$#,##0.00"


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _safe_float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _policy_slug(value: Any) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "", _safe_str(value))
    return (slug or "item")[:8]


def _policy_concept(row: dict[str, Any], policy_type: str) -> str:
    parts = [
        f"CxC {policy_type}",
        _safe_str(row.get("cfdi_uuid"))[:8],
        _safe_str(row.get("payer_name") or row.get("payer_rfc")),
        _safe_str(row.get("concept_name")),
    ]
    return " / ".join(part for part in parts if part)


def build_ar_coi_ready_policy_rows(
    rows: list[dict[str, Any]],
    payload: Optional[dict[str, Any]] = None,
) -> list[dict[str, Any]]:
    """Flatten ready CxC accounting previews into exportable policy rows."""

    ready_rows: list[dict[str, Any]] = []
    for row in rows:
        preview = build_ar_accounting_preview(row, payload)
        for policy_type, key, suffix in [
            ("factura", "invoice_policy_preview", "F"),
            ("cobro", "collection_policy_preview", "C"),
        ]:
            policy = preview.get(key) or {}
            if policy.get("status") != "lista":
                continue
            ar_item_id = _safe_str(row.get("ar_item_id"))
            policy_number = f"CXC-{_policy_slug(ar_item_id)}-{suffix}"
            concept = _policy_concept(row, policy_type)
            for line_no, line in enumerate(list(policy.get("lines") or []), start=1):
                side = _safe_str(line.get("side"))
                amount = _safe_float(line.get("amount"))
                ready_rows.append(
                    {
                        "ar_item_id": ar_item_id,
                        "policy_type": policy_type,
                        "policy_number": policy_number,
                        "cfdi_uuid": row.get("cfdi_uuid"),
                        "payer_name": row.get("payer_name"),
                        "payer_rfc": row.get("payer_rfc"),
                        "concept": concept,
                        "line_no": line_no,
                        "account_code": line.get("account_code"),
                        "debe": amount if side == "debe" else 0.0,
                        "haber": amount if side == "haber" else 0.0,
                    }
                )
    return ready_rows


def _group_by_policy(rows: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        key = (_safe_str(row.get("policy_number")), _safe_str(row.get("policy_type")))
        grouped.setdefault(key, []).append(row)
    return list(grouped.values())


def _coi_policy_type(policy_type: str) -> str:
    return "Ig" if policy_type == "cobro" else "Dr"


def _style_coi_sheet(sheet) -> None:
    for column_index in range(1, COI_COLUMNS + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
        is_movement_row = row[2].value == "0" and row[4].value == "1"
        if not is_movement_row:
            continue
        for amount_cell in (row[5], row[6]):
            if isinstance(amount_cell.value, (int, float)):
                amount_cell.number_format = COI_MONEY_NUMBER_FORMAT


def _style_trace_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 28
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generate_ar_coi_ready_xlsx(
    rows: list[dict[str, Any]],
    payload: Optional[dict[str, Any]] = None,
) -> bytes:
    """Generate an XLSX workbook for ready CxC policy previews."""

    ready_rows = build_ar_coi_ready_policy_rows(rows, payload)
    wb = Workbook()
    ws = wb.active
    ws.title = "COI CxC"
    if not ready_rows:
        ws.append(
            ["Sin prepólizas CxC listas para exportar", "", "", "", "", "", "", "", ""]
        )
    else:
        ws.append(["", "", "", "", "", "", "", "", ""])
        ws.append(["|||", "", "", "", "", "", "", "", ""])
        for index, policy_rows in enumerate(_group_by_policy(ready_rows), start=1):
            first = policy_rows[0]
            policy_type = _safe_str(first.get("policy_type"))
            concept = _safe_str(first.get("concept"))
            ws.append(
                [
                    _coi_policy_type(policy_type),
                    str(index),
                    concept,
                    str(len(policy_rows)),
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            for row in policy_rows:
                ws.append(
                    [
                        "",
                        row.get("account_code"),
                        "0",
                        row.get("concept"),
                        "1",
                        _safe_float(row.get("debe")) or "",
                        _safe_float(row.get("haber")) or "",
                        "",
                        "",
                    ]
                )
            if policy_type == "factura" and first.get("cfdi_uuid"):
                ws.append(["", "", "INICIO_CFDI", "", "", "", "", "", ""])
                ws.append(
                    [
                        "",
                        "",
                        "",
                        "",
                        "0",
                        "",
                        f" {first.get('payer_rfc') or ''}",
                        "",
                        first.get("cfdi_uuid") or "",
                    ]
                )
                ws.append(["", "", "FIN_CFDI", "", "", "", "", "", ""])
            ws.append(["", "FIN_PARTIDAS", "", "", "", "", "", "", ""])
    _style_coi_sheet(ws)

    trace = wb.create_sheet("Trazabilidad")
    trace.append(
        [
            "AR item",
            "Tipo póliza",
            "Número póliza",
            "CFDI UUID",
            "Cliente",
            "RFC",
            "Concepto",
            "Renglón",
            "Cuenta",
            "Debe",
            "Haber",
        ]
    )
    if ready_rows:
        for row in ready_rows:
            trace.append(
                [
                    row.get("ar_item_id"),
                    row.get("policy_type"),
                    row.get("policy_number"),
                    row.get("cfdi_uuid"),
                    row.get("payer_name"),
                    row.get("payer_rfc"),
                    row.get("concept"),
                    row.get("line_no"),
                    row.get("account_code"),
                    row.get("debe"),
                    row.get("haber"),
                ]
            )
    else:
        trace.append(
            ["", "", "", "", "", "", "Sin prepólizas CxC listas", "", "", "", ""]
        )
    _style_trace_sheet(trace)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
