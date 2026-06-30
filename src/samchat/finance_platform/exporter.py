"""Excel exports for the Finance Platform command center."""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_WHITE_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def _safe_float(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value)


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    start_row: int = 1,
) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, _header in enumerate(headers, start=1):
        max_len = max(
            len(_safe_text(ws.cell(row=row, column=col_idx).value))
            for row in range(start_row, start_row + len(rows) + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 12),
            55,
        )
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    return start_row + len(rows) + 2


def generate_finance_platform_xlsx(
    *,
    platform: dict[str, Any],
) -> bytes:
    """Build a finance operations workbook from the Finance Platform snapshot."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    period = platform.get("period") or {}
    summary = platform.get("summary") or {}
    action_queue = platform.get("action_queue") or {}
    finance_brief = platform.get("finance_brief") or {}
    cash_control = platform.get("cash_control_center") or {}
    accounting_close = platform.get("accounting_close_center") or {}
    tax_readiness = platform.get("tax_readiness") or {}
    payment_run = platform.get("payment_run") or {}

    generated_at = (
        datetime.now(UTC)
        .isoformat(timespec="seconds")
        .replace(
            "+00:00",
            "Z",
        )
    )
    ws["A1"] = "Finance Platform"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = f"Periodo: {period.get('month')}/{period.get('year')}"
    ws["A3"] = f"Generado: {generated_at}"

    _write_table(
        ws,
        ["Metrica", "Valor"],
        [
            ["Documentos", int(summary.get("documents") or 0)],
            ["Gastos", int(summary.get("expenses") or 0)],
            ["Polizas", int(summary.get("polizas") or 0)],
            ["Acciones abiertas", int(summary.get("open_actions") or 0)],
            ["Presion de pago", _safe_text(summary.get("payment_pressure"))],
            ["Estatus fiscal", _safe_text(summary.get("tax_status"))],
            [
                "Pagos pendientes",
                _safe_float(payment_run.get("payable_total")),
            ],
            [
                "Ingresos contabilizados",
                _safe_float(cash_control.get("income_total")),
            ],
            [
                "Polizas descuadradas",
                int(accounting_close.get("unbalanced_count") or 0),
            ],
            [
                "DIOT blockers",
                int(tax_readiness.get("diot_blockers_count") or 0),
            ],
        ],
        start_row=5,
    )

    ws_brief = wb.create_sheet("Brief")
    ws_brief["A1"] = "One-click Finance Brief"
    ws_brief["A1"].font = _BOLD_FONT
    ws_brief["A3"] = _safe_text(finance_brief.get("plain_text"))
    ws_brief["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_brief.column_dimensions["A"].width = 120
    ws_brief.row_dimensions[3].height = 140

    ws_actions = wb.create_sheet("Action Queue")
    _write_table(
        ws_actions,
        ["Severidad", "Modulo", "Accion", "Responsable", "Vence", "Detalle"],
        [
            [
                _safe_text(item.get("severity")),
                _safe_text(item.get("module")),
                _safe_text(item.get("title")),
                _safe_text(item.get("owner")),
                _safe_text(item.get("due")),
                _safe_text(item.get("detail")),
            ]
            for item in action_queue.get("actions") or []
        ],
    )

    ws_payments = wb.create_sheet("Payment Run")
    _write_table(
        ws_payments,
        ["Referencia", "Tipo", "Beneficiario/proveedor", "Monto", "Fecha", "ID"],
        [
            [
                _safe_text(item.get("numero_referencia")),
                _safe_text(item.get("tipo")),
                _safe_text(
                    item.get("beneficiario_nombre") or item.get("proveedor_nombre")
                ),
                _safe_float(item.get("monto_total") or item.get("monto_solicitado")),
                _safe_text(item.get("fecha_pago") or item.get("aprobado_en"))[:10],
                _safe_text(item.get("id")),
            ]
            for item in payment_run.get("items") or []
        ],
    )

    ws_coi = wb.create_sheet("COI pendientes")
    _write_table(
        ws_coi,
        [
            "Gasto",
            "Concepto",
            "Proyecto",
            "Monto",
            "Cuenta contable",
            "Contracuenta",
            "CFDI",
            "ID",
        ],
        [
            [
                _safe_text(item.get("numero_referencia")),
                _safe_text(item.get("concepto")),
                _safe_text(item.get("proyecto")),
                _safe_float(item.get("gasto_cantidad")),
                _safe_text(item.get("cuenta_contable_id")),
                _safe_text(item.get("contra_cuenta_contable_id")),
                (
                    "ok"
                    if item.get("cfdi_report_id") or item.get("cfdi_uuid_manual")
                    else "faltante"
                ),
                _safe_text(item.get("id")),
            ]
            for item in accounting_close.get("pending_coi_expenses") or []
        ],
    )

    ws_diot = wb.create_sheet("DIOT blockers")
    _write_table(
        ws_diot,
        [
            "Tipo",
            "Referencia",
            "Estado",
            "Monto",
            "Persona/proveedor",
            "UUID CFDI",
            "CFDI report id",
            "ID",
        ],
        [
            [
                _safe_text(item.get("entity_type")),
                _safe_text(item.get("numero_referencia")),
                _safe_text(item.get("estado") or item.get("estado_reembolso")),
                _safe_float(
                    item.get("monto_total")
                    or item.get("monto_solicitado")
                    or item.get("gasto_cantidad")
                ),
                _safe_text(
                    item.get("beneficiario_nombre")
                    or item.get("empleado_nombre")
                    or item.get("proveedor_nombre")
                ),
                _safe_text(item.get("cfdi_uuid_manual")),
                _safe_text(item.get("cfdi_report_id")),
                _safe_text(item.get("id")),
            ]
            for item in tax_readiness.get("blockers") or []
        ],
    )

    ws_polizas = wb.create_sheet("Polizas descuadradas")
    _write_table(
        ws_polizas,
        ["Poliza", "Beneficiario", "Origen", "Debe", "Haber", "Diferencia", "ID"],
        [
            [
                (
                    f"{_safe_text(item.get('tipo_poliza'))}-"
                    f"{_safe_text(item.get('numero_poliza'))}"
                ),
                _safe_text(item.get("beneficiario_nombre")),
                _safe_text(item.get("origen")),
                _safe_float(item.get("debe")),
                _safe_float(item.get("haber")),
                round(
                    _safe_float(item.get("debe")) - _safe_float(item.get("haber")), 2
                ),
                _safe_text(item.get("id")),
            ]
            for item in accounting_close.get("unbalanced_polizas") or []
        ],
    )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
