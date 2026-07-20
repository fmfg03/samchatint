"""
Excel and CSV export utilities for documentos.
"""

from pathlib import Path
from typing import Optional, List, Any, Tuple

import csv
import io
import openpyxl

from ..expense_metadata import format_solicitud_proyecto_display

# Template layout for docs/4599- Informe de gastos.xlsx (Hoja1):
# Header: F5=fecha, B8=nombre, C11:F11=tipo de gasto, G22=referencia operaciones (below Referencia)
# Expense table: row 26=header, data rows 27-46 (20 rows), cols B-G = concepto,fecha,no_factura,importe_sin_iva,iva,total
# Footer: row 47=TOTALES (B47,E47,F47,G47), 48=Cantidad Entregada (B48,G48), 49=Saldo (B49,F49,G49), 53=labels, 54=Autorizado name (D54)
INFORME_EXPENSE_FIRST_ROW = 27
INFORME_EXPENSE_ROWS_TEMPLATE = 20   # rows 27-46 in 4599 template
INFORME_TOTALES_ROW = 47
INFORME_CANTIDAD_ROW = 48
INFORME_SALDO_ROW = 49
INFORME_FIRMA_ROW = 53
INFORME_AUTORIZADO_ROW = 54


def _informe_template_tipo_labels(tipo_cuenta: Optional[str]) -> Tuple[str, str, str, str]:
    """Return row 11 labels with exactly one selected tipo shown as (X)."""
    t = (tipo_cuenta or "local").strip().lower()
    return (
        "Local (X)" if t == "local" else "Local ( )",
        "De viaje (X)" if t == "viaje" else "De viaje ( )",
        "Nacional (X)" if t == "nacional" else "Nacional ( )",
        "Extranjero (X)" if t == "extranjero" else "Extranjero ( )",
    )


def _informe_template_tipo_marks(tipo_cuenta: Optional[str]) -> Tuple[str, str]:
    """
    Map cuenta tipo onto the legacy CSV LOCAL / VIAJE columns.
    nacional -> LOCAL; extranjero -> VIAJE; local/viaje unchanged.
    """
    t = (tipo_cuenta or "local").strip().lower()
    local_x = "X" if t in ("local", "nacional") else ""
    viaje_x = "X" if t in ("viaje", "extranjero") else ""
    return local_x, viaje_x


def _get_base_docs_path() -> Path:
    """Return samchat/docs path (base for template files)."""
    base = Path(__file__).resolve().parent.parent.parent.parent.parent
    return base / "docs"


def _resolve_template_path(filename: str) -> Path:
    """Resolve Excel templates across runtime overlays and the source checkout."""
    candidates = [_get_base_docs_path() / filename]

    import os

    configured_docs = (os.getenv("SAMCHAT_TEMPLATE_DOCS_DIR") or "").strip()
    if configured_docs:
        candidates.append(Path(configured_docs) / filename)

    candidates.append(Path("/root/samchat/docs") / filename)

    seen = set()
    for candidate in candidates:
        candidate = candidate.resolve()
        if candidate in seen:
            continue
        seen.add(candidate)
        if candidate.exists():
            return candidate
    return candidates[0]


def _get_informe_template_path() -> Path:
    """Return path to the INFORME Excel template (docs/4599- Informe de gastos.xlsx)."""
    return _resolve_template_path("4599- Informe de gastos.xlsx")


def _get_solicitud_template_path() -> Path:
    """Return path to the SOLICITUD Excel template (docs/Sol trans.xlsx)."""
    return _resolve_template_path("Sol trans.xlsx")


def create_solicitud_excel(
    numero_referencia: str,
    fecha_documento: Optional[str] = None,
    fecha_enviado: Optional[str] = None,
    beneficiario: Optional[str] = None,
    banco: Optional[str] = None,
    cuenta: Optional[str] = None,
    cuenta_clabe: Optional[str] = None,
    cantidad_a_pagar: Optional[float] = None,
    proyecto: Optional[str] = None,
    fecha_pago: Optional[str] = None,
    concepto: Optional[str] = None,
    referencia_pago: Optional[str] = None,
    referencia_operaciones: Optional[str] = None,
    solicita: Optional[str] = None,
    autoriza: Optional[str] = None,
    aprueba: Optional[str] = None,
    ubicacion: Optional[str] = None,
    fase: Optional[str] = None,
    categorias: Optional[List[str]] = None,
    edicion: Optional[int] = None,
    currency: str = "MXN",
    **kwargs: Any,
) -> bytes:
    """
    Create Excel file for SOLICITUD documento by loading the template and filling
    only value cells. Merged cells, fonts, and layout are preserved from the
    template (Sol trans.xlsx), with small adjustments to match the expected
    header formatting.

    Returns:
        Excel file as bytes
    """
    template_path = _get_solicitud_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Solicitud template not found: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    from copy import copy
    from datetime import date as date_class, datetime
    from openpyxl.styles import Font

    # Header formatting adjustments (Hoja1):
    # - Keep row 1 blank (template already has it).
    # - A2: underline and merge only through column E.
    # - A3: slightly smaller, italic, merge only through column E.
    # - D4 = CDMX; E4 = document enviado date (or fecha_documento) in Spanish (same format as before).
    if "A2:F2" in {str(r) for r in ws.merged_cells.ranges}:
        ws.unmerge_cells("A2:F2")
    ws.merge_cells("A2:E2")
    a2 = ws["A2"]
    f2 = copy(a2.font)
    f2.underline = "single"
    a2.font = f2

    if "A3:F3" in {str(r) for r in ws.merged_cells.ranges}:
        ws.unmerge_cells("A3:F3")
    ws.merge_cells("A3:E3")
    a3 = ws["A3"]
    f3 = copy(a3.font)
    f3.italic = True
    if f3.size:
        f3.size = max(8, float(f3.size) - 2)
    a3.font = f3

    # Location + date block (row 4 per template expectations)
    ws["D4"].value = "CDMX"

    _weekday_es = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    _month_es = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    e4_date_str = ((fecha_enviado or "").strip() or (fecha_documento or "").strip())
    if e4_date_str:
        try:
            e4_dt = datetime.strptime(e4_date_str, "%Y-%m-%d")
            e4_weekday = _weekday_es[e4_dt.weekday()].capitalize()
            e4_month = _month_es[e4_dt.month - 1]
            ws["E4"].value = f"{e4_weekday}, {e4_dt.day} de {e4_month} del {e4_dt.year}"
        except Exception:
            ws["E4"].value = e4_date_str
    else:
        from ..utils.mexico_city_dates import today_mexico_city

        today = today_mexico_city()
        weekday = _weekday_es[today.weekday()].capitalize()
        month = _month_es[today.month - 1]
        # The template uses the merged cell E4:F4 for the date display; write into the anchor cell (E4).
        ws["E4"].value = f"{weekday}, {today.day} de {month} del {today.year}"

    # Body: only write into anchor cells of merged ranges so formatting is preserved.
    # Mapping follows docs/Sol trans.xlsx (Hoja1).
    if beneficiario is not None:
        ws["B5"].value = beneficiario
    if banco is not None:
        ws["B6"].value = banco
    if cuenta is not None:
        ws["B7"].value = cuenta
    if cuenta_clabe is not None:
        ws["E7"].value = cuenta_clabe

    if cantidad_a_pagar is not None and cantidad_a_pagar >= 0:
        ws["B8"].value = round(float(cantidad_a_pagar), 2)
    cantidad_en_letra_str = (
        _cantidad_en_letra(cantidad_a_pagar, currency=currency)
        if cantidad_a_pagar is not None
        else ""
    )
    ws["B9"].value = cantidad_en_letra_str or None

    if proyecto is not None or fase or categorias or edicion is not None:
        proyecto_cell = format_solicitud_proyecto_display(
            proyecto,
            fase=fase,
            categorias=categorias,
            edicion=edicion,
        )
        ws["B10"].value = proyecto_cell or None
    fecha_pago_display = None
    if fecha_pago:
        try:
            fecha_pago_dt = datetime.strptime(fecha_pago, "%Y-%m-%d")
            fecha_pago_weekday = _weekday_es[fecha_pago_dt.weekday()].capitalize()
            fecha_pago_month = _month_es[fecha_pago_dt.month - 1]
            fecha_pago_display = (
                f"{fecha_pago_weekday}, {fecha_pago_dt.day} de "
                f"{fecha_pago_month} del {fecha_pago_dt.year}"
            )
        except Exception:
            fecha_pago_display = fecha_pago
    ws["E10"].value = fecha_pago_display

    if concepto is not None:
        ws["B11"].value = concepto

    # numero_factura left blank
    ws["B12"].value = None

    # F12: Referencia Operaciones (informe de gastos); leave blank if missing
    ro_f12 = (referencia_operaciones or "").strip()
    ws["F12"].value = ro_f12 or None

    def _bold_cell_if_value(coord: str) -> None:
        cell = ws[coord]
        if cell.value is None:
            return
        bold_font = copy(cell.font)
        bold_font.bold = True
        cell.font = bold_font

    for _bold_coord in ("B5", "B6", "E7", "B8", "F12"):
        _bold_cell_if_value(_bold_coord)

    # SOLICITA / APRUEBA section (template uses C15 and E15)
    if solicita is not None:
        ws["C15"].value = solicita
    if aprueba is not None:
        ws["E15"].value = aprueba
    # autoriza not present in this template; ignore while keeping source extraction unchanged.

    # The source workbook contains historical sample rows below the signature
    # area, including stale contract totals and an IVA note. They are not part
    # of the generated solicitud and must not leak into downloads.
    from openpyxl.cell.cell import MergedCell

    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Template may ship with extra worksheets; export only the filled sheet (Sol trans layout).
    for _extra in list(wb.worksheets):
        if _extra is not ws:
            wb.remove(_extra)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _fmt_date_informe(s: Optional[str]) -> str:
    """Format date YYYY-MM-DD to DD/MM/YYYY for expense table (e.g. 08/01/2026)."""
    if not s:
        return ""
    try:
        from datetime import datetime
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return s


def _fmt_date_informe_header(s: Optional[str]) -> str:
    """Format date YYYY-MM-DD to Spanish long form for header (e.g. '2 de marzo de 2026')."""
    if not s:
        return ""
    try:
        from datetime import datetime
        _month_es = [
            "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
        ]
        dt = datetime.strptime(s, "%Y-%m-%d")
        return f"{dt.day} de {_month_es[dt.month - 1]} de {dt.year}"
    except Exception:
        return s


def _clear_cell(ws: Any, row: int, column: int) -> None:
    """Clear a cell so it is saved as empty. openpyxl's cell(row, col, value=None) does not
    clear existing values because None is the default; we must assign .value = None."""
    ws.cell(row=row, column=column).value = None


def _copy_cell_style(dest_cell, src_cell):
    """Copy font, border, alignment, fill from src_cell to dest_cell (openpyxl)."""
    from copy import copy
    if src_cell.font:
        dest_cell.font = copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dest_cell.alignment = copy(src_cell.alignment)
    if src_cell.fill:
        dest_cell.fill = copy(src_cell.fill)
    if src_cell.number_format:
        dest_cell.number_format = src_cell.number_format


def create_informe_excel(
    numero_referencia: str,
    empleado_nombre: Optional[str] = None,
    fecha_documento: Optional[str] = None,
    expenses: Optional[List[dict]] = None,
    cantidad_entregada: Optional[float] = None,
    saldo_cuenta: Optional[float] = None,
    tipo_cuenta: Optional[str] = None,
    referencia_operaciones: Optional[str] = None,
    motivo_del_gasto: Optional[str] = None,
    categorias: Optional[List[str]] = None,
    edicion: Optional[int] = None,
    currency: str = "MXN",
    **kwargs: Any,
) -> bytes:
    """
    Create Excel file for INFORME documento using template docs/4599- Informe de gastos.xlsx.
    Fills only value cells so merged cells, fonts, and layout match the template.
    If there are more than 20 expenses, inserts extra rows above TOTALES and copies
    cell format from the last template data row so new rows match the expense table style.

    Args:
        numero_referencia: Reference number (e.g. I-26000001)
        empleado_nombre: Name of person who verifies (Nombre de la persona que comprueba)
        fecha_documento: Document date string (YYYY-MM-DD)
        expenses: List of dicts with keys: concepto, fecha, no_factura (CFDI UUID), importe_sin_iva, iva, total
        cantidad_entregada: Sum of paid solicitudes (Cantidad Entregada)
        saldo_cuenta: Cuenta saldo (Dif. A Pagar y/o Reembolso / SALDO). G49: "a favor de empleado" if >= 0 else "a favor de empresa"
        tipo_cuenta: local|viaje|nacional|extranjero - row 11 selected label is shown as (X)
        referencia_operaciones: Document referencia operaciones (cell G22)
        motivo_del_gasto: Expense reason from cuenta nombre (merged cell B21:F22)
        **kwargs: Ignored

    Returns:
        Excel file as bytes
    """
    template_path = _get_informe_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Informe template not found: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    metadata_ws = wb.create_sheet("Metadata")
    metadata_ws.append(["Campo", "Valor"])
    metadata_ws.append(["Categorías", ", ".join(categorias or [])])
    metadata_ws.append(["Edición", edicion])
    metadata_ws.append(["Moneda", currency or "MXN"])

    # Clear template placeholders immediately (openpyxl cell(row, col, value=None) does not
    # clear existing values; we must assign .value = None on the cell object).
    _clear_cell(ws, 11, 3)   # C11 Local ( )
    _clear_cell(ws, 11, 4)   # D11 De viaje ( )
    _clear_cell(ws, 11, 5)   # E11 Nacional ( )
    _clear_cell(ws, 11, 6)   # F11 Extranjero ( )
    _clear_cell(ws, 21, 2)   # B21 anchor of B21:F22 (motivo del gasto)
    _clear_cell(ws, 22, 7)   # G22 reference number
    _clear_cell(ws, 21, 9)   # I21 anchor of I21:I22 (Gasto con American Express)
    for r in range(INFORME_EXPENSE_FIRST_ROW, INFORME_TOTALES_ROW):
        for c in range(2, 8):
            _clear_cell(ws, r, c)

    expenses = expenses or []
    num_expenses = len(expenses)
    # If more than template expense rows, insert rows above TOTALES so footer shifts down
    extra_rows = max(0, num_expenses - INFORME_EXPENSE_ROWS_TEMPLATE)
    if extra_rows > 0:
        ws.insert_rows(INFORME_TOTALES_ROW, amount=extra_rows)
        # Copy cell format from last template data row (row 46) to each new row so they match expense rows
        src_row = INFORME_TOTALES_ROW - 1  # 46 = last expense data row in template
        for i in range(extra_rows):
            new_row = INFORME_TOTALES_ROW + i
            for col in range(2, 8):  # columns B-G
                _copy_cell_style(ws.cell(row=new_row, column=col), ws.cell(row=src_row, column=col))

    row_totales = INFORME_TOTALES_ROW + extra_rows
    row_cantidad = INFORME_CANTIDAD_ROW + extra_rows
    row_saldo = INFORME_SALDO_ROW + extra_rows
    row_autorizado = INFORME_AUTORIZADO_ROW + extra_rows

    # Header per 4599: F5=date (anchor F5:I5) in Spanish long form; B8=nombre (anchor B8:F9)
    if fecha_documento:
        ws.cell(row=5, column=6, value=_fmt_date_informe_header(fecha_documento))
    if empleado_nombre:
        ws.cell(row=8, column=2, value=empleado_nombre)
    motivo_cell = (motivo_del_gasto or "").strip()
    if motivo_cell:
        ws.cell(row=21, column=2, value=motivo_cell)
    ro_cell = (referencia_operaciones or "").strip()
    if ro_cell:
        ws.cell(row=22, column=7, value=ro_cell)
    # Expense table: 4599 rows 27-46, columns B=2..G=7 (concepto, fecha, no_factura, importe_sin_iva, iva, total)
    for i, exp in enumerate(expenses):
        r = INFORME_EXPENSE_FIRST_ROW + i
        ws.cell(row=r, column=2, value=exp.get("concepto") or "")
        ws.cell(row=r, column=3, value=_fmt_date_informe(exp.get("fecha")) if exp.get("fecha") else "")
        ws.cell(row=r, column=4, value=exp.get("no_factura") or "")
        ws.cell(row=r, column=5, value=exp.get("importe_sin_iva"))
        ws.cell(row=r, column=6, value=exp.get("iva"))
        ws.cell(row=r, column=7, value=exp.get("total"))

    # Clear unused expense rows (value only)
    for r in range(INFORME_EXPENSE_FIRST_ROW + num_expenses, row_totales):
        for c in range(2, 8):
            _clear_cell(ws, r, c)

    # TOTALES row 47 (4599: B47="Total", E47,F47,G47 = sums)
    total_sub = sum((e.get("importe_sin_iva") or 0) for e in expenses)
    total_iva = sum((e.get("iva") or 0) for e in expenses)
    total_total = sum((e.get("total") or 0) for e in expenses)
    ws.cell(row=row_totales, column=2, value="Total")
    ws.cell(row=row_totales, column=5, value=round(total_sub, 2))
    ws.cell(row=row_totales, column=6, value=round(total_iva, 2))
    ws.cell(row=row_totales, column=7, value=round(total_total, 2))

    # Cantidad Entregada (4599: B48 label, G48 amount)
    if cantidad_entregada is not None:
        ws.cell(row=row_cantidad, column=7, value=round(cantidad_entregada, 2))

    # Dif. A Pagar y/o Reembolso / SALDO (4599: B49 label, F49 amount, G49 "a favor de...")
    if saldo_cuenta is not None:
        ws.cell(row=row_saldo, column=6, value=round(abs(saldo_cuenta), 2))
        ws.cell(row=row_saldo, column=7, value="a favor de empleado" if saldo_cuenta >= 0 else "a favor de empresa")

    # Autorizado por (4599: D54 = name)
    if empleado_nombre:
        ws.cell(row=row_autorizado, column=4, value=empleado_nombre)

    local_label, viaje_label, nacional_label, extranjero_label = _informe_template_tipo_labels(tipo_cuenta)
    ws.cell(row=11, column=3, value=local_label)
    ws.cell(row=11, column=4, value=viaje_label)
    ws.cell(row=11, column=5, value=nacional_label)
    ws.cell(row=11, column=6, value=extranjero_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def create_informe_csv(
    numero_referencia: str,
    empleado_nombre: Optional[str] = None,
    fecha_documento: Optional[str] = None,
    expenses: Optional[List[dict]] = None,
    cantidad_entregada: Optional[float] = None,
    saldo_cuenta: Optional[float] = None,
    tipo_cuenta: Optional[str] = None,
    **kwargs
) -> str:
    """
    Create CSV for INFORME with the same layout as the template (rows/columns match).
    Each CSV row = one template row; columns = A..H. Empty cells = empty field.

    tipo_cuenta: local|viaje|nacional|extranjero from Cuenta de Gastos. Populates the existing
    "TIPO DE GASTO" row (LOCAL/VIAJE); nacional uses LOCAL column, extranjero uses VIAJE.
    """
    expenses = expenses or []
    num_expenses = len(expenses)
    total_sub = sum((e.get("importe_sin_iva") or 0) for e in expenses)
    total_iva = sum((e.get("iva") or 0) for e in expenses)
    total_total = sum((e.get("total") or 0) for e in expenses)
    saldo_label = "a favor de empleado" if saldo_cuenta is not None and saldo_cuenta >= 0 else "a favor de empresa"
    local_x, viaje_x = _informe_template_tipo_marks(tipo_cuenta)

    def empty_row(cols: int = 8) -> List[str]:
        return [""] * cols

    def row(*values: Any) -> List[str]:
        row_list = list(values)
        while len(row_list) < 8:
            row_list.append("")
        return row_list[:8]

    rows: List[List[str]] = []
    # Rows 1-2
    rows.append(row("PLATAFORMA SPORTS"))
    rows.append(row("INFORME DE GASTOS"))
    rows.append(empty_row())
    # R4-5
    rows.append(row("FECHA:", fecha_documento or ""))
    rows.append(row("NOMBRE DE LA PERSONA QUE COMPRUEBA:", empleado_nombre or ""))
    # R6: TIPO DE GASTO row - existing column; we only populate Local/Viaje (no format change).
    rows.append(row("TIPO DE GASTO:", "", "LOCAL", local_x, "VIAJE", viaje_x))
    rows.append(row("PERSONAS QUE VIAJAN:"))
    rows.append(row("LUGAR A DONDE VIAJA:", "", "", "", "FECHAS DE VIAJE:", fecha_documento or ""))
    rows.append(row("MOTIVO DEL GASTO:"))
    # R10
    rows.append(row("REFERENCIA:", numero_referencia))
    rows.append(empty_row())
    # R12 header
    rows.append(row("CONCEPTO DEL GASTO", "", "FECHA", "NO. FACTURA", "IMPORTE S/IVA", "IVA", "TOTAL"))
    # Expense rows
    for exp in expenses:
        rows.append(row(
            exp.get("concepto") or "",
            "",
            exp.get("fecha") or "",
            exp.get("no_factura") or "",
            exp.get("importe_sin_iva") if exp.get("importe_sin_iva") is not None else "",
            exp.get("iva") if exp.get("iva") is not None else "",
            exp.get("total") if exp.get("total") is not None else "",
        ))
    # Empty rows up to where TOTALES would be (same vertical spacing as template: 1 blank after last expense then TOTALES)
    rows.append(empty_row())
    rows.append(row("TOTALES", "", "", "", round(total_sub, 2), round(total_iva, 2), round(total_total, 2)))
    rows.append(empty_row())
    rows.append(row("", "", "", "CANTIDAD ENTREGADA:", round(cantidad_entregada, 2) if cantidad_entregada is not None else ""))
    rows.append(row("", "", "", "DIFERENCIA / SALDO:", round(saldo_cuenta, 2) if saldo_cuenta is not None else "", saldo_label))
    rows.append(empty_row())
    rows.append(empty_row())
    rows.append(row("FIRMA DEL EMPLEADO:", "", "", "AUTORIZADO POR:", "", "", "ACUSE DE RECIBO:"))
    rows.append(row("", "", "", empleado_nombre or ""))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue()


def _cantidad_en_letra(monto: float, currency: str = "MXN") -> str:
    """Convert numeric amount to Spanish words for CANTIDAD EN LETRA."""
    from ..expense_metadata import cantidad_letra_currency_parts

    unit_label, currency_code = cantidad_letra_currency_parts(currency)
    if monto is None:
        return ""

    def _int_to_es(n: int) -> str:
        unidades = {
            0: "cero",
            1: "uno",
            2: "dos",
            3: "tres",
            4: "cuatro",
            5: "cinco",
            6: "seis",
            7: "siete",
            8: "ocho",
            9: "nueve",
            10: "diez",
            11: "once",
            12: "doce",
            13: "trece",
            14: "catorce",
            15: "quince",
            16: "dieciseis",
            17: "diecisiete",
            18: "dieciocho",
            19: "diecinueve",
            20: "veinte",
            21: "veintiuno",
            22: "veintidos",
            23: "veintitres",
            24: "veinticuatro",
            25: "veinticinco",
            26: "veintiseis",
            27: "veintisiete",
            28: "veintiocho",
            29: "veintinueve",
        }
        decenas = {
            30: "treinta",
            40: "cuarenta",
            50: "cincuenta",
            60: "sesenta",
            70: "setenta",
            80: "ochenta",
            90: "noventa",
        }
        centenas = {
            100: "cien",
            200: "doscientos",
            300: "trescientos",
            400: "cuatrocientos",
            500: "quinientos",
            600: "seiscientos",
            700: "setecientos",
            800: "ochocientos",
            900: "novecientos",
        }

        if n < 0:
            return "menos " + _int_to_es(abs(n))
        if n < 30:
            return unidades[n]
        if n < 100:
            d = (n // 10) * 10
            r = n % 10
            return decenas[d] if r == 0 else f"{decenas[d]} y {unidades[r]}"
        if n < 1000:
            if n == 100:
                return "cien"
            c = (n // 100) * 100
            r = n % 100
            pref = "ciento" if c == 100 else centenas[c]
            return pref if r == 0 else f"{pref} {_int_to_es(r)}"
        if n < 1_000_000:
            miles = n // 1000
            r = n % 1000
            if miles == 1:
                pref = "mil"
            else:
                pref = f"{_int_to_es(miles)} mil"
            return pref if r == 0 else f"{pref} {_int_to_es(r)}"
        if n < 1_000_000_000:
            millones = n // 1_000_000
            r = n % 1_000_000
            if millones == 1:
                pref = "un millon"
            else:
                pref = f"{_int_to_es(millones)} millones"
            return pref if r == 0 else f"{pref} {_int_to_es(r)}"
        miles_millones = n // 1_000_000_000
        r = n % 1_000_000_000
        if miles_millones == 1:
            pref = "mil millones"
        else:
            pref = f"{_int_to_es(miles_millones)} mil millones"
        return pref if r == 0 else f"{pref} {_int_to_es(r)}"

    try:
        from num2words import num2words  # type: ignore
        entero = int(float(monto))
        centavos = int(round((float(monto) - entero) * 100))
        if centavos >= 100:
            centavos = 0
            entero += 1
        letra_entero = num2words(entero, lang="es")
    except Exception:
        valor = round(float(monto or 0), 2)
        entero = int(valor)
        centavos = int(round((valor - entero) * 100))
        if centavos >= 100:
            centavos = 0
            entero += 1
        letra_entero = _int_to_es(entero)

    text = (letra_entero or "").strip()
    if text:
        text = text[0].upper() + text[1:]
    return f"{text} {unit_label} {centavos:02d}/100 {currency_code}".strip()


def create_solicitud_csv(
    numero_referencia: str,
    fecha_documento: Optional[str] = None,
    beneficiario: Optional[str] = None,
    banco: Optional[str] = None,
    cuenta: Optional[str] = None,
    cuenta_clabe: Optional[str] = None,
    cantidad_a_pagar: Optional[float] = None,
    proyecto: Optional[str] = None,
    fecha_pago: Optional[str] = None,
    concepto: Optional[str] = None,
    referencia_pago: Optional[str] = None,
    solicita: Optional[str] = None,
    autoriza: Optional[str] = None,
    aprueba: Optional[str] = None,
    ubicacion: Optional[str] = None,
    **kwargs
) -> str:
    """
    Create CSV for SOLICITUD DE TRANSFERENCIA with the same layout as the template.
    Works for both solicitudes de empleado (personal) and solicitudes a terceros.
    Each CSV row = one template row; columns match. Empty cells = empty field.
    """
    def row(*values: Any) -> List[str]:
        row_list = list(values)
        while len(row_list) < 8:
            row_list.append("")
        return row_list[:8]

    def empty_row(cols: int = 8) -> List[str]:
        return [""] * cols

    cantidad_str = ""
    if cantidad_a_pagar is not None and cantidad_a_pagar >= 0:
        cantidad_str = f"${cantidad_a_pagar:,.2f}"
    currency = kwargs.get("currency", "MXN")
    cantidad_en_letra_str = (
        _cantidad_en_letra(cantidad_a_pagar, currency=currency)
        if cantidad_a_pagar is not None
        else ""
    )

    rows: List[List[str]] = []
    # R1: PLATAFORMA SPORTS, col 7 = ubicación (e.g. CDMX)
    rows.append(row("PLATAFORMA SPORTS", "", "", "", "", "", ubicacion or ""))
    # R2: SOLICITUD DE TRANSFERENCIA, col 7 = fecha
    rows.append(row("SOLICITUD DE TRANSFERENCIA", "", "", "", "", "", fecha_documento or ""))
    rows.append(empty_row())
    rows.append(empty_row())
    # R5–R8: beneficiario, banco, cuenta, cuenta clabe
    rows.append(row("BENEFICIARIO:", beneficiario or ""))
    rows.append(row("BANCO:", banco or ""))
    rows.append(row("CUENTA:", cuenta or ""))
    rows.append(row("CUENTA CLABE:", cuenta_clabe or ""))
    rows.append(empty_row())
    # R10–R11: cantidad a pagar, cantidad en letra
    rows.append(row("CANTIDAD A PAGAR:", cantidad_str))
    rows.append(row("CANTIDAD EN LETRA:", cantidad_en_letra_str))
    rows.append(empty_row())
    # R13–R16: proyecto, fecha de pago, concepto, referencia de pago
    rows.append(row("PROYECTO:", proyecto or ""))
    rows.append(row("FECHA DE PAGO:", fecha_pago or ""))
    rows.append(row("CONCEPTO:", concepto or ""))
    rows.append(row("REFERENCIA DE PAGO:", referencia_pago or ""))
    rows.append(empty_row())
    rows.append(empty_row())
    # R19–R20: solicita, autoriza, aprueba
    rows.append(row("SOLICITA:", solicita or "", "", "AUTORIZA:", autoriza or ""))
    rows.append(row("APRUEBA:", aprueba or ""))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue()
