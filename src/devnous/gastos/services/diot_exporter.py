"""
DIOT export helpers.

Builds an auditable Excel workbook and the SAT TXT payload from expense/CFDI
data. The TXT is pipe-delimited and uses the expanded DIOT 2025 field order;
fields that Samchat cannot prove from CFDI/gasto data are emitted as zero/blank
instead of inferred aggressively.
"""

from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DIOT_2025_HEADERS: Tuple[str, ...] = (
    "tipo_tercero",
    "tipo_operacion",
    "rfc",
    "numero_id_fiscal",
    "nombre_extranjero",
    "pais_residencia",
    "lugar_jurisdiccion",
    "base_rfn",
    "bonificaciones_rfn",
    "base_rfs",
    "bonificaciones_rfs",
    "base_16",
    "bonificaciones_16",
    "base_importacion_tangible",
    "bonificaciones_importacion_tangible",
    "base_importacion_intangible",
    "bonificaciones_importacion_intangible",
    "iva_acreditable_rfn",
    "iva_acreditable_rfn_proporcion",
    "iva_acreditable_rfs",
    "iva_acreditable_rfs_proporcion",
    "iva_acreditable_16",
    "iva_acreditable_16_proporcion",
    "iva_acreditable_importacion_tangible",
    "iva_acreditable_importacion_tangible_proporcion",
    "iva_acreditable_importacion_intangible",
    "iva_acreditable_importacion_intangible_proporcion",
    "iva_no_acreditable_rfn_proporcion",
    "iva_no_acreditable_rfn_sin_requisitos",
    "iva_no_acreditable_rfn_exentas",
    "iva_no_acreditable_rfn_no_objeto",
    "iva_no_acreditable_rfs_proporcion",
    "iva_no_acreditable_rfs_sin_requisitos",
    "iva_no_acreditable_rfs_exentas",
    "iva_no_acreditable_rfs_no_objeto",
    "iva_no_acreditable_16_proporcion",
    "iva_no_acreditable_16_sin_requisitos",
    "iva_no_acreditable_16_exentas",
    "iva_no_acreditable_16_no_objeto",
    "iva_no_acreditable_importacion_tangible_proporcion",
    "iva_no_acreditable_importacion_tangible_sin_requisitos",
    "iva_no_acreditable_importacion_tangible_exentas",
    "iva_no_acreditable_importacion_tangible_no_objeto",
    "iva_no_acreditable_importacion_intangible_proporcion",
    "iva_no_acreditable_importacion_intangible_sin_requisitos",
    "iva_no_acreditable_importacion_intangible_exentas",
    "iva_no_acreditable_importacion_intangible_no_objeto",
    "iva_retenido",
    "base_iva_exento_importacion",
    "exentos",
    "base_cero",
    "no_objeto",
    "no_objeto_sin_establecimiento_nacional",
    "manifiesto",
)

DETAIL_HEADERS: Tuple[str, ...] = (
    "fecha_gasto",
    "referencia_gasto",
    "archivo",
    "uuid_cfdi",
    "rfc_emisor",
    "proveedor",
    "rfc_receptor",
    "concepto",
    "tipo_tercero",
    "tipo_operacion",
    "subtotal_cfdi",
    "base_16",
    "base_cero",
    "exentos",
    "no_objeto",
    "iva_trasladado",
    "iva_retenido",
    "total_cfdi",
    "total_gasto",
    "cuenta_contable",
    "advertencias",
)

CARGABATCH_HEADERS: Tuple[str, ...] = (
    "TIPO TERCERO",
    "TIPO DE OPERACIÓN",
    "RFC",
    "ID FISCAL",
    "NOMBRE DEL EXTRANJERO",
    "PAIS",
    "LUGAR DE JURISDICCION FISCAL",
    "ACTOS PAGADOS RFN",
    "DEV, DESC Y BON RFN",
    "ACTOS PAGADOS RFS",
    "DEV, DESC Y BON RFS",
    "BASE 16%",
    "DEV, DESC Y BON 16%",
    "IMPORTACIONES 16%",
    "DEV, DESC Y BON EN IMPORT. 16%",
    "IMPORT. INTANGIBLES 16%",
    "DEV, DESC Y BON IMPORT. INTANGIBLES 16%",
    "IVA ACRED. RFN",
    "IVA ACRED. RFN PROPORCIÓN",
    "IVA ACRED. RFS",
    "IVA ACRED. RFS PROPORCIÓN",
    "IVA ACRED. 16%",
    "IVA ACRED. 16% PROPORCIÓN",
    "IVA ACRED. IMPORT. TANGIBLE",
    "IVA ACRED. IMPORT. TANGIBLE PROPORCIÓN",
    "IVA ACRED. IMPORT. INTANGIBLE",
    "IVA ACRED. IMPORT. INTANGIBLE PROPORCIÓN",
    "IVA NO ACRED. RFN PROPORCIÓN",
    "IVA NO ACRED. RFN SIN REQUISITOS",
    "IVA NO ACRED. RFN EXENTAS",
    "IVA NO ACRED. RFN NO OBJETO",
    "IVA NO ACRED. RFS PROPORCIÓN",
    "IVA NO ACRED. RFS SIN REQUISITOS",
    "IVA NO ACRED. RFS EXENTAS",
    "IVA NO ACRED. RFS NO OBJETO",
    "IVA NO ACRED. 16% PROPORCIÓN",
    "IVA NO ACRED. 16% SIN REQUISITOS",
    "IVA NO ACRED. 16% EXENTAS",
    "IVA NO ACRED. 16% NO OBJETO",
    "IVA NO ACRED. IMPORT. TANGIBLE PROPORCIÓN",
    "IVA NO ACRED. IMPORT. TANGIBLE SIN REQUISITOS",
    "IVA NO ACRED. IMPORT. TANGIBLE EXENTAS",
    "IVA NO ACRED. IMPORT. TANGIBLE NO OBJETO",
    "IVA NO ACRED. IMPORT. INTANGIBLE PROPORCIÓN",
    "IVA NO ACRED. IMPORT. INTANGIBLE SIN REQUISITOS",
    "IVA NO ACRED. IMPORT. INTANGIBLE EXENTAS",
    "IVA NO ACRED. IMPORT. INTANGIBLE NO OBJETO",
    "IVA RETENIDO",
    "BASE IVA EXENTO IMPORTACIÓN",
    "EXENTOS",
    "BASE 0%",
    "NO OBJETO",
    "NO OBJETO SIN ESTABLECIMIENTO NACIONAL",
    "EFECTOS FISCALES A LOS COMPROBANTES",
)

CLAVES_TIPOS_ROWS: Tuple[Tuple[Optional[str], Optional[str]], ...] = (
    ("TIPO TERCERO", None),
    ("04", "PROVEEDOR NACIONAL"),
    ("05", "PROVEEDOR EXTRANJERO"),
    ("15", "PROVEEDOR GLOBAL"),
    (None, None),
    (None, None),
    ("TIPO DE OPERACIÓN", None),
    ("02", "ENAJENACION DE BIENES"),
    ("03", "PRESTACION DE SERV PROF"),
    ("06", "USO O GOCE TEMPORAL DE BIENES"),
    ("08", "IMPORTACION POR TRANSF VIRTUAL"),
    ("85", "OTROS"),
    ("07", "IMPORTACION DE BIENES O SERV"),
    ("87", "OPERACIONES GLOBALES"),
    (None, None),
    (None, None),
    ("EFECTOS FISCALES A LOS COMPROBANTES", None),
    ("01", "SI"),
    ("02", "NO"),
)


@dataclass
class DiotDetailRow:
    fecha_gasto: Optional[datetime]
    referencia_gasto: str
    archivo: str
    uuid_cfdi: str
    rfc_emisor: str
    proveedor: str
    rfc_receptor: str
    concepto: str
    tipo_tercero: str
    tipo_operacion: str
    subtotal_cfdi: Decimal = Decimal("0")
    base_16: Decimal = Decimal("0")
    base_cero: Decimal = Decimal("0")
    exentos: Decimal = Decimal("0")
    no_objeto: Decimal = Decimal("0")
    iva_trasladado: Decimal = Decimal("0")
    iva_retenido: Decimal = Decimal("0")
    total_cfdi: Decimal = Decimal("0")
    total_gasto: Decimal = Decimal("0")
    cuenta_contable: str = ""
    warnings: List[str] = field(default_factory=list)

    @property
    def group_key(self) -> Tuple[str, str, str, str, str, str, str]:
        return (
            self.tipo_tercero,
            self.tipo_operacion,
            self.rfc_emisor,
            "",
            "",
            "",
            "",
        )


@dataclass
class DiotSummaryRow:
    tipo_tercero: str
    tipo_operacion: str
    rfc: str
    numero_id_fiscal: str = ""
    nombre_extranjero: str = ""
    pais_residencia: str = ""
    lugar_jurisdiccion: str = ""
    amounts: Dict[str, Decimal] = field(default_factory=lambda: defaultdict(Decimal))

    def txt_values(self) -> List[str]:
        values: List[str] = [
            self.tipo_tercero,
            self.tipo_operacion,
            self.rfc,
            self.numero_id_fiscal,
            self.nombre_extranjero,
            self.pais_residencia,
            self.lugar_jurisdiccion,
        ]
        for header in DIOT_2025_HEADERS[7:]:
            if header == "manifiesto":
                values.append("02")
            else:
                values.append(_whole_pesos(self.amounts.get(header, Decimal("0"))))
        return values


@dataclass
class DiotExport:
    detail_rows: List[DiotDetailRow]
    summary_rows: List[DiotSummaryRow]
    warnings: List[str]


def _decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _money(value: Any) -> Decimal:
    return _decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _whole_pesos(value: Decimal) -> str:
    rounded = value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return str(max(Decimal("0"), rounded))


def _clean_text(value: Any) -> str:
    raw = str(value or "").strip()
    raw = raw.replace("\r", " ").replace("\n", " ").replace("|", " ")
    return re.sub(r"\s+", " ", raw)


def _clean_rfc(value: Any) -> str:
    raw = _clean_text(value).upper()
    return re.sub(r"[^A-Z0-9&Ñ]", "", raw)


def _fold(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()


def _infer_tipo_tercero(rfc: str) -> str:
    if rfc == "XAXX010101000":
        return "15"
    if rfc == "XEXX010101000":
        return "05"
    return "04"


def _infer_tipo_operacion(tipo_tercero: str, concepto: str) -> str:
    if tipo_tercero == "15":
        return "87"
    if tipo_tercero == "05":
        return "07"
    folded = _fold(concepto)
    if any(token in folded for token in ("honorario", "asesoria", "consultoria")):
        return "03"
    if any(token in folded for token in ("renta", "arrend", "hospedaje")):
        return "06"
    return "85"


def _iter_tax_items(cfdi_report: Any, key: str) -> Iterable[Dict[str, Any]]:
    detail = getattr(cfdi_report, "impuestos_detalle", None) or {}
    items = detail.get(key) or []
    return items if isinstance(items, list) else []


def _classify_bases(
    cfdi_report: Any, fallback_base: Decimal
) -> Tuple[Decimal, Decimal, Decimal, Decimal]:
    base_16 = Decimal("0")
    base_cero = Decimal("0")
    exentos = Decimal("0")
    no_objeto = Decimal("0")
    for traslado in _iter_tax_items(cfdi_report, "traslados"):
        impuesto = str(traslado.get("impuesto") or "")
        if impuesto and impuesto != "002":
            continue
        base = _money(traslado.get("base"))
        tipo_factor = _fold(str(traslado.get("tipo_factor") or ""))
        tasa = _decimal(traslado.get("tasa_o_cuota"))
        if tipo_factor == "exento":
            exentos += base
        elif tasa == Decimal("0"):
            base_cero += base
        elif tasa >= Decimal("0.15"):
            base_16 += base
    if base_16 == base_cero == exentos == Decimal("0"):
        iva = _money(getattr(cfdi_report, "total_impuestos_trasladados", None))
        if iva > 0:
            base_16 = fallback_base
        else:
            no_objeto = fallback_base
    return base_16, base_cero, exentos, no_objeto


def _iva_retenido(cfdi_report: Any) -> Decimal:
    total = Decimal("0")
    for retention in _iter_tax_items(cfdi_report, "retenciones"):
        impuesto = str(retention.get("impuesto") or "")
        if not impuesto or impuesto == "002":
            total += _money(retention.get("importe"))
    return total


def build_diot_export(expenses: Sequence[Any]) -> DiotExport:
    detail_rows: List[DiotDetailRow] = []
    warnings: List[str] = []

    for expense in expenses:
        cfdi = getattr(expense, "cfdi_report", None)
        total_gasto = _money(getattr(expense, "gasto_cantidad", 0))
        iva_fallback = _money(getattr(expense, "iva", 0))
        base_fallback = max(Decimal("0"), total_gasto - iva_fallback)
        concepto = _clean_text(
            getattr(cfdi, "descripcion_concepto_principal", None)
            or getattr(expense, "concepto", "")
            or "Gasto"
        )

        row_warnings: List[str] = []
        if cfdi is None:
            row_warnings.append("Sin CFDI vinculado; se usa total/IVA del gasto.")
            expense_ref = getattr(expense, "numero_referencia", expense.id)
            warnings.append(f"{expense_ref}: sin CFDI vinculado")

        rfc_emisor = _clean_rfc(getattr(cfdi, "emisor_rfc", ""))
        if not rfc_emisor:
            row_warnings.append("RFC emisor faltante.")
        tipo_tercero = _infer_tipo_tercero(rfc_emisor)
        tipo_operacion = _infer_tipo_operacion(tipo_tercero, concepto)

        subtotal_cfdi = (
            _money(getattr(cfdi, "subtotal", None)) if cfdi else base_fallback
        )
        total_cfdi = _money(getattr(cfdi, "total", None)) if cfdi else total_gasto
        base_16, base_cero, exentos, no_objeto = (
            _classify_bases(cfdi, subtotal_cfdi)
            if cfdi
            else (base_fallback, Decimal("0"), Decimal("0"), Decimal("0"))
        )
        iva_trasladado = (
            _money(getattr(cfdi, "total_impuestos_trasladados", None))
            if cfdi
            else iva_fallback
        )
        iva_ret = _iva_retenido(cfdi) if cfdi else Decimal("0")

        detail_rows.append(
            DiotDetailRow(
                fecha_gasto=getattr(expense, "fecha", None),
                referencia_gasto=_clean_text(getattr(expense, "numero_referencia", "")),
                archivo=_clean_text(
                    getattr(expense, "archivo_nombre", "")
                    or getattr(expense, "link_xml", "")
                    or getattr(cfdi, "cfdi_uuid", "")
                ),
                uuid_cfdi=_clean_text(getattr(cfdi, "cfdi_uuid", "")),
                rfc_emisor=rfc_emisor,
                proveedor=_clean_text(getattr(cfdi, "emisor_nombre", "")),
                rfc_receptor=_clean_rfc(getattr(cfdi, "receptor_rfc", "")),
                concepto=concepto,
                tipo_tercero=tipo_tercero,
                tipo_operacion=tipo_operacion,
                subtotal_cfdi=subtotal_cfdi,
                base_16=base_16,
                base_cero=base_cero,
                exentos=exentos,
                no_objeto=no_objeto,
                iva_trasladado=iva_trasladado,
                iva_retenido=iva_ret,
                total_cfdi=total_cfdi,
                total_gasto=total_gasto,
                cuenta_contable=_clean_text(
                    getattr(expense, "cuenta_contable_base", "")
                ),
                warnings=row_warnings,
            )
        )

    grouped: Dict[Tuple[str, str, str, str, str, str, str], DiotSummaryRow] = {}
    for row in detail_rows:
        if row.group_key not in grouped:
            grouped[row.group_key] = DiotSummaryRow(
                tipo_tercero=row.tipo_tercero,
                tipo_operacion=row.tipo_operacion,
                rfc=row.rfc_emisor,
            )
        summary = grouped[row.group_key]
        if row.tipo_tercero == "05":
            summary.amounts["base_importacion_intangible"] += row.base_16
            summary.amounts[
                "iva_acreditable_importacion_intangible"
            ] += row.iva_trasladado
        else:
            summary.amounts["base_16"] += row.base_16
            summary.amounts["iva_acreditable_16"] += row.iva_trasladado
        summary.amounts["base_cero"] += row.base_cero
        summary.amounts["exentos"] += row.exentos
        summary.amounts["no_objeto"] += row.no_objeto
        summary.amounts["iva_retenido"] += row.iva_retenido

    summary_rows = sorted(
        grouped.values(),
        key=lambda item: (item.tipo_tercero, item.tipo_operacion, item.rfc),
    )
    return DiotExport(
        detail_rows=detail_rows, summary_rows=summary_rows, warnings=warnings
    )


def generate_diot_txt(export: DiotExport) -> bytes:
    lines = ["|".join(row.txt_values()) for row in export.summary_rows]
    return ("\r\n".join(lines) + ("\r\n" if lines else "")).encode("utf-8")


def create_diot_excel(export: DiotExport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle CFDI"
    _write_sheet(
        ws, DETAIL_HEADERS, [_detail_values(row) for row in export.detail_rows]
    )

    txt_ws = wb.create_sheet("TXT DIOT")
    _write_sheet(
        txt_ws,
        DIOT_2025_HEADERS,
        [row.txt_values() for row in export.summary_rows],
    )

    warn_ws = wb.create_sheet("Advertencias")
    _write_sheet(warn_ws, ("advertencia",), [(warning,) for warning in export.warnings])

    cargabatch_ws = wb.create_sheet("CARGABATCH SAT")
    _write_cargabatch_sheet(cargabatch_ws, export.summary_rows)

    claves_ws = wb.create_sheet("CLAVES TIPOS")
    _write_claves_tipos_sheet(claves_ws)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _detail_values(row: DiotDetailRow) -> Tuple[Any, ...]:
    return (
        row.fecha_gasto.strftime("%Y-%m-%d") if row.fecha_gasto else "",
        row.referencia_gasto,
        row.archivo,
        row.uuid_cfdi,
        row.rfc_emisor,
        row.proveedor,
        row.rfc_receptor,
        row.concepto,
        row.tipo_tercero,
        row.tipo_operacion,
        float(row.subtotal_cfdi),
        float(row.base_16),
        float(row.base_cero),
        float(row.exentos),
        float(row.no_objeto),
        float(row.iva_trasladado),
        float(row.iva_retenido),
        float(row.total_cfdi),
        float(row.total_gasto),
        row.cuenta_contable,
        "; ".join(row.warnings),
    )


def _write_sheet(
    ws: Any, headers: Sequence[str], rows: Iterable[Sequence[Any]]
) -> None:
    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    for column_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(column_idx)
        width = min(max(len(str(header)) + 2, 12), 42)
        ws.column_dimensions[letter].width = width


def _write_cargabatch_sheet(ws: Any, rows: Sequence[DiotSummaryRow]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    title_fill = PatternFill("solid", fgColor="B6D7A8")
    field_count = len(CARGABATCH_HEADERS)

    for column_idx, title in enumerate(CARGABATCH_HEADERS, start=1):
        ws.cell(row=1, column=column_idx, value=column_idx).font = Font(bold=True)
        cell = ws.cell(row=2, column=column_idx, value=title)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    txt_title_col = field_count + 1
    ws.cell(row=1, column=txt_title_col, value="TXT").font = Font(bold=True)
    txt_cell = ws.cell(
        row=2,
        column=txt_title_col,
        value="COPIAR ESTA COLUMNA EN UN BLOC DE NOTAS (.TXT)",
    )
    txt_cell.fill = title_fill
    txt_cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=3):
        txt_values = row.txt_values()
        for column_idx, value in enumerate(txt_values, start=1):
            ws.cell(row=row_idx, column=column_idx, value=value)
        ws.cell(
            row=row_idx,
            column=txt_title_col,
            value=_cargabatch_formula(row_idx, field_count),
        )

    ws.freeze_panes = "A3"
    for column_idx, title in enumerate(CARGABATCH_HEADERS, start=1):
        letter = get_column_letter(column_idx)
        width = 18 if column_idx <= 7 else 20
        if len(title) > 28:
            width = 28
        ws.column_dimensions[letter].width = width
    ws.column_dimensions[get_column_letter(txt_title_col)].width = 44


def _cargabatch_formula(row_idx: int, field_count: int) -> str:
    tokens: List[str] = []
    for column_idx in range(1, field_count + 1):
        tokens.append(f"{get_column_letter(column_idx)}{row_idx}")
        if column_idx != field_count:
            tokens.append('"|"')
    return "=CONCATENATE(" + ",".join(tokens) + ")"


def _write_claves_tipos_sheet(ws: Any) -> None:
    for row in CLAVES_TIPOS_ROWS:
        ws.append(list(row))
    for cell_ref in ("A1", "A7", "A17"):
        ws[cell_ref].font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 42
