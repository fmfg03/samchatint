"""
Import helpers for proveedores/clientes uploads.

Supports:
- generic CSV uploads
- client RFC XLSX catalogs with beneficiary/bank/account/CLABE columns
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


VALID_TIPOS = {"proveedor", "cliente", "operadores_regionales"}


@dataclass
class ProveedorImportRow:
    tipo: str
    nombre: str
    rfc: Optional[str] = None
    banco: Optional[str] = None
    cuenta_clabe: Optional[str] = None
    cuenta_bancaria: Optional[str] = None
    entidad_region: Optional[str] = None
    activo: bool = True


def normalize_nombre(nombre: str) -> str:
    if not nombre:
        return ""
    normalized = unicodedata.normalize("NFKD", nombre.strip())
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.upper()


def normalize_header(value: object) -> str:
    return normalize_nombre(str(value or "")).lower()


def parse_activo(value: str) -> bool:
    lowered = (value or "").strip().lower()
    if lowered in {"false", "0", "no", "inactivo"}:
        return False
    return True


def clean_cuenta_bancaria(value: str) -> Optional[str]:
    if not value:
        return None
    text = value.strip()
    for separator in [" — ", " – ", " - ", "—", "–"]:
        if separator in text:
            text = text.split(separator)[0].strip()
            break
    return text or None


def normalize_clabe(value: str) -> Optional[str]:
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    if len(digits) == 18:
        return digits
    return None


def proveedor_match_key(row: ProveedorImportRow) -> Tuple[str, ...]:
    if row.rfc:
        return (row.tipo, row.rfc.upper())
    normalized_name = normalize_nombre(row.nombre)
    if row.cuenta_clabe:
        return (row.tipo, normalized_name, row.cuenta_clabe)
    if row.cuenta_bancaria:
        return (row.tipo, normalized_name, row.cuenta_bancaria)
    return (row.tipo, normalized_name)


def _decode_text(contents: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("No se pudo decodificar el archivo. Use UTF-8 o Latin-1.")


def _build_csv_row(fieldnames: Dict[str, str], row: Dict[str, str]) -> Optional[ProveedorImportRow]:
    tipo = str(row.get(fieldnames["tipo"]) or "").strip().lower()
    nombre = str(row.get(fieldnames["nombre"]) or "").strip()
    if not tipo or not nombre:
        return None
    if tipo not in VALID_TIPOS:
        return None

    rfc = str(row.get(fieldnames.get("rfc", "")) or "").strip() or None
    banco = str(row.get(fieldnames.get("banco", "")) or "").strip() or None
    cuenta_clabe = normalize_clabe(str(row.get(fieldnames.get("cuenta_clabe", "")) or "").strip())
    cuenta_bancaria = clean_cuenta_bancaria(str(row.get(fieldnames.get("cuenta_bancaria", "")) or row.get(fieldnames.get("cuenta_contable_codigo", "")) or "").strip())
    entidad_region = str(row.get(fieldnames.get("entidad_region", "")) or "").strip() or None
    activo = parse_activo(str(row.get(fieldnames.get("activo", "")) or "").strip())
    return ProveedorImportRow(
        tipo=tipo,
        nombre=nombre,
        rfc=rfc,
        banco=banco,
        cuenta_clabe=cuenta_clabe,
        cuenta_bancaria=cuenta_bancaria,
        entidad_region=entidad_region,
        activo=activo,
    )


def _parse_csv_rows(contents: bytes) -> List[ProveedorImportRow]:
    text = _decode_text(contents)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("El archivo CSV está vacío o no tiene encabezados.")

    fieldnames = {normalize_header(field): field for field in reader.fieldnames if field}
    required = ["tipo", "nombre"]
    missing = [name for name in required if name not in fieldnames]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    rows: List[ProveedorImportRow] = []
    seen = set()
    for row in reader:
        if not any((value or "").strip() for value in row.values() if isinstance(value, str)):
            continue
        built = _build_csv_row(fieldnames, row)
        if not built:
            continue
        key = proveedor_match_key(built)
        if key in seen:
            continue
        seen.add(key)
        rows.append(built)
    return rows


def _detect_rfc_xlsx_header(ws) -> Optional[int]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True), start=1):
        normalized = [normalize_header(cell) for cell in row]
        if "beneficiario" in normalized and ("clabe" in normalized or "cuenta clabe" in normalized):
            return row_idx
    return None


def _parse_rfc_xlsx_rows(contents: bytes) -> List[ProveedorImportRow]:
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("No se pudo leer el archivo XLSX.") from exc
    ws = wb[wb.sheetnames[0]]
    header_row = _detect_rfc_xlsx_header(ws)
    if not header_row:
        raise ValueError("No pude detectar encabezados tipo RFC (BENEFICIARIO/BANCOS/CUENTA/CLABE).")

    header_values = [normalize_header(cell) for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    index = {name: idx for idx, name in enumerate(header_values)}
    rows: List[ProveedorImportRow] = []
    seen = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        nombre = str(row[index.get("beneficiario", 0)] or "").strip()
        if not nombre:
            continue
        banco = str(row[index.get("bancos", 1)] or "").strip() or None
        cuenta_bancaria = clean_cuenta_bancaria(str(row[index.get("cuenta", 2)] or "").strip())
        clabe_raw = ""
        if "clabe" in index:
            clabe_raw = str(row[index["clabe"]] or "").strip()
        elif "cuenta clabe" in index:
            clabe_raw = str(row[index["cuenta clabe"]] or "").strip()
        cuenta_clabe = normalize_clabe(clabe_raw)
        built = ProveedorImportRow(
            tipo="proveedor",
            nombre=nombre,
            banco=banco,
            cuenta_clabe=cuenta_clabe,
            cuenta_bancaria=cuenta_bancaria,
            activo=True,
        )
        key = proveedor_match_key(built)
        if key in seen:
            continue
        seen.add(key)
        rows.append(built)
    return rows


def parse_proveedores_clientes_upload(filename: str, contents: bytes) -> List[ProveedorImportRow]:
    lower = (filename or "").strip().lower()
    if lower.endswith(".csv"):
        return _parse_csv_rows(contents)
    if lower.endswith(".xlsx"):
        return _parse_rfc_xlsx_rows(contents)
    if lower.endswith(".xls"):
        raise ValueError("Archivos .xls no soportados todavía. Convierte el archivo a .xlsx o CSV.")
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")
