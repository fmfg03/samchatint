"""
COI Poliza Exporter

Generates CSV/XLSX files in the COI poliza format for expenses. Supports both
CFDI-backed and non-CFDI expenses.
"""

import csv
import io
import logging
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .accounting_constants import DEFAULT_IVA_ACCOUNT_CODE

logger = logging.getLogger(__name__)

IVA_ACCOUNT = DEFAULT_IVA_ACCOUNT_CODE
PAYABLE_ACCOUNT = ""
COI_COLUMNS = 9
COI_MONEY_NUMBER_FORMAT = "$#,##0.00"


@dataclass
class ExpenseCFDI:
    """DTO for expenses with the fields needed by COI."""

    fecha: datetime
    total: float
    iva_amount: float
    subtotal_amount: float
    concepto: str
    cuenta_contable: str
    cuenta_contrapartida: str
    cfdi_uuid: Optional[str] = None
    cfdi_date: Optional[datetime] = None
    rfc_emisor: Optional[str] = None
    rfc_receptor: Optional[str] = None
    folio: Optional[str] = None
    nombre_emisor: Optional[str] = None
    cuenta_iva: Optional[str] = None
    retenciones: List[dict] = field(default_factory=list)
    impuestos_locales: List[dict] = field(default_factory=list)
    gastos_no_deducibles: List[dict] = field(default_factory=list)
    neto_contrapartida: Optional[float] = None
    base_amount: Optional[float] = None
    export_reference: Optional[str] = None
    receptor_uso_cfdi: Optional[str] = None
    cuenta_contable_nombre: Optional[str] = None
    allows_missing_cfdi: bool = False
    missing_cfdi_warning: Optional[str] = None


CoiCell = Union[str, int, float]
CoiRow = List[CoiCell]


def plain_amount_value(amount: float) -> CoiCell:
    """Return a plain numeric COI amount without currency formatting."""
    if not amount:
        return ""
    return round(float(amount), 2)


def format_date_dd_mm_yy(date: datetime) -> str:
    return date.strftime("%d/%m/%y")


def format_uso_factura_code(uso_cfdi: Optional[str]) -> str:
    """Return the two-digit COI uso de factura code from a CFDI UsoCFDI value."""
    text = (uso_cfdi or "").strip().upper()
    if not text:
        return ""
    match = re.search(r"(\d{2})$", text)
    if match:
        return match.group(1)
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    return digits[-2:].zfill(2)


def _expense_description(expense: ExpenseCFDI) -> str:
    if expense.cfdi_uuid:
        folio_str = expense.folio or expense.cfdi_uuid[:8]
        nombre_emisor_str = expense.nombre_emisor or "N/A"
    else:
        folio_str = ""
        nombre_emisor_str = ""
    uso_code = format_uso_factura_code(expense.receptor_uso_cfdi)
    prefix = f"{uso_code} / " if uso_code else ""
    return f"{prefix}{folio_str} / {nombre_emisor_str} / {expense.concepto or 'Gasto'}"


def _active_amount_lines(lines: List[dict]) -> List[dict]:
    return [item for item in lines if float(item.get("importe") or 0.0) > 0]


def _counterpart_net(expense: ExpenseCFDI, retention_lines: List[dict]) -> float:
    if expense.neto_contrapartida is not None:
        return expense.neto_contrapartida
    retention_total = sum(float(item.get("importe") or 0.0) for item in retention_lines)
    return max(0.0, expense.total - retention_total)


def _expense_base(
    expense: ExpenseCFDI,
    retention_lines: List[dict],
    local_tax_lines: Optional[List[dict]] = None,
    non_deductible_lines: Optional[List[dict]] = None,
) -> float:
    if expense.base_amount is not None:
        return expense.base_amount
    retention_total = sum(float(item.get("importe") or 0.0) for item in retention_lines)
    local_tax_total = sum(
        float(item.get("importe") or 0.0) for item in (local_tax_lines or [])
    )
    non_deductible_total = sum(
        float(item.get("importe") or 0.0) for item in (non_deductible_lines or [])
    )
    return (
        expense.total
        - expense.iva_amount
        - local_tax_total
        - non_deductible_total
        + retention_total
    )


def _expense_movements(expense: ExpenseCFDI) -> List[Dict[str, Any]]:
    desc = _expense_description(expense)
    retention_lines = _active_amount_lines(list(expense.retenciones or []))
    local_tax_lines = _active_amount_lines(list(expense.impuestos_locales or []))
    non_deductible_lines = _active_amount_lines(
        list(expense.gastos_no_deducibles or [])
    )
    neto_contrapartida = _counterpart_net(expense, retention_lines)
    expense_base = _expense_base(
        expense,
        retention_lines,
        local_tax_lines,
        non_deductible_lines,
    )

    movements: List[Dict[str, Any]] = [
        {
            "kind": "gasto_base",
            "cuenta": expense.cuenta_contable,
            "concepto": desc,
            "debe": expense_base,
            "haber": 0.0,
        }
    ]
    if expense.iva_amount > 0:
        movements.append(
            {
                "kind": "iva",
                "cuenta": expense.cuenta_iva or IVA_ACCOUNT,
                "concepto": desc,
                "debe": expense.iva_amount,
                "haber": 0.0,
            }
        )
    for local_tax in local_tax_lines:
        movements.append(
            {
                "kind": f"impuesto_local_{local_tax.get('kind') or 'tax'}",
                "cuenta": local_tax.get("cuenta_contable") or expense.cuenta_contable,
                "concepto": f"{desc} / {local_tax.get('label') or 'Impuesto local'}",
                "debe": float(local_tax.get("importe") or 0.0),
                "haber": 0.0,
                "label": local_tax.get("label") or "Impuesto local",
            }
        )
    for non_deductible in non_deductible_lines:
        movements.append(
            {
                "kind": f"no_deducible_{non_deductible.get('kind') or 'gasto'}",
                "cuenta": non_deductible.get("cuenta_contable")
                or expense.cuenta_contable,
                "concepto": (
                    f"{desc} / {non_deductible.get('label') or 'No deducible'}"
                ),
                "debe": float(non_deductible.get("importe") or 0.0),
                "haber": 0.0,
                "label": non_deductible.get("label") or "No deducible",
            }
        )
    for retention in retention_lines:
        movements.append(
            {
                "kind": "retencion",
                "cuenta": (
                    retention.get("cuenta_contable")
                    or PAYABLE_ACCOUNT
                    or expense.cuenta_contrapartida
                ),
                "concepto": (
                    f"{desc} / Retención {retention.get('label') or 'impuesto'}"
                ),
                "debe": 0.0,
                "haber": float(retention.get("importe") or 0.0),
                "label": retention.get("label") or "impuesto",
            }
        )
    movements.append(
        {
            "kind": "contrapartida",
            "cuenta": expense.cuenta_contrapartida
            or PAYABLE_ACCOUNT
            or expense.cuenta_contable,
            "concepto": desc,
            "debe": 0.0,
            "haber": neto_contrapartida,
        }
    )
    return movements


def _safe_export_filename(value: Optional[str], fallback: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip()).strip("._-")
    return text or fallback


def build_coi_poliza_preview(expenses: List[ExpenseCFDI]) -> List[Dict[str, Any]]:
    preview: List[Dict[str, Any]] = []
    for n, expense in enumerate(expenses, start=1):
        movements = _expense_movements(expense)
        preview.append(
            {
                "index": n,
                "description": _expense_description(expense),
                "num_movements": len(movements),
                "expense": expense,
                "has_cfdi": expense.cfdi_uuid is not None,
                "cfdi_date_str": (
                    format_date_dd_mm_yy(expense.cfdi_date or expense.fecha)
                    if expense.cfdi_uuid
                    else ""
                ),
                "movements": movements,
                "totals": {
                    "debe": round(sum(item["debe"] for item in movements), 2),
                    "haber": round(sum(item["haber"] for item in movements), 2),
                },
            }
        )
    return preview


def _movement_row(movement: Dict[str, Any]) -> CoiRow:
    return [
        "",
        movement["cuenta"],
        "0",
        movement["concepto"],
        "1",
        plain_amount_value(movement["debe"]),
        plain_amount_value(movement["haber"]),
        "",
        "",
    ]


def _cfdi_block_rows(expense: ExpenseCFDI) -> List[CoiRow]:
    cfdi_date = expense.cfdi_date or expense.fecha
    return [
        ["", "", "INICIO_CFDI", "", "", "", "", "", ""],
        [
            "",
            "",
            format_date_dd_mm_yy(cfdi_date),
            "",
            "0",
            f" {expense.rfc_emisor}" if expense.rfc_emisor else "  ",
            f" {expense.rfc_receptor}" if expense.rfc_receptor else "  ",
            plain_amount_value(expense.total),
            expense.cfdi_uuid or "",
        ],
        ["", "", "FIN_CFDI", "", "", "", "", "", ""],
    ]


def build_coi_poliza_rows(expenses: List[ExpenseCFDI]) -> List[CoiRow]:
    """Build the 9-column COI import layout."""
    rows: List[CoiRow] = [
        ["", "", "", "", "", "", "", "", ""],
        ["|||", "", "", "", "", "", "", "", ""],
    ]
    for n, expense in enumerate(expenses, start=1):
        desc = _expense_description(expense)
        movements = _expense_movements(expense)
        rows.append(["Eg", str(n), desc, str(len(movements)), "", "", "", "", ""])

        if movements:
            rows.append(_movement_row(movements[0]))

        if expense.cfdi_uuid:
            rows.extend(_cfdi_block_rows(expense))

        for movement in movements[1:]:
            rows.append(_movement_row(movement))

        rows.append(["", "FIN_PARTIDAS", "", "", "", "", "", "", ""])
    return rows


def validate_coi_poliza(expenses: List[ExpenseCFDI]) -> List[Dict[str, Any]]:
    """Return issues that should be reviewed before importing into COI."""
    if not expenses:
        return [
            {
                "severity": "error",
                "partida": "",
                "campo": "gastos",
                "mensaje": "No hay gastos activos para generar la poliza COI.",
            }
        ]

    issues: List[Dict[str, Any]] = []
    previews = build_coi_poliza_preview(expenses)
    for item in previews:
        expense = item["expense"]
        partida = item["index"]
        if not (expense.cuenta_contable or "").strip():
            issues.append(
                {
                    "severity": "error",
                    "partida": partida,
                    "campo": "cuenta_contable",
                    "mensaje": "Falta la cuenta contable del gasto.",
                }
            )
        if not (expense.cuenta_contrapartida or "").strip():
            issues.append(
                {
                    "severity": "error",
                    "partida": partida,
                    "campo": "cuenta_contrapartida",
                    "mensaje": "Falta la cuenta bancaria/contable de contrapartida.",
                }
            )
        for retention in expense.retenciones or []:
            if float(retention.get("importe") or 0.0) <= 0:
                continue
            if not (retention.get("cuenta_contable") or "").strip():
                issues.append(
                    {
                        "severity": "error",
                        "partida": partida,
                        "campo": "retencion",
                        "mensaje": (
                            "Una retencion no tiene cuenta contable asignada: "
                            f"{retention.get('label') or 'impuesto'}."
                        ),
                    }
                )
        for local_tax in expense.impuestos_locales or []:
            if float(local_tax.get("importe") or 0.0) <= 0:
                continue
            if not (local_tax.get("cuenta_contable") or "").strip():
                issues.append(
                    {
                        "severity": "warning",
                        "partida": partida,
                        "campo": "impuesto_local",
                        "mensaje": (
                            "Un impuesto local usara la cuenta del gasto por falta "
                            "de cuenta especifica: "
                            f"{local_tax.get('label') or 'impuesto local'}."
                        ),
                    }
                )
        for non_deductible in expense.gastos_no_deducibles or []:
            if float(non_deductible.get("importe") or 0.0) <= 0:
                continue
            if not (non_deductible.get("cuenta_contable") or "").strip():
                issues.append(
                    {
                        "severity": "warning",
                        "partida": partida,
                        "campo": "no_deducible",
                        "mensaje": (
                            "Un gasto no deducible usara la cuenta del gasto por "
                            "falta de cuenta especifica."
                        ),
                    }
                )
        if not (expense.cfdi_uuid or "").strip():
            issues.append(
                {
                    "severity": "warning",
                    "partida": partida,
                    "campo": "cfdi_uuid",
                    "mensaje": (
                        expense.missing_cfdi_warning
                        or (
                            "El gasto no tiene CFDI vinculado; el Excel no incluira "
                            "bloque Info-CFDI para esa partida."
                        )
                    ),
                }
            )

        debe = round(float(item["totals"]["debe"] or 0), 2)
        haber = round(float(item["totals"]["haber"] or 0), 2)
        if abs(debe - haber) >= 0.01:
            issues.append(
                {
                    "severity": "error",
                    "partida": partida,
                    "campo": "balance",
                    "mensaje": (
                        f"La partida no cuadra: debe {debe:.2f}, haber {haber:.2f}."
                    ),
                }
            )

    total_debe = round(sum(item["totals"]["debe"] for item in previews), 2)
    total_haber = round(sum(item["totals"]["haber"] for item in previews), 2)
    if abs(total_debe - total_haber) >= 0.01:
        issues.append(
            {
                "severity": "error",
                "partida": "total",
                "campo": "balance",
                "mensaje": (
                    f"La poliza completa no cuadra: debe {total_debe:.2f}, "
                    f"haber {total_haber:.2f}."
                ),
            }
        )
    return issues


def _style_review_sheet(sheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 28
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_coi_import_sheet(sheet) -> None:
    for column_index in range(1, COI_COLUMNS + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 22
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")

        is_movement_row = row[2].value == "0" and row[4].value == "1"
        if not is_movement_row:
            continue

        for amount_cell in (row[5], row[6]):
            if isinstance(amount_cell.value, (int, float)):
                amount_cell.number_format = COI_MONEY_NUMBER_FORMAT


def generate_coi_poliza_xlsx(
    expenses: List[ExpenseCFDI],
    *,
    lote_manifest_rows: Optional[List[List[str]]] = None,
) -> bytes:
    """Generate an XLSX workbook with COI import rows and validation details."""
    rows = build_coi_poliza_rows(expenses)
    issues = validate_coi_poliza(expenses)
    preview = build_coi_poliza_preview(expenses)

    wb = Workbook()
    ws = wb.active
    ws.title = "Poliza COI"
    for row in rows:
        ws.append(row[:COI_COLUMNS])
    _style_coi_import_sheet(ws)

    summary = wb.create_sheet("Resumen")
    total_debe = round(sum(item["totals"]["debe"] for item in preview), 2)
    total_haber = round(sum(item["totals"]["haber"] for item in preview), 2)
    for row in [
        ["Concepto", "Valor"],
        ["Partidas", len(expenses)],
        ["Renglones COI", len(rows)],
        ["Debe", total_debe],
        ["Haber", total_haber],
        ["Diferencia", round(total_debe - total_haber, 2)],
        ["Errores", sum(1 for item in issues if item["severity"] == "error")],
        ["Advertencias", sum(1 for item in issues if item["severity"] == "warning")],
    ]:
        summary.append(row)

    validation = wb.create_sheet("Validacion")
    validation.append(["Severidad", "Partida", "Campo", "Mensaje"])
    if issues:
        for issue in issues:
            validation.append(
                [
                    issue["severity"],
                    issue["partida"],
                    issue["campo"],
                    issue["mensaje"],
                ]
            )
    else:
        validation.append(["ok", "", "", "Sin hallazgos de validacion."])

    _style_review_sheet(summary)
    _style_review_sheet(validation)

    if lote_manifest_rows:
        manifest = wb.create_sheet("Manifest")
        for row in lote_manifest_rows:
            manifest.append(row)
        _style_review_sheet(manifest)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_coi_poliza_zip(
    expenses: List[ExpenseCFDI],
    *,
    filename_prefix: str = "Poliza_COI",
) -> bytes:
    """Generate a ZIP with one COI workbook per expense/poliza."""
    output = io.BytesIO()
    used_names: set[str] = set()

    with zipfile.ZipFile(
        output,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        for index, expense in enumerate(expenses, start=1):
            xlsx_bytes = generate_coi_poliza_xlsx([expense])
            reference = _safe_export_filename(
                expense.export_reference
                or expense.folio
                or expense.cfdi_uuid
                or expense.concepto,
                fallback=f"poliza_{index:03d}",
            )
            filename = f"{filename_prefix}_{reference}.xlsx"
            if filename in used_names:
                filename = f"{filename_prefix}_{index:03d}_{reference}.xlsx"
            used_names.add(filename)
            archive.writestr(filename, xlsx_bytes)

    output.seek(0)
    return output.getvalue()


def generate_coi_poliza_csv(expenses: List[ExpenseCFDI]) -> bytes:
    """
    Generate a CSV file in COI poliza format.

    The CSV has 9 columns and follows the same row layout used by the Excel
    export.
    """
    if not expenses:
        logger.warning("No expenses provided to generate COI poliza CSV")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(build_coi_poliza_rows(expenses))
    output.seek(0)
    return output.getvalue().encode("latin-1")
