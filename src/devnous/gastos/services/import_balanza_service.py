"""
Import helpers for chart of accounts uploads.

Supports:
- generic CSV uploads with headers like codigo/nombre[/activo][/tipo]
- client Balanza XLSX files with columns Cuenta / Descripción de la cuenta
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class CuentaContableImportRow:
    codigo: str
    nombre: str
    tipo: str
    activo: bool = True


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def _decode_text(contents: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("No se pudo decodificar el archivo. Use UTF-8 o Latin-1.")


def _parse_bool(value: object, default: bool = True) -> bool:
    normalized = str(value or "").strip().lower()
    if normalized in {"false", "0", "no", "inactivo"}:
        return False
    if normalized in {"true", "1", "si", "sí", "activo", ""}:
        return True
    return default


def infer_cuenta_tipo(codigo: str, nombre: str) -> str:
    codigo_norm = (codigo or "").strip().upper()
    nombre_norm = _normalize_header(nombre)

    if codigo_norm.startswith("1110"):
        return "caja"
    if codigo_norm.startswith(("1120", "1130")):
        return "banco"
    if codigo_norm.startswith(("1200", "1210")):
        return "iva"
    if codigo_norm.startswith(("2100", "2150", "2160")):
        return "pasivo"
    if codigo_norm.startswith(("5300", "5400", "5500", "5600")):
        return "gasto"

    if any(token in nombre_norm for token in ("banco", "santander", "banorte", "banamex", "bbva", "hsbc", "multiva")):
        return "banco"
    if any(token in nombre_norm for token in ("iva", "impuesto al valor agregado")):
        return "iva"
    if any(token in nombre_norm for token in ("retencion", "retención")):
        return "retencion"
    if any(token in nombre_norm for token in ("proveedor", "acreedor")):
        return "proveedor"
    if any(token in nombre_norm for token in ("anticipo", "deudor")):
        return "anticipo"
    if any(token in nombre_norm for token in ("gasto", "viatico", "viático", "hospedaje", "transporte", "alimentos", "medico", "médico")):
        return "gasto"
    return "otros"


def _parse_csv_rows(contents: bytes) -> List[CuentaContableImportRow]:
    text = _decode_text(contents)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("El archivo CSV está vacío o no tiene encabezados.")

    fields: Dict[str, str] = {_normalize_header(field): field for field in reader.fieldnames if field}
    codigo_col = fields.get("codigo")
    nombre_col = fields.get("nombre")
    if not codigo_col or not nombre_col:
        raise ValueError("Faltan columnas requeridas: codigo, nombre.")

    activo_col = fields.get("activo")
    tipo_col = fields.get("tipo")
    rows: List[CuentaContableImportRow] = []
    for row in reader:
        if not any((value or "").strip() for value in row.values() if isinstance(value, str)):
            continue
        codigo = str(row.get(codigo_col) or "").strip()
        nombre = str(row.get(nombre_col) or "").strip()
        if not codigo or not nombre:
            continue
        activo = _parse_bool(row.get(activo_col) if activo_col else "", default=True)
        tipo = str(row.get(tipo_col) or "").strip().lower() if tipo_col else ""
        rows.append(
            CuentaContableImportRow(
                codigo=codigo,
                nombre=nombre,
                tipo=tipo or infer_cuenta_tipo(codigo, nombre),
                activo=activo,
            )
        )
    return rows


def _detect_balanza_header(ws) -> Optional[int]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True), start=1):
        normalized = [_normalize_header(cell) for cell in row]
        if "cuenta" in normalized and any(val in normalized for val in ("descripcion de la cuenta", "descripción de la cuenta")):
            return row_idx
    return None


def _parse_balanza_xlsx_rows(contents: bytes) -> List[CuentaContableImportRow]:
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = _detect_balanza_header(ws)
    if not header_row:
        raise ValueError("No pude detectar encabezados tipo Balanza (Cuenta / Descripción de la cuenta).")

    rows: List[CuentaContableImportRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo = str(row[0] or "").strip()
        nombre = str(row[1] or "").strip() if len(row) > 1 else ""
        if not codigo or not nombre:
            continue
        rows.append(
            CuentaContableImportRow(
                codigo=codigo,
                nombre=nombre,
                tipo=infer_cuenta_tipo(codigo, nombre),
                activo=True,
            )
        )
    return rows


def parse_cuentas_contables_upload(filename: str, contents: bytes) -> List[CuentaContableImportRow]:
    suffix = (filename or "").strip().lower()
    if suffix.endswith(".csv"):
        return _parse_csv_rows(contents)
    if suffix.endswith(".xlsx"):
        return _parse_balanza_xlsx_rows(contents)
    if suffix.endswith(".xls"):
        raise ValueError("Archivos .xls no soportados todavía. Convierte el archivo a .xlsx o CSV.")
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")
