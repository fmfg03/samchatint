"""
Helpers to inspect and port legacy payroll workbook logic.

The workbook is not a normative source of truth.
It is treated as a legacy calculation spec whose formulas can be ported into
backend services while normative values come from the regulatory tables.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class WorkbookConceptSpec:
    concept_key: str
    display_name: str
    input_column: str
    taxable_column: str
    exempt_column: str
    input_header: str
    taxable_mode: str
    exempt_formula_key: str
    affects_sbc: bool
    notes: Optional[str] = None


@dataclass(frozen=True)
class WorkbookLogicSummary:
    workbook_name: str
    payroll_type: str
    smgv: float
    uma_daily: float
    payroll_days: float
    vacation_premium_rate: float
    aguinaldo_days: float
    concepts: List[WorkbookConceptSpec]
    vacation_schedule: Dict[int, int]
    warnings: List[str]


_CONCEPTS: List[WorkbookConceptSpec] = [
    WorkbookConceptSpec("salary", "Sueldo", "L", "Z", "AN", "Sueldo", "fully_taxable", "none", False),
    WorkbookConceptSpec("overtime_double", "Horas Extras dobles", "M", "AA", "AO", "Horas Extras dobles", "split_formula", "half_capped_5_uma_daily", False),
    WorkbookConceptSpec("overtime_triple", "Horas Extras triples", "N", "AB", "AP", "Horas Extras triples", "fully_taxable", "none", False),
    WorkbookConceptSpec("sunday_premium", "Prima Dominical", "O", "AC", "AQ", "Prima Dominical", "split_formula", "capped_1_uma_daily", False),
    WorkbookConceptSpec("food_vouchers", "Vales de Despensa", "P", "AD", "AR", "Vales de Despensa", "fully_taxable", "none", False),
    WorkbookConceptSpec("punctuality_bonus", "Premio Puntualidad", "Q", "AE", "AS", "Premio Puntualidad", "fully_taxable", "none", False),
    WorkbookConceptSpec("vacation_premium", "Prima Vacacional", "R", "AF", "AT", "Prima Vacacional", "split_formula", "capped_1_uma_daily", False),
    WorkbookConceptSpec("worked_rest_day", "Descanso laborado", "S", "AG", "AU", "Descanso laborado", "split_formula", "half_only_if_within_5_uma_daily", True),
    WorkbookConceptSpec("holiday_work", "Días Festivos", "T", "AH", "AV", "Días Festivos", "fully_taxable", "none", False),
    WorkbookConceptSpec("birth_aid", "Ayuda por Nacimiento", "U", "AI", "AW", "Ayuda por Nacimiento", "fully_taxable", "none", False),
    WorkbookConceptSpec(
        "death_aid_legacy_v",
        "Ayuda por Fallecimiento (legacy V)",
        "V",
        "AJ",
        "AX",
        "Premio Asistencia",
        "split_formula",
        "capped_365_uma_daily",
        False,
        "El header de entrada dice 'Premio Asistencia', pero la fórmula exenta corresponde a ayuda por fallecimiento.",
    ),
    WorkbookConceptSpec("school_aid", "Ayuda Escolar", "W", "AK", "AY", "Ayuda Escolar", "fully_taxable", "none", False),
    WorkbookConceptSpec("savings_fund", "Fondo de Ahorro", "X", "AL", "AZ", "Fondo de Ahorro", "fully_exempt", "full_amount", False),
    WorkbookConceptSpec("other_perception", "Otros", "Y", "AM", "BA", "Otros", "fully_taxable", "none", False),
]


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _find_index_value(ws, label: str) -> Any:
    wanted = label.strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
        first = str(row[0] or "").strip().lower() if row else ""
        if first == wanted:
            return row[1] if len(row) > 1 else None
    raise KeyError(f"Label not found in workbook index sheet: {label}")


def _lookup_payroll_days(tablas_ws, payroll_type: str) -> float:
    wanted = payroll_type.strip().lower()
    for row in tablas_ws.iter_rows(min_row=1, max_row=min(tablas_ws.max_row, 12), min_col=18, max_col=19, values_only=True):
        jornada = str(row[0] or "").strip().lower()
        if jornada == wanted:
            return _to_float(row[1])
    return 0.0


def parse_payroll_workbook(path: str | Path) -> WorkbookLogicSummary:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    idx = wb["Indice"]
    nom = wb["Nomina"]
    tablas = wb["Tablas"]

    warnings: List[str] = []

    payroll_type = str(_find_index_value(idx, "Tipo de Nómina") or "").strip()
    smgv = _to_float(_find_index_value(idx, "SMGV"))
    uma_daily = _to_float(_find_index_value(idx, "UMA"))
    vacation_premium_rate = _to_float(_find_index_value(idx, "Prima Vacacional (%)"))
    aguinaldo_days = _to_float(_find_index_value(idx, "Dias de Aguinaldo"))
    payroll_days = _lookup_payroll_days(tablas, payroll_type)

    header_row = next(nom.iter_rows(min_row=3, max_row=3, values_only=True))
    headers = {get_column_letter(idx): str(value or "").strip() for idx, value in enumerate(header_row, start=1)}

    if headers.get("V") != "Premio Asistencia":
        warnings.append("El header legacy de la columna V cambió; revisar mapping de ayuda por fallecimiento legacy.")
    else:
        warnings.append(
            "La columna V del workbook está etiquetada como 'Premio Asistencia', pero su fórmula exenta corresponde a ayuda por fallecimiento."
        )

    vacation_schedule: Dict[int, int] = {}
    for row_no in range(2, min(tablas.max_row, 40) + 1):
        years_value = tablas[f"U{row_no}"].value
        days_value = tablas[f"V{row_no}"].value
        if years_value is None or days_value is None:
            continue
        try:
            vacation_schedule[int(years_value)] = int(days_value)
        except (TypeError, ValueError):
            continue

    concepts: List[WorkbookConceptSpec] = []
    for spec in _CONCEPTS:
        input_header = headers.get(spec.input_column, "")
        if input_header and input_header != spec.input_header and spec.concept_key != "death_aid_legacy_v":
            warnings.append(
                f"Header inesperado en {spec.input_column}: '{input_header}' (esperado '{spec.input_header}')."
            )
        concepts.append(
            WorkbookConceptSpec(
                concept_key=spec.concept_key,
                display_name=spec.display_name,
                input_column=spec.input_column,
                taxable_column=spec.taxable_column,
                exempt_column=spec.exempt_column,
                input_header=input_header or spec.input_header,
                taxable_mode=spec.taxable_mode,
                exempt_formula_key=spec.exempt_formula_key,
                affects_sbc=spec.affects_sbc,
                notes=spec.notes,
            )
        )

    return WorkbookLogicSummary(
        workbook_name=workbook_path.name,
        payroll_type=payroll_type,
        smgv=smgv,
        uma_daily=uma_daily,
        payroll_days=payroll_days,
        vacation_premium_rate=vacation_premium_rate,
        aguinaldo_days=aguinaldo_days,
        concepts=concepts,
        vacation_schedule=vacation_schedule,
        warnings=warnings,
    )


def parse_payroll_workbook_as_dict(path: str | Path) -> Dict[str, Any]:
    summary = parse_payroll_workbook(path)
    return {
        "workbook_name": summary.workbook_name,
        "payroll_type": summary.payroll_type,
        "smgv": summary.smgv,
        "uma_daily": summary.uma_daily,
        "payroll_days": summary.payroll_days,
        "vacation_premium_rate": summary.vacation_premium_rate,
        "aguinaldo_days": summary.aguinaldo_days,
        "concepts": [asdict(item) for item in summary.concepts],
        "vacation_schedule": summary.vacation_schedule,
        "warnings": summary.warnings,
    }


__all__ = ["WorkbookConceptSpec", "WorkbookLogicSummary", "parse_payroll_workbook", "parse_payroll_workbook_as_dict"]
