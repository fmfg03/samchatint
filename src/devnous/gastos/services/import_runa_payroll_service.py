"""
Runa payroll layout import service.

Reads the Runa "Empleados" sheet and maps it into the normalized payroll
employee structure:
- payroll_employees
- payroll_employee_compensation_profiles
- payroll_employee_payment_profiles
- payroll_employee_deduction_profiles
- payroll_employee_benefit_profiles
- payroll_employee_address_profiles

Design constraints:
- dry-run first
- apply only when a row can be matched to an existing internal `empleado`
  or to an existing `payroll_employee`
- no phantom employee creation
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models import (
    Empleado,
    PayrollEmployee,
    PayrollEmployeeAddressProfile,
    PayrollEmployeeBenefitProfile,
    PayrollEmployeeCompensationProfile,
    PayrollEmployeeDeductionProfile,
    PayrollEmployeePaymentProfile,
)


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date_parts(day_value: Any, month_value: Any, year_value: Any) -> Optional[date]:
    if day_value in (None, "") or month_value in (None, "") or year_value in (None, ""):
        return None
    try:
        return date(int(float(year_value)), int(float(month_value)), int(float(day_value)))
    except (TypeError, ValueError):
        return None


def _parse_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Optional[Decimal]:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


@dataclass
class ParsedRunaEmployeeRow:
    source_row_number: int
    employee_code: Optional[str]
    first_name: Optional[str]
    paternal_last_name: Optional[str]
    maternal_last_name: Optional[str]
    birth_date: Optional[date]
    hire_date: Optional[date]
    gender: Optional[str]
    rfc: Optional[str]
    curp: Optional[str]
    nss: Optional[str]
    birth_place: Optional[str]
    tax_regime: Optional[str]
    personal_email: Optional[str]
    work_email: Optional[str]
    personal_postal_code: Optional[str]
    job_title: Optional[str]
    department_name: Optional[str]
    contract_start_date: Optional[date]
    contract_end_date: Optional[date]
    employment_state: Optional[str]
    employee_type: Optional[str]
    contract_type: Optional[str]
    policy_name: Optional[str]
    worker_type: Optional[str]
    geographic_area: Optional[str]
    schedule_scheme: Optional[str]
    reduced_workweek_type: Optional[str]
    worked_days_override: Optional[Decimal]
    compensation_regime: Optional[str]
    salary_type: Optional[str]
    monthly_net_salary: Optional[Decimal]
    daily_salary: Optional[Decimal]
    integrated_daily_salary: Optional[Decimal]
    severance_daily_salary: Optional[Decimal]
    payment_method: Optional[str]
    bank_name: Optional[str]
    account_number: Optional[str]
    clabe: Optional[str]
    customer_number: Optional[str]
    deduction_name: Optional[str]
    infonavit_discount_type: Optional[str]
    infonavit_discount_value: Optional[Decimal]
    infonavit_notice_folio: Optional[str]
    infonavit_credit_number: Optional[str]
    infonavit_start_date: Optional[date]
    loan_balance: Optional[Decimal]
    monthly_deduction_amount: Optional[Decimal]
    payroll_deduction_name: Optional[str]
    fonacot_credit_folio: Optional[str]
    alimony_percentage: Optional[Decimal]
    vacation_balance: Optional[Decimal]
    umf: Optional[str]
    street: Optional[str]
    exterior_number: Optional[str]
    interior_number: Optional[str]
    neighborhood: Optional[str]
    municipality: Optional[str]
    address_state: Optional[str]
    address_postal_code: Optional[str]
    voucher_provider: Optional[str]
    voucher_account_number: Optional[str]
    voucher_card_number: Optional[str]

    @property
    def full_name(self) -> str:
        return " ".join(
            part for part in [self.first_name, self.paternal_last_name, self.maternal_last_name] if part
        ).strip()


def parse_runa_workbook(filename: str, contents: bytes) -> List[ParsedRunaEmployeeRow]:
    lower_name = (filename or "").lower()
    if not lower_name.endswith(".xlsx"):
        raise ValueError("Debe seleccionar un archivo XLSX válido.")

    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    if "Empleados" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'Empleados' en el layout de Runa.")

    ws = wb["Empleados"]
    header_row_idx = 3
    headers = [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]
    normalized_to_index = {_normalize_header(value): idx for idx, value in enumerate(headers)}

    def get_value(row: tuple[Any, ...], header_name: str) -> Any:
        idx = normalized_to_index.get(_normalize_header(header_name))
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed_rows: List[ParsedRunaEmployeeRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        employee_code = _clean_text(get_value(row, "Código empleado"))
        first_name = _clean_text(get_value(row, "Nombre"))
        paternal_last_name = _clean_text(get_value(row, "Apellido paterno"))
        if not employee_code and not first_name:
            continue

        parsed_rows.append(
            ParsedRunaEmployeeRow(
                source_row_number=row_idx,
                employee_code=employee_code,
                first_name=first_name,
                paternal_last_name=paternal_last_name,
                maternal_last_name=_clean_text(get_value(row, "Apellido materno")),
                birth_date=_parse_date_parts(
                    get_value(row, "Fecha de nacimiento (Día)"),
                    get_value(row, "Fecha de nacimiento (Mes)"),
                    get_value(row, "Fecha de nacimiento (Año)"),
                ),
                hire_date=_parse_date_parts(
                    get_value(row, "Fecha de antigüedad (Día)"),
                    get_value(row, "Fecha de antigüedad (Mes)"),
                    get_value(row, "Fecha de antigüedad (Año)"),
                ),
                gender=_clean_text(get_value(row, "Género")),
                rfc=_clean_text(get_value(row, "RFC")),
                curp=_clean_text(get_value(row, "CURP")),
                nss=_clean_text(get_value(row, "Número de seguro social")),
                birth_place=_clean_text(get_value(row, "Lugar nacimiento")),
                tax_regime=_clean_text(get_value(row, "Régimen fiscal")),
                personal_email=_clean_text(get_value(row, "Email para notificaciones")),
                work_email=_clean_text(get_value(row, "Email laboral")),
                personal_postal_code=_clean_text(get_value(row, "Código postal")),
                job_title=_clean_text(get_value(row, "Puesto")),
                department_name=_clean_text(get_value(row, "Área")),
                contract_start_date=_parse_date_parts(
                    get_value(row, "Inicio de contrato (día)"),
                    get_value(row, "Inicio de contrato (mes)"),
                    get_value(row, "Inicio de contrato (año)"),
                ),
                contract_end_date=_parse_date_parts(
                    get_value(row, "Fin de contrato (día)"),
                    get_value(row, "Fin de contrato (mes)"),
                    get_value(row, "Fin de contrato (año)"),
                ),
                employment_state=_clean_text(get_value(row, "Estado")),
                employee_type=_clean_text(get_value(row, "Tipo de empleado")),
                contract_type=_clean_text(get_value(row, "Tipo de contrato")),
                policy_name=_clean_text(get_value(row, "Política")),
                worker_type=_clean_text(get_value(row, "Tipo trabajador")),
                geographic_area=_clean_text(get_value(row, "Área geográfica")),
                schedule_scheme=_clean_text(get_value(row, "Esquema jornada (pestaña anterior)")),
                reduced_workweek_type=_clean_text(get_value(row, "Semana/jornada reducida")),
                worked_days_override=_parse_decimal(get_value(row, "Número días laborados")),
                compensation_regime=_clean_text(get_value(row, "Sueldos y salarios/asimilado")),
                salary_type=_clean_text(get_value(row, "Tipo de sueldo")),
                monthly_net_salary=_parse_decimal(get_value(row, "Salario neto mensual")),
                daily_salary=_parse_decimal(get_value(row, "Salario diario")),
                integrated_daily_salary=_parse_decimal(get_value(row, "Salario base de cotización (SBC)")),
                severance_daily_salary=_parse_decimal(get_value(row, "Salario indemnización")),
                payment_method=_clean_text(get_value(row, "Método de pago")),
                bank_name=_clean_text(get_value(row, "Banco")),
                account_number=_clean_text(get_value(row, "Número de cuenta")),
                clabe=_clean_text(get_value(row, "CLABE")),
                customer_number=_clean_text(get_value(row, "Número de cliente\n(sólo si requerido)")),
                deduction_name=_clean_text(get_value(row, "Nombre del descuento")),
                infonavit_discount_type=_clean_text(get_value(row, "Tipo de descuento (INFONAVIT)")),
                infonavit_discount_value=_parse_decimal(get_value(row, "Valor de descuento (INFONAVIT)")),
                infonavit_notice_folio=_clean_text(get_value(row, "Folio del aviso (INFONAVIT)")),
                infonavit_credit_number=_clean_text(get_value(row, "No. crédito (INFONAVIT)")),
                infonavit_start_date=_parse_date(get_value(row, "Fecha de inicio (INFONAVIT)")),
                loan_balance=_parse_decimal(get_value(row, "Saldo (préstamo)")),
                monthly_deduction_amount=_parse_decimal(get_value(row, "Monto mensual (Caja de ahorro, préstamo, FONACOT)")),
                payroll_deduction_name=_clean_text(get_value(row, "Descuento nómina (Caja de ahorro, préstamo, FONACOT)")),
                fonacot_credit_folio=_clean_text(get_value(row, "Folio del crédito (FONACOT)")),
                alimony_percentage=_parse_decimal(get_value(row, "¿Pensión alimenticia?  Sí -> porcentaje")),
                vacation_balance=_parse_decimal(get_value(row, "Saldo vacaciones")),
                umf=_clean_text(get_value(row, "Unidad médico-familiar")),
                street=_clean_text(get_value(row, "Calle")),
                exterior_number=_clean_text(get_value(row, "Número exterior")),
                interior_number=_clean_text(get_value(row, "Número interior")),
                neighborhood=_clean_text(get_value(row, "Colonia")),
                municipality=_clean_text(get_value(row, "Municipio")),
                address_state=_clean_text(get_value(row, "Estado")),
                address_postal_code=_clean_text(get_value(row, "Código postal")),
                voucher_provider=_clean_text(get_value(row, "Proveedor")),
                voucher_account_number=_clean_text(get_value(row, "Número de cuenta")),
                voucher_card_number=_clean_text(get_value(row, "Número de tarjeta")),
            )
        )

    return parsed_rows


async def _build_runa_lookup_maps(
    session: AsyncSession,
    rows: List[ParsedRunaEmployeeRow],
) -> tuple[Dict[str, PayrollEmployee], Dict[str, PayrollEmployee], Dict[str, Empleado], Dict[str, PayrollEmployee]]:
    employee_numbers = sorted({row.employee_code for row in rows if row.employee_code})
    curps = sorted({row.curp for row in rows if row.curp})
    emails = sorted({email for row in rows for email in [row.work_email, row.personal_email] if email})

    payroll_stmt = select(PayrollEmployee).options(
        selectinload(PayrollEmployee.empleado),
        selectinload(PayrollEmployee.compensation_profile),
        selectinload(PayrollEmployee.payment_profile),
        selectinload(PayrollEmployee.deduction_profile),
        selectinload(PayrollEmployee.benefit_profile),
        selectinload(PayrollEmployee.address_profile),
    )
    payroll_filters = []
    if employee_numbers:
        payroll_filters.append(PayrollEmployee.employee_number.in_(employee_numbers))
    if curps:
        payroll_filters.append(PayrollEmployee.curp.in_(curps))
    payroll_rows: List[PayrollEmployee] = []
    if payroll_filters:
        payroll_result = await session.execute(payroll_stmt.where(or_(*payroll_filters)))
        payroll_rows = payroll_result.scalars().all()
    payroll_by_number: Dict[str, PayrollEmployee] = {}
    payroll_by_curp: Dict[str, PayrollEmployee] = {}
    payroll_by_empleado_id: Dict[str, PayrollEmployee] = {}
    for payroll in payroll_rows:
        if payroll.employee_number:
            payroll_by_number[payroll.employee_number] = payroll
        if payroll.curp:
            payroll_by_curp[payroll.curp] = payroll
        payroll_by_empleado_id[str(payroll.empleado_id)] = payroll

    empleados: List[Empleado] = []
    if emails:
        empleado_result = await session.execute(
            select(Empleado).where(
                Empleado.activo == True,
                Empleado.correo.in_(emails),
            )
        )
        empleados = empleado_result.scalars().all()
    empleado_by_email = {empleado.correo: empleado for empleado in empleados if empleado.correo}
    return payroll_by_number, payroll_by_curp, empleado_by_email, payroll_by_empleado_id


def _apply_runa_row_to_models(
    parsed: ParsedRunaEmployeeRow,
    payroll_employee: PayrollEmployee,
) -> None:
    payroll_employee.employee_number = parsed.employee_code
    payroll_employee.birth_date = parsed.birth_date
    payroll_employee.birth_place = parsed.birth_place
    payroll_employee.gender = parsed.gender
    payroll_employee.curp = parsed.curp
    payroll_employee.rfc = parsed.rfc
    payroll_employee.nss = parsed.nss
    payroll_employee.tax_regime = parsed.tax_regime
    payroll_employee.personal_email = parsed.personal_email
    payroll_employee.work_email = parsed.work_email
    payroll_employee.personal_postal_code = parsed.personal_postal_code
    payroll_employee.hire_date = parsed.hire_date
    payroll_employee.seniority_date = parsed.hire_date
    payroll_employee.contract_start_date = parsed.contract_start_date
    payroll_employee.contract_end_date = parsed.contract_end_date
    payroll_employee.employment_state = parsed.employment_state
    payroll_employee.employee_type = parsed.employee_type
    payroll_employee.contract_type = parsed.contract_type
    payroll_employee.policy_name = parsed.policy_name
    payroll_employee.worker_type = parsed.worker_type
    payroll_employee.geographic_area = parsed.geographic_area
    payroll_employee.schedule_scheme = parsed.schedule_scheme
    payroll_employee.reduced_workweek_type = parsed.reduced_workweek_type
    payroll_employee.worked_days_override = parsed.worked_days_override
    payroll_employee.job_title = parsed.job_title
    payroll_employee.department_name = parsed.department_name
    payroll_employee.daily_salary = parsed.daily_salary
    payroll_employee.integrated_daily_salary = parsed.integrated_daily_salary
    payroll_employee.payment_method = (parsed.payment_method or payroll_employee.payment_method or "transfer").lower()
    payroll_employee.bank_name = parsed.bank_name
    payroll_employee.bank_account_last4 = (parsed.account_number or "")[-4:] or None

    comp = payroll_employee.compensation_profile or PayrollEmployeeCompensationProfile(payroll_employee_id=payroll_employee.id)
    comp.compensation_regime = parsed.compensation_regime
    comp.salary_type = parsed.salary_type
    comp.monthly_net_salary = parsed.monthly_net_salary
    comp.daily_salary = parsed.daily_salary
    comp.integrated_daily_salary = parsed.integrated_daily_salary
    comp.variable_salary = None
    comp.severance_daily_salary = parsed.severance_daily_salary
    comp.work_risk_class = None
    payroll_employee.compensation_profile = comp

    payment = payroll_employee.payment_profile or PayrollEmployeePaymentProfile(payroll_employee_id=payroll_employee.id)
    payment.payment_method = parsed.payment_method
    payment.bank_name = parsed.bank_name
    payment.account_number = parsed.account_number
    payment.clabe = parsed.clabe
    payment.customer_number = parsed.customer_number
    payroll_employee.payment_profile = payment

    deduction = payroll_employee.deduction_profile or PayrollEmployeeDeductionProfile(payroll_employee_id=payroll_employee.id)
    deduction.deduction_name = parsed.deduction_name
    deduction.infonavit_discount_type = parsed.infonavit_discount_type
    deduction.infonavit_discount_value = parsed.infonavit_discount_value
    deduction.infonavit_notice_folio = parsed.infonavit_notice_folio
    deduction.infonavit_credit_number = parsed.infonavit_credit_number
    deduction.infonavit_start_date = parsed.infonavit_start_date
    deduction.loan_balance = parsed.loan_balance
    deduction.monthly_deduction_amount = parsed.monthly_deduction_amount
    deduction.payroll_deduction_name = parsed.payroll_deduction_name
    deduction.fonacot_credit_folio = parsed.fonacot_credit_folio
    deduction.alimony_percentage = parsed.alimony_percentage
    payroll_employee.deduction_profile = deduction

    benefit = payroll_employee.benefit_profile or PayrollEmployeeBenefitProfile(payroll_employee_id=payroll_employee.id)
    benefit.vacation_balance = parsed.vacation_balance
    benefit.umf = parsed.umf
    benefit.voucher_provider = parsed.voucher_provider
    benefit.voucher_account_number = parsed.voucher_account_number
    benefit.voucher_card_number = parsed.voucher_card_number
    payroll_employee.benefit_profile = benefit

    address = payroll_employee.address_profile or PayrollEmployeeAddressProfile(payroll_employee_id=payroll_employee.id)
    address.street = parsed.street
    address.exterior_number = parsed.exterior_number
    address.interior_number = parsed.interior_number
    address.neighborhood = parsed.neighborhood
    address.municipality = parsed.municipality
    address.state = parsed.address_state
    address.postal_code = parsed.address_postal_code
    payroll_employee.address_profile = address


async def import_runa_payroll_workbook(
    session: AsyncSession,
    *,
    filename: str,
    contents: bytes,
    apply_changes: bool,
) -> Dict[str, Any]:
    rows = parse_runa_workbook(filename, contents)
    (
        payroll_by_number,
        payroll_by_curp,
        empleado_by_email,
        payroll_by_empleado_id,
    ) = await _build_runa_lookup_maps(session, rows)
    created = 0
    updated = 0
    skipped = 0
    matched = 0
    warnings: List[Dict[str, Any]] = []
    samples: List[Dict[str, Any]] = []

    for parsed in rows:
        empleado = None
        payroll = None
        match_source = None

        if parsed.employee_code and parsed.employee_code in payroll_by_number:
            payroll = payroll_by_number[parsed.employee_code]
            empleado = payroll.empleado
            match_source = "payroll_employee_number"
        elif parsed.curp and parsed.curp in payroll_by_curp:
            payroll = payroll_by_curp[parsed.curp]
            empleado = payroll.empleado
            match_source = "payroll_curp"
        else:
            candidate_emails = [email for email in [parsed.work_email, parsed.personal_email] if email]
            matched_emails = [empleado_by_email[email] for email in candidate_emails if email in empleado_by_email]
            unique_empleados = {str(emp.id): emp for emp in matched_emails}
            if len(unique_empleados) == 1:
                empleado = next(iter(unique_empleados.values()))
                payroll = payroll_by_empleado_id.get(str(empleado.id))
                match_source = "empleado_email"
            elif len(unique_empleados) > 1:
                skipped += 1
                warnings.append(
                    {
                        "row": parsed.source_row_number,
                        "employee_code": parsed.employee_code,
                        "name": parsed.full_name,
                        "warning": "La fila coincide con más de un empleado interno por correo",
                    }
                )
                continue

        if empleado is None and payroll is None:
            skipped += 1
            warnings.append(
                {
                    "row": parsed.source_row_number,
                    "employee_code": parsed.employee_code,
                    "name": parsed.full_name,
                    "warning": "No se pudo vincular con un empleado interno activo",
                }
            )
            continue

        matched += 1
        action = "update"
        if payroll is None:
            created += 1
            action = "create"
            if apply_changes:
                payroll = PayrollEmployee(empleado_id=empleado.id)
                session.add(payroll)
                await session.flush()
                payroll_by_empleado_id[str(empleado.id)] = payroll
                if parsed.employee_code:
                    payroll_by_number[parsed.employee_code] = payroll
                if parsed.curp:
                    payroll_by_curp[parsed.curp] = payroll
        else:
            updated += 1

        if apply_changes and payroll is not None:
            _apply_runa_row_to_models(parsed, payroll)

        if len(samples) < 12:
            samples.append(
                {
                    "row": parsed.source_row_number,
                    "employee_code": parsed.employee_code,
                    "name": parsed.full_name,
                    "match_source": match_source,
                    "action": action,
                }
            )

    if apply_changes:
        await session.commit()
    else:
        await session.rollback()

    return {
        "mode": "apply" if apply_changes else "dry_run",
        "rows_seen": len(rows),
        "matched_internal_employee": matched,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "warnings": warnings[:50],
        "warning_count": len(warnings),
        "samples": samples,
    }
