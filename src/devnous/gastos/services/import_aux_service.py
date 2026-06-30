"""
Auxiliary ledger XLSX import service.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import (
    AccountingImportRun,
    AccountingPoliza,
    AuxLedgerEntry,
    CuentaContable,
)


@dataclass
class ParsedAuxEntry:
    source_sheet: str
    source_row_number: int
    cuenta_codigo: str
    cuenta_nombre: str
    tipo_poliza: Optional[str]
    numero_poliza: Optional[str]
    fecha: Optional[datetime]
    concepto: Optional[str]
    saldo_inicial: Optional[float]
    debe: Optional[float]
    haber: Optional[float]
    saldo: Optional[float]
    cfdi_uuid: Optional[str]
    raw_row: Dict[str, Any]


def _parse_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: Any) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _extract_account_header(text: str) -> tuple[str, str]:
    cleaned = str(text or "").strip()
    match = re.match(r"^Cuenta\s*:\s*([0-9\-]+)\s+(.+)$", cleaned, re.IGNORECASE)
    if not match:
        raise ValueError("No se pudo detectar el encabezado de cuenta del auxiliar")
    return match.group(1).strip(), match.group(2).strip()


def _extract_uuid(text: str) -> Optional[str]:
    if not text:
        return None
    match = re.search(
        r"\b[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}\b",
        str(text).upper(),
    )
    return match.group(0) if match else None


def parse_aux_workbook(filename: str, contents: bytes) -> List[ParsedAuxEntry]:
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    cuenta_codigo = ""
    cuenta_nombre = ""
    entries: List[ParsedAuxEntry] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [row[i] if i < len(row) else None for i in range(9)]
        values_str = [str(v).strip() if v is not None else "" for v in values]

        if row_idx == 2 and len(values_str) > 1 and values_str[1]:
            cuenta_codigo, cuenta_nombre = _extract_account_header(values_str[1])
            continue

        if row_idx <= 2:
            continue

        tipo_poliza = values_str[1]
        numero_poliza = values_str[2]
        if not any(values_str):
            continue
        if not tipo_poliza and not numero_poliza and not values_str[4]:
            continue

        concepto = values[4]
        entries.append(
            ParsedAuxEntry(
                source_sheet=ws.title,
                source_row_number=row_idx,
                cuenta_codigo=cuenta_codigo,
                cuenta_nombre=cuenta_nombre,
                tipo_poliza=tipo_poliza or None,
                numero_poliza=numero_poliza.strip() or None,
                fecha=_parse_date(values[3]),
                concepto=str(concepto).strip() if concepto is not None else None,
                saldo_inicial=_parse_float(values[5]),
                debe=_parse_float(values[6]),
                haber=_parse_float(values[7]),
                saldo=_parse_float(values[8]),
                cfdi_uuid=_extract_uuid(str(concepto or "")),
                raw_row={f"c{i+1}": values_str[i] for i in range(len(values_str))},
            )
        )

    return entries


async def import_aux_workbook(
    session: AsyncSession,
    *,
    filename: str,
    contents: bytes,
    apply_changes: bool,
    started_by_empleado_id: Optional[Any] = None,
) -> Dict[str, Any]:
    entries = parse_aux_workbook(filename, contents)
    file_sha = hashlib.sha256(contents).hexdigest()

    run = AccountingImportRun(
        id=uuid4(),
        source_type="auxiliar",
        filename=filename,
        source_sha256=file_sha,
        mode="apply" if apply_changes else "dry_run",
        status="completed",
        started_by_empleado_id=started_by_empleado_id,
        started_at=datetime.utcnow(),
    )
    session.add(run)
    await session.flush()

    created = 0
    updated = 0
    linked_polizas = 0
    linked_cfdi = 0
    samples: List[Dict[str, Any]] = []

    cuenta_cache: Dict[str, Optional[CuentaContable]] = {}

    for parsed in entries:
        existing = (
            await session.execute(
                select(AuxLedgerEntry).where(
                    AuxLedgerEntry.source_file == filename,
                    AuxLedgerEntry.cuenta_codigo == parsed.cuenta_codigo,
                    AuxLedgerEntry.source_row_number == parsed.source_row_number,
                )
            )
        ).scalar_one_or_none()

        cuenta = cuenta_cache.get(parsed.cuenta_codigo)
        if parsed.cuenta_codigo not in cuenta_cache:
            cuenta = (
                await session.execute(
                    select(CuentaContable).where(CuentaContable.codigo == parsed.cuenta_codigo)
                )
            ).scalar_one_or_none()
            cuenta_cache[parsed.cuenta_codigo] = cuenta

        related_poliza = None
        if parsed.tipo_poliza and parsed.numero_poliza:
            related_poliza = (
                await session.execute(
                    select(AccountingPoliza).where(
                        AccountingPoliza.tipo_poliza == parsed.tipo_poliza,
                        AccountingPoliza.numero_poliza == parsed.numero_poliza,
                    )
                )
            ).scalar_one_or_none()

        if related_poliza:
            linked_polizas += 1
        if parsed.cfdi_uuid:
            linked_cfdi += 1

        payload = {
            "import_run_id": run.id,
            "source_sheet": parsed.source_sheet,
            "cuenta_nombre": parsed.cuenta_nombre,
            "cuenta_contable_id": cuenta.id if cuenta else None,
            "tipo_poliza": parsed.tipo_poliza,
            "numero_poliza": parsed.numero_poliza,
            "fecha": parsed.fecha,
            "concepto": parsed.concepto,
            "saldo_inicial": parsed.saldo_inicial,
            "debe": parsed.debe,
            "haber": parsed.haber,
            "saldo": parsed.saldo,
            "cfdi_uuid": parsed.cfdi_uuid,
            "related_poliza_id": related_poliza.id if related_poliza else None,
            "raw_row_json": parsed.raw_row,
        }

        if existing:
            updated += 1
            if apply_changes:
                for key, value in payload.items():
                    setattr(existing, key, value)
        else:
            created += 1
            if apply_changes:
                session.add(
                    AuxLedgerEntry(
                        id=uuid4(),
                        source_file=filename,
                        source_row_number=parsed.source_row_number,
                        cuenta_codigo=parsed.cuenta_codigo,
                        **payload,
                    )
                )

        if len(samples) < 10:
            samples.append(
                {
                    "row": parsed.source_row_number,
                    "tipo": parsed.tipo_poliza,
                    "numero": parsed.numero_poliza,
                    "fecha": parsed.fecha.isoformat() if parsed.fecha else None,
                    "concepto": parsed.concepto,
                    "debe": parsed.debe,
                    "haber": parsed.haber,
                    "saldo": parsed.saldo,
                }
            )

    run.finished_at = datetime.utcnow()
    run.summary_json = {
        "entries": len(entries),
        "created": created,
        "updated": updated,
        "linked_polizas": linked_polizas,
        "linked_cfdi": linked_cfdi,
    }

    if apply_changes:
        await session.commit()
    else:
        await session.rollback()

    return {
        "mode": "apply" if apply_changes else "dry_run",
        "file": filename,
        "sha256": file_sha,
        "entries": len(entries),
        "created": created,
        "updated": updated,
        "linked_polizas": linked_polizas,
        "linked_cfdi": linked_cfdi,
        "samples": samples,
    }
