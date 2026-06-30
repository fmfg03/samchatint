"""
COI XLSX import service.

Parses COI-style workbook blocks:
- header row: Eg / number / description / line count
- accounting lines
- optional CFDI block:
  - INICIO_CFDI
  - detail row
  - FIN_CFDI
- FIN_PARTIDAS
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import uuid4

from openpyxl import load_workbook

logger = logging.getLogger(__name__)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .cfdi_expense_link_service import (
    bulk_link_pending_expenses_to_cfdi_reports,
    find_cfdi_report_by_fiscal_uuid,
    normalize_cfdi_uuid_to_canonical,
)
from ..models import (
    AccountingImportRun,
    AccountingPoliza,
    AccountingPolizaLine,
    CFDIReport,
    CuentaContable,
)


@dataclass
class ParsedCFDIBlock:
    fecha: Optional[datetime] = None
    serie: Optional[str] = None
    folio: Optional[str] = None
    emisor_rfc: Optional[str] = None
    receptor_rfc: Optional[str] = None
    total: Optional[float] = None
    cfdi_uuid: Optional[str] = None


@dataclass
class ParsedPolizaLine:
    line_no: int
    cuenta_codigo: str
    concepto: str
    movimiento_no: Optional[str]
    debe: Optional[float]
    haber: Optional[float]
    raw_row: Dict[str, Any]


@dataclass
class ParsedPoliza:
    source_sheet: str
    source_row_start: int
    tipo_poliza: str
    numero_poliza: str
    descripcion_raw: str
    line_count_declared: Optional[int]
    fecha_poliza: Optional[datetime] = None
    beneficiario_nombre: Optional[str] = None
    concepto_resumen: Optional[str] = None
    cfdi: Optional[ParsedCFDIBlock] = None
    lines: List[ParsedPolizaLine] = field(default_factory=list)


def _parse_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_coi_date(value: Any) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _extract_beneficiario_y_concepto(description: str) -> tuple[Optional[str], str]:
    text = str(description or "").strip()
    parts = [part.strip() for part in text.split(" / ") if part.strip()]
    beneficiario = parts[2] if len(parts) >= 3 else None
    trailing = parts[3] if len(parts) >= 4 else text
    trailing = re.sub(r"^\d+\.-\s*", "", trailing).strip()
    return beneficiario, trailing or text


def parse_coi_workbook(filename: str, contents: bytes) -> List[ParsedPoliza]:
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    polizas: List[ParsedPoliza] = []
    current: Optional[ParsedPoliza] = None
    pending_cfdi = False
    line_no = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [str(cell).strip() if cell is not None else "" for cell in row[:9]]
        if not any(values):
            continue

        first = values[0]
        second = values[1]
        third = values[2]

        if first == "Eg":
            if current is not None:
                polizas.append(current)
            declared = int(values[3]) if values[3].isdigit() else None
            beneficiario, resumen = _extract_beneficiario_y_concepto(values[2])
            current = ParsedPoliza(
                source_sheet=ws.title,
                source_row_start=row_idx,
                tipo_poliza=first,
                numero_poliza=second,
                descripcion_raw=values[2],
                line_count_declared=declared,
                beneficiario_nombre=beneficiario,
                concepto_resumen=resumen,
            )
            pending_cfdi = False
            line_no = 0
            continue

        if current is None:
            continue

        if third == "INICIO_CFDI":
            current.cfdi = ParsedCFDIBlock()
            pending_cfdi = True
            continue

        if third == "FIN_CFDI":
            pending_cfdi = False
            continue

        if second == "FIN_PARTIDAS":
            polizas.append(current)
            current = None
            pending_cfdi = False
            line_no = 0
            continue

        if pending_cfdi and current.cfdi is not None:
            current.cfdi.fecha = _parse_coi_date(values[2])
            current.cfdi.serie = values[3] or None
            current.cfdi.folio = values[4] or None
            current.cfdi.emisor_rfc = values[5] or None
            current.cfdi.receptor_rfc = values[6] or None
            current.cfdi.total = _parse_float(values[7])
            current.cfdi.cfdi_uuid = values[8] or None
            if current.cfdi.fecha and not current.fecha_poliza:
                current.fecha_poliza = current.cfdi.fecha
            continue

        if second and second != "FIN_PARTIDAS":
            line_no += 1
            current.lines.append(
                ParsedPolizaLine(
                    line_no=line_no,
                    cuenta_codigo=second,
                    concepto=values[3] or current.concepto_resumen or current.descripcion_raw,
                    movimiento_no=values[4] or None,
                    debe=_parse_float(values[5]),
                    haber=_parse_float(values[6]),
                    raw_row={f"c{i+1}": values[i] for i in range(len(values))},
                )
            )

    if current is not None:
        polizas.append(current)

    return polizas


async def import_coi_workbook(
    session: AsyncSession,
    *,
    filename: str,
    contents: bytes,
    apply_changes: bool,
    started_by_empleado_id: Optional[Any] = None,
) -> Dict[str, Any]:
    polizas = parse_coi_workbook(filename, contents)
    file_sha = hashlib.sha256(contents).hexdigest()

    run = AccountingImportRun(
        id=uuid4(),
        source_type="coi",
        filename=filename,
        source_sha256=file_sha,
        mode="apply" if apply_changes else "dry_run",
        status="completed",
        started_by_empleado_id=started_by_empleado_id,
        started_at=datetime.utcnow(),
    )
    session.add(run)
    await session.flush()

    created = 0
    updated = 0
    lines_created = 0
    cfdi_created = 0
    cfdi_reused = 0
    samples: List[Dict[str, Any]] = []

    for parsed in polizas:
        existing = (
            await session.execute(
                select(AccountingPoliza).where(
                    AccountingPoliza.source_file == filename,
                    AccountingPoliza.tipo_poliza == parsed.tipo_poliza,
                    AccountingPoliza.numero_poliza == parsed.numero_poliza,
                )
            )
        ).scalar_one_or_none()

        cfdi_report_id = None
        cfdi_uuid = parsed.cfdi.cfdi_uuid if parsed.cfdi else None
        canon_uid: Optional[str] = None
        if cfdi_uuid:
            try:
                canon_uid = normalize_cfdi_uuid_to_canonical(cfdi_uuid)
            except ValueError:
                canon_uid = None
        if canon_uid:
            existing_cfdi = await find_cfdi_report_by_fiscal_uuid(session, canon_uid)
            if existing_cfdi:
                cfdi_reused += 1
                if apply_changes:
                    if parsed.cfdi.fecha:
                        existing_cfdi.fecha = parsed.cfdi.fecha
                    existing_cfdi.folio = parsed.cfdi.folio
                    existing_cfdi.serie = parsed.cfdi.serie
                    existing_cfdi.emisor_rfc = parsed.cfdi.emisor_rfc
                    existing_cfdi.receptor_rfc = parsed.cfdi.receptor_rfc
                    existing_cfdi.total = parsed.cfdi.total
                    existing_cfdi.descripcion_concepto_principal = parsed.concepto_resumen
                    existing_cfdi.origen = "coi_xlsx"
                    if existing_cfdi.cfdi_uuid != canon_uid:
                        existing_cfdi.cfdi_uuid = canon_uid
                cfdi_report_id = existing_cfdi.id
            else:
                cfdi_created += 1
                if apply_changes:
                    new_cfdi = CFDIReport(
                        id=uuid4(),
                        fecha=parsed.cfdi.fecha,
                        serie=parsed.cfdi.serie,
                        folio=parsed.cfdi.folio,
                        emisor_rfc=parsed.cfdi.emisor_rfc,
                        receptor_rfc=parsed.cfdi.receptor_rfc,
                        total=parsed.cfdi.total,
                        cfdi_uuid=canon_uid,
                        descripcion_concepto_principal=parsed.concepto_resumen,
                        origen="coi_xlsx",
                        xml_parsed=False,
                    )
                    session.add(new_cfdi)
                    await session.flush()
                    cfdi_report_id = new_cfdi.id

        if existing:
            updated += 1
            if apply_changes:
                existing.import_run_id = run.id
                existing.source_sheet = parsed.source_sheet
                existing.source_row_start = parsed.source_row_start
                existing.fecha_poliza = parsed.fecha_poliza
                existing.beneficiario_nombre = parsed.beneficiario_nombre
                existing.concepto = parsed.descripcion_raw
                existing.concepto_resumen = parsed.concepto_resumen
                existing.line_count_declared = parsed.line_count_declared
                existing.line_count_actual = len(parsed.lines)
                existing.cfdi_uuid = cfdi_uuid
                existing.cfdi_report_id = cfdi_report_id
                existing.origen = "coi_xlsx"
                existing.updated_at = datetime.utcnow()
                # Rebuild lines to keep idempotent import behavior.
                for old_line in list(existing.lines):
                    await session.delete(old_line)
                await session.flush()
                target = existing
            else:
                target = existing
        else:
            created += 1
            if apply_changes:
                target = AccountingPoliza(
                    id=uuid4(),
                    import_run_id=run.id,
                    source_file=filename,
                    source_sheet=parsed.source_sheet,
                    source_row_start=parsed.source_row_start,
                    tipo_poliza=parsed.tipo_poliza,
                    numero_poliza=parsed.numero_poliza,
                    fecha_poliza=parsed.fecha_poliza,
                    beneficiario_nombre=parsed.beneficiario_nombre,
                    concepto=parsed.descripcion_raw,
                    concepto_resumen=parsed.concepto_resumen,
                    line_count_declared=parsed.line_count_declared,
                    line_count_actual=len(parsed.lines),
                    cfdi_uuid=cfdi_uuid,
                    cfdi_report_id=cfdi_report_id,
                    origen="coi_xlsx",
                )
                session.add(target)
                await session.flush()
            else:
                target = None

        if apply_changes and target is not None:
            for parsed_line in parsed.lines:
                cuenta = (
                    await session.execute(
                        select(CuentaContable).where(CuentaContable.codigo == parsed_line.cuenta_codigo)
                    )
                ).scalar_one_or_none()
                session.add(
                    AccountingPolizaLine(
                        id=uuid4(),
                        poliza_id=target.id,
                        line_no=parsed_line.line_no,
                        cuenta_codigo=parsed_line.cuenta_codigo,
                        cuenta_contable_id=cuenta.id if cuenta else None,
                        concepto=parsed_line.concepto,
                        movimiento_no=parsed_line.movimiento_no,
                        debe=parsed_line.debe,
                        haber=parsed_line.haber,
                        raw_row_json=parsed_line.raw_row,
                    )
                )
                lines_created += 1

        if len(samples) < 10:
            samples.append(
                {
                    "poliza": f"{parsed.tipo_poliza}-{parsed.numero_poliza}",
                    "beneficiario": parsed.beneficiario_nombre,
                    "concepto": parsed.concepto_resumen,
                    "cfdi_uuid": cfdi_uuid,
                    "lines": len(parsed.lines),
                }
            )

    run.finished_at = datetime.utcnow()
    run.summary_json = {
        "polizas": len(polizas),
        "created": created,
        "updated": updated,
        "lines_created": lines_created if apply_changes else sum(len(p.lines) for p in polizas),
        "cfdi_created": cfdi_created,
        "cfdi_reused": cfdi_reused,
    }

    if apply_changes:
        await session.commit()
        try:
            linked = await bulk_link_pending_expenses_to_cfdi_reports(session)
            if linked > 0:
                await session.commit()
        except Exception as exc:
            logger.warning("COI import: post-commit expense–CFDI link failed: %s", exc, exc_info=True)
    else:
        await session.rollback()

    return {
        "mode": "apply" if apply_changes else "dry_run",
        "file": filename,
        "sha256": file_sha,
        "polizas": len(polizas),
        "created": created,
        "updated": updated,
        "lines": lines_created if apply_changes else sum(len(p.lines) for p in polizas),
        "cfdi_created": cfdi_created,
        "cfdi_reused": cfdi_reused,
        "samples": samples,
    }
